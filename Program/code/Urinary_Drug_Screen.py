from openpyxl import load_workbook
import math
from openpyxl.utils.dataframe import dataframe_to_rows
from log_writer import log_writer
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings


def urinary_drug_screen(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Urinary drug screen
    '''

    df= df_root[df_root['name']== 'Urinary Drug Screen']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})
    # print(df_visit_done)
    # print('------------------')

    warnings.filterwarnings('ignore')

    lista_revision = []
    lista_logs = ['Urinary drug screen']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()

            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')

            if visita != 'Unscheduled Visits':
                pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
                pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                if visita == 'Unscheduled Visits':
                    was_DV_performed_pure = 1.0
                    date_of_visit = ''
                else:
                    was_DV_performed = row['was_DV_performed']
                    was_DV_performed_pure = was_DV_performed.split('|')[0]
                    was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                    date_of_visit = row['Date_of_visit']
                
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                
                if status != '':
                    try:
                        was_urine_test_performed = row['Was the urine test performed for drug screening?']
                        was_urine_test_performed_pure = was_urine_test_performed.split('|')[0]
                        was_urine_test_performed_form_field_isntance = was_urine_test_performed.split('|')[1]
                        was_urine_test_performed_disname = was_urine_test_performed.split('|')[2]
                    except Exception as e:
                        was_urine_test_performed_pure = math.nan
                        was_urine_test_performed_form_field_isntance = 'This field does not have any data'
                        was_urine_test_performed_disname = 'Empty'

                    try:
                        provide_reason = row['Please provide the reason']
                        provide_reason_pure = provide_reason.split('|')[0]
                        provide_reason_form_field_isntance = provide_reason.split('|')[1]
                    except Exception as e:
                        provide_reason_pure = ''
                        provide_reason_form_field_isntance = 'This field does not have any data'
                    
                    try:
                        date_of_test = row['Date of test performed']
                        date_of_test_pure = date_of_test.split('|')[0]
                        date_of_test_form_field_isntance = date_of_test.split('|')[1]
                    except Exception as e:
                        date_of_test_pure = ''
                        date_of_test_form_field_isntance = 'This field does not have any data'
                    
                    try:
                        check_below_trace = row['Check below trace/positive results']
                        check_below_trace_pure = check_below_trace.split('|')[0]
                        check_below_trace_form_field_isntance = check_below_trace.split('|')[1]
                        check_below_trace_disname = check_below_trace.split('|')[2]
                    except Exception as e:
                        check_below_trace_pure = ''
                        check_below_trace_form_field_isntance = 'This field does not have any data'
                        check_below_trace_disname = 'Empty'
                    
                    #----------------------------------------------------------------------------------------------

                    # Revision GE0070
                    if str(was_DV_performed_pure) !=  'unsch':
                        if float(was_DV_performed_pure) !=  1.0 :
                            error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                            lista_revision.append(error)

                    if date_of_test_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_of_test_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of examination performed', date_of_test_form_field_isntance ,f , date_of_test_pure, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UD0020
                    try:
                        if  float(was_urine_test_performed_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the urine test performed for drug screening?', \
                                         was_urine_test_performed_form_field_isntance, \
                                            'The "Not Required" option can only be selected if visit is D-1 and Screening visit date = D-1 date (screening done on D-1)', \
                                                was_urine_test_performed_disname, 'UD0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision UD0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UD0030
                    if check_below_trace_pure != '':
                        try:
                            lis_check_below = check_below_trace_pure.split(',')
 
                            if len(lis_check_below) > 1:
                                for i in lis_check_below:
                                    if float(i) == 0.0:
                                        error = [subject, visit, 'Check below trace/positive results ', check_below_trace_form_field_isntance, \
                                                'When the "None" option is selected, no other option can be selected, please review', \
                                                    check_below_trace_disname, 'UD0030']
                                        lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision UD0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UD0040
                    if str(was_DV_performed_pure) !=  'unsch':
                        if date_of_test_pure != '' and date_of_visit!='':
                            try:
                                date_format = '%d-%b-%Y'
                                date_of_test_f = datetime.strptime(date_of_test_pure, date_format)
                                date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                                if date_of_test_f != date_of_visit_f:
                                    error = [subject, visit, 'Date of test performed', date_of_test_form_field_isntance,\
                                            'The date must be the same as the date of visit date', \
                                                f'{date_of_test_pure} - {date_of_visit}', 'UD0040']
                                    lista_revision.append(error)
                                else:
                                    pass
                            except Exception as e:
                                lista_logs.append(f'Revision UD0040--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UD0050
                    if date_of_test_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_of_test_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of test performed', date_of_test_form_field_isntance, \
                                        'The date of test performed can not be before the informed consent date', \
                                            f'{date_of_test_pure} - {date_inform_consent}', 'UD0050']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision UD0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> UD0060
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_of_test_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_of_test_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of test performed', date_of_test_form_field_isntance ,'Date of test performed must be before the End of study/Early withdrawal date. ', date_of_test_pure, 'UD0060']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision UD0060 --> {e} - Subject: {subject},  Visit: {visit}  ')

        


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    urinary_drug_screen_output = pd.DataFrame(lista_revision, columns=column_names)
    urinary_drug_screen_output = urinary_drug_screen_output[~urinary_drug_screen_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]

    
    sheet = excel_writer.create_sheet("Urinary Drug Screen")

    for row in dataframe_to_rows(urinary_drug_screen_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return urinary_drug_screen_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    urinary_drug_screen(df_root, path_excel) 