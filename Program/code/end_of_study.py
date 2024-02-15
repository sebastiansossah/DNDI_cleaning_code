from datetime import datetime
import math

import numpy as np
from revision_fechas import revision_fecha
from log_writer import log_writer 
import warnings
import pandas as pd
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')


def end_of_study(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de End of Study Treatment
    '''

    df= df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'Variable' ,'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)
    df['Campo_Variable'] = df['Campo'].astype(str) + '|' + df['Variable'].astype(str)

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Inform_consent_date'})

    df_adverse_event_marked_yes = df_root[df_root['name']== 'Adverse Events']
    df_adverse_event_marked_yes = df_adverse_event_marked_yes[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_adverse_event_marked_yes = df_adverse_event_marked_yes[df_adverse_event_marked_yes['Variable'] == 'AESTDISYN']
    df_adverse_event_marked_yes = df_adverse_event_marked_yes[df_adverse_event_marked_yes['Valor'].astype('float') == 1.0]
    df_adverse_event_marked_yes = df_adverse_event_marked_yes[['Participante', 'Valor']]
    df_adverse_event_marked_yes = df_adverse_event_marked_yes.rename(columns={'Participante':'Subject', 'Valor':'caused_study_discontinuation'})

    df_adverse_event = df_root[df_root['name']== 'Adverse Events']
    df_adverse_event = df_adverse_event[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_adverse_event = df_adverse_event[df_adverse_event['Variable'] == 'AENO']
    df_adverse_event = df_adverse_event[['Participante', 'Valor']]
    df_adverse_event = df_adverse_event.rename(columns={'Participante':'Subject', 'Valor':'adverse_event_id'})

    df_miltefosine_administration = df_root[df_root['name']== 'Miltefosine Administration']
    df_miltefosine_administration = df_miltefosine_administration[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_miltefosine_administration = df_miltefosine_administration[df_miltefosine_administration['Variable'] == 'ECMLTCINTD']
    df_miltefosine_administration = df_miltefosine_administration[df_miltefosine_administration['Valor'].astype('float') == 3.0]
    df_miltefosine_administration = df_miltefosine_administration[['Participante', 'Valor']]
    df_miltefosine_administration = df_miltefosine_administration.rename(columns={'Participante':'Subject', 'Valor':'miltefosine_permanently_discontinued'})

    df_CPG_administration = df_root[df_root['name']== 'CpG ODN D35 Administration']
    df_CPG_administration = df_CPG_administration[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_CPG_administration = df_CPG_administration[df_CPG_administration['Variable'] == 'ECCPGCINTD']
    df_CPG_administration = df_CPG_administration[df_CPG_administration['Valor'].astype('float') == 3.0]
    df_CPG_administration = df_CPG_administration[['Participante', 'Valor']]
    df_CPG_administration = df_CPG_administration.rename(columns={'Participante':'Subject', 'Valor':'cpg_permanently_discontinued'})

    df_CPG_administration_max_date = df_root[df_root['name']== 'CpG ODN D35 Administration']
    df_CPG_administration_max_date = df_CPG_administration_max_date[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_CPG_administration_max_date = df_CPG_administration_max_date[df_CPG_administration_max_date['Variable'] == 'ECCPGDAT']
    df_CPG_administration_max_date['max_value_cpg_date'] = pd.to_datetime(df_CPG_administration_max_date['Valor'], format='%d-%b-%Y').max()
    df_CPG_administration_max_date = df_CPG_administration_max_date[['Participante', 'max_value_cpg_date']].drop_duplicates()
    df_CPG_administration_max_date = df_CPG_administration_max_date.rename(columns={'Participante':'Subject'})


    df_CPG_administration_max_date_visit = df_root[df_root['name']== 'CpG ODN D35 Administration']
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit[df_CPG_administration_max_date_visit['Variable'] == 'ECCPGDAT']
    df_CPG_administration_max_date_extract = df_CPG_administration_max_date_visit[df_CPG_administration_max_date_visit['Campo'] == 'Date of dosing']
    df_CPG_administration_max_date_visit['max_date'] = pd.to_datetime(df_CPG_administration_max_date_visit['Valor'], format='%d-%b-%Y').max()
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit[['Participante', 'max_date']].drop_duplicates()
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit.rename(columns={'max_date':'Valor'})
    df_visit_date = df_root[df_root['name']== 'Date of visit']
    df_visit_date = df_visit_date[['Participante', 'Campo', 'Valor', 'Visit' ]]
    df_visit_date = df_visit_date[df_visit_date['Campo'] == 'Visit Date']
    df_visit_date['Valor']  = pd.to_datetime(df_visit_date['Valor'], format='%d-%b-%Y')
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit.merge(df_visit_date, on=['Participante', 'Valor'])
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit[['Participante', 'Visit']]
    df_CPG_administration_max_date_visit = df_CPG_administration_max_date_visit.rename(columns={'Participante':'Subject', 'Visit':'Visita_maxima_fecha_cpg'})

    lista_revision = []
    lista_logs = ['End of Study Treatment']
    

    for sujeto in lista_sujetos:
        pru_1= df[df['Participante']==sujeto]
        pru = pru_1
        pru = pru[['Campo_Variable', 'Value_id']].T
        new_columns = pru.iloc[0]
        pru = pru[1:].set_axis(new_columns, axis=1)
        pru['Subject'] = sujeto
        pru['Visit'] = 'unscheduled'
        pru['status'] = pru_1['activityState'].unique()
        pru = pru.merge(df_informed, on=['Subject'], how='left')
        pru = pru.merge(df_adverse_event_marked_yes, on=['Subject'], how='left')
        pru = pru.merge(df_adverse_event, on=['Subject'], how='left')
        pru = pru.merge(df_miltefosine_administration, on=['Subject'], how='left')
        pru = pru.merge(df_CPG_administration, on=['Subject'], how='left')
        pru = pru.merge(df_CPG_administration_max_date, on=['Subject'], how='left')
        pru = pru.merge(df_CPG_administration_max_date_visit, on=['Subject'], how='left')
        
        for index, row in pru.iterrows():
            status = row['status']
            subject = row['Subject']
            visit = row['Visit']

            inform_consent_date = row['Inform_consent_date']
            caused_study_discontinuation = row['caused_study_discontinuation']
            adverse_event_id = row['adverse_event_id']
            miltefosine_permanently_discontinued = row['miltefosine_permanently_discontinued']
            cpg_permanently_discontinued = row['cpg_permanently_discontinued']
            max_value_cpg_date = row['max_value_cpg_date']
            Visita_maxima_fecha_cpg = row['Visita_maxima_fecha_cpg']
 
            
            columnas_dataset = pru.columns

            try:
                date_last_treatment_administration_miltefosine = row[next(filter(lambda x: x.split('|')[1] == 'DSMILDTC', columnas_dataset))]
                date_last_treatment_administration_miltefosine_pure = date_last_treatment_administration_miltefosine.split('|')[0]
                date_last_treatment_administration_miltefosine_form_field_instance = date_last_treatment_administration_miltefosine.split('|')[1]
            except:
                date_last_treatment_administration_miltefosine_pure = ''
                date_last_treatment_administration_miltefosine_form_field_instance = 'This field does not have any data'

            try:
                was_study_treatment_complited_protocol_miltefosine = row[next(filter(lambda x: x.split('|')[1] == 'DSMILCMP', columnas_dataset))]          
                was_study_treatment_complited_protocol_miltefosine_pure = was_study_treatment_complited_protocol_miltefosine.split('|')[0]
                was_study_treatment_complited_protocol_miltefosine_form_field_instance = was_study_treatment_complited_protocol_miltefosine.split('|')[1]
            except:
                was_study_treatment_complited_protocol_miltefosine_pure = math.nan
                was_study_treatment_complited_protocol_miltefosine_form_field_instance = 'This field does not have any data'

            try:
                primary_reason_not_completing_miltefosine = row[next(filter(lambda x: x.split('|')[1] == 'DSMILCAT', columnas_dataset))]
                primary_reason_not_completing_miltefosine_pure = primary_reason_not_completing_miltefosine.split('|')[0]
                primary_reason_not_completing_miltefosine_form_field_instance = primary_reason_not_completing_miltefosine.split('|')[1]
            except:
                primary_reason_not_completing_miltefosine_pure = math.nan
                primary_reason_not_completing_miltefosine_form_field_instance = 'This field does not have any data'
            
            # --------------------------------------------------------------------------------------
            try:
                date_last_treatment_administration_CPG = row[next(filter(lambda x: x.split('|')[1] == 'DSCPGDTC', columnas_dataset))]
                date_last_treatment_administration_CPG_pure = date_last_treatment_administration_CPG.split('|')[0]
                date_last_treatment_administration_CPG_form_field_instance = date_last_treatment_administration_CPG.split('|')[1]
            except:
                date_last_treatment_administration_CPG_pure = ''
                date_last_treatment_administration_CPG_form_field_instance = 'This field does not have any data'
            
            try:
                was_study_treatment_complited_protocol_CPG = row[next(filter(lambda x: x.split('|')[1] == 'DSCPGCMP', columnas_dataset))]
                was_study_treatment_complited_protocol_CPG_pure = was_study_treatment_complited_protocol_CPG.split('|')[0]
                was_study_treatment_complited_protocol_CPG_form_field_instance = was_study_treatment_complited_protocol_CPG.split('|')[1]
            except:
                was_study_treatment_complited_protocol_CPG_pure = math.nan
                was_study_treatment_complited_protocol_CPG_form_field_instance = 'This field does not have any data'
            
            try:
                primary_reason_not_completing_CPG =  row[next(filter(lambda x: x.split('|')[1] == 'DSCPGCAT', columnas_dataset))]
                primary_reason_not_completing_CPG_pure = primary_reason_not_completing_CPG.split('|')[0]
                primary_reason_not_completing_CPG_form_field_instance = primary_reason_not_completing_CPG.split('|')[1]
            except:
                primary_reason_not_completing_CPG_pure = math.nan
                primary_reason_not_completing_CPG_form_field_instance = 'This field does not have any data'
            # --------------------------------------------------------------------------------------
            try:
                end_of_study_date_general = row[next(filter(lambda x: x.split('|')[1] == 'DSDAT', columnas_dataset))]
                end_of_study_date_general_pure = end_of_study_date_general.split('|')[0]
                end_of_study_date_general_form_field_instance = end_of_study_date_general.split('|')[1]
            except:
                end_of_study_date_general_pure = ''
                end_of_study_date_general_form_field_instance = 'This field does not have any data'
            
            try:
                did_participant_complete_study = row[next(filter(lambda x: x.split('|')[1] == 'DSSTCMP', columnas_dataset))]
                did_participant_complete_study_pure = did_participant_complete_study.split('|')[0]
                did_participant_complete_study_form_field_instance = did_participant_complete_study.split('|')[1]
            except:
                did_participant_complete_study_pure = math.nan
                did_participant_complete_study_form_field_instance = 'This field does not have any data'
            
            try:
                primary_reason_not_completing_general = row[next(filter(lambda x: x.split('|')[1] == 'DSCAT', columnas_dataset))]
                primary_reason_not_completing_general_pure = primary_reason_not_completing_general.split('|')[0]
                primary_reason_not_completing_general_form_field_instance = primary_reason_not_completing_general.split('|')[1]
            except:
                primary_reason_not_completing_general_pure = math.nan
                primary_reason_not_completing_general_form_field_instance = 'This field does not have any data'

            # --------------------------------------------------------------------------------------------------------------------------------------------------
            if date_last_treatment_administration_miltefosine_pure == '':
                pass
            else:
                try:
                    # Primera  revision general de formato de fecha ->GE0020
                    f = revision_fecha(date_last_treatment_administration_miltefosine_pure)
                    if f == None:
                        pass
                    else:
                        error = [subject, visit, 'Date of last study treatment administration (Miltefosine)', \
                                    date_last_treatment_administration_miltefosine_form_field_instance,\
                                            f , date_last_treatment_administration_miltefosine_pure, 'GE0020']
                        lista_revision.append(error)     
                except Exception as e:
                    lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
            
            if date_last_treatment_administration_CPG_pure == '':
                pass
            else:
                try:
                    # Primera  revision general de formato de fecha ->GE0020
                    f = revision_fecha(date_last_treatment_administration_CPG_pure)
                    if f == None:
                        pass
                    else:
                        error = [subject, visit, 'Date of last study treatment administration (CpG ODN D35)', \
                                    date_last_treatment_administration_CPG_form_field_instance,\
                                            f , date_last_treatment_administration_CPG_pure, 'GE0020']
                        lista_revision.append(error)     
                except Exception as e:
                    lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

            if end_of_study_date_general_pure == '':
                pass
            else:

                try:
                    # Primera  revision general de formato de fecha ->GE0020
                    f = revision_fecha(end_of_study_date_general_pure)
                    if f == None:
                        pass
                    else:
                        error = [subject, visit, 'End of study/Early withdrawal date', \
                                    end_of_study_date_general_form_field_instance,\
                                            f , end_of_study_date_general_pure, 'GE0020']
                        lista_revision.append(error)     
                except Exception as e:
                    lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0030
            try:
                if datetime.strptime(str(end_of_study_date_general_pure), '%d-%b-%Y') < datetime.strptime(str(inform_consent_date), '%d-%b-%Y'):
                    error = [subject, visit, 'End of study/Early withdrawal date', end_of_study_date_general_form_field_instance, \
                            'End of study/Early withdrawal date can not be before the informed consent date', \
                                end_of_study_date_general_pure, 'ES0030']
                    lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0030 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0040
            try:
                if float(primary_reason_not_completing_general_pure) == 2.0:
                    if math.isnan(float(caused_study_discontinuation)):
                        error = [subject, visit, 'Primary reason for not completing the study', primary_reason_not_completing_general_form_field_instance, \
                            'The Primary reason for not completing the study is SAE or intolerable AE, that prevents the participant from continuing, but there are no  adverse events reported with Caused study discontinuation Yes.', \
                            primary_reason_not_completing_general_pure, 'ES0040']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0040 --> {e} - Subject: {subject},  Visit: {visit} ')
            
            # Revision ES0050
            try:
                if float(primary_reason_not_completing_general_pure) != 2.0:
                    if math.isnan(float(caused_study_discontinuation)):
                        pass
                    else:
                        error = [subject, visit, 'Primary reason for not completing the study', primary_reason_not_completing_general_form_field_instance, \
                            'The Primary reason for not completing the study is  not SAE or intolerable AE, that prevents the participant from continuing, therefore there should not be any adverse events reported with Caused study discontinuation Yes.', \
                            primary_reason_not_completing_general_pure, 'ES0050']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0050 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0090
            try:
                if float(primary_reason_not_completing_miltefosine_pure) == 2.0:
                     if math.isnan(float(adverse_event_id)):
                        error = [subject, visit, 'Primary reason for not completing the study treatment (miltefosine)', primary_reason_not_completing_miltefosine_form_field_instance, \
                            'The Primary reason for not completing the study is  not SAE or intolerable AE, that prevents the participant from continuing, therefore there should not be any adverse events reported with Caused study discontinuation Yes.', \
                            primary_reason_not_completing_miltefosine_pure, 'ES0090']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0090 --> {e} - Subject: {subject},  Visit: {visit} ')
            
            # Revision ES0120
            try:
                if float(primary_reason_not_completing_miltefosine_pure) == 2.0:
                    if math.isnan(float(miltefosine_permanently_discontinued)):
                        error = [subject, visit, 'Primary reason for not completing the study treatment (miltefosine)', primary_reason_not_completing_miltefosine_form_field_instance, \
                            'The Primary reason for not completing the study is  not SAE or intolerable AE, that prevents the participant from continuing, therefore there should not be any adverse events reported with Caused study discontinuation Yes.', \
                            primary_reason_not_completing_miltefosine_pure, 'ES0120']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0120 --> {e} - Subject: {subject},  Visit: {visit} ')
            

            # Revision ES0150
            try:
                if datetime.strptime(str(date_last_treatment_administration_CPG_pure), '%d-%b-%Y') != datetime.strptime(str(max_value_cpg_date), '%d-%b-%Y'):
                    error = [subject, visit, 'Date of last study treatment administration (CpG ODN D35)', \
                             date_last_treatment_administration_CPG_form_field_instance, \
                            'The date must be equal to the greatest date in the CpG ODN D35 administration forms where the dose is different from 0', \
                                f"{max_value_cpg_date} - {date_last_treatment_administration_CPG_pure}", 'ES0150']
                    lista_revision.append(error)
            except Exception as e:

                lista_logs.append(f'Revision ES0150 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0160
            try:
                if float(primary_reason_not_completing_CPG_pure) == 2.0:
                    if math.isnan(float(adverse_event_id)):
                        error = [subject, visit, 'Primary reason for not completing the study treatment (CpG ODN D35)', primary_reason_not_completing_CPG_form_field_instance, \
                            'There is no AE form completed, please verify', \
                            primary_reason_not_completing_CPG_pure, 'ES0160']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0160 --> {e} - Subject: {subject},  Visit: {visit} ')
            
            # Revision ES0170
            try:
                if float(was_study_treatment_complited_protocol_CPG_pure) == 1.0:
                    if Visita_maxima_fecha_cpg != 'D29':
                        error = [subject, visit, 'Was the study treatment completed per protocol? (CpG ODN D35)', was_study_treatment_complited_protocol_CPG_form_field_instance, \
                            'If "Was the study treatment completed per protocol? (CpG ODN D35)" is Yes , the last CpG ODN D35 study treatment administration date with dose different from 0 should be on D29. ', \
                            was_study_treatment_complited_protocol_CPG_pure, 'ES0170']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0170 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0180
            try:
                if float(was_study_treatment_complited_protocol_CPG_pure) == 0.0:
                    if Visita_maxima_fecha_cpg == 'D29':
                        error = [subject, visit, 'Was the study treatment completed per protocol? (CpG ODN D35)', was_study_treatment_complited_protocol_CPG_form_field_instance, \
                            'If "Was the study treatment completed per protocol? (CpG ODN D35)" is Yes , the last CpG ODN D35 study treatment administration date with dose different from 0 should be on D29. ', \
                            was_study_treatment_complited_protocol_CPG_pure, 'ES0180']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0180 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0190
            try:
                if float(primary_reason_not_completing_CPG_pure) == 2.0:
                    if math.isnan(float(cpg_permanently_discontinued)):
                        error = [subject, visit, 'Primary reason for not completing the study treatment (CpG ODN D35)', primary_reason_not_completing_CPG_form_field_instance, \
                            'If the "Primary reason for not completing the study treatment (CpG ODN D35)" is "SAE or intolerable AE", there should be at least one CpG ODN D35 study treatment record with Dosing Event equals to Permanently discontinued with Reason for dose adjustment reported Adverse event.', \
                            primary_reason_not_completing_CPG_pure, 'ES0190']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0190 --> {e} - Subject: {subject},  Visit: {visit} ')
            
            # Revision ES0210
            try:
                if float(was_study_treatment_complited_protocol_CPG_pure) ==  1.0:
                    if math.isnan(float(cpg_permanently_discontinued)):
                        pass
                    else:
                        error = [subject, visit, 'Was the study treatment completed per protocol? (CpG ODN D35)', was_study_treatment_complited_protocol_CPG_form_field_instance, \
                            'According to the CpG ODN D35 administration form, where there is a dosing event selected as "Permanently discontinued". the subject can not have completed the treatment per protocol', \
                            was_study_treatment_complited_protocol_CPG_pure, 'ES0210']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0210 --> {e} - Subject: {subject},  Visit: {visit} ')
            
            # Revision ES0220
            try:
                if float(cpg_permanently_discontinued) == 3.0:
                    if float(primary_reason_not_completing_CPG_pure) != 2.0:
                        error = [subject, visit, 'Primary reason for not completing the study treatment (CpG ODN D35)', primary_reason_not_completing_CPG_form_field_instance, \
                            'If on the CpG ODN D35 administration form any of the Dosing Events is Permanently discontinued and the Reason for dose adjustment is Adverse event The Primary reason for not completing the study treatment (CpG ODN D35) should be "SAE or intolerable AE, that prevents the participant from continuing."', \
                            primary_reason_not_completing_CPG_pure, 'ES0220']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0220 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0230
            try:
                if float(was_study_treatment_complited_protocol_miltefosine_pure) == 1.0:
                    if math.isnan(float(miltefosine_permanently_discontinued)):
                        pass
                    else:
                        error = [subject, visit, 'Was the study treatment completed per protocol? (Miltefosine)', was_study_treatment_complited_protocol_miltefosine_form_field_instance, \
                            'According to the Miltefosine administration form, where there is a dosing event selected as "Permanently discontinued". the subject can not have completed the treatment per protocol', \
                            was_study_treatment_complited_protocol_miltefosine_pure, 'ES0230']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0230 --> {e} - Subject: {subject},  Visit: {visit} ')

            # Revision ES0240
            try:
                if float(miltefosine_permanently_discontinued) == 3.0:
                    if float(primary_reason_not_completing_miltefosine_pure) != 2.0:
                        error = [subject, visit, 'Primary reason for not completing the study treatment (Miltefosine)', primary_reason_not_completing_miltefosine_form_field_instance, \
                            'If on the Miltefosine administration form any of the Dosing Events is Permanently discontinued and the Reason for dose adjustment is Adverse event The Primary reason for not completing the study treatment (Miltefosine) should be "SAE or intolerable AE, that prevents the participant from continuing."', \
                            primary_reason_not_completing_miltefosine_pure, 'ES0240']
                        lista_revision.append(error)
            except Exception as e:
                lista_logs.append(f'Revision ES0240 --> {e} - Subject: {subject},  Visit: {visit} ')



    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    end_of_study_output = pd.DataFrame(lista_revision, columns=column_names)

 
    sheet = excel_writer.create_sheet('End of Study Treatment')

    for row in dataframe_to_rows(end_of_study_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return end_of_study_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx")
    end_of_study(df_root, path_excel ) 

