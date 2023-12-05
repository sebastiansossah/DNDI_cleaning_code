import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


warnings.filterwarnings('ignore')

def demographic(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Demographic
    '''

    df= df_root[df_root['name']=='Demographics'] 
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)


    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Demographics']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]
        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            for index, row in pru.iterrows():
                status = row['status']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]

                if status != '':
                    subject = row['Subject']
                    visit = row['Visit']

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    try:
                        birth_year = row['Birth Year']
                        birth_year_pure = birth_year.split('|')[0]
                        birth_year_form_field_instance = birth_year.split('|')[1]
                        birth_year_disname = birth_year.split('|')[0]
                    except:
                        birth_year = ''
                        birth_year_form_field_instance = 'This field does not have any data'
                        birth_year_disname = 'Empty'

                    try:
                        age_at_consent =  row['Age at consent']
                        age_at_consent_pure = int(age_at_consent.split('|')[0])
                        age_at_consent_form_field_instance = age_at_consent.split('|')[1]
                        age_at_consent_disname = age_at_consent.split('|')[0]   
                    except:
                        age_at_consent_pure = ''
                        age_at_consent_form_field_instance = 'This field does not have any data'
                        age_at_consent_disname = 'Empty'

                    try:
                        año_visita = row['Valor'].split('-')[2]
                    except:
                        año_visita = ''
                    
                    # Revision for DM0030
                    try:
                        año_calculado = int(año_visita) - int(birth_year_pure)

                        if age_at_consent_pure >= año_calculado -1 and age_at_consent_pure <= año_calculado + 1:
                            pass
                        else:
                            error = [subject, visit, 'Age at consent', age_at_consent_form_field_instance ,'The subject AGE at consent does not match the AGE according to the month and year of birth' ,\
                                      f'{age_at_consent_disname} - {año_calculado}', 'DM0030']
                            lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision DM0030 --> {e} - Subject: {subject},  Visit: {visit} ')

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    Informed_Consent_output = pd.DataFrame(lista_revision, columns=column_names)

    
    sheet = excel_writer.create_sheet("Demographic")

    for row in dataframe_to_rows(Informed_Consent_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return Informed_Consent_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)


if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    demographic(df_root, path_excel ) 

