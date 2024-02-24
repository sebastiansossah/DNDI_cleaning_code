import math
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def PBMC_isolate(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de PBMC Isolate
    '''

    df= df_root[df_root['name']== 'PBMC Isolate']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    #----------------------------------------------------------- CPG --------------------------------------------------------------------------------------

    df_time_dosing1 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1 = df_time_dosing1[(df_time_dosing1['Campo']=='Date of dosing') | (df_time_dosing1['Campo']=='Time of Dosing')]
    df_time_dosing = df_time_dosing1[df_time_dosing1['Campo']=='Date of dosing']
    df_time_dosing['time_dosing_cpg_administration'] =  df_time_dosing1[df_time_dosing1['FormFieldInstance Id'].isin(df_time_dosing['FormFieldInstance Id'] + 1) & (df_time_dosing1['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing =df_time_dosing[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing = df_time_dosing.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})

    df_time_dosing12 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing12 = df_time_dosing12[(df_time_dosing12['Campo']=='Date of dosing') | (df_time_dosing12['Campo']=='Time of Dosing')]
    df_time_dosing2 = df_time_dosing12[df_time_dosing12['Campo']=='Date of dosing']
    df_time_dosing2['time_dosing_cpg_administration'] =  df_time_dosing12[df_time_dosing12['FormFieldInstance Id'].isin(df_time_dosing2['FormFieldInstance Id'] + 1) & (df_time_dosing12['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing2 =df_time_dosing2[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing2 = df_time_dosing2.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join2', 'time_dosing_cpg_administration': 'time_dosing_cpg_administration2'})

    df_time_dosing123 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing123 = df_time_dosing123[(df_time_dosing123['Campo']=='Date of dosing') | (df_time_dosing123['Campo']=='Time of Dosing')]
    df_time_dosing3 = df_time_dosing123[df_time_dosing123['Campo']=='Date of dosing']
    df_time_dosing3['time_dosing_cpg_administration'] =  df_time_dosing123[df_time_dosing123['FormFieldInstance Id'].isin(df_time_dosing3['FormFieldInstance Id'] + 1) & (df_time_dosing123['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing3 =df_time_dosing3[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing3 = df_time_dosing3.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join3', 'time_dosing_cpg_administration': 'time_dosing_cpg_administration3'})

    #----------------------------------------------------------- Miltefosine --------------------------------------------------------------------------------------

    df_time_dosing1_miltefosine = df_root[df_root['name']== 'Miltefosine Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1_miltefosine = df_time_dosing1_miltefosine[(df_time_dosing1_miltefosine['Campo']=='Date of dosing') | (df_time_dosing1_miltefosine['Campo']=='Time of Dosing')]
    df_time_dosing_miltefosine = df_time_dosing1_miltefosine[df_time_dosing1_miltefosine['Campo']=='Date of dosing']
    df_time_dosing_miltefosine['time_dosing_Miltefosine_administration'] =  df_time_dosing1_miltefosine[df_time_dosing1_miltefosine['FormFieldInstance Id'].isin(df_time_dosing_miltefosine['FormFieldInstance Id'] + 1) & (df_time_dosing1_miltefosine['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing_miltefosine =df_time_dosing_miltefosine[['Participante','Valor', 'time_dosing_Miltefosine_administration']]
    df_time_dosing_miltefosine = df_time_dosing_miltefosine.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})

    df_time_dosing12_Miltefosine = df_root[df_root['name']== 'Miltefosine Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing12_Miltefosine = df_time_dosing12_Miltefosine[(df_time_dosing12_Miltefosine['Campo']=='Date of dosing') | (df_time_dosing12_Miltefosine['Campo']=='Time of Dosing')]
    df_time_dosing2_Miltefosine = df_time_dosing12_Miltefosine[df_time_dosing12_Miltefosine['Campo']=='Date of dosing']
    df_time_dosing2_Miltefosine['time_dosing_Miltefosine_administration2'] =  df_time_dosing12_Miltefosine[df_time_dosing12_Miltefosine['FormFieldInstance Id'].isin(df_time_dosing2_Miltefosine['FormFieldInstance Id'] + 1) & (df_time_dosing12_Miltefosine['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing2_Miltefosine = df_time_dosing2_Miltefosine[['Participante','Valor', 'time_dosing_Miltefosine_administration2']]
    df_time_dosing2_Miltefosine = df_time_dosing2_Miltefosine.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join2'})

    df_time_dosing123_Miltefosine = df_root[df_root['name']== 'Miltefosine Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing123_Miltefosine = df_time_dosing123_Miltefosine[(df_time_dosing123_Miltefosine['Campo']=='Date of dosing') | (df_time_dosing123_Miltefosine['Campo']=='Time of Dosing')]
    df_time_dosing3_Miltefosine = df_time_dosing123_Miltefosine[df_time_dosing123_Miltefosine['Campo']=='Date of dosing']
    df_time_dosing3_Miltefosine['time_dosing_Miltefosine_administration3'] =  df_time_dosing123_Miltefosine[df_time_dosing123_Miltefosine['FormFieldInstance Id'].isin(df_time_dosing3_Miltefosine['FormFieldInstance Id'] + 1) & (df_time_dosing123_Miltefosine['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing3_Miltefosine =df_time_dosing3_Miltefosine[['Participante','Valor', 'time_dosing_Miltefosine_administration3']]
    df_time_dosing3_Miltefosine = df_time_dosing3_Miltefosine.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join3'})

    #------------------------------------------

    lista_logs = ['PBMC Isolate']
    lista_revision = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()

            try:
                pru['date_ex_to_join'] = pru['Date of the sample collected'].str.split('|',expand= True)[0]
            except:
                pru['date_ex_to_join'] = 'Nothing to join'

            #try:
            pru['date_ex_to_join2'] = pru['date_ex_to_join'].apply(
                        lambda x: (datetime.strptime(x, '%d-%b-%Y') - timedelta(days=1)).strftime('%d-%b-%Y'))
            pru['date_ex_to_join2'] = pru['date_ex_to_join2'].astype(str).str.upper()

            #try:
            pru['date_ex_to_join3'] = pru['date_ex_to_join'].apply(
                        lambda x: (datetime.strptime(x, '%d-%b-%Y') - timedelta(days=2)).strftime('%d-%b-%Y'))
            pru['date_ex_to_join3'] = pru['date_ex_to_join3'].astype(str).str.upper()





   
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            # GPC Join ------------------------------------------------------------------
            pru = pru.merge(df_time_dosing, on=['Subject', 'date_ex_to_join'], how='left')
            pru = pru.merge(df_time_dosing2, on=['Subject', 'date_ex_to_join2'], how='left')
            pru = pru.merge(df_time_dosing3, on=['Subject', 'date_ex_to_join3'], how='left')
            
            # Miltefosine Join ------------------------------------------------------------------
            pru = pru.merge(df_time_dosing_miltefosine, on=['Subject', 'date_ex_to_join'], how='left')
            pru = pru.merge(df_time_dosing2_Miltefosine, on=['Subject', 'date_ex_to_join2'], how='left')
            pru = pru.merge(df_time_dosing3_Miltefosine, on=['Subject', 'date_ex_to_join3'], how='left')
            


            #if sujeto =='011002':
                # print(pru)
                # print('---------------------')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]

                
                time_dosing_cpg_administration = row['time_dosing_cpg_administration']
                time_dosing_cpg_administration2 = row['time_dosing_cpg_administration2']
                time_dosing_cpg_administration3 = row['time_dosing_cpg_administration3']

                time_dosing_Miltefosine_administration = row['time_dosing_Miltefosine_administration']
                time_dosing_Miltefosine_administration2 = row['time_dosing_Miltefosine_administration2']
                time_dosing_Miltefosine_administration3 = row['time_dosing_Miltefosine_administration3']



                if status != '':

                    try:
                        was_sample_collected = row['Was the sample collected to investigate immunological marker in PBMCs?']
                        was_sample_collected_pure = was_sample_collected.split('|')[0]
                        was_sample_collected_form_field_instance = was_sample_collected.split('|')[1]
                    except Exception as e:
                        was_sample_collected_pure = ''
                        was_sample_collected_form_field_instance = 'This field does not have any data'

                    try:
                        provide_reason = row['Provide the reason']
                        provide_reason_pure = provide_reason.split('|')[0]
                        provide_reason_form_field_instance = provide_reason.split('|')[1]
                    except Exception as e:
                        provide_reason_pure = ''
                        provide_reason_form_field_instance = 'This field does not have any data'

                    try:
                        date_sample_collected = row['Date of the sample collected']
                        date_sample_collected_pure = date_sample_collected.split('|')[0]
                        date_sample_collected_form_field_instance = date_sample_collected.split('|')[1]
                        date_sample_collected_disname = date_sample_collected.split('|')[0]
                    except Exception as e:
                        date_sample_collected_pure = ''
                        date_sample_collected_form_field_instance = 'This field does not have any data'
                        date_sample_collected_disname = 'Empty'
                    
                    try:
                        Time_collected = row['Time collected'] 
                        Time_collected_pure = Time_collected.split('|')[0]
                        Time_collected_form_field_instance = Time_collected.split('|')[1]
                    except Exception as e:
                        Time_collected_pure = ''
                        Time_collected_form_field_instance = 'This field does not have any data'


                    # --------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_sample_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_sample_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of the sample collected', date_sample_collected_form_field_instance,\
                                        f , date_sample_collected_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision PB0010
                    if date_sample_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_sample_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of the sample collected', date_sample_collected_form_field_instance ,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form' , f'{date_sample_collected_disname} - {date_of_visit}', 'PB0010']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision PB0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision PB0030
                    if date_sample_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_sample_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of the sample collected', date_sample_collected_form_field_instance ,\
                                        'The date/time of sample collected can not  be before the informed consent date/time', f'{date_sample_collected_disname} - {date_inform_consent}', 'PB0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision PB0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> PB0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_sample_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_sample_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of the sample collected', date_sample_collected_form_field_instance ,\
                                        'Date of the sample collected must be before the End of study/Early withdrawal date. ', date_sample_collected_disname, 'PB0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PB0040 --> {e} - Subject: {subject},  Visit: {visit}  ')
 
                    # Revision CPG ------------------------------------------------------------------------------------------------------------------------------------------------------------

                    # Revision PB0050
                    if visit in ['D1', 'D15','D29' ]:
                        if str(time_dosing_cpg_administration) == 'nan':
                            if Time_collected_pure == '':
                                error = [subject, visit, 'Time Collected', Time_collected_form_field_instance ,\
                                        f'There should be a time on visit {visita}', Time_collected, 'PB0050']
                                lista_revision.append(error)
                        else:
                               
                                dif = float((datetime.strptime(time_dosing_cpg_administration, '%H:%M') - datetime.strptime(Time_collected_pure, '%H:%M')).total_seconds() / 60)
                                if dif >= 0.0 or dif <= 90.0:
                                        
                                    error = [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                                'The date and time collected must be between 0 and 90 minutes before the study treatment administration time', \
                                                    f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0050']
                                    lista_revision.append(error)

                    # Revision PB0060
                    if visit in ['D2', 'D16', 'D30']:
                        if str(time_dosing_cpg_administration2) == 'nan' and str(time_dosing_cpg_administration3) == 'nan':
                            error =  [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                             'The date and time collected must be between 24 and 25 hours  after the study treatment administration time of the day before', \
                                                f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0060']
                            lista_revision.append(error)
                        
                        if str(time_dosing_cpg_administration2) != 'nan':
                            time_date_compare_1_gcp =  row['date_ex_to_join2'] + ' ' + time_dosing_cpg_administration2
                            time_to_compare_pbmc_1 = date_sample_collected_pure + ' ' + Time_collected_pure

                            dif_25_1 = float((datetime.strptime(time_to_compare_pbmc_1, '%d-%b-%Y %H:%M') - datetime.strptime(time_date_compare_1_gcp, '%d-%b-%Y %H:%M')).total_seconds() / 60)
                            #print(dif_25_1)
                            if dif_25_1 < 1440 or dif_25_1 > 1500:
                                error =  [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                                'The date and time collected must be between 24 and 25 hours  after the study treatment administration time of the day before', \
                                                    f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0060']
                                lista_revision.append(error)
                        
                        if str(time_dosing_cpg_administration3) != 'nan':
                            time_date_compare_2_gcp =  row['date_ex_to_join3'] + ' ' + time_dosing_cpg_administration3
                            time_to_compare_pbmc_2 = date_sample_collected_pure + ' ' + Time_collected_pure

                            dif_25_2 = float((datetime.strptime(time_to_compare_pbmc_2, '%d-%b-%Y %H:%M') - datetime.strptime(time_date_compare_2_gcp, '%d-%b-%Y %H:%M')).total_seconds() / 60)
             
                            if dif_25_2 < 1440 or dif_25_2 > 1500:
                                error =  [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                                'The date and time collected must be between 24 and 25 hours  after the study treatment administration time of the day before', \
                                                    f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0060']
                                lista_revision.append(error)

                    # Revision Miltefosine ------------------------------------------------------------------------------------------------------------------------------------------------------------

                    # Revision PB0050
                    if visit in ['D1', 'D15','D29' ]:
                        if str(time_dosing_Miltefosine_administration) == 'nan':
                            if Time_collected_pure == '':
                                error = [subject, visit, 'Time Collected', Time_collected_form_field_instance ,\
                                        f'There should be a time on visit {visita}', Time_collected, 'PB0050']
                                lista_revision.append(error)
                        else:
                               
                                dif = float((datetime.strptime(time_dosing_Miltefosine_administration, '%H:%M') - datetime.strptime(Time_collected_pure, '%H:%M')).total_seconds() / 60)
                                if dif >= 0.0 or dif <= 90.0:
                                        
                                    error = [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                                'The date and time collected must be between 0 and 90 minutes before the study treatment administration time', \
                                                    f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_Miltefosine_administration}', 'PB0050']
                                    lista_revision.append(error)

                    # Revision PB0060
                    if visit in ['D2', 'D16', 'D30']:
                        if str(time_dosing_Miltefosine_administration2) == 'nan' and str(time_dosing_Miltefosine_administration3) == 'nan':
                            error =  [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                             'The date and time collected must be between 24 and 25 hours  after the study treatment administration time of the day before', \
                                                f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0060']
                            lista_revision.append(error)
                        
                        if str(time_dosing_Miltefosine_administration2) != 'nan':
                            time_date_compare_1_miltefosine =  row['date_ex_to_join2'] + ' ' + time_dosing_Miltefosine_administration2
                            time_to_compare_pbmc_1 = date_sample_collected_pure + ' ' + Time_collected_pure

                            dif_25_1_M = float((datetime.strptime(time_to_compare_pbmc_1, '%d-%b-%Y %H:%M') - datetime.strptime(time_date_compare_1_miltefosine, '%d-%b-%Y %H:%M')).total_seconds() / 60)
                           
                            if dif_25_1_M < 1440 or dif_25_1_M > 1500:
                                error =  [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                                'The date and time collected must be between 24 and 25 hours  after the study treatment administration time of the day before', \
                                                    f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0060']
                                lista_revision.append(error)
                        
                        if str(time_dosing_Miltefosine_administration3) != 'nan':
                            time_date_compare_2_miltefosine =  row['date_ex_to_join3'] + ' ' + time_dosing_Miltefosine_administration3
                            time_to_compare_pbmc_2 = date_sample_collected_pure + ' ' + Time_collected_pure

                            dif_25_2 = float((datetime.strptime(time_to_compare_pbmc_2, '%d-%b-%Y %H:%M') - datetime.strptime(time_date_compare_2_miltefosine, '%d-%b-%Y %H:%M')).total_seconds() / 60)
              
                            if dif_25_2 < 1440 or dif_25_2 > 1500:
                                error =  [subject, visit, 'Time Collected', Time_collected_form_field_instance,\
                                                'The date and time collected must be between 24 and 25 hours  after the study treatment administration time of the day before', \
                                                    f'Time Collected: {Time_collected_pure} - dose time administration{time_dosing_cpg_administration}', 'PB0060']
                                lista_revision.append(error)






    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    PBMC_isolate_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("PBMC Isolate")

    for row in dataframe_to_rows(PBMC_isolate_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return PBMC_isolate_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    PBMC_isolate(df_root, path_excel ) 