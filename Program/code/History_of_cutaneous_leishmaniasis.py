import numpy as np
import pandas as pd
import math
from datetime import datetime
from revision_fechas import revision_fecha, date_format
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def history_of_cutaneous_leishmaniasis(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de History of cutaneous leishmaniasis
    '''

    df= df_root[df_root['name']=='History of cutaneous leishmaniasis']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)


    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Visit','Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_demographic_age_year = df_root[df_root['name']=='Demographics']
    df_demographic_age_year = df_demographic_age_year[['Visit','Participante', 'Campo', 'Valor']]
    df_demographic_age_year = df_demographic_age_year[df_demographic_age_year['Campo']=='Birth Year']
    df_demographic_age_year = df_demographic_age_year[['Visit','Participante','Valor']]
    df_demographic_age_year = df_demographic_age_year.rename(columns={'Participante':'Subject', 'Valor':'Birth_year'})

    df_demographic_age_month = df_root[df_root['name']=='Demographics']
    df_demographic_age_month = df_demographic_age_month[['Visit','Participante', 'Campo', 'Valor']]
    df_demographic_age_month = df_demographic_age_month[df_demographic_age_month['Campo']=='Birth Month']
    df_demographic_age_month = df_demographic_age_month[['Visit','Participante','Valor']]
    df_demographic_age_month = df_demographic_age_month.rename(columns={'Participante':'Subject', 'Valor':'Birth_month'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_logs = ['History of cutaneous leishmaniasis']
    lista_revision = []
    lista_date_diagnosis = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in ['Screening Visit']:
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_demographic_age_month, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_demographic_age_year, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            lista_other_names = []

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                
                date_of_visit = row['Date_of_visit']

                year_birth = row['Birth_year']
                month_birth = row['Birth_month']

                date_inform_consent = row['Informed_consent_date']

                if status != '':
               
                    try:
                        date_confirmed_diagnosis = row['Date of confirmed diagnosis of CL']
                        date_confirmed_diagnosis_pure = date_confirmed_diagnosis.split('|')[0]
                        date_confirmed_diagnosis_form_field_instance = date_confirmed_diagnosis.split('|')[1]
                    except Exception as e:
                        date_confirmed_diagnosis_pure = ''
                        date_confirmed_diagnosis_form_field_instance = 'This field does not have any data'

                    try:
                        date_new_sample = row['Date of new sample taken']
                        date_new_sample_pure = date_new_sample.split('|')[0]
                        date_new_sample_form_field_instance = date_new_sample.split('|')[1]
                    except Exception as e:
                        date_new_sample_pure = ''
                        date_new_sample_form_field_instance = 'This field does not have any data'

                    try:
                        species_identification = row['Species identification']
                        species_identification_pure = species_identification.split('|')[0]
                        species_identification_form_field_instance = species_identification.split('|')[1]
                    except Exception as e:
                        species_identification_pure = math.nan
                        species_identification_form_field_instance = 'This field does not have any data'

                    try:
                        species_name = row['Species name']
                        species_name_pure = species_name.split('|')[0]
                        species_name_form_field_instance = species_name.split('|')[1]
                    except Exception as e:
                        species_name_pure = math.nan
                        species_name_form_field_instance = 'This field does not have any data'

                    try:
                        previous_history_leishmaniasis = row['Are there any previous history of leishmaniasis (with a diagnosis in the past)?']
                        previous_history_leishmaniasis_pure = previous_history_leishmaniasis.split('|')[0]
                        previous_history_leishmaniasis_form_field_instance = previous_history_leishmaniasis.split('|')[1]
                    except Exception as e:
                        previous_history_leishmaniasis_pure = math.nan
                        previous_history_leishmaniasis_form_field_instance = 'This field does not have any data'

                    try:
                        date_diagnosis = row['Date of Diagnosis']
                        date_diagnosis_pure = date_diagnosis.split('|')[0]
                        date_diagnosis_form_field_instance = date_diagnosis.split('|')[1]
                    except Exception as e:
                        date_diagnosis_pure = ''
                        date_diagnosis_form_field_instance = 'This field does not have any data'

                    #----------------------------------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)
                    
                    # Primera  revision general de formato de fecha ->GE0020
                    if date_confirmed_diagnosis_pure == '':
                        pass
                    else:
                        try:
                            f = revision_fecha(date_confirmed_diagnosis_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of confirmed diagnosis of CL', date_confirmed_diagnosis_form_field_instance ,f , date_confirmed_diagnosis_pure, 'GE0020']
                                lista_revision.append(error) 
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Primera  revision general de formato de fecha ->GE0020
                    if date_new_sample_pure == '':
                        pass
                    else:
                        try:
                            f = revision_fecha(date_new_sample_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit,'Date of new sample taken', date_new_sample_form_field_instance ,f , date_new_sample_pure, 'GE0020']
                                lista_revision.append(error) 
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision CL0010
                    if date_confirmed_diagnosis_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_format(str(date_confirmed_diagnosis_pure)), date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of confirmed diagnosis of CL', date_confirmed_diagnosis_form_field_instance ,\
                                        'The year of diagnosis of CL must be equal or after the year of birth in DEMOGRAPHIC' , f'{date_confirmed_diagnosis_pure} - {date_of_visit}', 'CL0010']
                                lista_revision.append(error)
                            else:
                                pass

                        except Exception as e:
                            lista_logs.append(f'Revision CL0010--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision CL0020
                    if date_confirmed_diagnosis_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_confirmed_diagnosis_f = datetime.strptime(date_format(str(date_confirmed_diagnosis_pure)), date_format)
                            date_new_sample_f = datetime.strptime(date_format(date_new_sample_pure), date_format)

                            if date_confirmed_diagnosis_f < date_new_sample_f:
                                error = [subject, visit, 'Date of Sample used for Diagnosis taken', date_new_sample_form_field_instance ,\
                                        'The date of sample must be before the diagnosis of CL date.' , f'{date_confirmed_diagnosis_pure} - {date_new_sample_pure}', 'CL0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision CL0020--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision CL0030
                    if date_new_sample_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_birth_cured = f'01-{month_birth}-{year_birth}'
                            date_birth_format = datetime.strptime(date_birth_cured, '%d-%m-%Y')
                            date_new_sample_f = datetime.strptime(date_format(str(date_new_sample_pure)), date_format)

                            if  date_new_sample_f <  date_birth_format:
                                error = [subject, visit, 'Date of Sample used for Diagnosis taken', date_new_sample_form_field_instance,  \
                                        'The year and month of Sample used for Diagnosis taken must be equal or after the year and month of birth in DEMOGRAPHIC' , date_new_sample_pure, 'CL0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision CL0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision CL0040
                    if math.isnan(float(species_identification_pure)):
                        pass
                    else:
                        try: 
                            if 99.0 in [float(i) for i in species_identification_pure.split(',')]:
                                if math.isnan(float(species_name_pure)) or str(species_name_pure) == '' or float(species_name_pure) == np.nan or  str(species_name_pure) == '-':
                                    error = [subject, visit, 'Species identification', species_identification_form_field_instance ,\
                                                'If "other" is selected, there must be at least one "other species" section added' , species_name_pure, 'CL0040']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision CL0040--> {e} - Subject: {subject},  Visit: {visit} ')
    
                    # Revision CL0050
                    if math.isnan(float(species_identification_pure)):
                        pass
                    else:
                        try: 
                            if 99.0 not in [float(i) for i in species_identification_pure.split(',')]:
                                pass 
                            else:
                                try:
                                    if math.isnan(float(species_name_pure)) or str(species_name_pure) != '' or float(species_name_pure) != np.nan or  str(species_name_pure) != '-':
                                        error = [subject, visit, 'Species identification', species_identification_form_field_instance, \
                                                'If at least one "other species" section is added, the "other" option must be selected' , f'{species_identification_pure} - {species_name_pure}', 'CL0050']
                                        lista_revision.append(error)
                                except Exception as e:
                                        lista_logs.append(f'Revision CL0050--> {e} - Subject: {subject},  Visit: {visit} ')
                        except Exception as e:
                            lista_logs.append(f'Revision CL0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision CL0070
                    if math.isnan(float(species_name_pure)):
                        pass
                    else:
                        try:
                            to_save =  species_name_pure.lower()
                            if to_save in lista_other_names:
                                error = [subject, visit, 'Other Species', species_name_form_field_instance ,'The error message should describe the inconsistency found', species_name_pure, 'CL0070']
                                lista_revision.append(error)
                            else:
                                lista_other_names.append(to_save)

                        except Exception as e:
                            lista_logs.append(f'Revision CL0070--> {e} - Subject: {subject},  Visit: {visit} ')

                    lista_validacion = [
                        'Type of Leishmaniasis history', 
                        'Date of Diagnosis',
                        'Is Condition Ongoing?', 
                        'Currently treated?'   
                    ]

                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try: 
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan

                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass

                    # Revision CL0080
                    if math.isnan(float(previous_history_leishmaniasis_pure)):
                        pass
                    else:
                        try:
                            if float(previous_history_leishmaniasis_pure) ==1.0: 
                                if mi_cuenta != 0:
                                    pass
                                else:
                                    error = [subject, visit, 'Are there any previous history of leishmaniasis (with a diagnosis in the past)?', previous_history_leishmaniasis_form_field_instance ,\
                                            'If "Are there any previous history of leishmaniasis (with a diagnosis in the past)?" = Yes at least one "History of Leishmaniasis Details" section must be added', \
                                                previous_history_leishmaniasis_pure, 'CL0080']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision CL0080--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision CL0090
                    if math.isnan(float(previous_history_leishmaniasis_pure)):
                        pass
                    else:
                        try:
                            if float(previous_history_leishmaniasis_pure) ==1.0: 
                                if mi_cuenta == 0:
                                    pass
                                else:
                                    error = [subject, visit, 'Are there any previous history of leishmaniasis (with a diagnosis in the past)?', previous_history_leishmaniasis_form_field_instance ,\
                                            'If "Are there any previous history of leishmaniasis (with a diagnosis in the past)?" = NO, there should be no "History of Leishmaniasis Details" section must be added', \
                                                previous_history_leishmaniasis_pure, 'CL0090']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision CL0090--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision CL0100
                    if date_diagnosis_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_birth_cured = f'01-{month_birth}-{year_birth}'
                            date_birth_format = datetime.strptime(date_birth_cured, '%d-%m-%Y')
                            date_diagnosis_f = datetime.strptime(date_format(date_diagnosis_pure), date_format)

                            if date_diagnosis_f <  date_birth_format:
                                error = [subject, visit, 'History of Leishmaniasis Details - Date of Diagnosis', date_diagnosis_form_field_instance,\
                                        'The year and month of  Date of Diagnosis taken must be equal or after the month and year of birth in DEMOGRAPHIC' , date_diagnosis_pure, 'CL0100']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision CL0100--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision CL0110
                    if date_diagnosis_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_diagnosis_f = datetime.strptime(date_format(date_diagnosis_pure), date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_diagnosis_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of Diagnosis', date_diagnosis_form_field_instance ,'The Date of Diagnosis should be before the informed consent date', \
                                        f'{date_diagnosis_pure} - {date_inform_consent}', 'CL0110']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision CL0110--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision CL0120
                    if date_diagnosis_pure != '':
                        try:
                            if date_diagnosis_pure in lista_date_diagnosis:
                                error = [subject, visit, 'Date of Diagnosis', date_diagnosis_form_field_instance, 'The Date of Diagnosis should not be repeated', date_diagnosis_pure, 'CL0120']
                                lista_revision.append(error)
                            else:
                                lista_date_diagnosis.append(date_diagnosis_pure)
                        except Exception as e:
                            lista_logs.append(f'Revision CL0120--> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    history_of_cutaneous_leishmaniasis_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("History cutane leishmaniasis")

    for row in dataframe_to_rows(history_of_cutaneous_leishmaniasis_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return history_of_cutaneous_leishmaniasis_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    history_of_cutaneous_leishmaniasis(df_root, path_excel) 