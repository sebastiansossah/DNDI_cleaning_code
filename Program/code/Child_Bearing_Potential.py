
from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import math

warnings.filterwarnings('ignore')

def child_bearing_potential(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Child Bearing Potential
    '''

    df= df_root[df_root['name']=='Child Bearing Potential']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject'})

    df_demographic = df_root[df_root['name']=='Demographics']
    df_demographic = df_demographic[['Visit','Participante', 'Campo', 'Valor']]
    df_demographic = df_demographic[df_demographic['Campo']=='Gender']
    df_demographic = df_demographic[['Visit','Participante','Valor']]
    df_demographic = df_demographic.rename(columns={'Participante':'Subject', 'Valor':'Genero'})

    df_demographic_age = df_root[df_root['name']=='Demographics']
    df_demographic_age = df_demographic_age[['Visit','Participante', 'Campo', 'Valor']]
    df_demographic_age = df_demographic_age[df_demographic_age['Campo']=='Birth Year']
    df_demographic_age = df_demographic_age[['Visit','Participante','Valor']]
    df_demographic_age = df_demographic_age.rename(columns={'Participante':'Subject', 'Valor':'Birth_year'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    df_contraception = df_root[df_root['name']== 'Prior And Concomitant Medications']
    df_contraception = df_contraception[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_contraception = df_contraception.sort_values(by=['FormFieldInstance Id'])
    df_contraception = df_contraception.reset_index(drop=True)
    date_indices = df_contraception.index[df_contraception['Campo'] == 'Concomitant Medication ID'].tolist()
    subdatasets = [df_contraception.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]
    df_to_join = pd.DataFrame()
    for sub in subdatasets:
        #if sub['Valor'].tolist()  in [5, 5.0, '5', '5.0', 9]:
        if len([float(x) for x in sub['Valor'].tolist() if x in [5, 5.0, '5', '5.0', 9]]) != 0:
            df_to_join = sub[(sub['Campo'] == 'Indication Category') | (sub['Campo'] == 'Start date')]
            sujeto = df_to_join.Participante.unique()
            df_to_join = df_to_join[['Campo', 'Valor']].T
            new_columns = df_to_join.iloc[0]
            df_to_join = df_to_join[1:].set_axis(new_columns, axis=1)
            df_to_join['Subject'] = sujeto
            break
    df_to_join = df_to_join.rename(columns={'Participante':'Subject', 'Start date':'start_date_combined_hormonal'})

    df_medical_surgical = df_root[df_root['name']== 'Medical Or Surgical History (other than Leishmaniasis)']
    df_medical_surgical = df_medical_surgical[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_medical_surgical = df_medical_surgical[df_medical_surgical['Campo']== 'Onset Date/First Diagnosis/Surgery']
    df_medical_surgical['Date of start of contraceptive method'] = df_medical_surgical['Valor']
    df_medical_surgical = df_medical_surgical[['Visit','Participante','Valor', 'Date of start of contraceptive method']]
    df_medical_surgical = df_medical_surgical.rename(columns={'Participante':'Subject', 'Valor':'onset_date_medical_contraceptive'})

    lista_revision = []
    lista_logs = ['Child Bearing Potential']


    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_demographic, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_demographic_age, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_to_join, on=['Subject'], how='left')
            pru = pru.merge(df_medical_surgical, on=['Subject', 'Date of start of contraceptive method' ], how='left')

            for index, row in pru.iterrows():
                status = row['status']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]

                start_date_combined_hormonal = row['start_date_combined_hormonal']
                onset_date_medical_contraceptive = row['onset_date_medical_contraceptive']

                if status != '':

                    try:
                        subject = row['Subject']
                    except Exception as e:
                        subject = ''

                    try:
                        visit = row['Visit']
                    except Exception as e:
                        visit = ''

                    try:
                        fecha_visita = row['Valor']
                    except Exception as e:
                        fecha_visita = ''

                    try:
                        genero = row['Genero']
                    except Exception as e:
                        genero = ''

                    try:
                        birth_year = row['Birth_year']
                    except Exception as e:
                        birth_year = ''
# --------------------------------------------------------------------------

                    try:
                        date_of_start_condom = row['Date of start of systematic use of condom']
                        date_of_start_condom_pure = date_of_start_condom.split('|')[0]
                        date_of_start_condom_form_field_instance = date_of_start_condom.split('|')[1]
                        date_of_start_condom_disname = date_of_start_condom.split('|')[0]
                    except:
                        date_of_start_condom_pure = ''
                        date_of_start_condom_form_field_instance = 'This field does not have any data'
                        date_of_start_condom_disname = 'Empty'

                    try:
                        date_start_contraceptive = row['Date of start of contraceptive method']
                        date_start_contraceptive_pure = date_start_contraceptive.split('|')[0]
                        date_start_contraceptive_form_field_instance = date_start_contraceptive.split('|')[1]
                        date_start_contraceptive_disname = date_start_contraceptive.split('|')[0]
                    except:
                        date_start_contraceptive_pure = ''
                        date_start_contraceptive_form_field_instance = 'This field does not have any data'
                        date_start_contraceptive_disname = 'Empty'

                    try:
                        last_mestruation_year = row['Year of Last Menstruation']
                        last_mestruation_year_pure = last_mestruation_year.split('|')[0]
                        last_mestruation_year_form_field_instance = last_mestruation_year.split('|')[1]
                        last_mestruation_year_disname = last_mestruation_year.split('|')[0]
                    except:
                        last_mestruation_year_pure = math.nan
                        last_mestruation_year_form_field_instance = 'This field does not have any data'
                        last_mestruation_year_disname = 'Empty'

                    try:
                        participant_postmenopausical = row['Is the participant postmenopausal?']
                        participant_postmenopausical_pure = participant_postmenopausical.split('|')[0]
                        participant_postmenopausical_form_field_instance = participant_postmenopausical.split('|')[1]
                        participant_postmenopausical_disname = participant_postmenopausical.split('|')[2]
                    except:
                        participant_postmenopausical_pure = math.nan
                        participant_postmenopausical_form_field_instance = 'This field does not have any data'
                        participant_postmenopausical_disname = 'Empty'

                    try:
                        last_mestruation_month = row['Month of Last Menstruation']
                        last_mestruation_month_pure = last_mestruation_month.split('|')[0]
                        last_mestruation_month_form_field_instance = last_mestruation_month.split('|')[1]
                        last_mestruation_month_disname = last_mestruation_month.split('|')[0]
                    except:
                        last_mestruation_month_pure = math.nan
                        last_mestruation_month_form_field_instance = 'This field does not have any data'
                        last_mestruation_month_disname = 'Empty'

                    try:
                        fsh_available = row['Is the FSH test result available and ≥ 40 IU/L?']
                        fsh_available_pure = fsh_available.split('|')[0]
                        fsh_available_form_field_instance = fsh_available.split('|')[1]
                        fsh_available_form_disname = fsh_available.split('|')[2]
                    except:
                        fsh_available_pure = math.nan
                        fsh_available_form_field_instance = 'This field does not have any data'
                        fsh_available_form_disname = 'Empty'

                    try:
                        contraception = row['Contraception of non post-menopausal woman']
                        contraception_pure = contraception.split('|')[0]
                        contraception_form_field_instance = contraception.split('|')[1]
                        contraception_disname = contraception.split('|')[2]
                    except:
                        contraception_pure = math.nan
                        contraception_form_field_instance = 'This field does not have any data'
                        contraception_disname = 'Empty'

                    try:
                        use_combined_hormonal = row['Use of combined (estrogen and progestogen-containing) hormonal contraception. associated with inhibition of ovulation']
                        use_combined_hormonal_pure = use_combined_hormonal.split('|')[0]
                        use_combined_hormonal_form_field_instance = use_combined_hormonal.split('|')[1]
                        use_combined_hormonal_disname = use_combined_hormonal.split('|')[2]
                    except:
                        use_combined_hormonal_pure = math.nan
                        use_combined_hormonal_form_field_instance = 'This field does not have any data'
                        use_combined_hormonal_disname = 'Empty'
                    
                    try:
                        progeston_hormonal = row['Use of progestogen-only hormonal contraception']
                        progeston_hormonal_pure = progeston_hormonal.split('|')[0]
                        progeston_hormonal_form_field_instance = progeston_hormonal.split('|')[1]
                        progeston_hormonal_disname = progeston_hormonal.split('|')[2]
                    except:
                        progeston_hormonal_pure = math.nan
                        progeston_hormonal_form_field_instance = 'This field does not have any data'
                        progeston_hormonal_disname = 'Empty'
                    # ----------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    # Revision para CB0010
                    if float(genero) == 1.0:
                        try:
                            if math.isnan(float(participant_postmenopausical_pure)) == False or str(participant_postmenopausical_pure) != ''  or str(participant_postmenopausical_pure) != 'N/A':
                                error = [subject, visit, 'Form Child Bearing Potential', participant_postmenopausical_form_field_instance ,'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , \
                                         participant_postmenopausical_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass
                      
                        try:
                            if math.isnan(float(last_mestruation_month_pure)) == False or str(last_mestruation_month_pure) != '' or str(last_mestruation_month_pure) != 'N/A':
                                error = [subject, visit, 'Form Child Bearing Potential', last_mestruation_month_form_field_instance,  \
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , last_mestruation_month_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass

                        try:
                            if math.isnan(float(last_mestruation_year_pure)) == False or str(last_mestruation_year_pure) != ''  or str(last_mestruation_year_pure) != 'N/A':
                                error = [subject, visit,'Form Child Bearing Potential', last_mestruation_year_form_field_instance, \
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , last_mestruation_year_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass

                        try:
                            if math.isnan(float(fsh_available_pure)) == False or str(fsh_available_pure) != ''  or str(fsh_available_pure) != 'N/A':
                                error = [subject, visit,'Form Child Bearing Potential', fsh_available_form_field_instance, \
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , fsh_available_form_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass

                        try:
                            if math.isnan(float(contraception_pure)) == False or str(contraception_pure) != '' or   str(contraception_pure) != 'N/A':
                                error = [subject, visit, 'Form Child Bearing Potential', contraception_form_field_instance, \
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty', contraception_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e: 
                            pass

                        try:
                            if math.isnan(float(use_combined_hormonal_pure)) == False or  str(use_combined_hormonal_pure) != ''  or str(use_combined_hormonal_pure) != 'N/A':
                                error = [subject, visit, 'Form Child Bearing Potential', use_combined_hormonal_form_field_instance, \
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , use_combined_hormonal_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass

                        try:
                            if math.isnan(float(progeston_hormonal_pure)) == False or str(progeston_hormonal_pure) != ''  or str(progeston_hormonal_pure) != 'N/A':
                                error = [subject, visit,'Form Child Bearing Potential', progeston_hormonal_form_field_instance ,\
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , progeston_hormonal_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass

                        try:
                            if math.isnan(float(date_start_contraceptive_pure)) == False or str(date_start_contraceptive_pure) != '' or str(date_start_contraceptive_pure) != 'N/A':
                                error = [subject, visit,'Form Child Bearing Potential', date_start_contraceptive_form_field_instance ,\
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , date_start_contraceptive_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass

                        try:
                            if math.isnan(float(date_of_start_condom_pure)) == False or str(date_of_start_condom_pure) != ''  or str(date_of_start_condom_pure) != 'N/A':
                                error = [subject, visit, 'Form Child Bearing Potential', date_of_start_condom_form_field_instance, \
                                         'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , date_of_start_condom_disname, 'CB0010']
                                lista_revision.append(error)
                        except Exception as e:
                            pass    
                    else:
# --------------------------------------------- Revision de edit check-----------------------------------------------------------

                        if date_of_start_condom_pure == '':
                            pass
                        else:
                            # Revision para ->GE0020
                            try:
                                f = revision_fecha(date_of_start_condom_pure)
                                if f == None:
                                    pass
                                else:
                                    error = [subject, visit, 'Date of start of systematic use of condom', date_of_start_condom_form_field_instance ,f , date_of_start_condom_disname, 'GE0020']
                                    lista_revision.append(error) 
                            except Exception as e:
                                lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        if date_start_contraceptive_pure == '':
                            pass
                        else:
                            # Revision para->GE0020
                            try:
                                f = revision_fecha(date_start_contraceptive_pure)
                                if f == None:
                                    pass
                                else:
                                    error = [subject, visit, 'Date of start of contraceptive method', date_start_contraceptive_form_field_instance ,f , date_start_contraceptive_disname, 'GE0020']
                                    lista_revision.append(error) 
                            except Exception as e:
                                lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                        

                        if math.isnan(float(last_mestruation_month_pure)):
                            pass
                        else:                      
                            # Revision para CB0020
                            try:        
                                vist_date_year =  fecha_visita.split('-')[2]
                                amount_years = int(vist_date_year) - int(last_mestruation_year_pure)
                                if amount_years >= 1:
                                    pass
                                else:
                                    error = [subject, visit, 'Year of Last Menstruation', last_mestruation_year_form_field_instance,\
                                            'There should be more than a year difference between the visit date and the year of last menstruation' , last_mestruation_year_disname, 'CB0020']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision CB0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para CB0030
                        if math.isnan(float(last_mestruation_month_pure)):
                            pass
                        else:       
                            try:                        
                                if int(last_mestruation_year_pure) <= 1968 or int(last_mestruation_year_pure) >= 2024:
                                    pass
                                else:
                                    error = [subject, visit, 'Year of Last Menstruation' , last_mestruation_year_form_field_instance ,'The year should be after 1968 and before 2024',\
                                            last_mestruation_year_disname, 'CB0030']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision CB0030 --> {e} - Subject: {subject},  Visit: {visit} ')

                        if date_start_contraceptive_pure == '':
                            pass
                        else:
                            try:
                                # Revision para CB0140
                                if int(birth_year) < int(date_start_contraceptive_pure.split('-')[2]):
                                    pass
                                else:
                                    error = [subject, visit, 'Date of start of contraceptive method', date_start_contraceptive_form_field_instance, \
                                            'The contraceptive method date can not be before the birth date' , date_start_contraceptive_disname, 'CB0140']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision CB0140 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision CB0070
                        if float(contraception_pure) == 1.0:
                            if str(date_start_contraceptive_pure) != str(start_date_combined_hormonal):
                                    error = [subject, visit, 'Contraception method used by Female of child-bearing potential', date_start_contraceptive_form_field_instance, \
                                            'The concomitant medication form related to this contraception method is not accurate (date) or the medication has not been added, please review', \
                                                date_start_contraceptive_disname, 'CB0070']
                                    lista_revision.append(error)

                        # Revision CB0080
                        if float(contraception_pure) == 2.0:
                            if str(date_start_contraceptive_pure) != str(start_date_combined_hormonal):
                                    error = [subject, visit, 'Contraception method used by Female of child-bearing potential', date_start_contraceptive_form_field_instance, \
                                            'The concomitant medication form related to this contraception method is not accurate (date) or the medication has not been added, please review', \
                                                date_start_contraceptive_disname, 'CB0080']
                                    lista_revision.append(error)
                       
                        # Revision CB0090
                        if float(contraception_pure) == 3.0:
                            if str(date_start_contraceptive_pure) != str(start_date_combined_hormonal):
                                    error = [subject, visit, 'Contraception method used by Female of child-bearing potential', date_start_contraceptive_form_field_instance, \
                                            'The concomitant medication form related to this contraception method is not accurate (date) or the medication has not been added, please review', \
                                                date_start_contraceptive_disname, 'CB0090']
                                    lista_revision.append(error)

                        # Revision CB0100
                        if float(contraception_pure) == 4.0:
                            if math.isnan(onset_date_medical_contraceptive):
                                    error = [subject, visit, 'Contraception method used by Female of child-bearing potential', date_start_contraceptive_form_field_instance, \
                                            'The concomitant medication form related to this contraception method is not accurate (date) or the medication has not been added, please review', \
                                                date_start_contraceptive_disname, 'CB0090']
                                    lista_revision.append(error)

                        # Revision CB0120
                        if float(contraception_pure) == 5.0:
                            if math.isnan(onset_date_medical_contraceptive):
                                    error = [subject, visit, 'Contraception method used by Female of child-bearing potential', date_start_contraceptive_form_field_instance, \
                                            'The concomitant medication form related to this contraception method is not accurate (date) or the medication has not been added, please review', \
                                                date_start_contraceptive_disname, 'CB0120']
                                    lista_revision.append(error)

    excel_writer = load_workbook(path_excel_writer)
    column_names =  ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    child_bearing_potential_output = pd.DataFrame(lista_revision, columns=column_names)
 
    sheet = excel_writer.create_sheet("Child Bearing Potential")

    for row in dataframe_to_rows(child_bearing_potential_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return child_bearing_potential_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    child_bearing_potential(df_root, path_excel ) 