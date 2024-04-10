from datetime import datetime
import math 
import numpy as np
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def clinical_laboratory_test_hematology(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Clinical Laboratory - Test Hematology
    '''

    df= df_root[df_root['name']== 'Clinical Laboratory - Test Hematology']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante',  'Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_demographic = df_root[df_root['name']=='Demographics']
    df_demographic = df_demographic[['Participante', 'Campo', 'Valor']]
    df_demographic = df_demographic[df_demographic['Campo']=='Gender']
    df_demographic = df_demographic[['Participante','Valor']]
    df_demographic = df_demographic.rename(columns={'Participante':'Subject', 'Valor':'Genero'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    warnings.filterwarnings('ignore')

    lista_revision = []
    lista_logs = ['Clinical Laboratory - Test Hematology']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_demographic, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            # print(pru)
            # print('----------------')


            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                genero = row['Genero']
                end_study_date = row['end_study_date']

                if status != '':
    
                    try:
                        absolute_Neutrophil = row['Absolute Neutrophil count, Out of normal range?']
                        absolute_Neutrophil_pure = absolute_Neutrophil.split('|')[0]
                        absolute_Neutrophil_form_field_instance = absolute_Neutrophil.split('|')[1]
                        absolute_Neutrophil_disname = absolute_Neutrophil.split('|')[2]
                    except Exception as e:
                        absolute_Neutrophil_pure = math.nan
                        absolute_Neutrophil_form_field_instance = 'This field does not have any data'
                        absolute_Neutrophil_disname = 'Empty'

                    try:
                        absolute_Neutrophil_result = row['Absolute Neutrophil count, Result (cels/uL)']
                        absolute_Neutrophil_result_pure = absolute_Neutrophil_result.split('|')[0]
                        absolute_Neutrophil_result_form_field_instance = absolute_Neutrophil_result.split('|')[1]
                        absolute_Neutrophil_result_disname = absolute_Neutrophil_result.split('|')[0]
                    except Exception as e:
                        absolute_Neutrophil_result_pure = math.nan
                        absolute_Neutrophil_result_form_field_instance = 'This field does not have any data'
                        absolute_Neutrophil_result_disname = 'Empty'

                    try:
                        absolute_basophil =  row['Absolute basophil count, Out of normal range?']
                        absolute_basophil_pure = absolute_basophil.split('|')[0]
                        absolute_basophil_form_field_isntance = absolute_basophil.split('|')[1]
                        absolute_basophil_disname = absolute_basophil.split('|')[2]
                    except Exception as e:
                        absolute_basophil_pure = math.nan
                        absolute_basophil_form_field_isntance = 'This field does not have any data'
                        absolute_basophil_disname = 'Empty'

                    try:
                        absolute_basophil_result =  row['Absolute basophil count, Result (cels/uL)']
                        absolute_basophil_result_pure = absolute_basophil_result.split('|')[0]
                        absolute_basophil_result_form_field_isntance = absolute_basophil_result.split('|')[1]
                        absolute_basophil_result_disname = absolute_basophil_result.split('|')[0]
                    except Exception as e:
                        absolute_basophil_result_pure = math.nan  
                        absolute_basophil_result_form_field_isntance   = 'This field does not have any data'                    
                        absolute_basophil_result_disname = 'Empty'

                    try:
                        absolute_eosinophil =row['Absolute eosinophil count, Out of normal range?']
                        absolute_eosinophil_pure = absolute_eosinophil.split('|')[0]
                        absolute_eosinophil_form_field_instance = absolute_eosinophil.split('|')[1]
                        absolute_eosinophil_disname = absolute_eosinophil.split('|')[2]
                    except Exception as e:
                        absolute_eosinophil_pure = math.nan
                        absolute_eosinophil_form_field_instance = 'This field does not have any data'
                        absolute_eosinophil_disname = 'Empty'

                    try:
                        absolute_eosinophil_result =row['Absolute eosinophil count, Result (cels/uL)']
                        absolute_eosinophil_result_pure = absolute_eosinophil_result.split('|')[0]
                        absolute_eosinophil_result_form_field_instance = absolute_eosinophil_result.split('|')[1]
                        absolute_eosinophil_result_disname = absolute_eosinophil_result.split('|')[0]
                    except Exception as e:
                        absolute_eosinophil_result_pure = math.nan
                        absolute_eosinophil_result_form_field_instance = 'This field does not have any data'
                        absolute_eosinophil_result_disname = 'Empty'

                    try:
                        absolute_lymphocyte = row['Absolute lymphocyte count, Out of normal range?']
                        absolute_lymphocyte_pure = absolute_lymphocyte.split('|')[0]
                        absolute_lymphocyte_form_field_instance = absolute_lymphocyte.split('|')[1]
                        absolute_lymphocyte_disname = absolute_lymphocyte.split('|')[0]
                    except Exception as e:
                        absolute_lymphocyte_pure = math.nan
                        absolute_lymphocyte_form_field_instance = 'This field does not have any data'
                        absolute_lymphocyte_disname = 'Empty'

                    try:
                        absolute_lymphocyte_result = row['Absolute lymphocyte count, Result (cels/uL)']
                        absolute_lymphocyte_result_pure = absolute_lymphocyte_result.split('|')[0]
                        absolute_lymphocyte_result_form_field_isntance = absolute_lymphocyte_result.split('|')[1]
                        absolute_lymphocyte_result_disname = absolute_lymphocyte_result.split('|')[0]
                    except Exception as e:
                        absolute_lymphocyte_result_pure  = math.nan
                        absolute_lymphocyte_result_form_field_isntance = 'This field does not have any data'
                        absolute_lymphocyte_result_disname = 'Empty'
                            
                    try:
                        absolute_monocyte = row['Absolute monocyte count, Out of normal range?']
                        absolute_monocyte_pure = absolute_monocyte.split('|')[0]
                        absolute_monocyte_form_field_instance = absolute_monocyte.split('|')[1]
                        absolute_monocyte_disname = absolute_monocyte.split('|')[2]
                    except Exception as e:
                        absolute_monocyte_pure = math.nan
                        absolute_monocyte_form_field_instance = 'This field does not have any data'
                        absolute_monocyte_disname = 'Empty'

                    try:
                        absolute_monocyte_result = row['Absolute monocyte count, Result (cels/uL)']
                        absolute_monocyte_result_pure = absolute_monocyte_result.split('|')[0]
                        absolute_monocyte_result_form_field_instance = absolute_monocyte_result.split('|')[1]
                        absolute_monocyte_result_disname = absolute_monocyte_result.split('|')[0]
                    except Exception as e:
                        absolute_monocyte_result_pure = math.nan
                        absolute_monocyte_result_form_field_instance = 'This field does not have any data'
                        absolute_monocyte_result_disname = 'Empty'

                    try:
                        Basophil_out_normal = row['Basophil, Out of normal range?']
                        Basophil_out_normal_pure = Basophil_out_normal.split('|')[0]
                        Basophil_out_normal_form_field_instance = Basophil_out_normal.split('|')[1]
                        Basophil_out_normal_disname = Basophil_out_normal.split('|')[0]
                    except Exception as e:
                        Basophil_out_normal_pure = math.nan
                        Basophil_out_normal_form_field_instance = 'This field does not have any data'
                        Basophil_out_normal_disname = 'Empty'
                    
                    try:
                        Basophil_result = row['Basophil, Result (%)']
                        Basophil_result_pure = Basophil_result.split('|')[0]
                        Basophil_result_form_field_instance = Basophil_result.split('|')[1]
                        Basophil_result_disname = Basophil_result.split('|')[0]
                    except Exception as e:
                        Basophil_result_pure = math.nan
                        Basophil_result_form_field_instance = 'This field does not have any data'
                        Basophil_result_disname = 'Empty'

                    try:
                        blood_sample_collected = row['Blood Sample Collected']
                        blood_sample_collected_pure = blood_sample_collected.split('|')[0]
                        blood_sample_collected_form_field_instance = blood_sample_collected.split('|')[1]
                        blood_sample_collected_disname = blood_sample_collected.split('|')[0]
                    except Exception as e:
                        blood_sample_collected_pure = math.nan
                        blood_sample_collected_form_field_instance = 'This field does not have any data'
                        blood_sample_collected_disname = 'Empty'

                    try:
                        date_collected = row['Date Collected']
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1]
                        date_collected_disname = date_collected.split('|')[0]
                    except Exception as e:
                        date_collected_pure = ''
                        date_collected_form_field_instance = 'This field does not have any data'
                        date_collected_disname = 'Empty'
                        
                    try:
                        Eosinophil_out_normal = row['Eosinophil, Out of normal range?']
                        Eosinophil_out_normal_pure = Eosinophil_out_normal.split('|')[0]
                        Eosinophil_out_normal_form_field_instance = Eosinophil_out_normal.split('|')[1]
                        Eosinophil_out_normal_disname = Eosinophil_out_normal.split('|')[0]
                    except Exception as e:
                        Eosinophil_out_normal_pure = math.nan
                        Eosinophil_out_normal_form_field_instance = 'This field does not have any data'
                        Eosinophil_out_normal_disname = 'Empty'

                    try:
                        Eosinophil_result = row['Eosinophil, Result (%)']
                        Eosinophil_result_pure = Eosinophil_result.split('|')[0]
                        Eosinophil_result_form_field_instance = Eosinophil_result.split('|')[1]
                        Eosinophil_result_disname = Eosinophil_result.split('|')[0]
                    except Exception as e:
                        Eosinophil_result_pure = math.nan
                        Eosinophil_result_form_field_instance = 'This field does not have any data'
                        Eosinophil_result_disname = 'Empty'

                    try:
                        Erythrocyte_out_normal = row['Erythrocyte sedimentation rate (ESR), Out of normal range?']
                        Erythrocyte_out_normal_pure = Erythrocyte_out_normal.split('|')[0]
                        Erythrocyte_out_normal_form_field_instance = Erythrocyte_out_normal.split('|')[1]
                        Erythrocyte_out_normal_disname = Erythrocyte_out_normal.split('|')[0]
                    except Exception as e:
                        Erythrocyte_out_normal_pure = math.nan
                        Erythrocyte_out_normal_form_field_instance = 'This field does not have any data'
                        Erythrocyte_out_normal_disname = 'Empty'

                    try:
                        Erythrocyte_result = row['Erythrocyte sedimentation rate (ESR), Result (mm/h)']
                        Erythrocyte_result_pure = Erythrocyte_result.split('|')[0]
                        Erythrocyte_result_form_field_instance = Erythrocyte_result.split('|')[1]
                        Erythrocyte_result_disname = Erythrocyte_result.split('|')[0]
                    except Exception as e:
                        Erythrocyte_result_pure = math.nan
                        Erythrocyte_result_form_field_instance = 'This field does not have any data'
                        Erythrocyte_result_disname = 'Empty'

                    try:
                        Haemoglobin_out_normal = row['Haemoglobin (Hgb), Out of normal range?']
                        Haemoglobin_out_normal_pure = Haemoglobin_out_normal.split('|')[0]
                        Haemoglobin_out_normal_form_field_isntance = Haemoglobin_out_normal.split('|')[1]
                        Haemoglobin_out_normal_disname = Haemoglobin_out_normal.split('|')[0]
                    except Exception as e:
                        Haemoglobin_out_normal_pure = math.nan
                        Haemoglobin_out_normal_form_field_isntance = 'This field does not have any data'
                        Haemoglobin_out_normal_disname = 'Empty'

                    try:
                        Haemoglobin_result = row['Haemoglobin (Hgb), Result (g/dL)']
                        Haemoglobin_result_pure = Haemoglobin_result.split('|')[0]
                        Haemoglobin_result_form_field_instance = Haemoglobin_result.split('|')[1]
                        Haemoglobin_result_disname = Haemoglobin_result.split('|')[0]
                    except Exception as e:
                        Haemoglobin_result_pure = math.nan
                        Haemoglobin_result_form_field_instance = 'This field does not have any data'
                        Haemoglobin_result_disname = 'Empty'

                    try:
                        Hematocrit_out_normal =  row['Hematocrit, Out of normal range?']
                        Hematocrit_out_normal_pure = Hematocrit_out_normal.split('|')[0]
                        Hematocrit_out_normal_form_field_isntance = Hematocrit_out_normal.split('|')[1]
                        Hematocrit_out_normal_disname = Hematocrit_out_normal.split('|')[0]
                    except Exception as e:
                        Hematocrit_out_normal_pure = math.nan
                        Hematocrit_out_normal_form_field_isntance = 'This field does not have any data'
                        Hematocrit_out_normal_disname = 'Empty'

                    try:
                        Hematocrit_result =  row['Hematocrit, Result (%)']
                        Hematocrit_result_pure = Hematocrit_result.split('|')[0]
                        Hematocrit_result_form_field_isntance = Hematocrit_result.split('|')[1]
                        Hematocrit_result_disname = Hematocrit_result.split('|')[0]
                    except Exception as e:
                        Hematocrit_result_pure = math.nan
                        Hematocrit_result_form_field_isntance = 'This field does not have any data'
                        Hematocrit_result_disname = 'Empty'

                    try:
                        Lymphocyte_out_normal =  row['Lymphocyte, Out of normal range?']
                        Lymphocyte_out_normal_pure = Lymphocyte_out_normal.split('|')[0]
                        Lymphocyte_out_normal_form_field_instance = Lymphocyte_out_normal.split('|')[1]
                        Lymphocyte_out_normal_disname = Lymphocyte_out_normal.split('|')[0]
                    except Exception as e:
                        Lymphocyte_out_normal_pure = math.nan
                        Lymphocyte_out_normal_form_field_instance = 'This field does not have any data'
                        Lymphocyte_out_normal_disname = 'Empty'

                    try:
                        Lymphocyte_result =  row['Lymphocyte, Result (%)']
                        Lymphocyte_result_pure = Lymphocyte_result.split('|')[0]
                        Lymphocyte_result_form_field_instance = Lymphocyte_result.split('|')[1]
                        Lymphocyte_result_disname = Lymphocyte_result.split('|')[0]
                    except Exception as e:
                        Lymphocyte_result_pure = math.nan
                        Lymphocyte_result_form_field_instance = 'This field does not have any data'
                        Lymphocyte_result_disname = 'Empty'

                    try:
                        MCH_out_normal = row['Mean Corpuscular Haemoglobin (MCH), Out of normal range?']
                        MCH_out_normal_pure = MCH_out_normal.split('|')[0]
                        MCH_out_normal_form_field_instance = MCH_out_normal.split('|')[1]
                        MCH_out_normal_disname = MCH_out_normal.split('|')[0]
                    except Exception as e:
                        MCH_out_normal_pure = math.nan
                        MCH_out_normal_form_field_instance = 'This field does not have any data'
                        MCH_out_normal_disname = 'Empty'

                    try:
                        MCH_result = row['Mean Corpuscular Haemoglobin (MCH), Result (pg)']
                        MCH_result_pure = MCH_result.split('|')[0]
                        MCH_result_form_field_instance = MCH_result.split('|')[1]
                        MCH_result_disname = MCH_result.split('|')[0]
                    except Exception as e:
                        MCH_result_pure = math.nan
                        MCH_result_form_field_instance = 'This field does not have any data'
                        MCH_result_disname = 'Empty'

                    try:
                        MCHC_out_normal = row['Mean Corpuscular Haemoglobin Concentration (MCHC), Out of normal range?']
                        MCHC_out_normal_pure = MCHC_out_normal.split('|')[0]
                        MCHC_out_normal_form_field_definition = MCHC_out_normal.split('|')[1]
                        MCHC_out_normal_disname = MCHC_out_normal.split('|')[0]
                    except Exception as e:
                        MCHC_out_normal_pure = math.nan
                        MCHC_out_normal_form_field_definition = 'This field does not have any data'
                        MCHC_out_normal_disname = 'Empty'

                    try:
                        MCHC_result = row['Mean Corpuscular Haemoglobin Concentration (MCHC), Result (g/dl)' ]
                        MCHC_result_pure = MCHC_result.split('|')[0]
                        MCHC_result_form_field_instance = MCHC_result.split('|')[1]
                        MCHC_result_disname = MCHC_result.split('|')[0]
                    except Exception as e:
                        MCHC_result_pure = math.nan
                        MCHC_result_form_field_instance = 'This field does not have any data'
                        MCHC_result_disname = 'Empty'

                    try:
                        MCV_out_normal = row['Mean Corpuscular Volume (MCV), Out of normal range?']
                        MCV_out_normal_pure = MCV_out_normal.split('|')[0]
                        MCV_out_normal_form_field_instance = MCV_out_normal.split('|')[1]
                        MCV_out_normal_disname = MCV_out_normal.split('|')[0]
                    except Exception as e:
                        MCV_out_normal_pure = math.nan
                        MCV_out_normal_form_field_instance = 'This field does not have any data'
                        MCV_out_normal_disname = 'Empty'

                    try:
                        MCV_result = row['Mean Corpuscular Volume (MCV), Result (fL)']
                        MCV_result_pure = MCV_result.split('|')[0]
                        MCV_result_form_field_instance = MCV_result.split('|')[1]
                        MCV_result_disname = MCV_result.split('|')[0]
                    except Exception as e:
                        MCV_result_pure = math.nan
                        MCV_result_form_field_instance = 'This field does not have any data'
                        MCV_result_disname = 'Empty'

                    try:
                        MPV_out_normal = row['Mean Platelet volume (MPV), Out of normal range?']
                        MPV_out_normal_pure = MPV_out_normal.split('|')[0]
                        MPV_out_normal_form_field_instance = MPV_out_normal.split('|')[1]
                        MPV_out_normal_disname = MPV_out_normal.split('|')[0]
                    except Exception as e:
                        MPV_out_normal_pure = math.nan
                        MPV_out_normal_form_field_instance = 'This field does not have any data'
                        MPV_out_normal_disname = 'Empty'
                    
                    try:
                        MPV_result = row['Mean Platelet volume (MPV), Result (fL)']
                        MPV_result_pure = MPV_result.split('|')[0]
                        MPV_result_form_field_instance = MPV_result.split('|')[1]
                        MPV_result_disname = MPV_result.split('|')[0]
                    except Exception as e:
                        MPV_result_pure = math.nan
                        MPV_result_form_field_instance = 'This field does not have any data'
                        MPV_result_disname = 'Empty'

                    try:
                        Monocytes_out_normal = row['Monocytes, Out of normal range?']
                        Monocytes_out_normal_pure = Monocytes_out_normal.split('|')[0]
                        Monocytes_out_normal_form_field_instance = Monocytes_out_normal.split('|')[1]
                        Monocytes_out_normal_disname = Monocytes_out_normal.split('|')[0]
                    except Exception as e:
                        Monocytes_out_normal_pure = math.nan
                        Monocytes_out_normal_form_field_instance = 'This field does not have any data'
                        Monocytes_out_normal_disname = 'Empty' 

                    try:
                        Monocytes_result = row['Monocytes, Result (%)']
                        Monocytes_result_pure = Monocytes_result.split('|')[0]
                        Monocytes_result_form_field_instance = Monocytes_result.split('|')[1]
                        Monocytes_result_disname = Monocytes_result.split('|')[0]
                    except Exception as e:
                        Monocytes_result_pure = math.nan
                        Monocytes_result_form_field_instance = 'This field does not have any data'
                        Monocytes_result_disname = 'Empty'

                    try:
                        Neutrophil_out_normal = row['Neutrophil, Out of normal range?']
                        Neutrophil_out_normal_pure = Neutrophil_out_normal.split('|')[0]
                        Neutrophil_out_normal_form_field_isntance = Neutrophil_out_normal.split('|')[1]
                        Neutrophil_out_normal_disname = Neutrophil_out_normal.split('|')[0]
                    except Exception as e:
                        Neutrophil_out_normal_pure = math.nan
                        Neutrophil_out_normal_form_field_isntance = 'This field does not have any data'
                        Neutrophil_out_normal_disname = 'Empty'

                    try:
                        Neutrophil_result = row['Neutrophil, Result (%)']
                        Neutrophil_result_pure = Neutrophil_result.split('|')[0]
                        Neutrophil_result_form_field_instance = Neutrophil_result.split('|')[1]
                        Neutrophil_result_disname = Neutrophil_result.split('|')[0]
                    except Exception as e:
                        Neutrophil_result_pure = math.nan
                        Neutrophil_result_form_field_instance = 'This field does not have any data'
                        Neutrophil_result_disname = 'Empty'

                    try:
                        platelet_count_out_normal = row['Platelet Count, Out of normal range?']
                        platelet_count_out_normal_pure = platelet_count_out_normal.split('|')[0]
                        platelet_count_out_normal_form_field_instance = platelet_count_out_normal.split('|')[1]
                        platelet_count_out_normal_disname = platelet_count_out_normal.split('|')[0]
                    except Exception as e:
                        platelet_count_out_normal_pure = math.nan
                        platelet_count_out_normal_form_field_instance = 'This field does not have any data'
                        platelet_count_out_normal_disname = 'Empty'

                    try:
                        platelet_count_result = row['Platelet Count, Result (x10^3 /uL)']
                        platelet_count_result_pure = platelet_count_result.split('|')[0]
                        platelet_count_result_form_field_instance = platelet_count_result.split('|')[1]
                        platelet_count_result_disname = platelet_count_result.split('|')[0]
                    except Exception as e:
                        platelet_count_result_pure = math.nan
                        platelet_count_result_form_field_instance = 'This field does not have any data'
                        platelet_count_result_disname = 'Empty'

                    try:
                        RBC_out_normal = row['Red Blood cell count (RBC), Out of normal range?']
                        RBC_out_normal_pure = RBC_out_normal.split('|')[0]
                        RBC_out_normal_form_field_instance = RBC_out_normal.split('|')[1]
                        RBC_out_normal_disname = RBC_out_normal.split('|')[0] 
                    except Exception as e:
                        RBC_out_normal_pure = math.nan       
                        RBC_out_normal_form_field_instance  = 'This field does not have any data'
                        RBC_out_normal_disname = 'Empty'

                    try:
                        RBC__result = row['Red Blood cell count (RBC), Result (mill/mm3)']
                        RBC__result_pure = RBC__result.split('|')[0]
                        RBC__result_form_field_isntance = RBC__result.split('|')[1]
                        RBC__result_disname = RBC__result.split('|')[0]
                    except Exception as e:
                        RBC__result_pure = math.nan
                        RBC__result_form_field_isntance = 'This field does not have any data'
                        RBC__result_disname = 'Empty'

                    try:
                        WBC_out_normal =row['White blood Cell count (WBC), Out of normal range?']
                        WBC_out_normal_pure = WBC_out_normal.split('|')[0]
                        WBC_out_normal_form_field_isntance = WBC_out_normal.split('|')[1]
                        WBC_out_normal_disname = WBC_out_normal.split('|')[0]
                    except Exception as e:
                        WBC_out_normal_pure = math.nan     
                        WBC_out_normal_form_field_isntance = 'This field does not have any data'
                        WBC_out_normal_disname = 'Empty'

                    try:
                        WBC_result =row['White blood Cell count (WBC), Result  (g/dL)']
                        WBC_result_pure = WBC_result.split('|')[0]
                        WBC_result_form_field_instance = WBC_result.split('|')[1]
                        WBC_result_disname = WBC_result.split('|')[0]
                    except Exception as e:
                        WBC_result_pure = math.nan
                        WBC_result_form_field_instance = 'This field does not have any data'
                        WBC_result_disname = 'Empty'
                        
                    # -----------------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)
                    
                    if date_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,f , date_collected_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')  

                    # Revision LBT0010
                    if date_collected_pure =='':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_collected_f != date_of_visit_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance, \
                                        'The date should be the same as the visit date in the "Date of Visit" Form',  \
                                            f'{date_collected_disname} - {date_of_visit}', 'LBT0010']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBT0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBT0030
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_collected_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance, \
                                        'The date/time of test performed can not be before the informed consent date/time', \
                                            f'{date_collected_disname} - {date_inform_consent}', 'LBT0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBT0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LBT0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,'Date Collected must be before the End of study/Early withdrawal date. ', date_collected_disname, 'LBT0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LBT0040 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision LBT0050
                    try:
                        if  float(blood_sample_collected_pure) == 9.0:
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and the D-1 visit date =Screening visit date or normal and done in the previous 10 days', \
                                            blood_sample_collected_disname, 'LBT0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0050--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    try:
                        # Revision LBT0060
                        if float(Erythrocyte_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Erythrocyte_result_pure) > 0.0 and float(Erythrocyte_result_pure) < 15.8:
                                    error = [subject, visit, 'Erythrocyte sedimentation rate (ESR), Out of normal range?', \
                                             Erythrocyte_result_form_field_instance, 'According to the result, the value is not out of range, please review.  (0.0 - 15.0)', \
                                                Erythrocyte_result_disname, 'LBT0060']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Erythrocyte_result_pure) > 0.0 and float(Erythrocyte_result_pure) < 20.0:
                                    error = [subject, visit, 'Erythrocyte sedimentation rate (ESR), Out of normal range?', \
                                             Erythrocyte_result_form_field_instance, \
                                                'According to the result, the value is not out of range, please review. (0.0 - 20.0)' , Erythrocyte_result_disname, 'LBT0060']
                                    lista_revision.append(error)

                        # Revision LBT0270
                        elif float(Erythrocyte_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Erythrocyte_result_pure) < 0.0 or float(Erythrocyte_result_pure) > 15.0:
                                    error = [subject, visit, 'Erythrocyte sedimentation rate (ESR), Out of normal range?', \
                                             Erythrocyte_result_form_field_instance,'According to the result, the value is out of range, please review. (0.0 - 15.0)', \
                                                Erythrocyte_result_disname, 'LBT0270']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Erythrocyte_result_pure) < 0.0 or float(Erythrocyte_result_pure) > 20.0:
                                    error = [subject, visit, 'Erythrocyte sedimentation rate (ESR), Out of normal range?', \
                                             Erythrocyte_result_form_field_instance, \
                                                'According to the result, the value is out of range, please review. (0.0 - 20.0)' , Erythrocyte_result_disname, 'LBT0270']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0060--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    try:
                        # Revision LBT0070
                        if float(WBC_out_normal_pure) == 1.0:
                            if float(WBC_result_pure) > 4.50 and float(WBC_result_pure) < 11.0:
                                error = [subject, visit, 'White blood Cell count (WBC), Out of normal range?', \
                                         WBC_result_form_field_instance,\
                                            'According to the result, the value is not out of range, please review. (4.50 - 11.0)' , WBC_result_disname, 'LBT0070']
                                lista_revision.append(error)

                        # Revision LBT0280
                        elif float(WBC_out_normal_pure) == 0.0:
                            if float(WBC_result_pure) < 4.50  or float(WBC_result_pure) > 11.0:
                                error = [subject, visit, 'White blood Cell count (WBC), Out of normal range?', \
                                         WBC_result_form_field_instance,\
                                            'According to the result, the value is out of range, please review.  (4.50 - 11.0)' , WBC_result_disname, 'LBT0280']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0070--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0080
                        if float(Neutrophil_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Neutrophil_result_pure) > 40.3 and float(Neutrophil_result_pure) < 74.8:
                                    error = [subject, visit, 'Neutrophil, Out of normal range?', Neutrophil_result_form_field_instance, \
                                             'According to the result, the value is not out of range, please review. (40.3 - 74.8)' , \
                                                Neutrophil_result_disname, 'LBT0080']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Neutrophil_result_pure) > 42.5 and float(Neutrophil_result_pure) < 73.2:
                                    error = [subject, visit, 'Neutrophil, Out of normal range?', Neutrophil_result_form_field_instance, \
                                             'According to the result, the value is not out of range, please review. (42.5 - 73.2)', \
                                                Neutrophil_result_disname, 'LBT0080']
                                    lista_revision.append(error)

                        # Revision LBT0290
                        elif float(Neutrophil_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Neutrophil_result_pure) < 40.3 or float(Neutrophil_result_pure) > 74.8:
                                    error = [subject, visit, 'Neutrophil, Out of normal range?', Neutrophil_result_form_field_instance, \
                                             'According to the result, the value is out of range, please review. (40.3 - 74.8)', \
                                                Neutrophil_result_disname, 'LBT0290']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Neutrophil_result_pure) < 42.5 or float(Neutrophil_result_pure) > 73.2:
                                    error = [subject, visit, 'Neutrophil, Out of normal range?', Neutrophil_result_form_field_instance, \
                                             'According to the result, the value is out of range, please review.( 42.5 - 73.2)',  \
                                                Neutrophil_result_disname, 'LBT0290']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0080--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0090
                        if float(Lymphocyte_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Lymphocyte_result_pure) > 12.2 and float(Lymphocyte_result_pure) < 47.1:
                                    error = [subject, visit, 'Lymphocyte, Out of normal range?', Lymphocyte_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (12.2 - 47.1)', \
                                                Lymphocyte_result_disname, 'LBT0090']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Lymphocyte_result_pure) > 18.2 and float(Lymphocyte_result_pure) < 47.4:
                                    error = [subject, visit, 'Lymphocyte, Out of normal range?', Lymphocyte_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (18.2 - 47.4)', \
                                                Lymphocyte_result_disname, 'LBT0090']
                                    lista_revision.append(error)

                        # Revision LBT0300
                        elif float(Lymphocyte_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Lymphocyte_result_pure) < 12.2 or float(Lymphocyte_result_pure) > 47.1:
                                    error = [subject, visit, 'Lymphocyte, Out of normal range?', \
                                             Lymphocyte_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (12.2 - 47.1)', \
                                                    Lymphocyte_result_disname, 'LBT0300']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Lymphocyte_result_pure) < 18.2 or float(Lymphocyte_result_pure) > 47.4:
                                    error = [subject, visit, 'Lymphocyte, Out of normal range?', \
                                             Lymphocyte_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (18.2 - 47.4)', \
                                                    Lymphocyte_result_disname, 'LBT0300']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0090--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0100
                        if float(Monocytes_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Monocytes_result_pure) > 4.4 and float(Monocytes_result_pure) < 12.3:
                                    error = [subject, visit, 'Monocytes, Out of normal range?', Monocytes_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (4.4 - 12.3)', \
                                                Monocytes_result_disname, 'LBT0100']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Monocytes_result_pure) > 4.3 and float(Monocytes_result_pure) < 11.0:
                                    error = [subject, visit, 'Monocytes, Out of normal range?', Monocytes_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (4.3 - 11.0)', \
                                                Monocytes_result_disname, 'LBT0100']
                                    lista_revision.append(error)

                        # Revision LBT0310
                        elif float(Monocytes_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Monocytes_result_pure) < 4.4 or float(Monocytes_result_pure) > 12.3:
                                    error = [subject, visit, 'Monocytes, Out of normal range?', Monocytes_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (4.4 - 12.3)', \
                                                Monocytes_result_disname, 'LBT0310']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Monocytes_result_pure) < 4.3 or float(Monocytes_result_pure) > 11.0:
                                    error = [subject, visit, 'Monocytes, Out of normal range?', Monocytes_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (4.3 - 11.0)', \
                                                Monocytes_result_disname, 'LBT0310']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0100--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0110
                        if float(Eosinophil_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Eosinophil_result_pure) > 0.0 and float(Eosinophil_result_pure) < 4.4 :
                                    error = [subject, visit, 'Eosinophil, Out of normal range?', Eosinophil_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (0.0 - 4.4)', \
                                                Eosinophil_result_disname, 'LBT0110']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Eosinophil_result_pure) > 0.0 and float(Eosinophil_result_pure) < 3.0:
                                    error = [subject, visit, 'Eosinophil, Out of normal range?', Eosinophil_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (0.0 - 3.0)', \
                                                Eosinophil_result_disname, 'LBT0110']
                                    lista_revision.append(error)
                    
                        # Revision LBT0320
                        elif float(Eosinophil_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Eosinophil_result_pure) < 0.0 or float(Eosinophil_result_pure) > 4.4 :
                                    error = [subject, visit, 'Eosinophil, Out of normal range?',Eosinophil_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (0.0 - 4.4)', \
                                                Eosinophil_result_disname, 'LBT0320']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Eosinophil_result_pure) < 0.0 or float(Eosinophil_result_pure) > 3.0:
                                    error = [subject, visit, 'Eosinophil, Out of normal range?', Eosinophil_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (0.0 - 3.0)', \
                                                Eosinophil_result_disname, 'LBT0320']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0100--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0120
                        if float(Basophil_out_normal_pure) == 1.0:
                            if float(Basophil_result_pure) > 0.0 and float(Basophil_result_pure) < 0.7:
                                error = [subject, visit, 'Basophil, Out of normal range?', Basophil_result_form_field_instance,\
                                         'According to the result, the value is not out of range, please review. (0.0 - 0.7)', Basophil_result_disname, 'LBT0120']
                                lista_revision.append(error)

                        # Revision LBT0330
                        elif float(Basophil_out_normal_pure) == 0.0:
                            if float(Basophil_result_pure) < 0.0 or float(Basophil_result_pure) > 0.7:
                                error = [subject, visit, 'Basophil, Out of normal range?', Basophil_result_form_field_instance,\
                                         'According to the result, the value is out of range, please review. (0.0 - 0.7)', Basophil_result_disname, 'LBT0330']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0120--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0130
                        if float(absolute_Neutrophil_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(absolute_Neutrophil_result_pure) > 1.82 and float(absolute_Neutrophil_result_pure) < 7.42 :
                                    error = [subject, visit, 'Absolute Neutrophil count, Out of normal range? ', \
                                             absolute_Neutrophil_result_form_field_instance, \
                                                'According to the result, the value is not out of range, please review. (1820.0 - 7150.0)', \
                                                    absolute_Neutrophil_result_disname, 'LBT0130']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_Neutrophil_result_pure) > 2.0 and float(absolute_Neutrophil_result_pure) <  7.15:
                                    error = [subject, visit, 'Absolute Neutrophil count, Out of normal range?', \
                                             absolute_Neutrophil_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (2000.0 - 7150.0)', \
                                                    absolute_Neutrophil_result_disname, 'LBT0130']
                                    lista_revision.append(error)

                        # Revision LBT0340
                        elif float(absolute_Neutrophil_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(absolute_Neutrophil_result_pure) < 1.82 or float(absolute_Neutrophil_result_pure) > 7.42 :
                                    error = [subject, visit, 'Absolute Neutrophil count, Out of normal range?',  \
                                             absolute_Neutrophil_result_form_field_instance, \
                                                'According to the result, the value is out of range, please review. (1820.0 - 7150.0)', \
                                                    absolute_Neutrophil_result_disname, 'LBT0340']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_Neutrophil_result_pure) < 2.0 or float(absolute_Neutrophil_result_pure) > 7.15:
                                    error = [subject, visit, 'Absolute Neutrophil count, Out of normal range?', \
                                             absolute_Neutrophil_result_form_field_instance, \
                                                'According to the result, the value is out of range, please review. (2000.0 - 7150.0)', \
                                                    absolute_Neutrophil_result_disname, 'LBT0340']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0130--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0140
                        if float(absolute_lymphocyte_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(absolute_lymphocyte_result_pure) > 0.85 and float(absolute_lymphocyte_result_pure) < 3.00 :
                                    error = [subject, visit, 'Absolute lymphocyte count, Out of normal range?', \
                                             absolute_lymphocyte_result_form_field_isntance,\
                                                'According to the result, the value is not out of range, please review. (0.85 - 3.00)', \
                                                    absolute_lymphocyte_result_disname, 'LBT0140']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_lymphocyte_result_pure) > 1.16 and float(absolute_lymphocyte_result_pure) < 3.18:
                                    error = [subject, visit, 'Absolute lymphocyte count, Out of normal range?', \
                                             absolute_lymphocyte_result_form_field_isntance,\
                                                'According to the result, the value is not out of range, please review. (1.16  - 3.18)', \
                                                    absolute_lymphocyte_result_disname, 'LBT0140']
                                    lista_revision.append(error)

                        # Revision LBT0350
                        elif float(absolute_lymphocyte_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(absolute_lymphocyte_result_pure) < 0.85 or float(absolute_lymphocyte_result_pure) > 3.00 :
                                    error = [subject, visit, 'Absolute lymphocyte count, Out of normal range?', \
                                             absolute_lymphocyte_result_form_field_isntance, \
                                                'According to the result, the value is out of range, please review. (0.85 - 3.00)', \
                                                    absolute_lymphocyte_result_disname, 'LBT0350']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_lymphocyte_result_pure) < 1.16 or float(absolute_lymphocyte_result_pure) > 3.18:
                                    error = [subject, visit, 'Absolute lymphocyte count, Out of normal range?', \
                                             absolute_lymphocyte_result_form_field_isntance, \
                                                'According to the result, the value is out of range, please review. (1.16  - 3.18)', \
                                                    absolute_lymphocyte_result_disname, 'LBT0350']
                                    lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBT0140--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0150
                        if float(absolute_monocyte_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(absolute_monocyte_result_pure) > 0.19 and float(absolute_monocyte_result_pure) < 0.77 :
                                    error = [subject, visit, 'Absolute monocyte count, Out of normal range?', \
                                             absolute_monocyte_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (0.19 - 0.77)', \
                                                    absolute_monocyte_result_disname, 'LBT0150']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_monocyte_result_pure) > 0.29 and float(absolute_monocyte_result_pure) < 0.71:
                                    error = [subject, visit, 'Absolute monocyte count, Out of normal range?', \
                                             absolute_monocyte_result_form_field_instance, \
                                                'According to the result, the value is not out of range, please review. (0.29 - 0.71)', \
                                                    absolute_monocyte_result_disname, 'LBT0150']
                                    lista_revision.append(error)

                        # Revision LBT0360
                        elif float(absolute_monocyte_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(absolute_monocyte_result_pure) < 0.19 or float(absolute_monocyte_result_pure) > 0.77 :
                                    error = [subject, visit, 'Absolute monocyte count, Out of normal range?', \
                                             absolute_monocyte_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (0.19 - 0.77)', \
                                                    absolute_monocyte_result_disname, 'LBT0360']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_monocyte_result_pure) < 0.29 or float(absolute_monocyte_result_pure) > 0.71:
                                    error = [subject, visit, 'Absolute monocyte count, Out of normal range?', \
                                             absolute_monocyte_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (0.29 - 0.71)', \
                                                    absolute_monocyte_result_disname, 'LBT0360']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0150--> {e} - Subject: {subject},  Visit: {visit} ')
 
                    try:
                        # Revision LBT0160 
                        if float(absolute_eosinophil_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(absolute_eosinophil_result_pure) > 0.03 and float(absolute_eosinophil_result_pure) < 0.44 :
                                    error = [subject, visit, 'Absolute eosinophil count, Out of normal range?',\
                                             absolute_eosinophil_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (0.03 - 0.44)', \
                                                    absolute_eosinophil_result_disname, 'LBT0160']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_eosinophil_result_pure) > 0.03 and float(absolute_eosinophil_result_pure) < 0.27:
                                    error = [subject, visit, 'Absolute eosinophil count, Out of normal range?'.\
                                             absolute_eosinophil_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (0.03 - 0.27 )', \
                                                    absolute_eosinophil_result_disname, 'LBT0160']
                                    lista_revision.append(error)

                        # Revision LBT0370
                        elif float(absolute_eosinophil_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(absolute_eosinophil_result_pure) < 0.03 or float(absolute_eosinophil_result_pure) >  0.44:
                                    error = [subject, visit, 'Absolute eosinophil count, Out of normal range?',\
                                             absolute_eosinophil_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (0.03 - 0.44)', \
                                                    absolute_eosinophil_result_disname, 'LBT0370']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(absolute_eosinophil_result_pure) < 0.03 or float(absolute_eosinophil_result_pure) > 0.27:
                                    error = [subject, visit, 'Absolute eosinophil count, Out of normal range?', \
                                             absolute_eosinophil_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (0.03 - 0.27 )', \
                                                    absolute_eosinophil_result_disname, 'LBT0370']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0160--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0170
                        if float(absolute_basophil_pure) == 1.0:
                            if float(absolute_basophil_result_pure) > 0.01 and float(absolute_basophil_result_pure) < 0.05:
                                error = [subject, visit, 'Absolute basophil count, Out of normal range?',\
                                         absolute_basophil_result_form_field_isntance,\
                                            'According to the result, the value is out of range, please review. (0.01 - 0.05 )',\
                                                  absolute_basophil_result_disname, 'LBT0170']
                                lista_revision.append(error)

                        # Revision LBT0380
                        elif float(absolute_basophil_pure) == 0.0:
                            if float(absolute_basophil_result_pure) < 0.01 or float(absolute_basophil_result_pure) > 0.05:
                                error = [subject, visit, 'Absolute basophil count, Out of normal range?', \
                                         absolute_basophil_result_form_field_isntance,\
                                            'According to the result, the value is out of range, please review. (0.01 - 0.05 )', \
                                                absolute_basophil_result_disname, 'LBT0380']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0170--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0180
                        if float(RBC_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(RBC__result_pure) > 4.6 and float(RBC__result_pure) < 6.2 :
                                    error = [subject, visit, 'Red Blood cell count (RBC), Out of normal range?',\
                                             RBC__result_form_field_isntance,\
                                                'According to the result, the value is not out of range, please review. (4.6 - 6.2)', RBC__result_disname, 'LBT0180']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(RBC__result_pure) > 4.2 and float(RBC__result_pure) < 5.4:
                                    error = [subject, visit, 'Red Blood cell count (RBC), Out of normal range?',\
                                              RBC__result_form_field_isntance,\
                                                'According to the result, the value is not out of range, please review. (4.2 - 5.4)' , RBC__result_disname, 'LBT0180']
                                    lista_revision.append(error)

                        # Revision LBT0390
                        elif float(RBC_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(RBC__result_pure) < 4.6 or float(RBC__result_pure) > 6.2 :
                                    error = [subject, visit, 'Red Blood cell count (RBC), Out of normal range?',\
                                             RBC__result_form_field_isntance,\
                                                'According to the result, the value is out of range, please review. (4.6 - 6.2)', RBC__result_disname, 'LBT0390']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(RBC__result_pure) < 4.2 or float(RBC__result_pure) > 5.4:
                                    error = [subject, visit, 'Red Blood cell count (RBC), Out of normal range?',\
                                             RBC__result_form_field_isntance,\
                                                'According to the result, the value is out of range, please review. (4.2 - 5.4)' , RBC__result_disname, 'LBT0390']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0180--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0190 
                        if float(Haemoglobin_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Haemoglobin_result_pure) > 13.5 and float(Haemoglobin_result_pure) < 18.0 :
                                    error = [subject, visit, 'Haemoglobin (Hgb), Out of normal range?', \
                                             Haemoglobin_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (13.5 - 18.0)', \
                                                    Haemoglobin_result_disname, 'LBT0190']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Haemoglobin_result_pure) > 12.0 and float(Haemoglobin_result_pure) < 16.0:
                                    error = [subject, visit, 'Haemoglobin (Hgb), Out of normal range?', \
                                             Haemoglobin_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (12.0 - 16.0 )', \
                                                    Haemoglobin_result_disname, 'LBT0190']
                                    lista_revision.append(error)

                        # Revision LBT0400
                        elif float(Haemoglobin_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Haemoglobin_result_pure) < 13.5 or float(Haemoglobin_result_pure) > 18.0 :
                                    error = [subject, visit, 'Haemoglobin (Hgb), Out of normal range?', \
                                             Haemoglobin_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (13.5 - 18.0 )', \
                                                Haemoglobin_result_disname, 'LBT0400']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Haemoglobin_result_pure) < 12.0 or float(Haemoglobin_result_pure) > 16.0:
                                    error = [subject, visit, 'Haemoglobin (Hgb), Out of normal range?', \
                                             Haemoglobin_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (12.0 - 16.0)', 
                                                Haemoglobin_result_disname, 'LBT0400']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0190--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0200 
                        if float(Hematocrit_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Hematocrit_result_pure) > 40.0 and float(Hematocrit_result_pure) < 54.0 :
                                    error = [subject, visit, 'Hematocrit, Out of normal range?',\
                                             Hematocrit_result_form_field_isntance,\
                                                'According to the result, the value is not out of range, please review. (40.0 - 54.0)', \
                                                    Hematocrit_result_disname, 'LBT0200']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Hematocrit_result_pure) > 38.0 and float(Hematocrit_result_pure) < 48.0:
                                    error = [subject, visit, 'Hematocrit, Out of normal range?', \
                                             Hematocrit_result_form_field_isntance,\
                                                'According to the result, the value is not out of range, please review. (38.0 - 48.0)', \
                                                    Hematocrit_result_disname, 'LBT0200']
                                    lista_revision.append(error)

                        # Revision LBT0410
                        elif float(Hematocrit_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Hematocrit_result_pure) < 40.0 or float(Hematocrit_result_pure) > 54.0 :
                                    error = [subject, visit, 'Hematocrit, Out of normal range?', \
                                             Hematocrit_result_form_field_isntance,\
                                                'According to the result, the value is out of range, please review. (40.0 - 54.0)', \
                                                    Hematocrit_result_disname, 'LBT0410']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Hematocrit_result_pure) < 38.0 or float(Hematocrit_result_pure) > 48.0:
                                    error = [subject, visit, 'Hematocrit, Out of normal range?', \
                                             Hematocrit_result_form_field_isntance,\
                                                'According to the result, the value is out of range, please review. (38.0 - 48.0 )', \
                                                    Hematocrit_result_disname, 'LBT0410']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0200--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0210 
                        if float(MCV_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(MCV_result_pure) > 86.0 and float(MCV_result_pure) < 96.0 :
                                    error = [subject, visit, 'Mean Corpuscular Volume (MCV), Out of normal range?', \
                                             MCV_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (86.0 - 96.0)', \
                                                    MCV_result_disname, 'LBT0210']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MCV_result_pure) > 86.0 and float(MCV_result_pure) < 96.0:
                                    error = [subject, visit, 'Mean Corpuscular Volume (MCV), Out of normal range?',\
                                             MCV_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (86.0 - 96.0)', \
                                                    MCV_result_disname, 'LBT0210']
                                    lista_revision.append(error)

                        # Revision LBT0420
                        elif float(MCV_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(MCV_result_pure) < 86.0 or float(MCV_result_pure) > 96.0 :
                                    error = [subject, visit, 'Mean Corpuscular Volume (MCV), Out of normal range?', \
                                             MCV_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (86.0 - 96.0 )', \
                                                    MCV_result_disname, 'LBT0420']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MCV_result_pure) < 86.0 or float(MCV_result_pure) > 96.0:
                                    error = [subject, visit, 'Mean Corpuscular Volume (MCV), Out of normal range?',\
                                             MCV_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (86.0 - 96.0)' , 
                                                MCV_result_form_field_instance, 'LBT0420']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0210--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0220 
                        if float(MCH_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(MCH_result_pure) > 25.0 and float(MCH_result_pure) < 31.0 :
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin (MCH), Out of normal range?',
                                             MCH_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (86.0 - 96.0)', \
                                                    MCH_result_disname, 'LBT0220']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MCH_result_pure) > 25.0 and float(MCH_result_pure) < 31.0:
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin (MCH), Out of normal range? ',\
                                             MCH_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (86.0 - 96.0)', \
                                                    MCH_result_disname, 'LBT0220']
                                    lista_revision.append(error)

                        # Revision LBT0430
                        elif float(MCH_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(MCH_result_pure) < 25.0 or float(MCH_result_pure) > 31.0 :
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin (MCH), Out of normal range? ',\
                                             MCH_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (86.0 - 96.0)', \
                                                    MCH_result_disname, 'LBT0430']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MCH_result_pure) < 25.0 or float(MCH_result_pure) > 31.0:
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin (MCH), Out of normal range? ', \
                                             MCH_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (86.0 - 96.0)' , \
                                                    MCH_result_disname, 'LBT0430']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0220--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0230
                        if float(MCHC_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(MCHC_result_pure) > 32.0 and float(MCHC_result_pure) < 38.0 :
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin Concentration (MCHC), Out of normal range? ',\
                                             MCHC_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (32.0 - 38.0 )', \
                                                    MCHC_result_disname, 'LBT0230']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MCHC_result_pure) > 32.0 and float(MCHC_result_pure) < 38.0:
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin Concentration (MCHC), Out of normal range?',\
                                             MCHC_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (32.0 - 38.0)', \
                                                    MCHC_result_disname, 'LBT0230']
                                    lista_revision.append(error)

                        # Revision LBT0440
                        elif float(MCHC_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(MCHC_result_pure) < 32.0 or float(MCHC_result_pure) > 38.0 :
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin Concentration (MCHC), Out of normal range? ',\
                                             MCHC_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (32.0 - 38.0)', \
                                                    MCHC_result_disname, 'LBT0440']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MCHC_result_pure) < 32.0 or float(MCHC_result_pure) > 38.0:
                                    error = [subject, visit, 'Mean Corpuscular Haemoglobin Concentration (MCHC), Out of normal range?',\
                                             MCHC_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (32.0 - 38.0)',\
                                                      MCHC_result_disname, 'LBT0440']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0230--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0240 
                        if float(platelet_count_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(platelet_count_result_pure) > 150.0 and float(platelet_count_result_pure) < 400.0:
                                    error = [subject, visit, 'Platelet Count, Out of normal range? ', \
                                             platelet_count_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (150.0 - 400.0)',\
                                                      platelet_count_result_disname, 'LBT0240']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(platelet_count_result_pure) > 186.0 and float(platelet_count_result_pure) < 353.0:
                                    error = [subject, visit, 'Platelet Count, Out of normal range? ', \
                                             platelet_count_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (186.0 - 353.0)', \
                                                    platelet_count_result_disname, 'LBT0240']
                                    lista_revision.append(error)

                        # Revision LBT0450 
                        elif float(platelet_count_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(platelet_count_result_pure) < 150.0 or float(platelet_count_result_pure) > 400.0:
                                    error = [subject, visit, 'Platelet Count, Out of normal range? ', \
                                             platelet_count_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (150.0 - 400.0)', \
                                                    platelet_count_result_disname, 'LBT0450']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(platelet_count_result_pure) < 186.0 or float(platelet_count_result_pure) > 353.0:
                                    error = [subject, visit, 'Platelet Count, Out of normal range? ', \
                                             platelet_count_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (186.0 - 353.0)', \
                                                    platelet_count_result_disname, 'LBT0450']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0240--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBT0250
                        if float(MPV_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(MPV_result_pure) > 9.7 and float(MPV_result_pure) < 11.9:
                                    error = [subject, visit, 'Mean Platelet volume (MPV), Out of normal range? ',\
                                             MPV_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (9.7 - 11.9)', \
                                                    MPV_result_disname, 'LBT0250']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MPV_result_pure) > 9.6 and float(MPV_result_pure) < 12.0:
                                    error = [subject, visit, 'Mean Platelet volume (MPV), Out of normal range? ', \
                                             MPV_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (9.6 - 12.0)', \
                                                    MPV_result_disname, 'LBT0250']
                                    lista_revision.append(error)

                        # Revision LBT0460
                        elif float(MPV_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(MPV_result_pure) < 9.7 or float(MPV_result_pure) > 11.9:
                                    error = [subject, visit, 'Mean Platelet volume (MPV), Out of normal range? ',\
                                             MPV_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (9.7 - 11.9)', \
                                                    MPV_result_disname, 'LBT0460']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(MPV_result_pure) < 9.6 or float(MPV_result_pure) > 12.0:
                                    error = [subject, visit, 'Mean Platelet volume (MPV), Out of normal range? ',\
                                             MPV_result_form_field_instance,\
                                                'According to the result, the value is out of range, please review. (9.6 - 12.0)',\
                                                      MPV_result_disname, 'LBT0460']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0250--> {e} - Subject: {subject},  Visit: {visit} ')

                    lista_validacion = ['Erythrocyte sedimentation rate (ESR)',
                        'White blood Cell count (WBC)',
                        'Neutrophil',
                        'Lymphocyte',
                        'Monocytes\n',
                        'Eosinophil\n',
                        'Basophil\n',
                        'Absolute Neutrophil count',
                        'Absolute lymphocyte count\n',
                        'Absolute Monocyte count',
                        'Absolute monocyte count, Result (cels/uL)',
                        'Absolute monocyte count, Out of normal range?',
                        'Absolute monocyte count, If abnormal, Specify',
                        'Absolute eosinophil count',
                        'Absolute eosinophil count, Result (cels/uL)',
                        'Absolute eosinophil count, Out of normal range?',
                        'Absolute eosinophil count, If abnormal, Specify',
                        'Absolute basophil count',
                        'Absolute basophil count, Result (cels/uL)',
                        'Absolute basophil count, Out of normal range?',
                        'Absolute basophil count, If abnormal, Specify',
                        'Red Blood cell count (RBC)\n',
                        'Red Blood cell count (RBC), Result (mill/mm3)',
                        'Red Blood cell count (RBC), Out of normal range?',
                        'Red Blood cell count (RBC), If abnormal, Specify',
                        'Haemoglobin (Hgb)\n',
                        'Haemoglobin (Hgb), Result (g/dL)',
                        'Haemoglobin (Hgb), Out of normal range?',
                        'Haemoglobin (Hgb), If abnormal, Specify',
                        'Hematocrit\n',
                        'Hematocrit, Result (%)',
                        'Hematocrit, Out of normal range?',
                        'Hematocrit, If abnormal, Specify',
                        'Mean Corpuscular Volume (MCV)',
                        'Mean Corpuscular Volume (MCV), Result (fL)',
                        'Mean Corpuscular Volume (MCV), Out of normal range?',
                        'Mean Corpuscular Volume (MCV), If abnormal, Specify',
                        'Mean Corpuscular Haemoglobin (MCH)',
                        'Mean Corpuscular Haemoglobin (MCH), Result (pg)',
                        'Mean Corpuscular Haemoglobin (MCH), Out of normal range?',
                        'Mean Corpuscular Haemoglobin (MCH), If abnormal, Specify',
                        'Mean Corpuscular Haemoglobin Concentration (MCHC)',
                        'Mean Corpuscular Haemoglobin Concentration (MCHC), Result (g/dl)',
                        'Mean Corpuscular Haemoglobin Concentration (MCHC), Out of normal range?',
                        'Mean Corpuscular Haemoglobin Concentration (MCHC), If abnormal, Specify',
                        'Platelet Count',
                        'Platelet Count, Result (x10^3 /uL)',
                        'Platelet Count, Out of normal range?',
                        'Platelet Count, If abnormal, Specify',
                        'Mean Platelet volume (MPV)',
                        'Mean Platelet volume (MPV), Result (fL)',
                        'Mean Platelet volume (MPV), Out of normal range?',
                        'Mean Platelet volume (MPV), If abnormal, Specify']
                    
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try: 
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
                        
                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass

                    # Revision LBT0260
                    try:
                        if float(blood_sample_collected_pure) ==1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', \
                                         blood_sample_collected_form_field_instance,\
                                            'If Blood Sample Collected is checked as "Yes", not all laboratory tests can be "not done"', \
                                                blood_sample_collected_disname, 'LBT0260']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBT0260--> {e} - Subject: {subject},  Visit: {visit} ')

    excel_writer = load_workbook(path_excel_writer)
    column_names =  ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    clinical_laboratory_test_hematology_output = pd.DataFrame(lista_revision, columns=column_names)
    clinical_laboratory_test_hematology_output = clinical_laboratory_test_hematology_output[~clinical_laboratory_test_hematology_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]

    sheet = excel_writer.create_sheet("CL - Hematology")

    for row in dataframe_to_rows(clinical_laboratory_test_hematology_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return clinical_laboratory_test_hematology_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)
if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    clinical_laboratory_test_hematology(df_root, path_excel) 