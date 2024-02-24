import numpy as np
import math
import pandas as pd
from datetime import datetime
from log_writer import log_writer
from revision_fechas import revision_fecha
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def covid_19_testing(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Covid 19 testing 
    '''

    df= df_root[df_root['name']=='Covid 19 testing']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Covid 19 testing']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            # print(pru)
            # print('------------------')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                
                if status != '':

                    try:
                        was_antigen_performed = row['Was the SARS-CoV-2 antigen test performed?']
                        was_antigen_performed_pure = was_antigen_performed.split('|')[0]
                        was_antigen_performed_form_field_instance = was_antigen_performed.split('|')[1]
                        was_antigen_performed_disname = was_antigen_performed.split('|')[2]
                    except Exception as e:
                        was_antigen_performed_pure = math.nan
                        was_antigen_performed_form_field_instance = 'This field does not have any data'
                        was_antigen_performed_disname = 'Empty'

                    try:
                        provide_reason = row['Provide the reason']
                        provide_reason_pure = provide_reason.split('|')[0]
                        provide_reason_form_field_instance = provide_reason.split('|')[1]
                        provide_reason_disname = provide_reason.split('|')[0]
                    except Exception as e:
                        provide_reason_pure = math.nan
                        provide_reason_form_field_instance = 'This field does not have any data'
                        provide_reason_disname = 'Empty'

                    try:
                        date_test_performed = row['Date of test performed']
                        date_test_performed_pure = date_test_performed.split('|')[0]
                        date_test_performed_form_field_instance = date_test_performed.split('|')[1]
                        date_test_performed_disname = date_test_performed.split('|')[0] 
                    except Exception as e:
                        date_test_performed_pure = ''
                        date_test_performed_form_field_instance = 'This field does not have any data'
                        date_test_performed_disname = 'Empty'

                    try:
                        result = row['Result']
                        result_pure = result.split('|')[0]
                        result_form_field_instance = result.split('|')[1]
                        result_disname = result.split('|')[0]
                    except Exception as e:
                        result_pure = math.nan
                        result_form_field_instance = 'This field does not have any data'
                        result_disname = 'Empty'

                    # ------------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_test_performed_pure == '' :
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_test_performed_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance ,f , date_test_performed_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBCOV0010 
                    try:
                        if float(was_antigen_performed_pure) == 9.0: 
                            if visita =='D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the SARS-CoV-2 antigen test performed?', was_antigen_performed_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and Screening visit date = D-1 date (screening done on D-1)' ,\
                                              was_antigen_performed_disname, 'LBCOV0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBCOV0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBCOV0030
                    if date_test_performed_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_test_performed_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance, \
                                        'The date should be the same as the visit date in the "Date of Visit" Form' ,\
                                            f'{date_test_performed_disname} - {date_of_visit}', 'LBCOV0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBCOV0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBCOV0040
                    if  date_test_performed_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_test_performed_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance,\
                                        'Date of test performed should be after the date of informed consent.', \
                                            f'{date_test_performed_disname} - {date_inform_consent}', 'LBCOV0040']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBCOV0040--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LBCOV0050
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_test_performed_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_test_performed_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance ,'Date of test performed must be before the End of study/Early withdrawal date. ', date_test_performed_disname, 'LBCOV0050']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LBCOV0050 --> {e} - Subject: {subject},  Visit: {visit}  ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    covid_19_testing_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("Covid 19 testing")

    for row in dataframe_to_rows(covid_19_testing_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return covid_19_testing_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
     path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
     df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
     covid_19_testing(df_root, path_excel ) 
