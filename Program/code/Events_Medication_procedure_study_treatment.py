import numpy as np
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def ev_med_proce_treatment(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Ev, Med, Proc And Study Treatment Summary
    '''

    df = df_root[df_root['name']== 'Ev, Med, Proc And Study Treatment Summary'] 
    lista_sujetos = df['Participante'].unique()

    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)
    
    df_adverse = df_root[df_root['name']=='Adverse Events']
    df_adverse = df_adverse[['Visit','Participante', 'Campo', 'Valor']]
    df_adverse = df_adverse[df_adverse['Campo']== 'Adverse Event ID']
    df_adverse = df_adverse[['Participante', 'Valor']]
    df_adverse = df_adverse.rename(columns={'Participante':'Subject', 'Valor':'adverse_event_id'})
    
    df_prior_medication = df_root[df_root['name']== 'Prior And Concomitant Medications']
    df_prior_medication = df_prior_medication[['Visit','Participante', 'Campo', 'Valor']]
    df_prior_medication = df_prior_medication[df_prior_medication['Campo']== 'Concomitant Medication ID']
    df_prior_medication = df_prior_medication[['Participante' ,'Valor']]
    df_prior_medication = df_prior_medication.rename(columns={'Participante':'Subject', 'Valor':'prior_medication'})
    
    df_prior_procedure = df_root[df_root['name']== 'Prior And Concomitant Procedures']
    df_prior_procedure = df_prior_procedure[['Visit','Participante', 'Campo', 'Valor']]
    df_prior_procedure = df_prior_procedure[df_prior_procedure['Campo']== 'Procedure ID']
    df_prior_procedure = df_prior_procedure[['Participante' ,'Valor']]
    df_prior_procedure = df_prior_procedure.rename(columns={'Participante':'Subject', 'Valor': 'prior_procedure'})

    df_CpG_administration = df_root[df_root['name']== 'CpG ODN D35 Administration']
    df_CpG_administration = df_CpG_administration[['Visit','Participante', 'Campo', 'Valor']]
    df_CpG_administration = df_CpG_administration[df_CpG_administration['Campo']== 'Date of dosing']
    df_CpG_administration = df_CpG_administration[['Participante' ,'Valor']]
    df_CpG_administration = df_CpG_administration.rename(columns={'Participante':'Subject', 'Valor': 'date_cpg_adminsitration'})

    df_miltefosine_administration = df_root[df_root['name']== 'Miltefosine Administration']
    df_miltefosine_administration = df_miltefosine_administration[['Visit','Participante', 'Campo', 'Valor']]
    df_miltefosine_administration = df_miltefosine_administration[df_miltefosine_administration['Campo']== 'Date of dosing']
    df_miltefosine_administration = df_miltefosine_administration[['Participante' ,'Valor']]
    df_miltefosine_administration = df_miltefosine_administration.rename(columns={'Participante':'Subject', 'Valor': 'date_miltefosine_adminsitration'})

    lista_revision = []
    lista_logs = ['Ev, Med, Proc And Study Treatment Summary']

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']== sujeto]
        sujeto_principal = sujeto_principal.sort_values(by=['FormFieldInstance Id'], ascending=True)
        sujeto_principal = sujeto_principal.reset_index(drop=True)

        # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, por lo que usamos el siguiente codigo para realizar la particion
        date_indices = sujeto_principal.index[sujeto_principal['Campo'] == 'Were any adverse events experienced since Informed Consent?'].tolist()
        subdatasets = [sujeto_principal.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]

        for subdataset in subdatasets:

            pru = subdataset
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = 'unscheduled'
            pru['status'] = subdataset['activityState'].unique()
            pru = pru.merge(df_adverse, on=['Subject'], how='left')
            pru = pru.merge(df_prior_medication, on=['Subject'], how='left')
            pru = pru.merge(df_prior_procedure, on=['Subject'], how='left')
            pru = pru.merge(df_CpG_administration, on=['Subject'], how='left')
            pru = pru.merge(df_miltefosine_administration, on=['Subject'], how='left')

            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                adverse_event_id = row['adverse_event_id']
                prior_medication_id = row['prior_medication']
                prior_procedure_id = row['prior_procedure']
                date_cpg_administration = row['date_cpg_adminsitration']
                date_miltefosine_administration = row['date_miltefosine_adminsitration']

                if status != '':
                    try:
                        were_AE_experienced_since_informed_consent = row['Were any adverse events experienced since Informed Consent?']
                        were_AE_experienced_since_informed_consent_pure = were_AE_experienced_since_informed_consent.split('|')[0]
                        were_AE_experienced_since_informed_consent_form_field_instance = were_AE_experienced_since_informed_consent.split('|')[1]
                        were_AE_experienced_since_informed_consent_disname = were_AE_experienced_since_informed_consent.split('|')[2]
                    except:
                        were_AE_experienced_since_informed_consent_pure = ''
                        were_AE_experienced_since_informed_consent_form_field_instance = 'This field does not have any data'
                        were_AE_experienced_since_informed_consent_disname = 'Empty'

                    try:
                        were_any_concomitant_medication_8_weeks_before = row['Were any concomitant medications taken within 8 weeks before start of screening or during the study?']
                        were_any_concomitant_medication_8_weeks_before_pure = were_any_concomitant_medication_8_weeks_before.split('|')[0]
                        were_any_concomitant_medication_8_weeks_before_form_field_instance = were_any_concomitant_medication_8_weeks_before.split('|')[1]
                        were_any_concomitant_medication_8_weeks_before_disname = were_any_concomitant_medication_8_weeks_before.split('|')[2]
                    except:
                        were_any_concomitant_medication_8_weeks_before_pure = ''
                        were_any_concomitant_medication_8_weeks_before_form_field_instance = 'This field does not have any data'
                        were_any_concomitant_medication_8_weeks_before_disname = 'Empty'
                    
                    try:
                        were_concomitant_procedure_performed_during_study = row['Were any concomitant procedures/surgeries performed during the study?']
                        were_concomitant_procedure_performed_during_study_pure = were_concomitant_procedure_performed_during_study.split('|')[0]
                        were_concomitant_procedure_performed_during_study_form_field_instance = were_concomitant_procedure_performed_during_study.split('|')[1]
                        were_concomitant_procedure_performed_during_study_disname = were_concomitant_procedure_performed_during_study.split('|')[2]
                    except:
                        were_concomitant_procedure_performed_during_study_pure = ''
                        were_concomitant_procedure_performed_during_study_form_field_instance = 'This field does not have any data'
                        were_concomitant_procedure_performed_during_study_disname = 'Empty'

                    try:
                        has_subject_taken_cpg_ODN = row['Has the subject taken at least one dose of CpG ODN D35 study treatment?']
                        has_subject_taken_cpg_ODN_pure = has_subject_taken_cpg_ODN.split('|')[0]
                        has_subject_taken_cpg_ODN_form_field_instance = has_subject_taken_cpg_ODN.split('|')[1]
                        has_subject_taken_cpg_ODN_disname = has_subject_taken_cpg_ODN.split('|')[2]
                    except:
                        has_subject_taken_cpg_ODN_pure = ''
                        has_subject_taken_cpg_ODN_form_field_instance = 'This field does not have any data'
                        has_subject_taken_cpg_ODN_disname = 'Empty'
                    
                    try:
                        has_subject_taken_miltefosine = row['Has the subject taken at least one dose of Miltefosine study treatment?']
                        has_subject_taken_miltefosine_pure = has_subject_taken_miltefosine.split('|')[0]
                        has_subject_taken_miltefosine_form_field_instance = has_subject_taken_miltefosine.split('|')[1]
                        has_subject_taken_miltefosine_disname = has_subject_taken_miltefosine.split('|')[2]
                    except:
                        has_subject_taken_miltefosine_pure = ''
                        has_subject_taken_miltefosine_form_field_instance = 'This field does not have any data'
                        has_subject_taken_miltefosine_disname = 'Empty'

                    # -----------------------------------------------------------------------------------------------------------------------------------
                    # Revision SM0010 
                    try:
                        if float(were_AE_experienced_since_informed_consent_pure) == 1.0:
                            if float(adverse_event_id) == np.nan or str(adverse_event_id) == 'nan' or str(adverse_event_id) == '' or str(adverse_event_id) == '-':

                                error =  [subject, visit, 'Were any adverse events experienced since Informed Consent?', were_AE_experienced_since_informed_consent_form_field_instance, \
                                            'If Yes, at least one adverse event form must be completed', \
                                                were_AE_experienced_since_informed_consent_disname, 'SM0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0010 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0020 
                    try:
                        if float(were_AE_experienced_since_informed_consent_pure) == 0.0:
                            if float(adverse_event_id) == np.nan or str(adverse_event_id) == 'nan' or str(adverse_event_id) == '' or str(adverse_event_id) == '-':
                                pass
                            else:
                                error =  [subject, visit, 'Were any adverse events experienced since Informed Consent?', were_AE_experienced_since_informed_consent_form_field_instance, \
                                            'If No, no adverse event forms should be completed', \
                                                were_AE_experienced_since_informed_consent_disname, 'SM0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision SM0030
                    try:
                        if float(were_any_concomitant_medication_8_weeks_before_pure) == 1.0:
                            if float(prior_medication_id) == np.nan or str(prior_medication_id) == 'nan' or str(prior_medication_id) == '' or str(prior_medication_id) == '-':
                                error =  [subject, visit, 'Were any concomitant medications taken within 8 weeks before start of screening or during the study?', \
                                          were_any_concomitant_medication_8_weeks_before_form_field_instance, 'If Yes, at least one adverse event form must be completed', \
                                                were_any_concomitant_medication_8_weeks_before_disname, 'SM0030']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0030 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0040
                    try:
                        if float(were_any_concomitant_medication_8_weeks_before_pure) == 0.0:
                            if float(prior_medication_id) == np.nan or str(prior_medication_id) == 'nan' or str(prior_medication_id) == '' or str(prior_medication_id) == '-':
                                pass
                            else:
                                error =  [subject, visit, 'Were any concomitant medications taken within 8 weeks before start of screening or during the study?', \
                                          were_any_concomitant_medication_8_weeks_before_form_field_instance, 'If No, no concomitant medication forms should be completed', \
                                                were_any_concomitant_medication_8_weeks_before_disname, 'SM0040']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0040 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision SM0050
                    try:
                        if float(were_concomitant_procedure_performed_during_study_pure) == 1.0:
                            if float(prior_procedure_id) == np.nan or str(prior_procedure_id) == 'nan' or str(prior_procedure_id) == '' or str(prior_procedure_id) == '-':
                                error =  [subject, visit, 'Were any concomitant procedures/surgeries performed during the study?', \
                                          were_concomitant_procedure_performed_during_study_form_field_instance, 'If Yes, at least one adverse event form must be completed', \
                                                were_concomitant_procedure_performed_during_study_disname, 'SM0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0050 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0060
                    try:
                        if float(were_concomitant_procedure_performed_during_study_pure) == 0.0:
                            if float(prior_procedure_id) == np.nan or str(prior_procedure_id) == 'nan' or str(prior_procedure_id) == '' or str(prior_procedure_id) == '-':
                                pass
                            else:
                                error =  [subject, visit, 'Were any concomitant procedures/surgeries performed during the study?', \
                                          were_concomitant_procedure_performed_during_study_form_field_instance, 'If No, no concomitant medication forms should be completed', \
                                                were_concomitant_procedure_performed_during_study_disname, 'SM0060']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0060 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0070
                    try:
                        if float(has_subject_taken_cpg_ODN_pure) == 1.0:
                            if str(date_cpg_administration) == 'nan' or str(date_cpg_administration) == '' or str(date_cpg_administration) == '-':
                                error =  [subject, visit, 'Has the subject taken at least one dose of CpG ODN D35 study treatment?', \
                                          has_subject_taken_cpg_ODN_form_field_instance, 'If Yes, at least one adverse event form must be completed', \
                                                has_subject_taken_cpg_ODN_disname, 'SM0070']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0070 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0080
                    try:
                        if float(has_subject_taken_cpg_ODN_pure) == 0.0:
                            if  str(date_cpg_administration) == 'nan' or str(date_cpg_administration) == '' or str(date_cpg_administration) == '-':
                                pass
                            else:
                                error =  [subject, visit, 'Has the subject taken at least one dose of CpG ODN D35 study treatment?', \
                                          has_subject_taken_cpg_ODN_form_field_instance, 'If No, no CpG ODN D35 administration forms should be completed', \
                                                has_subject_taken_cpg_ODN_disname, 'SM0080']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0080 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0110
                    try:
                        if float(has_subject_taken_miltefosine_pure) == 1.0:
                            if str(date_miltefosine_administration) == 'nan' or str(date_miltefosine_administration) == '' or str(date_miltefosine_administration) == '-':
                                error =  [subject, visit, 'Has  subject taken at least one dose of Miltefosine study treatment?', \
                                          has_subject_taken_miltefosine_form_field_instance, 'If Yes, at least one adverse event form must be completed', \
                                                has_subject_taken_miltefosine_disname, 'SM0110']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0110 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision SM0120
                    try:
                        if float(has_subject_taken_miltefosine_pure) == 0.0:
                            pass
                        else:
                            if str(date_miltefosine_administration) == 'nan' or str(date_miltefosine_administration) == '' or str(date_miltefosine_administration) == '-':
                                error =  [subject, visit, 'Has  subject taken at least one dose of Miltefosine study treatment?', \
                                          has_subject_taken_miltefosine_form_field_instance, 'If No, no Miltefosine administration forms should be completed', \
                                                has_subject_taken_miltefosine_disname, 'SM0120']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision SM0120 --> {e} - Subject: {subject},  Visit: {visit} ')

    
    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    ev_med_proce_treatment_output = pd.DataFrame(lista_revision, columns=column_names)
    ev_med_proce_treatment_output = ev_med_proce_treatment_output[~ev_med_proce_treatment_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
 
    sheet = excel_writer.create_sheet('Ev, Med, Proc, Study')

    for row in dataframe_to_rows(ev_med_proce_treatment_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return ev_med_proce_treatment_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx")
    ev_med_proce_treatment(df_root, path_excel ) 

