import numpy as np
import math
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def Pharmacokinetic_blood_sampling(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Pharmacokinetic Blood Sampling (PK)
    '''

    df= df_root[df_root['name']== 'Pharmacokinetic Blood Sampling (PK)']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)


    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante',  'Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    df_time_dosing1 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1 = df_time_dosing1[(df_time_dosing1['Campo']=='Date of dosing') | (df_time_dosing1['Campo']=='Time of Dosing')]
    df_time_dosing = df_time_dosing1[df_time_dosing1['Campo']=='Date of dosing']
    df_time_dosing['time_dosing_cpg_administration'] =  df_time_dosing1[df_time_dosing1['FormFieldInstance Id'].isin(df_time_dosing['FormFieldInstance Id'] + 1) & (df_time_dosing1['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing =df_time_dosing[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing = df_time_dosing.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})

    lista_logs = ['Pharmacokinetic Blood Sampling (PK)']
    lista_revision = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()

            try:
                pru['date_ex_to_join'] = pru['Date of blood sample collected'].str.split('|',expand= True)[0]
            except:
                pru['date_ex_to_join'] = 'Nothing to join'

            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_time_dosing, on=['Subject', 'date_ex_to_join'], how='left')
            # print(pru)
            # print('--------------------------')

            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']
                time_dosing_cpg_administration = row['time_dosing_cpg_administration']


                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                if status != '':
                    try:
                        Was_any_pharmacokinetic_blood_sample_collected = row["Was any pharmacokinetic blood sample collected?"]
                        Was_any_pharmacokinetic_blood_sample_collected_pure = Was_any_pharmacokinetic_blood_sample_collected.split('|')[0]
                        Was_any_pharmacokinetic_blood_sample_collected_form_field_instance = Was_any_pharmacokinetic_blood_sample_collected.split('|')[1]
                        Was_any_pharmacokinetic_blood_sample_collected_disname = Was_any_pharmacokinetic_blood_sample_collected.split('|')[2]
                    except Exception as e:
                        Was_any_pharmacokinetic_blood_sample_collected_pure = math.nan
                        Was_any_pharmacokinetic_blood_sample_collected_form_field_instance = 'This field does not have any data'
                        Was_any_pharmacokinetic_blood_sample_collected_disname = 'Empty'

                    # try:
                    #     Provide_the_reason = row["Provide the reason"]
                    # except Exception as e:
                    #     pass 

                    try:
                        Date_of_blood_sample_collected = row["Date of blood sample collected"]
                        Date_of_blood_sample_collected_pure = Date_of_blood_sample_collected.split('|')[0]
                        Date_of_blood_sample_collected_form_field_instance = Date_of_blood_sample_collected.split('|')[1]
                        Date_of_blood_sample_collected_disname = Date_of_blood_sample_collected.split('|')[0]
                    except Exception as e:
                        Date_of_blood_sample_collected_pure = ''
                        Date_of_blood_sample_collected_form_field_instance = 'This field does not have any data'
                        Date_of_blood_sample_collected_disname = 'Empty'

                    try:
                        Pre_dose = row['Pre dose, Time']
                        Pre_dose_pure = Pre_dose.split('|')[0]
                        Pre_dose_form_field_instance = Pre_dose.split('|')[1]
                    except Exception as e:
                        Pre_dose_pure = ''
                        Pre_dose_form_field_instance = 'Empty'


                    try:
                        min_05_post_dose = row['05-min post dose, Time']
                        min_05_post_dose_pure = min_05_post_dose.split('|')[0]
                        min_05_post_dose_form_field_isntance = min_05_post_dose.split('|')[1]
                    except Exception as e:
                        min_05_post_dose_pure = ''
                        min_05_post_dose_form_field_isntance = 'Empty'


                    try:
                        min_10_post_dose = row['10-min post dose, Time']
                        min_10_post_dose_pure = min_10_post_dose.split('|')[0]
                        min_10_post_dose_form_field_instance = min_10_post_dose.split('|')[1]
                    except Exception as e:
                        min_10_post_dose_pure = ''
                        min_10_post_dose_form_field_instance =  'Empty'


                    try:
                        min_15_post_dose = row['15-min post dose, Time']
                        min_15_post_dose_pure = min_15_post_dose.split('|')[0]
                        min_15_post_dose_form_field_instance = min_15_post_dose.split('|')[1]
                    except Exception as e:
                        min_15_post_dose_pure = ''
                        min_15_post_dose_form_field_instance = 'Empty'
     
                    try:
                        min_20_post_dose = row['20-min post dose, Time']
                        min_20_post_dose_pure = min_20_post_dose.split('|')[0]
                        min_20_post_dose_form_field_isntance = min_20_post_dose.split('|')[1]
                    except Exception as e:
                        min_20_post_dose_pure = ''
                        min_20_post_dose_form_field_isntance = 'Empty'

                    try:
                        min_25_post_dose = row['25-min post dose, Time']
                        min_25_post_dose_pure = min_25_post_dose.split('|')[0]
                        min_25_post_dose_form_field_instance = min_25_post_dose.split('|')[1]
                    except Exception as e:
                        min_25_post_dose_pure = ''
                        min_25_post_dose_form_field_instance = 'Empty'


                    try:
                        min_30_post_dose = row['30-min post dose, Time']
                        min_30_post_dose_pure = min_30_post_dose.split('|')[0]
                        min_30_post_dose_form_field_instance = min_30_post_dose.split('|')[1]
                    except Exception as e:
                        min_30_post_dose_pure = ''
                        min_30_post_dose_form_field_instance = 'Empty'


                    try:
                        min_45_post_dose = row['45-min post dose, Time']
                        min_45_post_dose_pure = min_45_post_dose.split('|')[0]
                        min_45_post_dose_form_field_instance = min_45_post_dose.split('|')[1]
                    except Exception as e:
                        min_45_post_dose_pure = ''
                        min_45_post_dose_form_field_instance = 'Empty'


                    try:
                        min_60_post_dose = row['60-min post dose, Time']
                        min_60_post_dose_pure = min_60_post_dose.split('|')[0]
                        min_60_post_dose_form_field_instance = min_60_post_dose.split('|')[1]
                    except Exception as e:
                        min_60_post_dose_pure = ''
                        min_60_post_dose_form_field_instance = 'Empty'


                    try:
                        min_75_post_dose = row['75-min post dose, Time']
                        min_75_post_dose_pure = min_75_post_dose.split('|')[0]
                        min_75_post_dose_form_field_instance = min_75_post_dose.split('|')[1]
                    except Exception as e:
                        min_75_post_dose_pure = ''
                        min_75_post_dose_form_field_instance = 'Empty'
                        
                    # --------------------------------------------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if Date_of_blood_sample_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(Date_of_blood_sample_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance,\
                                        f , Date_of_blood_sample_collected_disname, 'GE0020']
                                lista_revision.append(error)     
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision PK0010
                    if Date_of_blood_sample_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(Date_of_blood_sample_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)
                          
                            if date_of_test_f != date_of_visit_f:
                                # print(date_of_test_f, date_of_visit_f)
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance ,\
                                        'The date should be the same as the visit date in the "Date of Visit" form' , f'{Date_of_blood_sample_collected_pure} - {date_of_visit}', 'PK0010']
                                lista_revision.append(error)
                            else:
                                pass

                        except Exception as e:
                            lista_logs.append(f'Revision PK0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision PK0030
                    if Date_of_blood_sample_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(Date_of_blood_sample_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance ,\
                                        'The date of sample collected can not be before the informed consent date', f'{Date_of_blood_sample_collected_disname} - {date_inform_consent}', 'PK0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision PK0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> PK0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or Date_of_blood_sample_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(Date_of_blood_sample_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance,\
                                        'Date of blood sample collected must be before the End of study/Early withdrawal date. ', Date_of_blood_sample_collected_disname, 'PK0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PK0040 --> {e} - Subject: {subject},  Visit: {visit}  ')


                    lista_validacion = [
                            'Pre dose, Time',
                            '05-min post dose, Time',
                            '10-min post dose, Time',
                            '15-min post dose, Time',
                            '20-min post dose, Time',
                            '25-min post dose, Time',
                            '30-min post dose, Time',
                            '45-min post dose, Time',
                            '60-min post dose, Time',
                            '75-min post dose, Time',
                    ]

                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        

                        try:    
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador=''
                        if validador!='':
                            mi_cuenta+=1
                            

                    # Revision PK0050
                    try:
                        if float(Was_any_pharmacokinetic_blood_sample_collected_pure) == 1.0:
                            if mi_cuenta > 0:
                                pass
                            else:
                                error = [subject, visit, 'Was blood sample collected?', Was_any_pharmacokinetic_blood_sample_collected_form_field_instance ,\
                                        'If the sample was collected, not all sections can be "not done"', Was_any_pharmacokinetic_blood_sample_collected_disname, 'PK0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision PK0050--> {e} - Subject: {subject},  Visit: {visit} ')
                    

                    # Revision PK0060
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif = float((datetime.strptime(time_dosing_cpg_administration , '%H:%M') - datetime.strptime(Pre_dose_pure, '%H:%M')).total_seconds() / 60)
                            if dif < 0.0 or dif > 90.0:
                                    
                                error = [subject, visit, 'Pre dose, Time', Pre_dose_form_field_instance,\
                                             'Pre dose Time is not within 90 minutes before the study treatment administration time.', \
                                                f'Pre dose, Time: {Pre_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0060']
                                lista_revision.append(error)

                        except Exception as e:
                            lista_logs.append(f'Revision PK0060 --> {e} - Subject: {subject},  Visit: {visit} ')  


                    # Revision PK0070
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_5m = float((datetime.strptime(min_05_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_5m > 6.0 or dif_5m < 4.0:
                                    
                                error = [subject, visit, '05-min post dose, Time', min_05_post_dose_form_field_isntance,\
                                             'The time selected should be less than 6 min and greater than 4 min after the study treatment administration', \
                                                f'05-min post dose, Time: {min_05_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0070']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0070 --> {e} - Subject: {subject},  Visit: {visit} ')  


                    # Revision PK0080
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_10m = float((datetime.strptime(min_10_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_10m > 11.0 or dif_10m < 9.0:
                                    
                                error = [subject, visit, '10-min post dose, Time', min_10_post_dose_form_field_instance,\
                                             'The time selected should be less than 11 min and greater than 9 min after the study treatment administration', \
                                                f'10-min post dose, Time: {min_10_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0080']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0080 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0090
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_15m = float((datetime.strptime(min_15_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_15m > 16.0 or dif_15m < 14.0:
                                    
                                error = [subject, visit, '15-min post dose, Time', min_15_post_dose_form_field_instance,\
                                             'The time selected should be less than 16 min and greater than 14 min after the study treatment administration', \
                                                f'15-min post dose, Time: {min_15_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0090']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0090 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0100
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_20m = float((datetime.strptime(min_20_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_20m > 21.0 or dif_20m < 19.0:
                                    
                                error = [subject, visit, '20-min post dose, Time', min_20_post_dose_form_field_isntance,\
                                             'The time selected should be less than 21 min and greater than 19 min after the study treatment administration', \
                                                f'20-min post dose, Time: {min_20_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0100']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0100 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0110
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_25m = float((datetime.strptime(min_25_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_25m > 26.0 or dif_25m < 24.0:
                                    
                                error = [subject, visit, '25-min post dose, Time', min_25_post_dose_form_field_instance,\
                                             'The time selected should be less than 26 min and greater than 24 min after the study treatment administration', \
                                                f'25-min post dose, Time: {min_25_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0110']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0110 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0120
                    if str(time_dosing_cpg_administration) != 'nan' and str(min_30_post_dose_pure) != '':
                            
                        try:
                            dif_30m = float((datetime.strptime(min_30_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_30m > 31.0 or dif_30m < 29.0:
                                    
                                error = [subject, visit, '30-min post dose, Time', min_30_post_dose_form_field_instance,\
                                             'The time selected should be less than 31 min and greater than 29 min after the study treatment administration', \
                                                f'30-min post dose, Time: {min_30_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0120']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0120 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0130
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_45m = float((datetime.strptime(min_45_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_45m > 48.0 or dif_45m < 42.0:
                                    
                                error = [subject, visit, '45-min post dose, Time', min_45_post_dose_form_field_instance,\
                                             'The time selected should be less than 48 min and greater than 42 min after the study treatment administration', \
                                                f'45-min post dose, Time: {min_45_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0130']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0130 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0140
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_60m = float((datetime.strptime(min_60_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_60m > 63.0 or dif_60m < 57.0:
                                    
                                error = [subject, visit, '60-min post dose, Time', min_60_post_dose_form_field_instance,\
                                             'The time selected should be less than 63 min and greater than 57 min after the study treatment administration', \
                                                f'60-min post dose, Time: {min_60_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0140']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0140 --> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision PK0150
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_75m = float((datetime.strptime(min_75_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_75m > 78.0 or dif_75m < 72.0:
                                    
                                error = [subject, visit, '75-min post dose, Time', min_75_post_dose_form_field_instance,\
                                             'The time selected should be less than 78 min and greater than 72 min after the study treatment administration', \
                                                f'75-min post dose, Time: {min_75_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'PK0150']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision PK0150 --> {e} - Subject: {subject},  Visit: {visit} ') 



                            

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    Pharmacokinetic_blood_sampling_output = pd.DataFrame(lista_revision, columns=column_names)
    Pharmacokinetic_blood_sampling_output = Pharmacokinetic_blood_sampling_output[~Pharmacokinetic_blood_sampling_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
    
    sheet = excel_writer.create_sheet("Pharmacokinetic BS(PK)")

    for row in dataframe_to_rows(Pharmacokinetic_blood_sampling_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return Pharmacokinetic_blood_sampling_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    Pharmacokinetic_blood_sampling(df_root, path_excel ) 