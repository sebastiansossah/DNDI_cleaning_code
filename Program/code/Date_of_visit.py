import math
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



def date_of_visit(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Date of Visit
    '''

    df = df_root[df_root['name']=='Date of visit'] 
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)


    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    warnings.filterwarnings('ignore')

    lista_revision = []
    lista_logs = ['Date of visit']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        visit_dictionary = {'Screening Visit': '', 'D-1':'', 'D1':'', 'D2':'', 'D3':'', 'D4':'', 'D7':'', 'D14':'', 
        'D15':'', 'D16':'', 'D17':'', 'D18':'', 'D21':'', 'D28':'', 'D29':'', 'D30':'',
        'D31':'', 'D32':'', 'D35':'', 'D42':'', 'D63':'', 'D90':'', 'D105':''}

        for visita in sujeto_principal.Visit.unique():
        
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            

            for index, row in pru.iterrows():

                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                informed_consent_date = row['Valor']
                end_study_date = row['end_study_date']

                try:
                    visit_date = row['Visit Date']
                    visit_date_pure = visit_date.split('|')[0]
                    visit_date_form_field_instance = visit_date.split('|')[1]
                except:
                    visit_date = ''
                    visit_date_form_field_instance = 'This field does not have any data'

                
                if status != '':
                    try:
                        # Primera  revision general de formato de fecha ->GE0020
                        f = revision_fecha(visit_date_pure)
                        # Agregamos la fecha de visit date al diccionario, para poder ser consultado posteriormente
                        if f == None:
                            pass
                        else:
                            error = [subject, visit, 'Visit Date', visit_date_form_field_instance ,f , visit_date_pure, 'GE0020']
                            lista_revision.append(error)     

                        visit_dictionary[visit] = visit_date

                    except Exception as e:
                        lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision -> VS0050
                    try:
                        if datetime.strptime(str(visit_date_pure), '%d-%b-%Y') >= datetime.strptime(str(informed_consent_date), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'Visit Date', visit_date_form_field_instance ,'Visit date must be equal or greater than the inform consent date', visit_date_pure, 'VS0050']
                            lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision VS0050 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision -> VS0030
                    if  str(end_study_date) == '' or   str(end_study_date) == 'nan': 
                        pass
                    else:
                        try:
                            if datetime.strptime(str(visit_date_pure), '%d-%b-%Y') >= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Visit Date', visit_date_form_field_instance ,'Visit Date must be before the End of study/Early withdrawal date. ', visit_date_pure, 'VS0030']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision VS0030 --> {e} - Subject: {subject},  Visit: {visit}  ')

        
        # Codigo para verificar si las fechas de las visitas, son mayores consecutivamente -> VS0020
        lista_valores_diccionario_visitas = list(visit_dictionary.values())
        lista_keys_diccionario_visitas = list(visit_dictionary.keys())

        for i in range(len(lista_valores_diccionario_visitas)):

            if  lista_valores_diccionario_visitas[i].split('|')[0] == 0:
                pass
            elif lista_valores_diccionario_visitas[i].split('|')[0] == '' :
                break

            elif lista_valores_diccionario_visitas[i] == 'D-1' :
                    if datetime.strptime(lista_valores_diccionario_visitas[i].split('|')[0], '%d-%b-%Y')  < datetime.strptime(lista_valores_diccionario_visitas[i-1].split('|')[0], '%d-%b-%Y'):
                        error = [sujeto, lista_keys_diccionario_visitas[i], 'Visit Date', lista_valores_diccionario_visitas[i].split('|')[1] ,'Visit date must be greater than the previous visit date', \
                                  lista_valores_diccionario_visitas[i].split('|')[0], 'VS0020']
                        lista_revision.append(error)
                    else:
                        pass

            else:
                try:
                    if datetime.strptime(lista_valores_diccionario_visitas[i].split('|')[0], '%d-%b-%Y')  <= datetime.strptime(lista_valores_diccionario_visitas[i-1].split('|')[0], '%d-%b-%Y'):
                        error = [sujeto, lista_keys_diccionario_visitas[i], 'Visit Date', lista_valores_diccionario_visitas[i].split('|')[1] ,'Visit date must be greater than the previous visit date' , \
                                  lista_valores_diccionario_visitas[i].split('|')[0], 'VS0020']
                        lista_revision.append(error)
                    else:
                        pass
                except Exception as e:
                    lista_logs.append(f'Revision VS0050 --> {e} - Subject: {subject},  Visit: {visit}  ')
                    
    
    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    date_of_visit_output = pd.DataFrame(lista_revision, columns=column_names)

 
    sheet = excel_writer.create_sheet("Date of visit")

    for row in dataframe_to_rows(date_of_visit_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return date_of_visit_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    date_of_visit(df_root, path_excel ) 

