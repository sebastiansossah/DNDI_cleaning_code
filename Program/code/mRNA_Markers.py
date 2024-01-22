import numpy as np
import math
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def mRNA_markers(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de mRNA Markers
    '''

    df= df_root[df_root['name']== 'mRNA Markers']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    df_time_dosing1 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1 = df_time_dosing1[(df_time_dosing1['Campo']=='Date of dosing') | (df_time_dosing1['Campo']=='Time of Dosing')]
    df_time_dosing = df_time_dosing1[df_time_dosing1['Campo']=='Date of dosing']
    df_time_dosing['time_dosing_cpg_administration'] =  df_time_dosing1[df_time_dosing1['FormFieldInstance Id'].isin(df_time_dosing['FormFieldInstance Id'] + 1) & (df_time_dosing1['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing =df_time_dosing[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing = df_time_dosing.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})

    lista_logs = ['mRNA Markers']
    lista_revision = []

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            try:
                pru['date_ex_to_join'] = pru['Date of blood sample collected'].str.split('|',expand= True)[0]
            except:
                pru['date_ex_to_join'] = 'Nothing to join'

            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_time_dosing, on=['Subject', 'date_ex_to_join'], how='left')

            # print(pru)
            # print('---------------------')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]

                time_dosing_cpg_administration = row['time_dosing_cpg_administration']
   
                if status != '':
                    try:
                        Was_blood_sample_collected = row["Was blood sample collected?"]
                        Was_blood_sample_collected_pure = Was_blood_sample_collected.split('|')[0]
                        Was_blood_sample_collected_form_field_instance = Was_blood_sample_collected.split('|')[1]
                        Was_blood_sample_collected_disname = Was_blood_sample_collected.split('|')[2]
                    except Exception as e:
                        Was_blood_sample_collected_pure = math.nan
                        Was_blood_sample_collected_form_field_instance = 'This field does not have any data'
                        Was_blood_sample_collected_disname = 'Empty'

                    try:
                        Provide_the_reason = row["Provide the reason"]
                        Provide_the_reason_pure = Provide_the_reason.split('|')[0]
                        Provide_the_reason_form_field_instance = Provide_the_reason.split('|')[1]
                        Provide_the_reason_disname = Provide_the_reason.split('|')[0]
                    except Exception as e:
                        Provide_the_reason_pure = ''
                        Provide_the_reason_form_field_instance = 'This field does not have any data'
                        Provide_the_reason_disname = 'Empty'
                    
                    try:
                        Date_of_blood_sample_collected = row["Date of blood sample collected"]
                        Date_of_blood_sample_collected_pure = Date_of_blood_sample_collected.split('|')[0]
                        Date_of_blood_sample_collected_form_field_instance = Date_of_blood_sample_collected.split('|')[1]
                        Date_of_blood_sample_collected_disname = Date_of_blood_sample_collected.split('|')[0]
                    except Exception as e:
                        Date_of_blood_sample_collected_pure = ''
                        Date_of_blood_sample_collected_form_field_instance = 'This field does not have any data'
                        Date_of_blood_sample_collected_disname = 'Empty'
                    
                    try:
                        Pre_dose = row["Pre-dose"]
                        Pre_dose_pure = Pre_dose.split('|')[0]
                        Pre_dose_form_field_instance = Pre_dose.split('|')[1]
                        Pre_dose_disname = Pre_dose.split('|')[0]
                    except Exception as e:
                        Pre_dose_pure = ''
                        Pre_dose_form_field_instance = 'This field does not have any data'
                        Pre_dose_disname = 'Empty'
             
                    try:
                        Pre_dose_time = row['Pre-dose, Time']
                        Pre_dose_time_pure = Pre_dose_time.split('|')[0]
                        Pre_dose_time_form_field_instance = Pre_dose_time.split('|')[1]
                    except Exception as e:
                        Pre_dose_time_pure = ''
                        Pre_dose_time_form_field_instance = 'This field does not have any data'
    
                    
                    try:
                        Pre_dose_Reason_not_done = row["Pre-dose, Reason not done"]
                        Pre_dose_Reason_not_done_pure = Pre_dose_Reason_not_done.split('|')[0]
                        Pre_dose_Reason_not_done_form_field_instance = Pre_dose_Reason_not_done.split('|')[1]
                        Pre_dose_Reason_not_done_disname = Pre_dose_Reason_not_done.split('|')[0]
                    except Exception as e:
                        Pre_dose_Reason_not_done_pure = ''
                        Pre_dose_Reason_not_done_form_field_instance = 'This field does not have any data'
                        Pre_dose_Reason_not_done_disname = 'Empty'

                    try:
                        hours_04_post_dose = row['04-hours post dose, Time']
                        hours_04_post_dose_pure = hours_04_post_dose.split('|')[0]
                        hours_04_post_dose_form_field_instance = hours_04_post_dose.split('|')[1]
                        hours_04_post_dose_disname = hours_04_post_dose.split('|')[0]
                    except Exception as e:
                        hours_04_post_dose_pure = ''
                        hours_04_post_dose_form_field_instance = 'This field does not have any data'
                        hours_04_post_dose_disname = 'Empty'

                    
                    try:
                        hours_04_post_dose_Reason_not_done = row["04-hours post dose, Reason not done"]
                        hours_04_post_dose_Reason_not_done_pure = hours_04_post_dose_Reason_not_done.split('|')[0]
                        hours_04_post_dose_Reason_not_done_form_field_instance = hours_04_post_dose_Reason_not_done.split('|')[1]
                        hours_04_post_dose_Reason_not_done_disname = hours_04_post_dose_Reason_not_done.split('|')[0]
                    except Exception as e:
                        hours_04_post_dose_Reason_not_done_pure = ''
                        hours_04_post_dose_Reason_not_done_form_field_instance = 'This field does not have any data'
                        hours_04_post_dose_Reason_not_done_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose = row['12-hours post dose, Time']
                        hours_12_post_dose_pure = hours_12_post_dose.split('|')[0]
                        hours_12_post_dose_form_field_instance = hours_12_post_dose.split('|')[1]
                    except Exception as e:
                        hours_12_post_dose_pure = ''
                        hours_12_post_dose_form_field_instance = 'This field does not have any data'
                    
                    try:
                        hours_12_post_dose_Reason_not_done = row["12-hours post dose, Reason not done"]
                        hours_12_post_dose_Reason_not_done_pure = hours_12_post_dose_Reason_not_done.split('|')[0]
                        hours_12_post_dose_Reason_not_done_form_field_instance = hours_12_post_dose_Reason_not_done.split('|')[1]
                        hours_12_post_dose_Reason_not_done_disname = hours_12_post_dose_Reason_not_done.split('|')[0]
                    except Exception as e:
                        hours_12_post_dose_Reason_not_done_pure = ''
                        hours_12_post_dose_Reason_not_done_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Reason_not_done_disname = 'Empty'
                    
                    # --------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if Date_of_blood_sample_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(Date_of_blood_sample_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance,\
                                        f , Date_of_blood_sample_collected_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision MR0010
                    if Date_of_blood_sample_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(Date_of_blood_sample_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance ,\
                                        'The date should be the same as the visit date in the "Date of Visit" form' , f'{Date_of_blood_sample_collected_disname} - {date_of_visit}', 'MR0010']
                                lista_revision.append(error)
                            else:
                                pass

                        except Exception as e:
                            lista_logs.append(f'Revision MR0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision MR0020
                    if Date_of_blood_sample_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(Date_of_blood_sample_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance ,\
                                        'The date of sample collected can not  be before the informed consent date', f'{Date_of_blood_sample_collected_disname} - {date_inform_consent}', 'MR0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision MR0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> MR0030
                    if  str(end_study_date) == 'nan' or end_study_date == '' or Date_of_blood_sample_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(Date_of_blood_sample_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of blood sample collected', Date_of_blood_sample_collected_form_field_instance,\
                                        'Date of blood sample collected must be before the End of study/Early withdrawal date. ', Date_of_blood_sample_collected_disname, 'MR0030']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision MR0030 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    lista_validacion =[
                        'Pre-dose, Time',
                        '04-hours post dose, Time',
                        '12-hours post dose, Time',
                    ]
                    cuenta_validar = 0
                        
                    for validador_raw in lista_validacion:
                        try: 
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
       
                        if validador == '':
                            pass
                        else:
                            cuenta_validar += 1
                    
                    # Revision MR0050
                    if visit in ['D1', 'D15' , 'D29']:
                        try:
                            if float(Was_blood_sample_collected_pure) == 1.0:
                                if cuenta_validar > 0:
                                    pass
                                else:
                                    error = [subject, visit, 'Was blood sample collected?', Was_blood_sample_collected_form_field_instance ,\
                                            'If the sample was collected, not all sections can be "not done"', Was_blood_sample_collected_disname, 'MR0050']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision MR0050--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision MR0060
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif = float((datetime.strptime(time_dosing_cpg_administration , '%H:%M') - datetime.strptime(Pre_dose_time_pure, '%H:%M')).total_seconds() / 60)
                            if dif < 0.0 or dif > 60.0:
                                    
                                error = [subject, visit, 'Pre dose, Time', Pre_dose_time_form_field_instance,\
                                             'Pre dose Time is not within 60 minutes before the study treatment administration time.', \
                                                f'Pre dose, Time: {Pre_dose_time_pure} - dose time administration{time_dosing_cpg_administration}', 'MR0060']
                                lista_revision.append(error)

                        except Exception as e:
                            lista_logs.append(f'Revision MR0060 --> {e} - Subject: {subject},  Visit: {visit} ')  


                    # Revision MR0070
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_4h = float((datetime.strptime(hours_04_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_4h > 255.0 or dif_4h < 225.0:
                                    
                                error = [subject, visit, '04-hours post dose, Time', hours_04_post_dose_form_field_instance,\
                                             '4-hours post dose, Time  is not within 4 hours (+/- 15 minutes) minutes after the study treatment administration time.', \
                                                f'4-hours post dose: {hours_04_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'MR0070']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision MR0070 --> {e} - Subject: {subject},  Visit: {visit} ')  


                    # Revision MR0080
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif_12h = float((datetime.strptime(hours_12_post_dose_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                            if dif_12h > 735.0 or dif_12h < 705.0:
                                    
                                error = [subject, visit, '12-hours post dose', hours_12_post_dose_form_field_instance,\
                                             '12-hours post dose, Time  is not within 12 hours (+/- 15 minutes) minutes after the study treatment administration time.', \
                                                f'12-hours post dose: {hours_12_post_dose_pure} - dose time administration{time_dosing_cpg_administration}', 'MR0080']
                                lista_revision.append(error)

                        except Exception as e:
                                lista_logs.append(f'Revision MR0080 --> {e} - Subject: {subject},  Visit: {visit} ')  



    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    mRNA_markers_output = pd.DataFrame(lista_revision, columns=column_names)

    sheet = excel_writer.create_sheet("mRNA Markers")

    for row in dataframe_to_rows(mRNA_markers_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return mRNA_markers_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    mRNA_markers(df_root, path_excel )