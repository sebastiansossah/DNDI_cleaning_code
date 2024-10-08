import math
from datetime import datetime
from revision_fechas import revision_fecha, date_format
from log_writer import log_writer
import warnings
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

def Medical_or_surgical_history(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Medical Or Surgical History (other than Leishmaniasis)
    '''

    df= df_root[df_root['name']== 'Medical Or Surgical History (other than Leishmaniasis)']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_demographic_age = df_root[df_root['name']=='Demographics']
    df_demographic_age = df_demographic_age[['Participante', 'Campo', 'Valor']]
    df_demographic_age = df_demographic_age[df_demographic_age['Campo']=='Birth Year']
    df_demographic_age = df_demographic_age[['Participante','Valor']]
    df_demographic_age = df_demographic_age.rename(columns={'Participante':'Subject', 'Valor':'Birth_Year'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Medical Or Surgical History (other than Leishmaniasis)']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        lista_comprobacion_overlap = []

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_demographic_age, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            # print(pru)
            # print('-----------------')

            for index, row in pru.iterrows():

                status = row['status']

                try:
                    was_DV_performed = row['was_DV_performed']
                    was_DV_performed_pure = was_DV_performed.split('|')[0]
                    was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                except:
                    was_DV_performed_pure = ''
                if status != '':
                
                    subject = row['Subject']
                    visit = row['Visit']
                    demographic_year = row['Birth_Year']

                    try:
                        any_relevant_medical = row['Are there any relevant medical history or surgical history ?']
                        any_relevant_medical_pure = any_relevant_medical.split('|')[0]
                        any_relevant_medical_form_field_instance = any_relevant_medical.split('|')[1]
                        any_relevant_medical_disname = any_relevant_medical.split('|')[2]
                    except:
                        any_relevant_medical_pure = math.nan
                        any_relevant_medical_form_field_instance = 'This field does not have any data'
                        any_relevant_medical_disname = 'Empty'

                    # if float(any_relevant_medical_pure) == 1.0:

                    #     pru2 = df[df['Participante']==sujeto]
                    #     pru2 = pru2.sort_values(by=['FormFieldInstance Id'], ascending=True)
                    #     pru2 = pru2.reset_index(drop=True)

                    #     # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, 
                    #     # por lo que usamos el siguiente codigo para realizar la particion

                    #     date_indices = pru2.index[pru2['Campo'] == 'Medical/Surgical History/Current Condition'].tolist()
                    #     subdatasets = [pru2.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]
                    #     for subdataset in subdatasets:
                    #         pru_pri = subdataset
                    #         pru_sec = pru_pri
                    #         pru_sec = pru_sec[['Campo', 'Value_id']].T
                    #         new_columns = pru_sec.iloc[0]
                    #         pru_sec = pru_sec[1:].set_axis(new_columns, axis=1)
                    #         pru_sec['Subject'] = sujeto
                    #         pru_sec['Visit'] = visita
                    #         pru_sec['status'] = pru_pri['activityState'].unique()
                    #         pru_sec = pru_sec.merge(df_demographic_age, on=['Subject'], how='left')
                    #         pru_sec = pru_sec.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
                    #         # print(pru_sec)
                    #         # print('------------------')

                    #         for index_ulti, row in pru_sec.iterrows():
                    if index != 0:
                        lista_logs.append('Duplicados en la data, revisar subdataset')

                    try:
                        medical_surgical = row['Medical/Surgical History/Current Condition']
                        medical_surgical_pure = medical_surgical.split('|')[0]
                        medical_surgical_form_field_instance = medical_surgical.split('|')[1]
                        medical_surgical_disname = medical_surgical.split('|')[2]
                    except Exception as e:
                        medical_surgical_pure = math.nan
                        medical_surgical_form_field_instance = 'This field does not have any data'
                        medical_surgical_disname = 'Empty'

                    try:
                        onset_date = row['Onset Date/First Diagnosis/Surgery']
                        onset_date_pure = onset_date.split('|')[0]
                        onset_date_form_field_instance = onset_date.split('|')[1]
                        onset_date_disname = onset_date.split('|')[0]
                    except Exception as e:
                        onset_date_pure = ''
                        onset_date_form_field_instance = 'This field does not have any data'
                        onset_date_disname = 'Empty'

                    try:
                        end_date = row['End Date Medical/Surgical History/Current Condition']
                        end_date_pure = end_date.split('|')[0]
                        end_date_form_field_instance = end_date.split('|')[1]
                        end_date_disname = end_date.split('|')[0]
                    except Exception as e:
                        end_date_pure = ''
                        end_date_form_field_instance = 'This field does not have any data'
                        end_date_disname = 'Empty'

                    # --------------------------------------------------------------------------------------

                    # Revision GE0070
                    try:
                        if float(was_DV_performed_pure) != 1.0:
                            error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance, 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                            lista_revision.append(error)
                    except:
                        pass
                    if onset_date_pure == '':
                        pass
                    else:
                        try:
                            # Primera revision general de formato de fecha ->GE0020
                            f = revision_fecha(onset_date_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Onset Date/First Diagnosis/Surgery', onset_date_form_field_instance, f, onset_date_disname, 'GE0020']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject}, Visit: {visit} ')

                    if end_date_pure == '':
                        pass
                    else:
                        try:
                            # Primera revision general de formato de fecha ->GE0020
                            f = revision_fecha(end_date_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'End Date', end_date_form_field_instance, f, end_date_disname, 'GE0020']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject}, Visit: {visit} ')

                    lista_validacion = [
                        'Medical/Surgical History/Current Condition',
                        'Onset Date/First Diagnosis/Surgery',
                        'Medical/Surgical History/Current Condition',
                        'Onset Date/First Diagnosis/Surgery',
                        'Is Condition Ongoing?',
                        'End Date',
                        'Medical/Surgical History/Current Condition',
                        'Onset Date/First Diagnosis/Surgery',
                        'Is Condition Ongoing?',
                        'Severity',
                        'Frequency',
                        'Currently treated?'
                    ]
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        
                        try:    
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador=''
                        if validador!='':
                            mi_cuenta+=1

                    # Revision MS0010
                    try:
                        if float(any_relevant_medical_pure) == 1.0:
                            if mi_cuenta == 0 :
                                error = [subject, visit, 'Are there any relevant medical history or surgical history?', any_relevant_medical_form_field_instance,
                                        'If the answer is Yes, at least one section of Medical or Surgical History Detail should be added', 'no fields founded', 'MS0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision MS0020 --> {e} - Subject: {subject}, Visit: {visit} ')

                    # Revision MS0020
                    try:
                        if float(any_relevant_medical_pure) == 0.0:
                            if mi_cuenta > 0:
                                error = [subject, visit, 'Are there any relevant medical history or surgical history?', any_relevant_medical_form_field_instance,
                                        'If the answer is No, No sections of Medical or Surgical History Detail should be added', 'fields added', 'MS0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision MS0020 --> {e} - Subject: {subject}, Visit: {visit} ')

                    if float(any_relevant_medical_pure) == 1.0 and end_date_pure != '':
                        # Revision MS0040
                        try:
                            # date_format = '%d-%b-%Y'
                            onset_date_f = datetime.strptime(onset_date_pure, '%d-%b-%Y')
                            end_date_f = datetime.strptime(end_date_pure, '%d-%b-%Y')

                            if onset_date_f > end_date_f:
                                error = [subject, visit, 'End Date', end_date_form_field_instance,
                                        'End date must be after Onset Date/First Diagnosis/Surgery.', end_date_disname, 'MS0040']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision MS0040 --> {e} - Subject: {subject}, Visit: {visit} ')

                    # Revision MS0050
                    if onset_date_pure != '':
                        try:
                            onset_date_year = str(onset_date_pure).split('-')[2]

                            if int(onset_date_year) < int(demographic_year):
                                error = [subject, visit, 'Onset Date/First Diagnosis/Surgery', onset_date_form_field_instance,
                                        'The year and month of Onset Date/First taken must be equal or after the month and year of birth in DEMOGRAPHIC Diagnosis/Surgery.', onset_date_year, 'MS0050']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision MS0050 --> {e} - Subject: {subject}, Visit: {visit} ')

                    # Revision MS0060
                    if medical_surgical_pure != '':
                        try:
                            medical_date_history = (medical_surgical_pure, onset_date_pure, end_date_pure)

                            if medical_date_history in lista_comprobacion_overlap:
                                error = [subject, visit, 'Medical/Surgical History/ Current Condition', medical_surgical_form_field_instance,
                                        'The Medica/Surgical History/ Current Condition should not be entered twice if the dates overlap', medical_surgical_disname, 'MS0060']
                                lista_revision.append(error)
                            else:
                                lista_comprobacion_overlap.append(medical_date_history)
                        except Exception as e:
                            lista_logs.append(f'Revision MS0060 --> {e} - Subject: {subject}, Visit: {visit} ')

        
                    


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    medical_surgical_output = pd.DataFrame(lista_revision, columns=column_names)
    medical_surgical_output = medical_surgical_output[~medical_surgical_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
 
    sheet = excel_writer.create_sheet("Medical Or Surgical History")

    for row in dataframe_to_rows(medical_surgical_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return medical_surgical_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    Medical_or_surgical_history(df_root, path_excel ) 