from datetime import datetime
import math

import numpy as np
from revision_fechas import revision_fecha
from log_writer import log_writer 
import warnings
import pandas as pd
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')


def adverse_events(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Adverse Events
    '''

    df= df_root[df_root['name']=='Adverse Events']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Inform_consent_date'})

    df_medical_surgical = df_root[df_root['name']== 'Medical Or Surgical History (other than Leishmaniasis)']
    df_medical_surgical = df_medical_surgical[['Visit','Participante', 'Campo', 'Valor']]
    df_medical_surgical = df_medical_surgical[df_medical_surgical['Campo'] == 'Are there any relevant medical history or surgical history ?']
    df_medical_surgical = df_medical_surgical[['Participante','Valor']]
    df_medical_surgical = df_medical_surgical.rename(columns={'Participante':'Subject', 'Valor':'relevant_medical_condition'})

    df_medical_eligibility = df_root[df_root['name']== 'Eligibility']
    df_medical_eligibility = df_medical_eligibility[['Visit','Participante', 'Campo', 'Valor']]
    df_medical_eligibility = df_medical_eligibility[df_medical_eligibility['Campo'] == 'Will the participant be randomized?']
    df_medical_eligibility = df_medical_eligibility[['Participante','Valor']]
    df_medical_eligibility = df_medical_eligibility.rename(columns={'Participante':'Subject', 'Valor':'subject_randomized'})

    df_medical_eligibility_date = df_root[df_root['name']== 'Eligibility']
    df_medical_eligibility_date = df_medical_eligibility_date[['Visit','Participante', 'Campo', 'Valor']]
    df_medical_eligibility_date = df_medical_eligibility_date[df_medical_eligibility_date['Campo'] == 'Date of decision to not randomize the participant']
    df_medical_eligibility_date = df_medical_eligibility_date[['Participante','Valor']]
    df_medical_eligibility_date = df_medical_eligibility_date.rename(columns={'Participante':'Subject', 'Valor':'date_decision_not_randomize'})

    df_administration_CPG = df_root[df_root['name']== 'CpG ODN D35 Administration']
    df_administration_CPG = df_administration_CPG[['Visit','Participante', 'Campo', 'Valor']]
    df_administration_CPG = df_administration_CPG[df_administration_CPG['Campo'] == 'Date of dosing']
    df_administration_CPG = df_administration_CPG[['Participante','Valor']]
    df_administration_CPG = df_administration_CPG.rename(columns={'Participante':'Subject', 'Valor':'date_dosing_CPG'})

    df_administration_miltefosine = df_root[df_root['name']==  'Miltefosine Administration']
    df_administration_miltefosine = df_administration_miltefosine[['Visit','Participante', 'Campo', 'Valor']]
    df_administration_miltefosine = df_administration_miltefosine[df_administration_miltefosine['Campo'] == 'Date of dosing']
    df_administration_miltefosine['Valor'] = pd.to_datetime(df_administration_miltefosine['Valor'], format='%d-%b-%Y')
    max_per_participante = df_administration_miltefosine.groupby('Participante')['Valor'].transform('min')
    df_administration_miltefosine = df_administration_miltefosine[df_administration_miltefosine['Valor'] == max_per_participante]
    df_administration_miltefosine['Valor'] = pd.to_datetime(df_administration_miltefosine['Valor']).dt.strftime('%d-%b-%Y').str.upper()
    df_administration_miltefosine = df_administration_miltefosine[['Participante','Valor']].drop_duplicates()
    df_administration_miltefosine = df_administration_miltefosine.rename(columns={'Participante':'Subject', 'Valor':'date_dosing_miltefosine'})

    # df_date_visit_maxima = df_root[df_root['name']== 'Date of visit']
    # df_date_visit_maxima = df_date_visit_maxima[['Visit','Participante', 'Campo', 'Valor']]
    # df_date_visit_maxima = df_date_visit_maxima[df_date_visit_maxima['Campo']== 'Visit Date']
    # df_date_visit_maxima['Valor'] = pd.to_datetime(df_date_visit_maxima['Valor'], format='%d-%b-%Y')
    # max_per_participante = df_date_visit_maxima.groupby('Participante')['Valor'].transform('max')
    # df_date_visit_maxima = df_date_visit_maxima[df_date_visit_maxima['Valor'] == max_per_participante]
    # df_date_visit_maxima['Valor'] = pd.to_datetime(df_date_visit_maxima['Valor']).dt.strftime('%d-%b-%Y').str.upper()
    # df_date_visit_maxima = df_date_visit_maxima[['Participante', 'Valor' ]]
    # df_date_visit_maxima = df_date_visit_maxima.rename(columns={'Participante':'Subject', 'Valor':'fecha_visita_maxima'})


    df_was_completed_cpg = df_root[df_root['name']=='End of Study Treatment (Miltefosine)']
    df_was_completed_cpg = df_was_completed_cpg[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_was_completed_cpg = df_was_completed_cpg[df_was_completed_cpg['Variable'] == 'DSCPGCMP']
    df_was_completed_cpg = df_was_completed_cpg[['Participante', 'Valor']]
    df_was_completed_cpg = df_was_completed_cpg.rename(columns={'Participante':'Subject', 'Valor':'was_completed_cpg'})

    df_was_completed_cpg_reason = df_root[df_root['name']=='End of Study Treatment (Miltefosine)']
    df_was_completed_cpg_reason = df_was_completed_cpg_reason[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_was_completed_cpg_reason = df_was_completed_cpg_reason[df_was_completed_cpg_reason['Variable'] == 'DSCPGCAT']
    df_was_completed_cpg_reason = df_was_completed_cpg_reason[['Participante', 'Valor']]
    df_was_completed_cpg_reason = df_was_completed_cpg_reason.rename(columns={'Participante':'Subject', 'Valor':'was_completed_cpg_reason'})

    df_cpg_dosing_event_permanentely = df_root[df_root['name']=='CpG ODN D35 Administration']
    df_cpg_dosing_event_permanentely = df_cpg_dosing_event_permanentely[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_cpg_dosing_event_permanentely = df_cpg_dosing_event_permanentely[df_cpg_dosing_event_permanentely['Variable'] == 'ECCPGCINTD']
    df_cpg_dosing_event_permanentely = df_cpg_dosing_event_permanentely[df_cpg_dosing_event_permanentely['Valor'].isin([3, 3.0, '3'])]
    df_cpg_dosing_event_permanentely = df_cpg_dosing_event_permanentely[['Participante', 'Valor']]
    df_cpg_dosing_event_permanentely = df_cpg_dosing_event_permanentely.rename(columns={'Participante':'Subject', 'Valor':'cpg_dosing_event_permanentely'})

    df_cpg_dosing_event_temporarily = df_root[df_root['name']=='CpG ODN D35 Administration']
    df_cpg_dosing_event_temporarily = df_cpg_dosing_event_temporarily[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_cpg_dosing_event_temporarily = df_cpg_dosing_event_temporarily[df_cpg_dosing_event_temporarily['Variable'] == 'ECCPGCINTD']
    df_cpg_dosing_event_temporarily = df_cpg_dosing_event_temporarily[df_cpg_dosing_event_temporarily['Valor'].isin([2, 2.0, '2'])]
    df_cpg_dosing_event_temporarily = df_cpg_dosing_event_temporarily[['Participante', 'Valor']]
    df_cpg_dosing_event_temporarily = df_cpg_dosing_event_temporarily.rename(columns={'Participante':'Subject', 'Valor':'cpg_dosing_event_temporarily'})

    df_cpg_dose_mg = df_root[df_root['name']=='CpG ODN D35 Administration']
    df_cpg_dose_mg = df_cpg_dose_mg[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_cpg_dose_mg = df_cpg_dose_mg[df_cpg_dose_mg['Variable'] == 'ECCPGDOS']
    df_cpg_dose_mg = df_cpg_dose_mg[df_cpg_dose_mg['Valor'].astype('float') != 0.0]
    df_cpg_dose_mg = df_cpg_dose_mg[['Participante', 'Valor']]
    df_cpg_dose_mg = df_cpg_dose_mg.rename(columns={'Participante':'Subject', 'Valor':'cpg_dose_mg'})

    df_was_completed_miltefosine = df_root[df_root['name']=='End of Study Treatment (Miltefosine)']
    df_was_completed_miltefosine = df_was_completed_miltefosine[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_was_completed_miltefosine = df_was_completed_miltefosine[df_was_completed_miltefosine['Variable'] == 'DSMILCMP']
    df_was_completed_miltefosine = df_was_completed_miltefosine[['Participante', 'Valor']]
    df_was_completed_miltefosine = df_was_completed_miltefosine.rename(columns={'Participante':'Subject', 'Valor':'was_completed_miltefosine'})

    df_was_completed_miltefosine_reason = df_root[df_root['name']=='End of Study Treatment (Miltefosine)']
    df_was_completed_miltefosine_reason = df_was_completed_miltefosine_reason[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_was_completed_miltefosine_reason = df_was_completed_miltefosine_reason[df_was_completed_miltefosine_reason['Variable'] == 'DSMILCAT']
    df_was_completed_miltefosine_reason = df_was_completed_miltefosine_reason[['Participante', 'Valor']]
    df_was_completed_miltefosine_reason = df_was_completed_miltefosine_reason.rename(columns={'Participante':'Subject', 'Valor':'was_completed_miltefosine_reason'})
        
    df_miltefosine_dosing_event_permanentely = df_root[df_root['name']== 'Miltefosine Administration']
    df_miltefosine_dosing_event_permanentely = df_miltefosine_dosing_event_permanentely[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_miltefosine_dosing_event_permanentely = df_miltefosine_dosing_event_permanentely[df_miltefosine_dosing_event_permanentely['Variable'] == 'ECMLTCINTD']
    df_miltefosine_dosing_event_permanentely = df_miltefosine_dosing_event_permanentely[df_miltefosine_dosing_event_permanentely['Valor'].isin([3, 3.0, '3'])]
    df_miltefosine_dosing_event_permanentely = df_miltefosine_dosing_event_permanentely[['Participante', 'Valor']]
    df_miltefosine_dosing_event_permanentely = df_miltefosine_dosing_event_permanentely.rename(columns={'Participante':'Subject', 'Valor':'miltefosine_dosing_event_permanentely'})

    df_miltefosine_dosing_event_temporarily = df_root[df_root['name']== 'Miltefosine Administration']
    df_miltefosine_dosing_event_temporarily = df_miltefosine_dosing_event_temporarily[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_miltefosine_dosing_event_temporarily = df_miltefosine_dosing_event_temporarily[df_miltefosine_dosing_event_temporarily['Variable'] == 'ECMLTCINTD']
    df_miltefosine_dosing_event_temporarily = df_miltefosine_dosing_event_temporarily[df_miltefosine_dosing_event_temporarily['Valor'].isin([2, 2.0, '2'])]
    df_miltefosine_dosing_event_temporarily = df_miltefosine_dosing_event_temporarily[['Participante', 'Valor']]
    df_miltefosine_dosing_event_temporarily = df_miltefosine_dosing_event_temporarily.rename(columns={'Participante':'Subject', 'Valor':'miltefosine_dosing_event_temporarily'})

    df_miltefosine_dose_mg = df_root[df_root['name']== 'Miltefosine Administration']
    df_miltefosine_dose_mg = df_miltefosine_dose_mg[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_miltefosine_dose_mg = df_miltefosine_dose_mg[df_miltefosine_dose_mg['Variable'] == 'ECMLTDOS']
    df_miltefosine_dose_mg = df_miltefosine_dose_mg[df_miltefosine_dose_mg['Valor'].astype('float') != 0.0]
    df_miltefosine_dose_mg = df_miltefosine_dose_mg[['Participante', 'Valor']]
    df_miltefosine_dose_mg['miltefosine_dose_mg'] = df_miltefosine_dose_mg['Valor'].min()
    df_miltefosine_dose_mg = df_miltefosine_dose_mg[['Participante' ,'miltefosine_dose_mg']].drop_duplicates()
    df_miltefosine_dose_mg = df_miltefosine_dose_mg.rename(columns={'Participante':'Subject'})

    # df_concomitant_medication = df_root[df_root['name']== 'Prior And Concomitant Medications']
    # df_concomitant_medication = df_concomitant_medication[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    # df_concomitant_medication = df_concomitant_medication[df_concomitant_medication['Variable'] == 'CMTRT']
    # df_concomitant_medication = df_concomitant_medication[['Participante', 'Valor']].drop_duplicates()
    # df_concomitant_medication = df_concomitant_medication.rename(columns={'Participante':'Subject', 'Valor':'concomitant_medication_name'})

    df_concomitant_medication = df_root[df_root['name']== 'Prior And Concomitant Medications']
    df_concomitant_medication = df_concomitant_medication[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_concomitant_medication = df_concomitant_medication[df_concomitant_medication['Variable'].isin(['CMTRT'])]
    df_concomitant_medication['concomitant_medication_name'] = df_concomitant_medication.groupby('Participante')['Valor'].transform('count')
    df_concomitant_medication = df_concomitant_medication[['Participante', 'concomitant_medication_name']].drop_duplicates()
    df_concomitant_medication = df_concomitant_medication.rename(columns={'Participante':'Subject'})


    df_concomitant_procedures = df_root[df_root['name']== 'Prior And Concomitant Procedures']
    df_concomitant_procedures = df_concomitant_procedures[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_concomitant_procedures = df_concomitant_procedures[df_concomitant_procedures['Variable'] == 'PRTRT']
    df_concomitant_procedures = df_concomitant_procedures[['Participante', 'Valor']]
    df_concomitant_procedures = df_concomitant_procedures.rename(columns={'Participante':'Subject', 'Valor':'concomitant_procedure_name'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSSTCMP']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_general'})

    df_end_study_general_primary_reason = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general_primary_reason = df_end_study_general_primary_reason[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general_primary_reason = df_end_study_general_primary_reason[df_end_study_general_primary_reason['Variable'] == 'DSCAT']
    df_end_study_general_primary_reason = df_end_study_general_primary_reason[['Participante', 'Valor']]
    df_end_study_general_primary_reason = df_end_study_general_primary_reason.rename(columns={'Participante':'Subject', 'Valor':'end_study_general_primary_reason'})

    df_end_study_general_early = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general_early = df_end_study_general_early[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general_early = df_end_study_general_early[df_end_study_general_early['Variable'] == 'DSDAT']
    df_end_study_general_early = df_end_study_general_early[['Participante', 'Valor']]
    df_end_study_general_early = df_end_study_general_early.rename(columns={'Participante':'Subject', 'Valor':'end_study_date_early'})

    lista_revision = []
    lista_logs = ['Adverse Events']
    

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]
        sujeto_principal = sujeto_principal.sort_values(by=['FormFieldInstance Id'], ascending=True)
        sujeto_principal = sujeto_principal.reset_index(drop=True)

        adverse_events_id_review = []
        list_of_tuples_adverse_id = []

        # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, 
        # por lo que usamos el siguiente codigo para realizar la particion

        date_indices = sujeto_principal.index[sujeto_principal['Campo'] == 'Adverse Event ID'].tolist()
        subdatasets = [sujeto_principal.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]

        for subdataset in subdatasets:
            pru_1 = subdataset
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = 'unscheduled'
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_medical_surgical, on=['Subject'], how='left')
            pru = pru.merge(df_medical_eligibility, on=['Subject'], how='left')
            pru = pru.merge(df_medical_eligibility_date, on=['Subject'], how='left')
            pru = pru.merge(df_administration_CPG, on=['Subject'], how='left')
            pru = pru.merge(df_administration_miltefosine, on=['Subject'], how='left')
            pru = pru.merge(df_was_completed_cpg, on=['Subject'], how='left')  
            pru = pru.merge(df_was_completed_cpg_reason, on=['Subject'], how='left')
            pru = pru.merge(df_cpg_dosing_event_permanentely, on=['Subject'], how='left')
            pru = pru.merge(df_cpg_dosing_event_temporarily, on=['Subject'], how='left')
            pru = pru.merge(df_cpg_dose_mg, on=['Subject'], how='left')
            pru = pru.merge(df_was_completed_miltefosine, on=['Subject'], how='left') 
            pru = pru.merge(df_was_completed_miltefosine_reason, on=['Subject'], how='left')
            pru = pru.merge(df_miltefosine_dosing_event_permanentely, on=['Subject'], how='left')
            pru = pru.merge(df_miltefosine_dosing_event_temporarily, on=['Subject'], how='left') #
            pru = pru.merge(df_miltefosine_dose_mg, on=['Subject'], how='left') #
            pru = pru.merge(df_concomitant_medication, on=['Subject'], how='left')
            pru = pru.merge(df_concomitant_procedures, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general_primary_reason, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general_early, on=['Subject'], how='left')

        

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                inform_consent_date = row['Inform_consent_date']
                relevant_medical_condition = row['relevant_medical_condition']
                subject_randomized = row['subject_randomized']
                date_decision_no_randomize = str(row['date_decision_not_randomize']).split(' ')[0]
                date_dosing_CPG = str(row['date_dosing_CPG']).split(' ')[0]
                date_dosing_miltefosine = str(row['date_dosing_miltefosine']).split(' ')[0]
                was_completed_CPG_y_n = row['was_completed_cpg']
                was_completed_cpg_reason = row['was_completed_cpg_reason']
                cpg_dosing_event_permanentely = row['cpg_dosing_event_permanentely']
                cpg_dosing_event_temporarily = row['cpg_dosing_event_temporarily']
                cpg_dose_mg = row['cpg_dose_mg']
                was_completed_miltefosine = row['was_completed_miltefosine']
                was_completed_miltefosine_reason = row['was_completed_miltefosine_reason']
                miltefosine_dosing_event_permanentely = row['miltefosine_dosing_event_permanentely']
                miltefosine_dosing_event_temporarily = row['miltefosine_dosing_event_temporarily']
                miltefosine_dose_mg = row['miltefosine_dose_mg']
                concomitant_medication_name = row['concomitant_medication_name']
                concomitant_procedure_name = row['concomitant_procedure_name']
                end_study_general = row['end_study_general']
                end_study_general_primary_reason = row['end_study_general_primary_reason']
                end_study_date_early = row['end_study_date_early']
                

                #if status == 'DATA_ENTRY_COMPLETE':
                if status != '':
  
                    try:
                        adverse_event_id = row['Adverse Event ID']
                        adverse_event_id_pure = adverse_event_id.split('|')[0]
                        adverse_event_id_form_field_instance = adverse_event_id.split('|')[1]
                        adverse_event_id_disname = adverse_event_id.split('|')[0]
                    except:
                        adverse_event_id_pure = ''
                        adverse_event_id_form_field_instance = 'This field does not have any data'
                        adverse_event_id_disname = 'Empty'

                    try:
                        adverse_events_reported_term = row['Adverse Event Reported Term']
                        adverse_events_reported_term_pure = adverse_events_reported_term.split('|')[0]
                        adverse_events_reported_term_form_field_instance = adverse_events_reported_term.split('|')[1]
                        adverse_events_reported_term_disname = adverse_events_reported_term.split('|')[2]
                    except:
                        adverse_events_reported_term_pure = ''
                        adverse_events_reported_term_form_field_instance = 'This field does not have any data'
                        adverse_events_reported_term_disname = 'Empty'
                    
                    try:
                        start_date = row['Start Date']
                        start_date_pure = str(start_date.split('|')[0]).split(' ')[0]
                        start_date_form_field_instnace = start_date.split('|')[1]
                        start_date_disname = start_date.split('|')[0]
                    except:
                        start_date_pure = ''
                        start_date_form_field_instnace = 'This field does not have any data'
                        start_date_disname = 'Empty'
                    # start_date = row['Start Date']
                    # start_date_pure = str(start_date.split('|')[0]).split(' ')[0]
                    # start_date_form_field_instnace = start_date.split('|')[1]
                    # start_date_disname = start_date.split('|')[0]
                    # print(start_date_pure)

                    try:
                        outcome = row['Outcome']
                        outcome_pure = outcome.split('|')[0]
                        outcome_form_field_intance = outcome.split('|')[1]
                        outcome_form_disname = outcome.split('|')[2]
                    except:
                        outcome_pure = ''
                        outcome_form_field_intance = 'This field does not have any data'
                        outcome_form_disname = 'Empty'
                    
                    try:
                        severity = row['Severity']
                        severity_pure = severity.split('|')[0]
                        severity_form_field_instance = severity.split('|')[1]
                        severity_disname = severity.split('|')[2]
                    except:
                        severity_pure = math.nan
                        severity_form_field_instance = 'This field does not have any data'
                        severity_disname = 'Empty'
                    
                    try:
                        frequency = row['Frequency']
                        frequency_pure = frequency.split('|')[0]
                        frequency_form_field_instance = frequency.split('|')[1]
                        frequency_form_disname = frequency.split('|')[0]
                    except:
                        frequency_pure = ''
                        frequency_form_field_instance = 'This field does not have any data'
                        frequency_form_disname = 'Empty'
                    
                    try:
                        serious = row['Serious']
                        serious_pure = serious.split('|')[0]
                        serious_form_field_instance = serious.split('|')[1]
                        serious_form_disname = serious.split('|')[2]
                    except:
                        serious_pure = math.nan
                        serious_form_field_instance = 'This field does not have any data'
                        serious_form_disname = 'Empty'

                    try:
                        seriousness_criteria = row["Seriousness Criteria"]
                        seriousness_criteria_pure = seriousness_criteria.split('|')[0]
                        seriousness_criteria_form_field_instance = seriousness_criteria.split('|')[1]
                        seriousness_criteria_disname = seriousness_criteria.split('|')[2]
                    except:
                        seriousness_criteria_pure = ''
                        seriousness_criteria_form_field_instance = 'This field does not have any data'
                        seriousness_criteria_disname = 'Empty'

                    try:
                        causal_relation_cpg = row['Causal relationship with study treatment (CPG ODN D35)']
                        causal_relation_cpg_pure = causal_relation_cpg.split('|')[0]
                        causal_relation_cpg_form_field_instance = causal_relation_cpg.split('|')[1]
                        causal_relation_cpg_disname = causal_relation_cpg.split('|')[2]
                    except:
                        causal_relation_cpg_pure = ''
                        causal_relation_cpg_form_field_instance = 'This field does not have any data'
                        causal_relation_cpg_disname = 'Empty'

                    try:
                        causal_realation_miltefosine = row['Causal relationship with study treatment (Miltefosine)']
                        causal_realation_miltefosine_pure = causal_realation_miltefosine.split('|')[0]
                        causal_realation_miltefosine_form_field_instance = causal_realation_miltefosine.split('|')[1]
                        causal_realation_miltefosine_disname = causal_realation_miltefosine.split('|')[2]
                    except:
                        causal_realation_miltefosine_pure = ''
                        causal_realation_miltefosine_form_field_instance = 'This field does not have any data'
                        causal_realation_miltefosine_disname = 'Empty'
                    try:
                        actions_taken_cpg = row['Action taken with study treatment (CPG ODN D35)']
                        actions_taken_cpg_pure = actions_taken_cpg.split('|')[0]
                        actions_taken_cpg_form_field_instance = actions_taken_cpg.split('|')[1]
                        actions_taken_cpg_disname = actions_taken_cpg.split('|')[2]
                    except:
                        actions_taken_cpg_pure = ''
                        actions_taken_cpg_form_field_instance = 'This field does not have any data'
                        actions_taken_cpg_disname = 'Empty'

                    try:
                        actions_taken_miltefosine = row['Action taken with study treatment (Miltefosine)']
                        actions_taken_miltefosine_pure = actions_taken_miltefosine.split('|')[0]
                        actions_taken_miltefosine_form_field_instance = actions_taken_miltefosine.split('|')[1]
                        actions_taken_miltefosine_disname = actions_taken_miltefosine.split('|')[2]
                    except:
                        actions_taken_miltefosine_pure = ''
                        actions_taken_miltefosine_form_field_instance = 'This field does not have any data'
                        actions_taken_miltefosine_disname = 'Empty'

                    try:
                        other_action_taken = row['Other action taken']
                        other_action_taken_pure = other_action_taken.split('|')[0]
                        other_action_taken_form_field_instance = other_action_taken.split('|')[1]
                        other_action_taken_disname = other_action_taken.split('|')[2]
                    except:
                        other_action_taken_pure = ''
                        other_action_taken_form_field_instance = 'This field does not have any data'
                        other_action_taken_disname = 'Empty'
                    
                    try:
                        study_discontinued = row['Study discontinued']
                        study_discontinued_pure = study_discontinued.split('|')[0]
                        study_discontinued_form_field_instance = study_discontinued.split('|')[1]
                        study_discontinued_disname = study_discontinued.split('|')[2]
                    except:
                        study_discontinued_pure = ''
                        study_discontinued_form_field_instance = 'This field does not have any data'
                        study_discontinued_disname = 'Empty'
                    
                    try:
                        start_time = row['Start time']
                        start_time_pure = start_time.split('|')[0]
                        start_time_form_field_instance = start_time.split('|')[1]
                        start_time_disname = start_time.split('|')[0]
                    except:
                        start_time_pure = ''
                        start_time_form_field_instance = 'This field does not have any data'
                        start_time_disname = 'Empty'
                    
                    try:
                        end_date = row['End date']
                        end_date_pure = str(end_date.split('|')[0]).split(' ')[0]
                        end_date_form_field_instance = end_date.split('|')[1]
                        end_date_disname = end_date.split('|')[0]
                    except:
                        end_date_pure = ''
                        end_date_form_field_instance = 'This field does not have any data'
                        end_date_disname = 'Empty'
                    
                    try:
                        sae_start_date_AE_became_serious = row['SAE Start date (AE became serious)']
                        sae_start_date_AE_became_serious_pure = sae_start_date_AE_became_serious.split('|')[0]
                        sae_start_date_AE_became_serious_form_field_instance = sae_start_date_AE_became_serious.split('|')[1]
                        sae_start_date_AE_became_serious_disname = sae_start_date_AE_became_serious.split('|')[0]
                    except:
                        sae_start_date_AE_became_serious_pure = ''
                        sae_start_date_AE_became_serious_form_field_instance = 'This field does not have any data'
                        sae_start_date_AE_became_serious_disname = 'Empty'
                        # print(sae_start_date_AE_became_serious_pure)
                        # print('------------------')

                    # ---------------------------------------------------------------------------------------
                    if start_date_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(start_date_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Start date', start_date_form_field_instnace,\
                                        f , start_date_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    if end_date_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(end_date_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'End date', end_date_form_field_instance,\
                                        f , end_date_disname, 'GE0020']
                                lista_revision.append(error)     
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision AE0010
                    try:
                        if float(adverse_event_id_pure) in adverse_events_id_review:
                            error = [subject, visit, 'Adverse Event ID', adverse_event_id_form_field_instance, \
                                                'This value should be unique, it can not be repeated', \
                                                    adverse_event_id_disname, 'AE0010']
                            lista_revision.append(error)
                        else:
                            adverse_events_id_review.append(adverse_event_id_pure)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0010 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0020
                    try:
                        tuple_to_compare = (adverse_events_reported_term_pure, start_date_pure, end_date_pure)

                        if tuple_to_compare in list_of_tuples_adverse_id:
                            error =  [subject, visit, 'Adverse Event Reported Term', adverse_events_reported_term_form_field_instance, \
                                        'There are two adverse events that have the same term, and the dates overlap', \
                                            adverse_events_reported_term_disname, 'AE0020']
                            lista_revision.append(error)
                        else:
                            list_of_tuples_adverse_id.append(tuple_to_compare)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0040
                    try:
                        if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(inform_consent_date), '%d-%b-%Y'):
                            if float(relevant_medical_condition) == 1.0:
                                pass
                            else:
                                error = [subject, visit, 'Start Date', start_date_form_field_instnace, \
                                        'The start date is before the informed consent date. Please check and report in medical history if applicable.', \
                                            start_date_disname, 'AE0040']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0040 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0050
                    try:
                        if float(subject_randomized) == 0.0:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') >= datetime.strptime(str(date_decision_no_randomize), '%d-%b-%Y'):
                                error = [subject, visit, 'Start Date', start_date_form_field_instnace, \
                                        'The subject was not randomized, therefore the AE date can not be after the date of decision to not randomize the participant', \
                                            start_date_disname, 'AE0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0050 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0060
                    if str(end_date_pure) != 'nan' and str(end_date_pure) != '':
                        try:
                            if datetime.strptime(str(end_date_pure), '%d-%b-%Y') >= datetime.strptime(str(start_date_pure), '%d-%b-%Y'):
                                pass
                            else:
                                error = [subject, visit, 'End Date', end_date_form_field_instance, \
                                        'The date should be equal or grater than the start date', \
                                            end_date_disname, 'AE0060']
                                lista_revision.append(error) 
                        except Exception as e:
                            lista_logs.append(f'Revision AE0060 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> AE0070
                    if str(end_date_pure) != 'nan' and str(end_date_pure) != '':
                        try:
                            if datetime.strptime(str(end_date_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date_early), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'End Date', end_date_form_field_instance,\
                                        'End Date must be before the End of study/Early withdrawal date. ', end_date_disname, 'AE0070']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0070 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                    # Revision AE0100
                    if str(sae_start_date_AE_became_serious_pure) != 'nan' and str(sae_start_date_AE_became_serious_pure) != '':                            
                        try:
                            if datetime.strptime(str(sae_start_date_AE_became_serious_pure), '%d-%b-%Y') >= datetime.strptime(str(start_date_pure), '%d-%b-%Y'):
                                pass
                            else:
                                error = [subject, visit, 'SAE Start date (AE became serious)', sae_start_date_AE_became_serious_form_field_instance, \
                                        'The date must be the equal or later than the adverse event start date. ', \
                                            sae_start_date_AE_became_serious_disname, 'AE0100']
                                lista_revision.append(error) 
                        except Exception as e:
                            lista_logs.append(f'Revision AE0100 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0120
                    if str(date_dosing_CPG) != 'nan' and str(date_dosing_CPG) != '':       
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_CPG), '%d-%b-%Y'):
                                if float(causal_relation_cpg_pure) != 1.0:
                                    error = [subject, visit, 'Causal relationship with study treatment (CPG ODN D35)', causal_relation_cpg_form_field_instance, \
                                        'The AE started before the first administration of CpG ODN D35, therefore the causal relationship can not be different from "Not Related"', \
                                            causal_relation_cpg_disname, 'AE0120']
                                    lista_revision.append(error) 
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0120 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision AE0130
                    try:
                        if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_miltefosine), '%d-%b-%Y'):
                            if float(causal_realation_miltefosine_pure) != 1.0:
                                error = [subject, visit, 'Causal relationship with study treatment (Miltefosine)', causal_realation_miltefosine_form_field_instance, \
                                    'The AE started before the first administration of CpG ODN D35, therefore the causal relationship can not be different from "Not Related"', \
                                        causal_realation_miltefosine_disname, 'AE0130']
                                lista_revision.append(error) 
                            else:
                                pass
                    except Exception as e:
                        lista_logs.append(f'Revision AE0130 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0140
                    if str(date_dosing_CPG) != 'nan' and str(date_dosing_CPG) != '':    
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_CPG), '%d-%b-%Y'):
                                if float(actions_taken_cpg_pure) != 5.0:
                                    error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'The AE started before the first administration of CpG ODN D35, therefore the action taken can not be different from "not applicable"', \
                                            actions_taken_cpg_disname, 'AE0140']
                                    lista_revision.append(error) 
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0140 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0150
                    if isinstance(actions_taken_miltefosine_pure, (int, float, complex)):
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_miltefosine), '%d-%b-%Y'):
                                if float(actions_taken_miltefosine_pure) != 5.0:
                                    error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                        'The AE started before the first administration of Miltefosine, therefore the action taken can not be different from "not applicable"', \
                                            actions_taken_miltefosine_disname, 'AE0150']
                                    lista_revision.append(error) 
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0150 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0160
                    if str(date_dosing_CPG) != 'nan' and str(date_dosing_CPG) != '':    
                        try:
                            if datetime.strptime(str(end_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_CPG), '%d-%b-%Y'):
                                if float(actions_taken_cpg_pure) != 5.0:
                                    error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'The AE ended before the first administration of CpG ODN D35, therefore the action taken can not be different from "not applicable"', \
                                            actions_taken_cpg_disname, 'AE0160']
                                    lista_revision.append(error) 
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0160 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0170
                    if str(end_date_pure) != 'nan' and str(end_date_pure) != '' and str(date_dosing_miltefosine) != 'nan' and str(date_dosing_miltefosine) != '':    
                        try:
                            if datetime.strptime(str(end_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_miltefosine), '%d-%b-%Y'):
                                if float(actions_taken_miltefosine_pure) != 5.0:
                                    error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                        'The AE ended before the first administration of Miltefosine, therefore the action taken can not be different from "not applicable"', \
                                            actions_taken_miltefosine_disname, 'AE0170']
                                    lista_revision.append(error) 
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0170 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision AE0180
                    if str(date_dosing_CPG) != 'nan' and str(date_dosing_CPG) != '':    
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_CPG), '%d-%b-%Y'):
                                if end_date_pure == '':
                                    if float(actions_taken_cpg_pure) == 5.0:
                                        error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                            'The AE started before the first administration of CpG ODN D35, but its ongoing therefore the action taken can not be "not applicable"', \
                                                actions_taken_cpg_disname, 'AE0180']
                                        lista_revision.append(error) 
                                    else:
                                        pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0180 --> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision AE0190
                    if str(end_date_pure) != 'nan' and str(end_date_pure) != '' and str(date_dosing_miltefosine) != 'nan' and str(date_dosing_miltefosine) != '':    
                        try:
                            if datetime.strptime(str(end_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_miltefosine), '%d-%b-%Y'):
                                if end_date_pure == '':
                                    if float(actions_taken_miltefosine_pure) == 5.0:
                                        error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'The AE started before the first administration of Miltefosine but its ongoing, therefore the action taken can not be "not applicable"', \
                                                actions_taken_miltefosine_disname, 'AE0190']
                                        lista_revision.append(error) 
                                    else:
                                        pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0190 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0200
                    if str(date_dosing_CPG) != 'nan' and str(date_dosing_CPG) != '':    
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_CPG), '%d-%b-%Y'):
                                if datetime.strptime(str(end_date_pure), '%d-%b-%Y') >= datetime.strptime(str(date_dosing_CPG), '%d-%b-%Y'):
                                    if float(actions_taken_cpg_pure) == 5.0:
                                        error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                            'The first administration of CpG ODN D35 was done during the AE, therefore the action taken can not be "not applicable"', \
                                                actions_taken_cpg_disname, 'AE0200']
                                        lista_revision.append(error) 
                                    else:
                                        pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0200 --> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision AE0210
                    if str(date_dosing_miltefosine) != 'nan' and str(date_dosing_miltefosine) != '' and str(end_date_pure) != 'nan' and str(end_date_pure) != '':     
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(date_dosing_miltefosine), '%d-%b-%Y'):
                                if datetime.strptime(str(end_date_pure), '%d-%b-%Y') >= datetime.strptime(str(date_dosing_miltefosine), '%d-%b-%Y'):
                                    if float(actions_taken_miltefosine_pure) == 5.0:
                                        error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'The AE started before the first administration of Miltefosine but its ongoing, therefore the action taken can not be "not applicable"', \
                                                actions_taken_miltefosine_disname, 'AE0210']
                                        lista_revision.append(error) 
                                    else:
                                        pass
                        except Exception as e:
                            lista_logs.append(f'Revision AE0210 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0220
                    try:
                        if float(causal_relation_cpg_pure) == 1.0 or float(causal_relation_cpg_pure) == 2.0:
                            if float(actions_taken_cpg_pure) == 2.0 or float(actions_taken_cpg_pure) == 3.0 or float(actions_taken_cpg_pure) == 4.0:
                                error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'Action taken with study treatment (CPG ODN D35) is Dose Reduced or CT drug interrupted (temporarily) or CT drug stopped (permanently),  causal relationship with the study treatment (CPG ODN D35) can not be  related or probably not related.', \
                                            actions_taken_cpg_disname, 'AE0220']
                                lista_revision.append(error) 
                    except Exception as e:
                        lista_logs.append(f'Revision AE0220 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0230
                    try:
                        if float(causal_realation_miltefosine_pure) == 1.0 or float(causal_realation_miltefosine_pure) == 2.0:
                            if float(actions_taken_miltefosine_pure) == 2.0 or float(actions_taken_miltefosine_pure) == 3.0 or float(actions_taken_miltefosine_pure) == 4.0:
                                error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                        'Action taken with study treatment (CPG ODN D35) is Dose Reduced or CT drug interrupted (temporarily) or CT drug stopped (permanently),  causal relationship with the study treatment (CPG ODN D35) can not be  related or probably not related.', \
                                            actions_taken_miltefosine_disname, 'AE0230']
                                lista_revision.append(error) 
                    except Exception as e:
                        lista_logs.append(f'Revision AE0230 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0240
                    try:
                        if float(actions_taken_cpg_pure) == 4.0:
                            if float(was_completed_CPG_y_n) == 1.0:
                                error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'If the action taken was to stop permanently the CT drug, the subject could not have completed the study per protocol in the end of study form', \
                                            actions_taken_cpg_disname, 'AE0240']
                                lista_revision.append(error) 

                    except Exception as e:
                        lista_logs.append(f'Revision AE0240 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0250
                    try:
                        if float(actions_taken_cpg_pure) == 4.0:
                            if float(was_completed_cpg_reason) != 2.0:
                                error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'If the action taken was to stop permanently the CT drug, on the end of study form, the "Primary reason for not completing the CPG ODN D35 study treatment" should be "SAE or intolerable AE." ', \
                                            actions_taken_cpg_disname, 'AE0250']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision AE0250 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0260
                    try:
                        if float(actions_taken_cpg_pure) == 4.0:
                            if float(cpg_dosing_event_permanentely) != 3.0:
                                error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'If Action taken with study treatment (CPG ODN D35) is CT drug stopped (permanently) , on the CPG ODN D35 study treatment administration form there should be at least one Dosing Event as permanently discontinued and with reason for dose adjustment as Adverse event', \
                                            actions_taken_cpg_disname, 'AE0260']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0260 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0270
                    try:
                        if float(actions_taken_cpg_pure) == 3.0:
                            if float(cpg_dosing_event_temporarily) != 2.0:
                                error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'If Action taken with study treatment (CPG ODN D35) is CT drug stopped (temporarily) , on the CPG ODN D35 study treatment administration form there should be at least one Dosing Event as temporarily discontinued and with reason for dose adjustment as Adverse event', \
                                            actions_taken_cpg_disname, 'AE0270']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0270 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0280
                    try:
                        if float(actions_taken_cpg_pure) == 2.0:
                            if math.isnan(float(cpg_dose_mg)):
                                error = [subject, visit, 'Action taken with study treatment (CPG ODN D35)', actions_taken_cpg_form_field_instance, \
                                        'If Action taken with study treatment (CPG ODN D35) is Dose Reduced, on the CPG ODN D35 study treatment administration form at least one Dose  should be different from 0 and with Reason for dose adjustment reported as Adverse event.', \
                                            actions_taken_cpg_disname, 'AE0280']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0280 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0290
                    if str(was_completed_miltefosine) == 'nan' and str(was_completed_miltefosine) == '':
                        try:
                            if float(actions_taken_miltefosine_pure) == 4.0:
                                if float(was_completed_miltefosine) ==1.0:
                                    error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'If Action taken with study treatment (CPG ODN D35) is Dose Reduced, on the CPG ODN D35 study treatment administration form at least one Dose  should be different from 0 and with Reason for dose adjustment reported as Adverse event.', \
                                                actions_taken_miltefosine_disname, 'AE0290']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0290 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0300
                    if str(was_completed_miltefosine) == 'nan' and str(was_completed_miltefosine) == '':
                        try:
                            if float(actions_taken_miltefosine_pure) == 4.0:
                                if float(was_completed_miltefosine_reason) != 2.0:
                                    error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'If Action taken with study treatment (Miltefosine) is CT drug stopped (permanently) , on the end of study form, the "Primary reason for not completing the Miltefosine study treatment" should be "SAE or intolerable AE." ', \
                                                actions_taken_miltefosine_disname, 'AE0300']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0300 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0310
                    if str(miltefosine_dosing_event_permanentely) == 'nan' and str(miltefosine_dosing_event_permanentely) == '':
                        try:
                            if float(actions_taken_miltefosine_pure) == 4.0:
                                if float(miltefosine_dosing_event_permanentely) != 3.0:
                                    error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'If Action taken with study treatment (Miltefosine) is CT drug stopped (permanently) , on the Miltefosine study treatment administration form there should be at least one Dosing Event as permanently discontinued and with reason for dose adjustment as Adverse event', \
                                                actions_taken_miltefosine_disname, 'AE0310']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0310 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0320
                    if str(miltefosine_dosing_event_temporarily) == 'nan' and str(miltefosine_dosing_event_temporarily) == '':
                        try:
                            if float(actions_taken_miltefosine_pure) == 3.0:
                                if float(miltefosine_dosing_event_temporarily) != 2.0:
                                    error =  [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'Action taken with study treatment (Miltefosine) is CT drug stopped (temporarily) and there is no Miltefosine study treatment administration records with Dosing Event equals to Temporarily discontinued with Reason for dose adjustment reported Adverse event.', \
                                                actions_taken_miltefosine_disname, 'AE0320']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0320 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0330
                    if isinstance(actions_taken_miltefosine_pure, (int, float, complex)):
                        try:
                            if float(actions_taken_miltefosine_pure) == 2.0:
                                if math.isnan(float(miltefosine_dose_mg)):
                                    error = [subject, visit, 'Action taken with study treatment (Miltefosine)', actions_taken_miltefosine_form_field_instance, \
                                            'Action taken with study treatment (Miltefosine) is Dose Reduced and there is no Miltefosine study treatment administration records with Dose different from 0 with Reason for dose adjustment reported Adverse event.', \
                                                actions_taken_miltefosine_disname, 'AE0330']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0330 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0340
                    if isinstance(other_action_taken_pure, (int, float, complex)):
                        try:
                            if float(other_action_taken_pure) == 2.0:
                                # if str(concomitant_medication_name) != 'nan' and str(concomitant_medication_name) != '':
                                if float(concomitant_medication_name) == 0.0 or math.isnan(float(concomitant_medication_name)):

                                    error = [subject, visit, 'Other action taken', other_action_taken_form_field_instance, \
                                            'If Other action taken="Concomitant medication" at least one concomitant medication form must be completed', \
                                                other_action_taken_disname, 'AE0340']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0340 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision AE0350
                    if isinstance(other_action_taken_pure, (int, float, complex)):
                        try:
                            if float(other_action_taken_pure) == 3.0:
                                if math.isnan(float(concomitant_procedure_name)):
                                    error = [subject, visit, 'Other action taken', other_action_taken_form_field_instance, \
                                            'If Other action taken="Concomitant procedure" at least one concomitant procedure form must be completed', \
                                                other_action_taken_disname, 'AE0350']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0350 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0360
                    try:
                        if float(outcome_pure) == 5.0:
                            if float(serious_pure) != 1.0:
                                error = [subject, visit, 'Outcome ', outcome_form_field_intance, \
                                        'If the outcome is fatal, the AE must be serious', \
                                            outcome_form_disname, 'AE0360']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0360 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0370
                    try: 
                        if float(outcome_pure) == 5.0:
                            if float(seriousness_criteria_pure) != 1:
                                error = [subject, visit, 'Outcome ', outcome_form_field_intance, \
                                        'If value is "Fatal", The "Seriousness criteria" should be "results in death" ', \
                                            outcome_form_disname, 'AE0370']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0370 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0380
                    try:
                        if float(outcome_pure) == 5.0:
                            if float(severity_pure) != 5.0:
                                error = [subject, visit, 'Outcome ', outcome_form_field_intance, \
                                        'If value is "Fatal", The "Severity" should be "Death" ', \
                                            outcome_form_disname, 'AE0380']  
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0380 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0390
                    try:
                        if float(outcome_pure) == 6.0:
                            if float(end_study_general) != 1.0:
                                error = [subject, visit, 'Outcome ', outcome_form_field_intance, \
                                        'Taking into account that the subject has finished the study please verify if the outcome is truly "Unknown" or not', \
                                            outcome_form_disname, 'AE0390']  
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0390 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision AE0400
                    if str(seriousness_criteria_pure) != 'nan' and str(seriousness_criteria_pure) != '' and str(outcome_pure) != 'nan' and str(outcome_pure) != '':
                        try:
                            if float(seriousness_criteria_pure) == 1.0:
                                if float(outcome_pure) != 5.0:
                                    error = [subject, visit, 'Seriousness Criteria ', seriousness_criteria_form_field_instance, \
                                            'If value is "results in death", the "Outcome" should be "Fatal"', \
                                                seriousness_criteria_disname, 'AE0400']  
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision AE0400 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision AE0410
                    try:
                        if float(severity_pure) == 5.0:
                            if float(outcome_pure) != 5.0:
                                error = [subject, visit, 'Severity', severity_form_field_instance, \
                                        'If value is "Death", the "outcome" should be "Fatal"', \
                                            severity_disname, 'AE0410']  
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0410 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision AE0420
                    try:
                        if float(study_discontinued_pure) == 1.0:
                            if float(end_study_general) == 1.0:
                                error = [subject, visit, 'Study discontinued', study_discontinued_form_field_instance, \
                                        'The Study was discontinued therefore ,the study must not have been completed per protocol in the end of study form', \
                                            study_discontinued_disname, 'AE0420']  
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0420 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision AE0430
                    try:
                        if float(study_discontinued_pure) == 1.0:
                            if float(end_study_general_primary_reason) != 2.0:
                                error = [subject, visit, 'Study discontinued', study_discontinued_form_field_instance, \
                                        'The Study was discontinued therefore ,the primary reason for not completing the study should be "SAE or intolerable AE"', \
                                            study_discontinued_disname, 'AE0430']  
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision AE0430 --> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    adverse_events_output = pd.DataFrame(lista_revision, columns=column_names)

 
    sheet = excel_writer.create_sheet('Adverse Events')

    for row in dataframe_to_rows(adverse_events_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return adverse_events_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\data\88781e53-0c6e-42bc-8d84-c01f1015cb4f.xlsx")
    df_root.rename(columns = {'Instancia':'FormFieldInstance Id'}, inplace = True)
    adverse_events(df_root, path_excel ) 

