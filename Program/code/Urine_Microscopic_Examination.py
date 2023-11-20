from datetime import datetime
from log_writer import log_writer
import numpy as np
from revision_fechas import revision_fecha
import warnings
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

def urine_microscopic_examination(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Urine Microscopic Examination
    '''

    df= df_root[df_root['name']== 'Urine Microscopic Examination']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)

    lista_revision = []
    lista_logs = ['Urine Microscopic Examination']
    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                if status != '':
                    try:
                        was_performed = row['Was the urine microscopic examination performed?']
                        was_performed_pure = was_performed.split('|')[0]
                        was_performed_form_field_instance = was_performed.split('|')[1]
                    except Exception as e:
                        was_performed_pure = math.nan
                        was_performed_form_field_instance  = 'This field does not have any data'


                    # ------------------------------------------------------- 

                    lista_validacion = [
                        'RBC',
                        'WBC',
                        'Epithelial Cells',
                        'Crystals',
                        'Casts',
                        'Bacteria',
                    ]
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try: 
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = ''

                        if str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass
                    
                    # Revision URM0010
                    try:
                        if float(was_performed_pure) == 1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Was the urine microscopic examination performed?', was_performed_form_field_instance,\
                                         'If Urine Sample Collected is checked as "Yes", not all laboratory tests can be "not done"', \
                                            was_performed_pure, 'URM0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision URM0010--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    lista_validacion_resultados = [
                    'RBC, Result (/hpf)',
                    'WBC, Result (/hpf)',
                    'Epithelial Cells, Result (/hpf)',
                    'Crystals, Result',
                    'Casts, Result (/lpf)',
                    'Bacteria, Result',
                    ]
                    mi_cuenta_resultados = 0

                    for validador_raw_resultados in lista_validacion_resultados:
                        try: 
                            validador_contador = row[validador_raw_resultados].split('|')[0]
                        except:
                            validador_contador = ''

                        if validador_contador:
                            mi_cuenta_resultados +=1


                    # Revision URM0020
                    try:
                        if float(was_performed_pure) == 1.0: 
                            if mi_cuenta_resultados != 0:
                                pass
                            else:
                                error = [subject, visit, 'Was the urine microscopic examination performed?', was_performed_form_field_instance,\
                                         'None of the result of the urinalysis form is >=+, therefore this examination is not required', \
                                            was_performed_pure, 'URM0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision URM0020--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision URM0030
                    try:
                        if float(was_performed_pure) == 0.0: 
                            if mi_cuenta_resultados == 0:
                                pass
                            else:
                                error = [subject, visit, 'Was the urine microscopic examination performed?', was_performed_form_field_instance,\
                                         'if Was the urine microscopic examination performed? ="No" and any of the results from the urinalysis at the same visit is >= +', \
                                            was_performed_pure, 'URM0030']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision URM0030--> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    urine_microscopic_examination_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("Urine Microscopic Examination")

    for row in dataframe_to_rows(urine_microscopic_examination_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return urine_microscopic_examination_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx')
    urine_microscopic_examination(df_root, path_excel) 