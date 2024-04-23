from datetime import datetime
import os
import math
from log_writer import log_writer
import numpy as np
from revision_fechas import revision_fecha
import warnings
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')



def clinical_laboratory_test_coagulation(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Clinical Laboratory Test - Coagulation
    '''
    # Normals ranges file

    script_directory = os.path.dirname(os.path.abspath(__file__)) if '__file__' in locals() else os.getcwd()
    relative_folder_path = r"data\rangos_normales"
    folder_path = os.path.join(script_directory.replace('\code', ''), relative_folder_path)
    file = os.listdir(folder_path)
    path = f"{folder_path}\{[x for x in file if 'cl_coagulation' in x][0]}" 
    df_normal_ranges = pd.read_csv(path, sep=';')


    df= df_root[df_root['name']== 'Clinical Laboratory Test - Coagulation']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)


    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante',  'Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Clinical Laboratory Test - Coagulation']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            #print(pru)
            # print('------------')

            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                subject = row['Subject']
                visit = row['Visit']
                status = row['status']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                if status != '':
                    try:
                        blood_sample_collected = row['Blood Sample Collected']
                        blood_sample_collected_pure = blood_sample_collected.split('|')[0]
                        blood_sample_collected_form_field_instance = blood_sample_collected.split('|')[1]
                        blood_sample_collected_disname = blood_sample_collected.split('|')[2]
                    except Exception as e:
                        blood_sample_collected_pure = math.nan
                        blood_sample_collected_form_field_instance = 'This field does not have any data'
                        blood_sample_collected_disname = 'Empty'

                    try:
                        date_collected = row['Date Collected']
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1] 
                        date_collected_disname = date_collected.split('|')[0] 
                    except Exception as e:
                        date_collected_pure = ''
                        blood_sample_collected_form_field_instance = 'This field does not have any data'
                        date_collected_disname = 'Empty'

                    # try:
                    #     provide_reason = row['Provide the reason']
                    # except Exception as e:
                    #     pass        

                    try:
                        INR = row['INR']
                        INR_pure = INR.split('|')[0]
                        INR_form_field_instance = INR.split('|')[1]
                        INR_disname = INR.split('|')[0]
                    except Exception as e:
                        INR_pure = math.nan
                        INR_form_field_instance = 'This field does not have any data'
                        INR_disname = 'Empty'

                    # try:
                    #     INR_specify = row['INR, If abnormal, Specify']
                    # except Exception as e:
                    #     pass        

                    try:
                        INR_out_normal_range = row['INR, Out of normal range?']
                        INR_out_normal_range_pure = INR_out_normal_range.split('|')[0]
                        INR_out_normal_range_form_field_instance = INR_out_normal_range.split('|')[1]
                        INR_out_normal_range_disname = INR_out_normal_range.split('|')[2]
                    except Exception as e:
                        INR_out_normal_range_pure = math.nan
                        INR_out_normal_range_form_field_instance = 'This field does not have any data'
                        INR_out_normal_range_disname = 'Empty'

                    try:
                        INR_result = row['INR, Result']
                        INR_result_pure = INR_result.split('|')[0]
                        INR_result_form_field_instance = INR_result.split('|')[1]
                        INR_result_disname = INR_result.split('|')[0]
                    except Exception as e:
                        INR_result_pure = ''
                        INR_result_form_field_instance = 'This field does not have any data'
                        INR_result_disname = 'Empty'

                    try:
                        PT = row['PT']
                        PT_pure = PT.split('|')[0]
                        PT_form_field_isntance = PT.split('|')[1]
                        PT_disname = PT.split('|')[0]
                    except Exception as e:
                        PT_pure = math.nan
                        PT_form_field_isntance = 'This field does not have any data'
                        PT_disname = 'Empty'
                        
                    # try:
                    #     PT_specify = row['PT, If abnormal, Specify']
                    # except Exception as e:
                    #     pass        

                    try:
                        PT_out_normal_range = row['PT, Out of normal range?']
                        PT_out_normal_range_pure = PT_out_normal_range.split('|')[0]
                        PT_out_normal_range_form_field_instance = PT_out_normal_range.split('|')[1]
                        PT_out_normal_range_disname = PT_out_normal_range.split('|')[2]
                    except Exception as e:
                        PT_out_normal_range_pure = math.nan
                        PT_out_normal_range_form_field_instance   = 'This field does not have any data'
                        PT_out_normal_range_disname = 'Empty'

                    try:
                        PT_result = row['PT, Result (Seconds)']
                        PT_result_pure = PT_result.split('|')[0]
                        PT_result_form_field_instance = PT_result.split('|')[1]
                        PT_result_disname = PT_result.split('|')[0]
                    except Exception as e:
                        PT_result_pure = math.nan     
                        PT_result_form_field_instance = 'This field does not have any data'
                        PT_result_disname = 'Empty'

                    try:
                        aPTT = row['aPTT']
                        aPTT_pure = aPTT.split('|')[0]
                        aPTT_form_field_instance = aPTT.split('|')[1]
                        aPTT_disname = aPTT.split('|')[0]
                    except Exception as e:
                        aPTT_pure = math.nan
                        aPTT_form_field_instance = 'This field does not have any data'
                        aPTT_disname = 'Empty'

                    # try:
                    #     aPTT_specify = row['aPTT, If abnormal, Specify']
                    # except Exception as e:
                    #     pass        

                    try:
                        aPTT_out_normal_range =  row['aPTT, Out of normal range?']
                        aPTT_out_normal_range_pure = aPTT_out_normal_range.split('|')[0]
                        aPTT_out_normal_range_form_field_instance = aPTT_out_normal_range.split('|')[1]
                        aPTT_out_normal_range_disname = aPTT_out_normal_range.split('|')[2]
                    except Exception as e:
                        aPTT_out_normal_range_pure = math.nan
                        aPTT_out_normal_range_form_field_instance = 'This field does not have any data'
                        aPTT_out_normal_range_disname = 'Empty'

                    try:
                        aPTT_result = row['aPTT, Result (Seconds)']
                        aPTT_result_pure = aPTT_result.split('|')[0]
                        aPTT_result_form_field_instance = aPTT_result.split('|')[1]
                        aPTT_result_disname = aPTT_result.split('|')[0]
                    except Exception as e:
                        aPTT_result_pure = math.nan
                        aPTT_result_form_field_instance = 'This field does not have any data'
                        aPTT_result_disname = 'Empty'

                    # ------------------------------------------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date collected', date_collected_form_field_instance ,f , date_collected_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBO0010
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_collected_f != date_of_visit_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form', \
                                            f'{date_collected_disname} - {date_of_visit}', 'LBO0010']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBO0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBO0020
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_collected_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance,\
                                        'The date/time of test performed can not be before the informed consent date/time',\
                                            f'{date_collected_disname} - {date_inform_consent}', 'LBO0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBO0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LBO0030
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Visit Date', date_collected_form_field_instance ,'Date Collected must be before the End of study/Early withdrawal date. ', date_collected_disname, 'LBO0030']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LBO0030 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision LBO0050
                    try:
                        if  float(blood_sample_collected_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and the D-1 visit date =Screening visit date or normal and done in the previous 10 days', \
                                            blood_sample_collected_disname, 'LBO0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBO0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    lista_validacion =[
                        'INR',
                        'INR, If abnormal, Specify',
                        'INR, Out of normal range?',
                        'INR, Result',
                        'PT',
                        'PT, If abnormal, Specify',
                        'PT, Out of normal range?',
                        'PT, Result (Seconds)',
                        'Provide the reason',
                        'aPTT',
                        'aPTT, If abnormal, Specify',
                        'aPTT, Out of normal range?',
                        'aPTT, Result (Seconds)'
                    ]
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try:
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
                        
                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass
                    
                    # Revision LBO0060
                    try:
                        if float(blood_sample_collected_pure) ==1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance,\
                                         'If Blood Sample Collected is checked as "Yes", not all laboratory tests can be "not done"', \
                                            blood_sample_collected_disname, 'LBO0060']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBO0060--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBO0080
                        if float(aPTT_out_normal_range_pure) == 1.0:
                            #if float(aPTT_result_pure) > 23.6 and float(aPTT_result_pure) < 34.8 :
                            if float(aPTT_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="aPTT, Result (Seconds)"]['min'].iloc[0]) \
                                and float(aPTT_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="aPTT, Result (Seconds)"]['max'].iloc[0]) :
                                error = [subject, visit, 'aPTT, Out of normal range?', aPTT_out_normal_range_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review.',\
                                              f"aPPT Out Of Normal: {aPTT_result_disname} - aPPT Result: {aPTT_result_pure}", 'LBO0080']
                                lista_revision.append(error)

                        # Revision LBO0100
                        elif float(aPTT_out_normal_range_pure) == 0.0:
                            #if float(aPTT_result_pure) <  23.6  or float(aPTT_result_pure) > 34.8 :
                            if float(aPTT_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="aPTT, Result (Seconds)"]['min'].iloc[0]) \
                                or float(aPTT_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="aPTT, Result (Seconds)"]['max'].iloc[0]) :
                                error = [subject, visit, 'aPTT, Out of normal range?', aPTT_out_normal_range_form_field_instance,\
                                         'According to the result, the value is out of range, please review.',\
                                              f"aPPT Out Of Normal: {aPTT_result_disname} - aPPT Result: {aPTT_result_pure}", 'LBO0100']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBO0080--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBO0090
                        if float(PT_out_normal_range_pure) == 1.0:
                            #if float(PT_result_pure) > 11.7 and float(PT_result_pure) < 15.3 :
                            if float(PT_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="PT, Result (Seconds)"]['min'].iloc[0]) \
                                and float(PT_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="PT, Result (Seconds)"]['max'].iloc[0]) :
                                error = [subject, visit, 'PT, Out of normal range?', PT_out_normal_range_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review.',\
                                             f"PT Out Of Normal: {PT_result_disname} - PT Result: {PT_result_pure}", 'LBO0090']
                                              
                                lista_revision.append(error)

                        # Revision LBO0110
                        elif float(PT_out_normal_range_pure) == 0.0:
                            #if float(PT_result_pure) <  11.7  or float(PT_result_pure) > 15.3 :
                            if float(PT_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="PT, Result (Seconds)"]['min'].iloc[0]) \
                                or float(PT_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="PT, Result (Seconds)"]['max'].iloc[0]) :
                                error = [subject, visit, 'PT, Out of normal range? ', PT_out_normal_range_form_field_instance ,\
                                         'According to the result, the value is out of range, please review.',\
                                              f"PT Out Of Normal: {PT_result_disname} - PT Result: {PT_result_pure}", 'LBO0110']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBO0110--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBO0120
                        if float(INR_out_normal_range_pure) == 1.0:
                            #if  float(INR_result_pure) <= 1.1 :
                            if float(INR_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="INR, Result"]['min'].iloc[0]) \
                                and float(INR_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="INR, Result"]['max'].iloc[0]) :
                                error = [subject, visit, 'INR, Out of normal range?', INR_out_normal_range_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review.',\
                                             f"INR Out Of Normal: {INR_result_disname} - INR Result: {INR_result_pure}", 'LBO0120']
                                lista_revision.append(error)

                        # Revision LBO0130
                        elif float(INR_out_normal_range_pure) == 0.0:
                            #if float(INR_result_pure) >= 1.1 :
                            if float(INR_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="INR, Result"]['min'].iloc[0]) \
                                or float(INR_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="INR, Result"]['max'].iloc[0]) :
                                error = [subject, visit, 'INR, Out of normal range?', INR_out_normal_range_form_field_instance ,\
                                         'According to the result, the value is out of range, please review.',\
                                             f"INR Out Of Normal: {INR_result_disname} - INR Result: {INR_result_pure}", 'LBO0130']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBO0130--> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    clinical_laboratory_test_coagulation_output = pd.DataFrame(lista_revision, columns=column_names)
    clinical_laboratory_test_coagulation_output = clinical_laboratory_test_coagulation_output[~clinical_laboratory_test_coagulation_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]

    sheet = excel_writer.create_sheet("CL - Coagulation")

    for row in dataframe_to_rows(clinical_laboratory_test_coagulation_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return clinical_laboratory_test_coagulation_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    clinical_laboratory_test_coagulation(df_root, path_excel) 