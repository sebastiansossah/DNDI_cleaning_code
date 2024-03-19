from datetime import datetime
from log_writer import log_writer
import numpy as np
import math
from revision_fechas import revision_fecha
import warnings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

def lesion_measurement(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Lesion Measurement
    '''

    df= df_root[df_root['name']== 'Lesion Measurement']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Lesion Measurement']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            
            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                if status != '':
                    try:
                        Was_the_lesion_measurement_performed = row["Was the lesion measurement performed?"]
                        Was_the_lesion_measurement_performed_pure = Was_the_lesion_measurement_performed.split('|')[0]
                        Was_the_lesion_measurement_performed_form_field_instance = Was_the_lesion_measurement_performed.split('|')[1] 
                        Was_the_lesion_measurement_performed_disname = Was_the_lesion_measurement_performed.split('|')[2] 
                    except Exception as e:
                        Was_the_lesion_measurement_performed_pure = math.nan
                        Was_the_lesion_measurement_performed_form_field_instance = 'This field does not have any data'
                        Was_the_lesion_measurement_performed_disname = 'Empty'
    
                    try:
                        Date_of_assessment_performed = row["Date of assessment performed"]
                        Date_of_assessment_performed_pure = Date_of_assessment_performed.split('|')[0]
                        Date_of_assessment_performed_form_field_instance = Date_of_assessment_performed.split('|')[1]
                        Date_of_assessment_performed_disname = Date_of_assessment_performed.split('|')[0]
                    except Exception as e:
                        Date_of_assessment_performed_pure = ''
                        Date_of_assessment_performed_form_field_instance = 'This field does not have any data'
                        Date_of_assessment_performed_disname = 'Empty'

                    # try:
                    #     Provide_the_reason = row["Provide the reason"]
                    # except Exception as e:
                    #     pass   

                    # try:
                    #     Anatomical_Location = row["Anatomical Location"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Side = row["Side"]
                    # except Exception as e:
                    #     pass   
                            
                    # try:
                    #     Sagital_axis = row["Sagital axis"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Longest_diameter_of_lesion_in_mm = row["Longest diameter of lesion in mm"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Perpendicular_to_longest_diameter_of_lesion_in_mm = row["Perpendicular to longest diameter of lesion in mm"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Type_of_lesion = row["Type of lesion"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Mucosa_involved_in_lesion = row["Mucosa involved in lesion"]
                    # except Exception as e:
                    #     pass   
                           
                    # try:
                    #     Duration_of_the_lesion_in_weeks = row["Duration of the lesion in weeks"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Lesion_Photograph_taken = row["Lesion Photograph taken?"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Was_the_lesion_failure_criteria = row["Was the lesion failure criteria met (see section 12.2 of the study protocol)?"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Re_epithelialization_of_ulcerated_lesions_compared_to_the_Day = row["Re-epithelialization of ulcerated lesions compared to the Day-1"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Re_epithelialization_of_ulcerated_lesions_compared_to_the_Day_pertentaje = row["Re-epithelialization of ulcerated lesions compared to the Day-1 (%)"]
                    # except Exception as e:
                    #     pass   
    
                    # try:
                    #     Flattening_of_Non_Ulcerated_compared_to_the_Day = row["Flattening of Non –Ulcerated compared to the Day-1 (%)"]
                    # except Exception as e:
                    #     pass   

                    # ----------------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if Date_of_assessment_performed_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(Date_of_assessment_performed_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of assessment performed', Date_of_assessment_performed_form_field_instance ,f , Date_of_assessment_performed_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LM0010
                    try:
                        if  float(Was_the_lesion_measurement_performed_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the lesion measurement performed?', Was_the_lesion_measurement_performed_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and Screening visit date = D-1 date (screening done on D-1)', \
                                            Was_the_lesion_measurement_performed_disname, 'LM0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LM0010--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision LM0030
                    if Date_of_assessment_performed_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(Date_of_assessment_performed_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of assessment performed', Date_of_assessment_performed_form_field_instance,\
                                        'The date of assessment can not be before the informed consent date' , \
                                            f'{Date_of_assessment_performed_disname} - {date_of_visit}', 'LM0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LM0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LM0040
                    if Date_of_assessment_performed_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(Date_of_assessment_performed_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of assessment performed', Date_of_assessment_performed_form_field_instance, \
                                        'The date of assessment can not be before the informed consent date',f'{Date_of_assessment_performed_disname} - {date_inform_consent}', 'LM0040']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LM0040--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LM0050
                    if  str(end_study_date) == 'nan' or end_study_date == '' or Date_of_assessment_performed_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(Date_of_assessment_performed_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of assessment performed', Date_of_assessment_performed_form_field_instance ,'Date of assessment performed must be before the End of study/Early withdrawal date. ', Date_of_assessment_performed_disname, 'LM0050']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LM0050 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    lista_validacion = [
                                    'Anatomical Location',
                                    'Side',
                                    'Sagital axis',
                                    'Longest diameter of lesion in mm',
                                    'Perpendicular to longest diameter of lesion in mm',
                                    'Type of lesion',
                                    'Mucosa involved in lesion',
                                    'Duration of the lesion in weeks',
                                    'Lesion Photograph taken?',
                                    'Was the lesion failure criteria met (see section 12.2 of the study protocol)?',
                                    'Re-epithelialization of ulcerated lesions compared to the Day-1',
                                    'Re-epithelialization of ulcerated lesions compared to the Day-1 (%)']
                    
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try: 
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
                        
                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass
                    
                    # Revision LM0060
                    try:
                        if float(Was_the_lesion_measurement_performed_pure) == 1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Was the lesion measurement performed?', Was_the_lesion_measurement_performed_form_field_instance,\
                                         'If answer is Yes at least one "Lesion" section must be added', \
                                            Was_the_lesion_measurement_performed_disname, 'LM0060']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LM0060--> {e} - Subject: {subject},  Visit: {visit} ')

                    
                    # Revision LM0070
                    try:
                        if float(Was_the_lesion_measurement_performed_pure) == 0.0: 
                            if mi_cuenta == 0:
                                pass
                            else:
                                error = [subject, visit, 'Was the lesion measurement performed?', Was_the_lesion_measurement_performed_form_field_instance,\
                                         'If answer is "No", no "Lesion" section must be added', Was_the_lesion_measurement_performed_disname, 'LM0070']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LM0070--> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    lesion_measurement_output = pd.DataFrame(lista_revision, columns=column_names)
    lesion_measurement_output = lesion_measurement_output[~lesion_measurement_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
    
    sheet = excel_writer.create_sheet("Lesion Measurement")

    for row in dataframe_to_rows(lesion_measurement_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return lesion_measurement_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    lesion_measurement(df_root, path_excel) 