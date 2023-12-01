import pandas as pd
import math
from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


warnings.filterwarnings('ignore')

def vein_assesment(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Vein Assesment
    '''

    df= df_root[df_root['name']== 'Vein assessment']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})


    lista_logs = ['Vein Assesment']
    lista_revision = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
    
            for index, row in pru.iterrows():
                status = row['status']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                
        
                if status != '':
                    subject = row['Subject']
                    visit = row['Visit']
                    date_of_visit = row['Date_of_visit']
                    date_inform_consent = row['Informed_consent_date']
                    end_study_date = row['end_study_date']

                    try:
                        vein_assessment_performed = row['Was the vein assessment performed?']
                        vein_assessment_performed_pure = vein_assessment_performed.split('|')[0]
                        vein_assessment_performed_form_field_instance =  vein_assessment_performed.split('|')[1]
                        vein_assessment_performed_disname =  vein_assessment_performed.split('|')[2]
                    except Exception as e:
                        vein_assessment_performed_pure = math.nan
                        vein_assessment_performed_form_field_instance = 'This field does not have any data'
                        vein_assessment_performed_disname = 'Empty'

                    # try:
                    #     provide_reason = row['Provide the reason']
                    #     provide_reason_pure = provide_reason.split('|')[0]
                    #     provide_reason_form_field_intance = provide_reason.split('|')[1]
                    # except Exception as e:
                    #     provide_reason_pure = ''
                    #     provide_reason_form_field_intance = 'This field does not have any data'

                    try:
                        date_of_assesment = row['Date of assessment performed']
                        date_of_assesment_pure = date_of_assesment.split('|')[0]
                        date_of_assesment_form_field_instance = date_of_assesment.split('|')[1]
                    except Exception as e:
                        date_of_assesment_pure = ''
                        date_of_assesment_form_field_instance = 'This field does not have any data'

                    # try:
                    #     suitable_veins = row['Suitable veins for multiple venepunctures/cannulations found?']
                    #     suitable_veins_pure = suitable_veins.split('|')[0]
                    #     suitable_veins_form_field_isntance = suitable_veins.split('|')[1]
                    # except Exception as e:
                    #     suitable_veins_pure = ''
                    #     suitable_veins_form_field_isntance = 'This field does not have any data'
                    
                    # ----------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_of_assesment_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha -> GE0020
                            f = revision_fecha(date_of_assesment_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of assessment performed', date_of_assesment_form_field_instance ,f , date_of_assesment_pure, 'GE0020']
                                lista_revision.append(error)     
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> VA0050
                    if date_of_assesment_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_of_assesment_pure), '%d-%b-%Y') >= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of assessment performed', date_of_assesment_form_field_instance ,'Date of assessment performed must be before the End of study/Early withdrawal date. ', date_of_assesment_pure, 'VA0050']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision VA0050 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision VA0020
                    try:
                        if float(vein_assessment_performed_pure) == 9.0: 
                            if visita =='D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the vein assessment performed?', vein_assessment_performed_form_field_instance, \
                                         'The "Not Required" option can only be selected if visit is D-1 and Screening visit date = D-1 date (screening done on D-1)',\
                                              vein_assessment_performed_disname, 'VA0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision VA0020--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision VA0030
                    if date_of_assesment_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_assesment_f = datetime.strptime(date_of_assesment_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_assesment_f != date_of_visit_f:
                                error = [subject, visit, 'Date of assessment performed', date_of_assesment_form_field_instance, \
                                        'The date must be the same as the date of visit date', \
                                            f'{date_of_assesment_pure} - {date_of_visit}', 'VA0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision VA0030--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision VA0040
                    if date_of_assesment_pure == '':
                        pass
                    else:
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_assesment_f = datetime.strptime(date_of_assesment_pure, date_format)
                            date_inform_consentf = datetime.strptime(date_inform_consent, date_format)

                            if date_of_assesment_f < date_inform_consentf:
                                error = [subject, visit, 'Date of assessment performed', date_of_assesment_form_field_instance, \
                                        'The date of assessment can not be before the informed consent date',\
                                            f'{date_of_assesment_pure} - {date_inform_consent}', 'VA0040']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision VA0040--> {e} - Subject: {subject},  Visit: {visit} ')
                else:
                    pass

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    vein_assesment_output = pd.DataFrame(lista_revision, columns=column_names)

    
    sheet = excel_writer.create_sheet("Vein assessment")

    for row in dataframe_to_rows(vein_assesment_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return vein_assesment_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    vein_assesment(df_root, path_excel ) 