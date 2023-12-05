import pandas as pd
import math
from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings
pd.set_option('display.max_columns', None)
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



def adminsitration_CpG_ODN(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de CpG ODN D35 Administration
    '''

    df = df_root[df_root['name']=='CpG ODN D35 Administration'] 
    lista_sujetos = df['Participante'].unique()

    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_date_visit = df_root[df_root['name']== 'Date of visit']
    df_date_visit = df_date_visit[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit = df_date_visit[df_date_visit['Campo']== 'Visit Date']
    df_date_visit['to_join'] = df_date_visit['Valor']
    df_date_visit = df_date_visit[['Participante', 'to_join','Visit', 'Valor']]
    df_date_visit = df_date_visit.rename(columns={'Participante':'Subject' ,'Visit':'visita_para_comparar', 'Valor':'Date_of_visit_value'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Inform_consent_date'})

    df_date_visit_randomization = df_root[df_root['name']== 'Date of visit']
    df_date_visit_randomization = df_date_visit_randomization[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit_randomization = df_date_visit_randomization[df_date_visit_randomization['Campo']== 'Visit Date']
    df_date_visit_randomization = df_date_visit_randomization[df_date_visit_randomization['Visit']== 'D-1']
    df_date_visit_randomization = df_date_visit_randomization[['Participante','Valor']]
    df_date_visit_randomization = df_date_visit_randomization.rename(columns={'Participante':'Subject', 'Valor':'Visita_randomization'})

    df_adverse = df_root[df_root['name']=='Adverse Events']
    df_adverse = df_adverse[['Visit','Participante', 'Campo', 'Valor']]
    df_adverse = df_adverse[df_adverse['Campo']== 'Action taken with study treatment (CPG ODN D35)']
    df_adverse = df_adverse[['Participante','Valor']]
    df_adverse = df_adverse.rename(columns={'Participante':'Subject', 'Valor':'action_taken_study_treatment'})

    lista_revision = []
    lista_logs = ['CpG ODN D35 Administration']

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']== sujeto]
        sujeto_principal = sujeto_principal.sort_values(by=['FormFieldInstance Id'], ascending=True)
        sujeto_principal = sujeto_principal.reset_index(drop=True)

        # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, por lo que usamos el siguiente codigo para realizar la particion
        #p
        date_indices = sujeto_principal.index[sujeto_principal['Campo'] == 'Date of dosing'].tolist()
        subdatasets = [sujeto_principal.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]

        date_dosing_list_review = []

        for subdataset in subdatasets:

            pru = subdataset
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = 'unscheduled'
            pru['status'] = 'doesnt matter'
            pru['to_join'] = pru['Date of dosing'].str.split('|').str[0]
            pru = pru.merge(df_date_visit, on=['Subject', 'to_join'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_date_visit_randomization, on=['Subject'], how='left')
            pru = pru.merge(df_adverse, on=['Subject'], how='left')


            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                visita_comparar = row['visita_para_comparar']
                inform_consent_date = row['Inform_consent_date']
                visita_randomization = row['Visita_randomization']
                action_taken_CpG = row['action_taken_study_treatment']


                if status != '':
                    try:
                        date_dosing = row['Date of dosing']
                        date_dosing_pure = date_dosing.split('|')[0]
                        date_dosing_form_field_instance = date_dosing.split('|')[1]
                        date_dosing_disname =  date_dosing.split('|')[0]
                    except:
                        date_dosing_pure = ''
                        date_dosing_form_field_instance = 'This field doesnt have any data'
                        date_dosing_disname = 'Empty'

                    try:
                        reason_dose_adjustment = row['Reason for dose adjustment']
                        reason_dose_adjustment_pure = reason_dose_adjustment.split('|')[0]
                        reason_dose_adjustment_form_field_instance = reason_dose_adjustment.split('|')[1]
                        reason_dose_adjustment_disname = reason_dose_adjustment.split('|')[2]
                    except:
                        reason_dose_adjustment_pure = math.nan
                        reason_dose_adjustment_form_field_instance = 'This field doesnt have any data'
                        reason_dose_adjustment_disname = 'Empty'
                    
                    try:
                        dosing_event = row['Dosing Event']
                        dosing_event_pure = dosing_event.split('|')[0]
                        dosing_event_form_field_instance = dosing_event.split('|')[1]
                        dosing_event_disname = dosing_event.split('|')[2]
                    except:
                        dosing_event_pure =  math.nan
                        dosing_event_form_field_instance = 'This field doesnt have any data'
                        dosing_event_disname = 'Empty'
                    # ---------------------------------------------------------------------------------------
                    if date_dosing_pure == '':
                         pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_dosing_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of dosing', date_dosing_form_field_instance,\
                                        f , date_dosing_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IMP0020
                    try:
                        lista_permitidos_visita_dosing = ['D1', 'D15', 'D29']
                        if visita_comparar not in lista_permitidos_visita_dosing:
                            error = [subject, visit, 'Date of dosing', date_dosing_form_field_instance, \
                                        'The date must be equal to the D1, D15 or D29 date of visit', visita_comparar, 'IMP0020']
                            lista_revision.append(error)
                    except Exception as e:
                            lista_logs.append(f'Revision IMP0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IMP0040
                    try:
                        if datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') >= datetime.strptime(str(inform_consent_date), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'Date of dosing', date_dosing_form_field_instance, \
                                        'The date/time of dosing can not  be before the informed consent date/time', date_dosing_disname, 'IMP0040']
                            lista_revision.append(error)
                    except Exception as e:
                            lista_logs.append(f'Revision IMP0040 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IMP0050
                    try:
                        if datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') >= datetime.strptime(str(visita_randomization), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'Date of dosing', \
                                     date_dosing_form_field_instance, \
                                        'The date/time of dosing can not  be before the randomization date/time', \
                                            f'{date_dosing_disname} - {visita_randomization}', 'IMP0050']
                            lista_revision.append(error)
                    except Exception as e:
                            lista_logs.append(f'Revision IMP0050 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    if date_dosing == '':
                         pass
                    else:
                        # Revision IMP0060
                        try:
                            if date_dosing_pure in date_dosing_list_review:
                                error = [subject, visit, 'Date of dosing', \
                                        date_dosing_form_field_instance, \
                                        'The dosing date can not  be repeated', date_dosing_disname, 'IMP0060']
                                lista_revision.append(error)
                            else: 
                                date_dosing_list_review.append(date_dosing_pure)
                        except Exception as e:
                                lista_logs.append(f'Revision IMP0060 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IMP0080
                    try: 
                        if float(dosing_event_pure) == 2.0: 
                            if float(reason_dose_adjustment_pure) == 1.0:
                                 if float(action_taken_CpG) == 3.0:
                                      pass
                                 else:
                                      error = [subject, visit, 'Reason for dose adjustment', dosing_event_form_field_instance, \
                                                    'If dosing event is Temporarily discontinued and the reason for adjustment is "Adverse event" there should be an adverse event created where the action taken (CPG ODN 035) should be CT  drug stopped (temporarily)', \
                                                        dosing_event_disname, 'IMP0080']
                                      lista_revision.append(error)
                    except Exception as e:
                         lista_logs.append(f'Revision IMP0080 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision IMP0090
                    try:
                        if float(dosing_event_pure) == 3.0:
                            if float(reason_dose_adjustment_form_field_instance) == 1.0:
                                if float(action_taken_CpG) ==  4.0:
                                     pass
                                else:
                                      error = [subject, visit, 'Reason for dose adjustment', dosing_event_form_field_instance, \
                                                    'If dosing event is Permanently discontinued and the reason for adjustment is "Adverse event" there should be an adverse event created where the action taken (CPG ODN 035) should be CT  drug stopped (permanently)', \
                                                        dosing_event_disname, 'IMP0090']
                                      lista_revision.append(error)
                    except Exception as e:
                         lista_logs.append(f'Revision IMP0090 --> {e} - Subject: {subject},  Visit: {visit} ')

    
    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    adminsitration_CpG_ODN_output = pd.DataFrame(lista_revision, columns=column_names)


 
    sheet = excel_writer.create_sheet('CpG ODN D35 Administration')

    for row in dataframe_to_rows(adminsitration_CpG_ODN_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return adminsitration_CpG_ODN_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)


if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx")
    adminsitration_CpG_ODN(df_root, path_excel ) 







                