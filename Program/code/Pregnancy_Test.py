import pandas as pd
from datetime import datetime
from log_writer import log_writer
from revision_fechas import revision_fecha
import warnings

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


warnings.filterwarnings('ignore')


def pregnancy_test(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Pregnancy test
    '''

    df= df_root[df_root['name']== 'Pregnancy Test']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)

    df_child_bearing  = df_root[df_root['name']=='Child Bearing Potential']
    df_child_bearing = df_child_bearing[['Visit','Participante', 'Campo', 'Valor']]
    df_child_bearing = df_child_bearing[df_child_bearing['Campo']=='Is the participant postmenopausal?']
    df_child_bearing = df_child_bearing[['Visit','Participante','Valor']]
    df_child_bearing = df_child_bearing.rename(columns={'Participante':'Subject', 'Valor': 'post_menopausal'})

    df_demographic = df_root[df_root['name']=='Demographics']
    df_demographic = df_demographic[['Visit','Participante', 'Campo', 'Valor']]
    df_demographic = df_demographic[df_demographic['Campo']=='Gender']
    df_demographic = df_demographic[['Visit','Participante','Valor']]
    df_demographic = df_demographic.rename(columns={'Participante':'Subject', 'Valor':'Genero'})

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Visit','Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    lista_revision = []
    lista_logs = ['Pregnancy test']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_child_bearing, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_demographic, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']
                post_menopausal = row['post_menopausal']
                genero = row['Genero']
                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                if status == 'DATA_ENTRY_COMPLETE':

                    try:
                        was_pregnancy_test_performed = row['Was the Pregnancy Test performed?']
                        was_pregnancy_test_performed_pure = was_pregnancy_test_performed.split('|')[0]
                        was_pregnancy_test_performed_form_field_isntance = was_pregnancy_test_performed.split('|')[1]
                    except Exception as e:
                        was_pregnancy_test_performed_pure =''
                        was_pregnancy_test_performed_form_field_isntance = 'This field doesnt have any data'

                    try:
                        provide_reason = row['Provide the reason']
                        provide_reason_pure = provide_reason.split('|')[0]
                        provide_reason_form_field_instance = provide_reason.split('|')[1]
                    except Exception as e:
                        provide_reason_pure = ''
                        provide_reason_form_field_instance = 'This field doesnt have any data'

                    try:
                        type_pregnancy_test = row['Type of pregnancy test']
                        type_pregnancy_test_pure = type_pregnancy_test.split('|')[0]
                        type_pregnancy_test_form_field_isntance = type_pregnancy_test.split('|')[1]
                    except Exception as e:
                        type_pregnancy_test_pure = ''
                        type_pregnancy_test_form_field_isntance = 'This field doesnt have any data'

                    try:
                        date_test_performed = row['Date of test performed']
                        date_test_performed_pure = date_test_performed.split('|')[0]
                        date_test_performed_form_field_instance = date_test_performed.split('|')[1]
                    except Exception as e:
                        date_test_performed_pure = ''
                        date_test_performed_form_field_instance = 'This field doesnt have any data'

                    try:
                        pregnancy_test_result = row['Pregnancy Test Results']
                        pregnancy_test_result_pure = pregnancy_test_result.split('|')[0]
                        pregnancy_test_result_form_field_instance = pregnancy_test_result.split('|')[1]
                    except Exception as e:
                        pregnancy_test_result_pure = ''
                        pregnancy_test_result_form_field_instance = 'This field doesnt have any data'

                    # ------------------------------------------------------------------------------------------

                    try:
                        # Primera  revision general de formato de fecha ->GE0020
                        f = revision_fecha(date_test_performed_pure)
                        if f == None:
                            pass
                        else:
                            error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance ,f , date_test_performed_pure, 'GE0020']
                            lista_revision.append(error)     

                    except Exception as e:
                        lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision RP0010
                    try:
                        if float(post_menopausal) == 1.0:
                            error = [subject, visit, 'Pregnancy Test Form', pregnancy_test_result_form_field_instance,\
                                     'Check that women with post menopausal=No in "Child Bearing Potential" Form' , \
                                        post_menopausal, 'RP0010']
                            lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision RP0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision RP0030
                    try:
                        if float(genero) == 1.0:
                            error = [subject, visit, 'Pregnancy Test Form', pregnancy_test_result_form_field_instance, \
                                     'If Subjects Gender is "Male" in DEMOGRAPHIC, form should be left empty' , post_menopausal, 'RP0030']
                            lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision RP0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision RP0040
                    try:
                        if  float(was_pregnancy_test_performed_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the Pregnancy Test performed?', \
                                         was_pregnancy_test_performed_form_field_isntance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and Screening visit date = D-1 date (screening done on D-1)' , \
                                            was_pregnancy_test_performed_pure, 'RP0040']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision RP0040--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision RP0050
                    try:
                        date_format = '%d-%b-%Y'
                        date_of_test_f = datetime.strptime(date_test_performed_pure, date_format)
                        date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                        if date_of_test_f != date_of_visit_f:
                            error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance, \
                                     'The date must be the same as the date of visit date', \
                                        f'{date_test_performed_pure} - {date_of_visit}', 'RP0050']
                            lista_revision.append(error)
                        else:
                            pass
                    except Exception as e:
                        lista_logs.append(f'Revision RP0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision RP0060
                    try:
                        date_format = '%d-%b-%Y'
                        date_of_test_f = datetime.strptime(date_test_performed_pure, date_format)
                        date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                        if date_of_test_f < date_inform_consent_f:
                            error = [subject, visit, 'Date of test performed', date_test_performed_form_field_instance, \
                                     'The date of test performed cant be before the informed consent date', \
                                        f'{date_test_performed_pure} - {date_inform_consent}', 'RP0060']
                            lista_revision.append(error)
                        else:
                            pass
                    except Exception as e:
                        lista_logs.append(f'Revision RP0060--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> RP0070
                    try:
                        if datetime.strptime(str(date_test_performed_pure), '%d-%b-%Y') >= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'date of test performed', date_test_performed_form_field_instance ,'The date of test performed cant be after the study/Early withdrawal date.', date_test_performed_pure, 'RP0070']
                            lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision RP0070 --> {e} - Subject: {subject},  Visit: {visit}  ')

    excel_writer = load_workbook(path_excel_writer)
    column_names =  ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    pregnancy_test_output = pd.DataFrame(lista_revision, columns=column_names)

    
    sheet = excel_writer.create_sheet("Pregnancy Test")

    for row in dataframe_to_rows(pregnancy_test_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return pregnancy_test_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    pregnancy_test(df_root, path_excel) 