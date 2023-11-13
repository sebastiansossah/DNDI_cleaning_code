from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer 
import warnings
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')


def informed_consent_revision(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de informed consent
    '''

    df= df_root[df_root['name']=='Informed Consent']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Informed Consent']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    lista_validacion_prior_screening = []
    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]
        
        
        for visita in sujeto_principal.Visit.unique():
            
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                fecha_visita = row['Valor']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                try:
                    signature_date =  row['Informed consent signature date']
                    signature_date_pure = signature_date.split('|')[0]
                    signature_date_form_field_instance = signature_date.split('|')[1]
                except:
                    signature_date_pure = ''
                    signature_date_form_field_instance = 'This field doesnt have any data'

                try:
                    prior_screening_number = row['Prior screening number']
                    prior_screening_number_pure = prior_screening_number.split('|')[0]
                    prior_screening_number_form_field_instance = prior_screening_number.split('|')[1]
                except:
                    prior_screening_number_pure = ''
                    prior_screening_number_form_field_instance = 'This field doesnt have any data'


                date_format = '%d-%b-%Y'

                if status == 'DATA_ENTRY_COMPLETE':

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    # Revision general de la fehcha GE0020
                    try:

                        f = revision_fecha(signature_date_pure)
                        if f == None:
                            pass
                        else:
                            error = [subject, visit, 'Informed consent signature date' ,signature_date_form_field_instance , f , signature_date_pure, 'GE0020']
                            lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    
                    # Revision para IC0030
                    try:
                        if  prior_screening_number_pure in lista_validacion_prior_screening:
                            error = [subject, visit, 'Prior screening number', prior_screening_number_form_field_instance  ,'The entered number should be a non existing subject number' , \
                                      prior_screening_number_pure, 'IC0020']
                            lista_revision.append(error)
                        else:
                            pass

                    except Exception as e:
                        lista_logs.append(f'Revision IC0030 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                    lista_validacion_prior_screening.append(prior_screening_number_pure)

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    Informed_Consent_output = pd.DataFrame(lista_revision, columns=column_names)

    
    sheet = excel_writer.create_sheet("Informed Consent")

    for row in dataframe_to_rows(Informed_Consent_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return Informed_Consent_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\data\6da79231-2439-4881-aeca-81cf5e9cd052.xlsx")
    informed_consent_revision(df_root, path_excel ) 

