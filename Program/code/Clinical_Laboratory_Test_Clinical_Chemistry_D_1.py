import numpy as np
import pandas as pd
import math
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def clinical_laboratory_test_clinical_chemistry_D_1(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Clinical Laboratory Test - Clinical Chemistry D-1
    '''
    df= df_root[df_root['name']== 'Clinical Laboratory Test - Clinical Chemistry D-1']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Visit','Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_demographic = df_root[df_root['name']=='Demographics']
    df_demographic = df_demographic[['Visit','Participante', 'Campo', 'Valor']]
    df_demographic = df_demographic[df_demographic['Campo']=='Gender']
    df_demographic = df_demographic[['Visit','Participante','Valor']]
    df_demographic = df_demographic.rename(columns={'Participante':'Subject', 'Valor':'Genero'})

    clicinical_chemestry_principal = df_root[df_root['name']=='Clinical Laboratory Test - Clinical Chemistry']
    clicinical_chemestry_principal = clicinical_chemestry_principal[['Visit','Participante', 'Campo', 'Valor']]
    clicinical_chemestry_principal = clicinical_chemestry_principal[clicinical_chemestry_principal['Campo']=='Blood Sample Collected']
    clicinical_chemestry_principal = clicinical_chemestry_principal[['Visit','Participante','Valor']]
    clicinical_chemestry_principal = clicinical_chemestry_principal.rename(columns={'Participante':'Subject', 'Valor':'blood_sample_principal'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Clinical Laboratory Test - Clinical Chemistry D-1']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_demographic, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(clicinical_chemestry_principal, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            
            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                genero = row['Genero']
                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                clinical_chemestry_principal = row['blood_sample_principal']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                
                if status != '':
                    try:
                        alanine_aminotransferase_out_normal = row['Alanine Aminotransferase (ALT), Out of normal range?']
                        alanine_aminotransferase_out_normal_pure = alanine_aminotransferase_out_normal.split('|')[0]
                        alanine_aminotransferase_out_normal_form_field_instance = alanine_aminotransferase_out_normal.split('|')[1]
                    except Exception as e:
                        alanine_aminotransferase_out_normal_pure = math.nan
                        alanine_aminotransferase_out_normal_form_field_instance = 'This field does not have any data'

                    try:
                        alanine_aminotransferase_result = row['Alanine Aminotransferase (ALT), Result (U/L)']
                        alanine_aminotransferase_result_pure = alanine_aminotransferase_result.split('|')[0]
                        alanine_aminotransferase_result_form_field_instance = alanine_aminotransferase_result.splti('|')[1]
                    except Exception as e:
                        alanine_aminotransferase_result_pure = math.nan
                        alanine_aminotransferase_result_form_field_instance = 'This field does not have any data'

                    try:
                        aspartate_aminotransferase_out_normal = row['Aspartate Aminotransferase (AST), Out of normal range?']
                        aspartate_aminotransferase_out_normal_pure = aspartate_aminotransferase_out_normal.split('|')[0]
                        aspartate_aminotransferase_out_normal_form_field_instance = aspartate_aminotransferase_out_normal.split('|')[1]
                    except Exception as e:
                        aspartate_aminotransferase_out_normal_pure = math.nan
                        aspartate_aminotransferase_out_normal_form_field_instance = 'This field does not have any data'

                    try:
                        aspartate_aminotransferase_result = row['Aspartate Aminotransferase (AST), Result (U/L)']
                        aspartate_aminotransferase_result_pure = aspartate_aminotransferase_result.split('|')[0]
                        aspartate_aminotransferase_result_form_field_instance = aspartate_aminotransferase_result.split('|')[1]
                    except Exception as e:
                        aspartate_aminotransferase_result_pure = math.nan
                        aspartate_aminotransferase_result_form_field_instance = 'This field does not have any data'

                    try:
                        blood_sample_collected = row['Blood Sample Collected']
                        blood_sample_collected_pure = blood_sample_collected.split('|')[0]
                        blood_sample_collected_form_field_instance = blood_sample_collected.split('|')[1]
                    except Exception as e:
                        blood_sample_collected_pure = math.nan
                        blood_sample_collected_form_field_instance = 'This field does not have any data'

                    try:
                        creatine_kinase_out_normal = row['Creatine Kinase (CK), Out of normal range?']
                        creatine_kinase_out_normal_pure = creatine_kinase_out_normal.split('|')[0]
                        creatine_kinase_out_normal_form_field_instance = creatine_kinase_out_normal.split('|')[1]
                    except Exception as e:
                        creatine_kinase_out_normal_pure = math.nan
                        creatine_kinase_out_normal_form_field_instance = 'This field does not have any data'

                    try:
                        creatine_kinase_result = row['Creatine Kinase (CK), Result (U/L)']
                        creatine_kinase_result_pure = creatine_kinase_result.split('|')[0]
                        creatine_kinase_result_form_field_instance = creatine_kinase_result.split('|')[1]
                    except Exception as e:
                        creatine_kinase_result_pure = math.nan
                        creatine_kinase_result_form_field_instance = 'This field does not have any data'

                    try:
                        date_collected = row['Date Collected']
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1]
                    except Exception as e:
                        date_collected_pure = ''
                        date_collected_form_field_instance = 'This field does not have any data'

                    try:
                        serum_creatine_out_normal = row['Serum Creatinine, Out of normal range?']
                        serum_creatine_out_normal_pure = serum_creatine_out_normal.split('|')[0]
                        serum_creatine_out_normal_form_field_instance =serum_creatine_out_normal.split('|')[1]
                    except Exception as e:
                        serum_creatine_out_normal_pure = math.nan 
                        serum_creatine_out_normal_form_field_instance = 'This field does not have any data'

                    try:
                        serum_creatine_result = row['Serum Creatinine, Result (mg/dL)']
                        serum_creatine_result_pure = serum_creatine_result.split('|')[0]
                        serum_creatine_result_form_field_instance = serum_creatine_result.split('|')[1]
                    except Exception as e:
                        serum_creatine_result_pure = math.nan
                        serum_creatine_result_form_field_instance = 'This field does not have any data'
                    
                    # -----------------------------------------------------------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_collected_pure:
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,\
                                        f , date_collected_pure, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                    # Revision LBD0010
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form' , \
                                            f'{date_collected_pure} - {date_of_visit}', 'LBD0010']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBD0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBD0020
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,\
                                        'The date/time of test performed can not be before the informed consent date/time' ,\
                                            f'{date_collected_pure} - {date_inform_consent}', 'LBD0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBD0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LBD0030
                    if date_collected_pure != '':
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') >= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,\
                                        'Date Collected must be before the End of study/Early withdrawal date. ', date_collected_pure, 'LBD0030']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LBD0030 --> {e} - Subject: {subject},  Visit: {visit}  ')
                    
                    # Revision LBD0050
                    try:
                        if float(aspartate_aminotransferase_out_normal_pure) == 1.0:
                            if float(aspartate_aminotransferase_result_pure) > 5.0 and \
                            float(aspartate_aminotransferase_result_pure) < 34.0:
                                
                                error = [subject, visit, 'Aspartate Aminotransferase (AST), Out of normal range? ', aspartate_aminotransferase_result_form_field_instance,\
                                        'According to the result, the value is not out of range, please review' , \
                                            aspartate_aminotransferase_result_pure, 'LBD0050']
                                lista_revision.append(error)
                    
                        # Revision LBD0090
                        elif float(aspartate_aminotransferase_out_normal_pure) == 0.0:
                            if float(aspartate_aminotransferase_result_pure) < 5.0 or \
                            float(aspartate_aminotransferase_result_pure) > 34.0:
                                
                                error = [subject, visit, 'Aspartate Aminotransferase (AST), Out of normal range? ', aspartate_aminotransferase_result_form_field_instance,\
                                        'According to the result, the value is out of range, please review' , \
                                            aspartate_aminotransferase_result_pure, 'LBD0090']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBD0050--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision LBD0060
                    try:
                        if float(alanine_aminotransferase_out_normal_pure) == 1.0:
                            if float(alanine_aminotransferase_result_pure) > 0.0 and \
                            float(alanine_aminotransferase_result_pure) < 55.0:
                                
                                error = [subject, visit, 'Alanine Aminotransferase (ALT), Out of normal range?', alanine_aminotransferase_result_form_field_instance ,\
                                        'According to the result, the value is not out of range, please review' , \
                                            alanine_aminotransferase_result_pure, 'LBD0060']
                                lista_revision.append(error)
                        
                        # Revision LBD0100
                        elif float(alanine_aminotransferase_out_normal_pure) == 0.0:
                            if float(alanine_aminotransferase_result_pure) < 0.0 or \
                            float(alanine_aminotransferase_result_pure) > 55.0:
                                
                                error = [subject, visit, 'Alanine Aminotransferase (ALT), Out of normal range?', alanine_aminotransferase_result_form_field_instance ,\
                                        'According to the result, the value is out of range, please review' , \
                                            alanine_aminotransferase_result_pure, 'LBD0100']
                                lista_revision.append(error)                                
                    except Exception as e:
                        lista_logs.append(f'Revision LBD0060--> {e} - Subject: {subject},  Visit: {visit} ')


                    try:
                        # Revision LBD0070
                        if float(serum_creatine_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(serum_creatine_result_pure) > 0.73 and float(serum_creatine_result_pure) <  1.18  :
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', serum_creatine_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review.' , serum_creatine_result_pure, 'LBD0070']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(serum_creatine_result_pure) >  0.55  and float(serum_creatine_result_pure) < 1.02 :
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', serum_creatine_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review.' , serum_creatine_result_pure, 'LBD0070']
                                    lista_revision.append(error)

                        # Revision LBD0110
                        elif float(serum_creatine_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(serum_creatine_result_pure) < 0.73 or float(serum_creatine_result_pure) > 1.18 :
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', serum_creatine_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review.', serum_creatine_result_pure, 'LBD0110']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(serum_creatine_result_pure) < 0.55 or float(serum_creatine_result_pure) > 1.02:
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', serum_creatine_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review.' , serum_creatine_result_pure, 'LBD0110']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBD0070 --> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBD0080
                        if float(creatine_kinase_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(creatine_kinase_result_pure) > 30.0 and float(creatine_kinase_result_pure) <  200.0 :
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', creatine_kinase_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review.' , creatine_kinase_result_pure, 'LBD0080']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(creatine_kinase_result_pure) >  29.0  and float(creatine_kinase_result_pure) < 168.0 :
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', creatine_kinase_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review.' , creatine_kinase_result_pure, 'LBD0080']
                                    lista_revision.append(error)

                        # Revision LBD0120
                        elif float(creatine_kinase_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(creatine_kinase_result_pure) < 30.0 or float(creatine_kinase_result_pure) > 200.0 :
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', creatine_kinase_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review.' , creatine_kinase_result_pure, 'LBD0120']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(creatine_kinase_result_pure) < 29.0 or float(creatine_kinase_result_pure) > 168.0 :
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', creatine_kinase_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review.' , creatine_kinase_result_pure, 'LBD0120']
                                    lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBD0080 --> {e} - Subject: {subject},  Visit: {visit} ')


                    lista_validacion = [
                        'Aspartate Aminotransferase (AST)',
                        'Aspartate Aminotransferase (AST), Result (U/L)',
                        'Alanine Aminotransferase (ALT)',
                        'Alanine Aminotransferase (ALT), Result (U/L)',
                        'Serum Creatinine',
                        'Serum Creatinine, Result (mg/dL)',
                        'Creatine Kinase (CK)',
                        'Creatine Kinase (CK), Result (U/L)',
                    ]


                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try:
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
     
                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass

                    
                    # Revision LBD0130
                    try:
                        if float(blood_sample_collected_pure) ==1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance ,\
                                         'If Blood Sample Collected is checked as "Yes", not all laboratory tests can be "not done"' , blood_sample_collected_pure, 'LBD0130']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBD0130--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBD0140
                    try:
                        if float(clinical_chemestry_principal) == 0.0:
                            if float(blood_sample_collected_pure) == 0.0:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance ,\
                                         'The Clinical chemistry  D-1 form should be completed if the Clinical chemistry form was not completed' , blood_sample_collected_pure, 'LBD0140']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBD0140--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBD0150
                    try:
                        if float(clinical_chemestry_principal) == 1.0:
                            if float(blood_sample_collected_pure) == 1.0:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance ,\
                                         'The Clinical chemistry  D-1 form should not  be completed if the Clinical chemistry form was completed' , blood_sample_collected_pure, 'LBD0150']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBD0150--> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    clinical_laboratory_test_clinical_chemistry_D_1_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("CL- Clinical Chemistry D-1")

    for row in dataframe_to_rows(clinical_laboratory_test_clinical_chemistry_D_1_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return clinical_laboratory_test_clinical_chemistry_D_1_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)
if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    clinical_laboratory_test_clinical_chemistry_D_1(df_root, path_excel ) 