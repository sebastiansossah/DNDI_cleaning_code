import numpy as np
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from log_writer import log_writer
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import math
import warnings

def physical_examination(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Physical Examination
    '''

    df= df_root[df_root['name']== 'Physical Examination']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})


    df_time_dosing1 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1 = df_time_dosing1[(df_time_dosing1['Campo']=='Date of dosing') | (df_time_dosing1['Campo']=='Time of Dosing')]
    df_time_dosing = df_time_dosing1[df_time_dosing1['Campo']=='Date of dosing']
    df_time_dosing['time_dosing_cpg_administration'] =  df_time_dosing1[df_time_dosing1['FormFieldInstance Id'].isin(df_time_dosing['FormFieldInstance Id'] + 1) & (df_time_dosing1['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing =df_time_dosing[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing = df_time_dosing.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})

    df_time_milteosine1 = df_root[df_root['name']== 'Miltefosine Administration'].sort_values(by='FormFieldInstance Id')
    df_time_milteosine1 = df_time_milteosine1[(df_time_milteosine1['Campo']=='Date of dosing') | (df_time_milteosine1['Campo']=='Time of Dosing') | (df_time_milteosine1['Campo']=='Dose (mg)')]
    df_time_milteosine = df_time_milteosine1[df_time_milteosine1['Campo']=='Time of Dosing']
    df_time_milteosine['date_ex_to_join'] =  df_time_milteosine1[df_time_milteosine1['FormFieldInstance Id'].isin(df_time_milteosine['FormFieldInstance Id'] - 1) & (df_time_milteosine1['Campo'] == 'Date of dosing')]['Valor'].values
    df_time_milteosine = df_time_milteosine[df_time_milteosine['Valor'].str.split(':').str[0] != '00']
    df_time_milteosine['time_dosing_miltefosine_administration'] = df_time_milteosine.groupby(['Participante', 'date_ex_to_join'])['Valor'].transform(lambda x: x.min())
    df_time_milteosine =df_time_milteosine[['Participante','date_ex_to_join', 'time_dosing_miltefosine_administration']].drop_duplicates()
    df_time_milteosine = df_time_milteosine.rename(columns={'Participante':'Subject'})


    lista_revision = []
    lista_logs = ['Physical Examination']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru['date_ex_to_join'] = pru['Date of examination performed'].str.split('|',expand= True)[0]
            
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            if visita != 'Unscheduled Visits':
                pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
                pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')

            pru = pru.merge(df_time_dosing, on=['Subject', 'date_ex_to_join'], how='left')
            pru = pru.merge(df_time_milteosine, on=['Subject', 'date_ex_to_join'], how='left')
            # if sujeto =='011002':



            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    print(pru)
                    print('---------------------')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                time_dosing_cpg_administration = row['time_dosing_cpg_administration']
                time_dosing_miltefosine_administration = row['time_dosing_miltefosine_administration']
        
                
                if visita == 'Unscheduled Visits':
                    was_DV_performed_pure = 1.0
                    date_of_visit = ''
                else:
                    was_DV_performed = row['was_DV_performed']
                    was_DV_performed_pure = was_DV_performed.split('|')[0]
                    was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                    date_of_visit = row['Date_of_visit']
                
                if status != '':
                    try:
                        was_physical_performed = row['Was the physical examination performed?']
                        was_physical_performed_pure = was_physical_performed.split('|')[0]
                        was_physical_performed_form_field_instance = was_physical_performed.split('|')[1]
                        was_physical_performed_disname = was_physical_performed.split('|')[2]
                    except Exception as e:
                        was_physical_performed_pure = math.nan
                        was_physical_performed_form_field_instance = 'This field does not have any data'
                        was_physical_performed_disname = 'Empty'
                    
                    try:
                        date_examination_performed = row['Date of examination performed']
                        date_examination_performed_pure = date_examination_performed.split('|')[0]
                        date_examination_performed_form_field_instance = date_examination_performed.split('|')[1]
                        date_examination_performed_disname = date_examination_performed.split('|')[0] 
                    except Exception as e:
                        date_examination_performed_pure = ''
                        date_examination_performed_form_field_instance = 'This field does not have any data'
                        date_examination_performed_disname = 'Empty'

                    try:
                        undefined_clinical = row['Undefined, Clinical interpretation?']
                        undefined_clinical_pure = undefined_clinical.split('|')[0]
                        undefined_clinical_form_field_instance = undefined_clinical.split('|')[1]
                        undefined_clinical_disname = undefined_clinical.split('|')[2]
                    except Exception as e:
                        undefined_clinical_pure = math.nan
                        undefined_clinical_form_field_instance = 'This field does not have any data'
                        undefined_clinical_disname = 'Empty'

                    try:
                        undefined_body_system = row['Undefined, Body System']
                        undefined_body_system_pure = undefined_body_system.split('|')[0]
                        undefined_body_system_form_field_instance = undefined_body_system.split('|')[1]
                        undefined_body_system_disname = undefined_body_system.split('|')[2]
                    except Exception as e:
                        undefined_body_system_pure = math.nan
                        undefined_body_system_form_field_instance = 'This field does not have any data'
                        undefined_body_system_disname = 'Empty'

                    try:
                        predose_clinical = row['Pre dose, Clinical interpretation?']
                        predose_clinical_pure = predose_clinical.split('|')[0]
                        predose_clinical_form_field_instance = predose_clinical.split('|')[1]
                        predose_clinical_disname = predose_clinical.split('|')[2] 
                    except Exception as e:
                        predose_clinical_pure = math.nan
                        predose_clinical_form_field_instance = 'This field does not have any data'
                        predose_clinical_disname = 'Empty'

                    try:
                        predose_clinical_time = row['Pre dose, Time']
                        predose_clinical_time_pure = predose_clinical_time.split('|')[0]
                        predose_clinical_time_form_field_instnance = predose_clinical_time.split('|')[1]
                        predose_clinical_time_disname = predose_clinical_time.split('|')[2] 
                    except Exception as e:
                        predose_clinical_time_pure = math.nan
                        predose_clinical_time_form_field_instnance = 'This field does not have any data'
                        predose_clinical_time_disname = 'Empty'

                    try:
                        two_hours = row['2-hours post dose, Clinical interpretation?']
                        two_hours_pure = two_hours.split('|')[0]
                        two_hours_form_field_instance = two_hours.split('|')[1]
                        two_hours_form_disname = two_hours.split('|')[2]
                    except Exception as e:
                        two_hours_pure = math.nan
                        two_hours_form_field_instance = 'This field does not have any data'
                        two_hours_form_disname = 'Empty'

                    try:
                        two_hours_time = row['2-hours post dose, Time']
                        two_hours_time_pure = two_hours_time.split('|')[0]
                        two_hours_time_fomr_field_instance = two_hours_time.split('|')[1]
                        two_hours_time_disname = two_hours_time.split('|')[2]
                    except Exception as e:
                        two_hours_time_pure = math.nan
                        two_hours_time_fomr_field_instance = 'This field does not have any data'
                        two_hours_time_disname = 'Empty'

                    try:
                        four_hours = row['4-hours post dose, Clinical interpretation?']
                        four_hours_pure = four_hours.split('|')[0]
                        four_hours_form_field_instance = four_hours.split('|')[1]
                        four_hours_disname = four_hours.split('|')[2]
                    except Exception as e:
                        four_hours_pure = math.nan
                        four_hours_form_field_instance = 'This field does not have any data'
                        four_hours_disname = 'Empty'

                    try:
                        four_hours_time = row['4-hours post dose, Time']
                        four_hours_time_pure = four_hours_time.split('|')[0]
                        four_hours_time_form_field_isntance = four_hours_time.split('|')[1]
                        four_hours_time_disname = four_hours_time.split('|')[2]
                    except Exception as e:
                        four_hours_time_pure = math.nan
                        four_hours_time_form_field_isntance = 'This field does not have any data'
                        four_hours_time_disname = 'Empty'

                    try:
                        eight_hours = row['8-hours post dose, Clinical interpretation?']
                        eight_hours_pure = eight_hours.split('|')[0]
                        eight_hours_form_field_instance = eight_hours.split('|')[1]
                        eight_hours_disname = eight_hours.split('|')[2]
                    except Exception as e:
                        eight_hours_pure = math.nan
                        eight_hours_form_field_instance = 'This field does not have any data'
                        eight_hours_disname = 'Empty'

                    try:
                        eight_hours_time = row['8-hours post dose, Time']
                        eight_hours_time_pure = eight_hours_time.split('|')[0]
                        eight_hours_time_form_field_instance = eight_hours_time.split('|')[1]
                        eight_hours_time_disname = eight_hours_time.split('|')[2]
                    except Exception as e:
                        eight_hours_time_pure = math.nan
                        eight_hours_time_form_field_instance = 'This field does not have any data'
                        eight_hours_time_disname = 'Empty'
                                        
                    # ----------------------------------------------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_examination_performed_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_examination_performed_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of examination performed', date_examination_performed_form_field_instance ,f , date_examination_performed_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision PE0020
                    if date_examination_performed != '' and date_of_visit != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_examination_performed_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of examination performed', date_examination_performed_form_field_instance,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form' , f'{date_examination_performed_disname} - {date_of_visit}', 'PE0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision PE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision PE0030
                    if date_examination_performed_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_examination_performed_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of examination performed', date_examination_performed_form_field_instance,\
                                        'The date/time of test performed can not be before the informed consent date/time' ,f'{date_examination_performed_disname} - {date_inform_consent}', 'PE0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision PE0030 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> PE0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_examination_performed_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_examination_performed_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of examination performed', date_examination_performed_form_field_instance ,'Date of examination performed must be before the End of study/Early withdrawal date. ', date_examination_performed_disname, 'PE0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PE0040 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    
                    if visit not in lista_revision:

                        lista_validacion_uno =[
                            'Undefined, If abnormal Any change in abnormality from previous physical examination?',
                            'Undefined, Body System',
                            'Undefined, Abnormality',
                            'Undefined, If abnormal, specify'
                                    ]
                        
                        mi_cuenta= 0

                        for validador_raw_uno in lista_validacion_uno:
                            try: 
                                validador_uno = row[validador_raw_uno].split('|')[0]
                                
                            except:
                                validador_uno = ''

                            if   validador_uno != '-' or validador_uno != np.nan or  str(validador_uno) != 'nan' or  str(validador_uno) != '':
                                mi_cuenta+=1
                            else:
                                pass
                            
                        # Revision PE0090
                        try:
                            if float(undefined_clinical_pure) == 2.0: 
                                if mi_cuenta != 0:
                                    pass
                                else:
                                    error = [subject, visit, 'Undefined, Clinical interpretation?', undefined_clinical_form_field_instance,\
                                             'If abnormal, the abnormality section must be added at least once', undefined_clinical_disname, 'PE0090']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PE0090 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision PE0100
                    try:
                        if  float(was_physical_performed_pure) == 9.0:
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the physical examination performed?', was_physical_performed_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and D-1 date=Screening visit date', was_physical_performed_disname, 'PE0100']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision PE0100--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision PE0110
                    lista_body_system = [
                            1.0, 7.0, 8.0, 9.0
                        ]
                    
                    try:
                        if float(undefined_body_system_pure) in lista_body_system:
                            if visit == 'Screening Visit':
                                error = [subject, visit, 'Undefined, Body System ', undefined_body_system_form_field_instance,\
                                         'General appearance, Neurological, Musculo-skeletal, Lymphatic should only be selected at the screening visit', undefined_body_system_disname, 'PE0110']
                                lista_revision.append(error)        
                    except Exception as e:
                        lista_logs.append(f'Revision PE0110 --> {e} - Subject: {subject},  Visit: {visit} ')

                    lista_visita_revision = [
                        'D1', 'D15', 'D29'
                    ]

                    if visita in lista_visita_revision:
                        # Revision PE0050

                        lista_revision_predose = [
                            'Pre dose, If abnormal Any change in abnormality from previous physical examination?',
                            'Pre dose abnormalities, Body System',
                            'Pre dose abnormalities, Abnormality',
                            'Pre dose abnormalities, If abnormal, specify'
                        ]

                        mi_cuenta_pre_dose = 0
                        for validador_predose_raw in lista_revision_predose:
                            try:
                                validador_predose = row[validador_predose_raw].split('|')[0]
                            except:
                                validador_predose = math.nan

                            if math.isnan(float(validador_predose))  or validador_predose != '-' or validador_predose != np.nan or  str(validador_predose) != 'nan' or float(validador_predose) !=0.0 or str(validador_predose) != '':
                                mi_cuenta_pre_dose+=1
                            else:
                                pass

                        try:
                            if float(predose_clinical_pure) == 2.0:
                                if mi_cuenta_pre_dose != 0:
                                    pass
                                else:
                                    error = [subject, visit, 'Pre dose, Clinical interpretation?', predose_clinical_form_field_instance,\
                                             'If abnormal, the abnormality section must be added at least once', predose_clinical_disname, 'PE0050']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PE0050 --> {e} - Subject: {subject},  Visit: {visit} ')


                        # Revision PE0060
                        lista_revision_2_hours = [
                        '2-hours post dose, If abnormal Any change in abnormality from previous physical examination?',
                        '2-hours post dose abnormalities, Body System',
                        '2-hours post dose abnormalities, Abnormality',
                        '2-hours post dose abnormalities, If abnormal, specify'
                        ]

                        mi_cuenta_two_hours = 0
                        for validador_two_raw in lista_revision_2_hours:
                            try:
                                validador_two = row[validador_two_raw].split('|')[0]
                            except:
                                validador_two = math.nan

                            if math.isnan(float(validador_two)) or validador_two != '-' or validador_two != np.nan or  str(validador_two) != 'nan' or float(validador_two) !=0.0 or str(validador_two) != '':
                                mi_cuenta_two_hours+=1
                            else:
                                pass

                        try:
                            if float(two_hours_pure) == 2.0:
                                if mi_cuenta_two_hours != 0:
                                    pass
                                else:
                                    error = [subject, visit, '2-hours post dose, Clinical interpretation?', two_hours_form_field_instance,\
                                             'If abnormal, the abnormality section must be added at least once' , two_hours_form_disname, 'PE0060']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PE0060 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision PE0070
                        lista_revision_4_hours = [
                        '4-hours post dose, If abnormal Any change in abnormality from previous physical examination?',
                        '4-hours post dose abnormalities, Body System',
                        '4-hours post dose abnormalities, Abnormality',
                        '4-hours post dose abnormalities, If abnormal, specify'
                        ]

                        mi_cuenta_four_hours = 0
                        for validador_four_raw in lista_revision_4_hours:
                            try:
                                validador_four = row[validador_four_raw].split('|')[0]
                            except:
                                validador_four = math.nan
                            
                            if math.isnan(float(validador_four)) or validador_four != '-' or validador_four != np.nan or  str(validador_four) != 'nan' or float(validador_four) !=0.0 or str(validador_four) != '':
                                mi_cuenta_four_hours+=1
                            else:
                                pass
                            
                        try:
                            if float(four_hours_pure) == 2.0:
                                if mi_cuenta_four_hours != 0:
                                    pass
                                else:
                                    error = [subject, visit, '4-hours post dose, Clinical interpretation?', four_hours_form_field_instance,\
                                             'If abnormal, the abnormality section must be added at least once' , four_hours_disname, 'PE0070']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PE0070 --> {e} - Subject: {subject},  Visit: {visit} ')

# ----------------------------------------------
                        # Revision PE0080
                        lista_revision_8_hours = [
                        '4-hours post dose, If abnormal Any change in abnormality from previous physical examination?',
                        '4-hours post dose abnormalities, Body System',
                        '4-hours post dose abnormalities, Abnormality',
                        '4-hours post dose abnormalities, If abnormal, specify'
                        ]

                        mi_cuenta_eight_hours = 0
                        for validador_eight_raw in lista_revision_8_hours:
                            try:
                                validador_eight = row[validador_eight_raw].split('|')[0]
                            except:
                                validador_eight = math.nan
                            
                            if math.isnan(float(validador_eight)) or validador_eight != '-' or validador_eight != np.nan or  str(validador_eight) != 'nan' or float(validador_eight) !=0.0 or str(validador_eight) != '':
                                mi_cuenta_eight_hours+=1
                            else:
                                pass

                        try:
                            if float(eight_hours_pure) == 2.0:
                                if mi_cuenta_eight_hours != 0:
                                    pass
                                else:
                                    error = [subject, visit, '8-hours post dose, Clinical interpretation?', eight_hours_form_field_instance,\
                                             'If abnormal, the abnormality section must be added at least once', eight_hours_disname, 'PE0080']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PE0080 --> {e} - Subject: {subject},  Visit: {visit} ')

                        #--------------------------------------------------------------------- Medicamento CPG ------------------------------    
                        #Revision PE0120
                        if str(time_dosing_cpg_administration) != 'nan':
                            try:
                                dif = float((datetime.strptime(time_dosing_cpg_administration, '%H:%M') - datetime.strptime(predose_clinical_time_pure, '%H:%M')).total_seconds() / 60)
                                if dif < 0.0 or dif > 90.0:
                                    #print(dif)
                                    error = [subject, visit, 'Pre dose, Time', predose_clinical_time_form_field_instnance,\
                                             'The time selected should be less than 90 min before the study treatment administration', \
                                                f'Pre dose, Time: {predose_clinical_time_pure} - dose time administration: {time_dosing_cpg_administration}', 'PE0120']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision PE0120 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # Revision PE0130
                        if str(time_dosing_cpg_administration) != 'nan':

                            try:
                                dif_two = float((datetime.strptime(two_hours_time_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                        
                                if  dif_two > 135.0 or dif_two < 105.0:
                                    error = [subject, visit, '2-hours post dose, Time', two_hours_time_fomr_field_instance,\
                                             'The time selected should be less than 2h15 and greater than 1h45 after the study treatment administration', \
                                                f'2-hours post dose,Time: {two_hours_time_pure} - dose time administration: {time_dosing_cpg_administration}', 'PE0130']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision PE0130 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # Revision PE0140
                        if str(time_dosing_cpg_administration) != 'nan':

                            try:
                                dif_four = float((datetime.strptime(four_hours_time_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if  dif_four > 255.0 or dif_four < 225.0:
                                    
                                    error = [subject, visit, '4-hours post dose, Time', four_hours_time_form_field_isntance,\
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', \
                                                f'4-hours post dose,Time: {four_hours_time_pure} - dose time administration: {time_dosing_cpg_administration}', 'PE0140']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision PE0140 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # Revision PE0150
                        if str(time_dosing_cpg_administration) != 'nan':
                            
                            try:
                                dif_eight = float((datetime.strptime(eight_hours_time_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_eight > 495.0 or dif_eight < 465.0:
                                    
                                    error = [subject, visit, '8-hours post dose, Time', eight_hours_time_form_field_instance,\
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', \
                                                f'8-hours post dose,Time: {eight_hours_time_pure} - dose time administration: {time_dosing_cpg_administration}', 'PE0150']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision PE0150 --> {e} - Subject: {subject},  Visit: {visit} ')  
                        
                        #--------------------------------------------------------------------- Medicamento Miltefosine  ------------------------------    

                        #Revision PE0120
                        if str(time_dosing_miltefosine_administration) != 'nan':
                            try:
                                dif_M = float((datetime.strptime(time_dosing_miltefosine_administration, '%H:%M') - datetime.strptime(predose_clinical_time_pure, '%H:%M')).total_seconds() / 60)
                                
                                if dif_M < 0.0 or dif_M > 90.0:
                                    # print(dif_M)
                                    error = [subject, visit, 'Pre dose, Time', predose_clinical_time_form_field_instnance,\
                                             'The time selected should be less than 60 min before the study treatment administration', \
                                                f'Pre dose, Time: {predose_clinical_time_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'PE0120']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision PE0120 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # Revision PE0130
                        if str(time_dosing_miltefosine_administration) != 'nan':

                            try:
                                dif_two_M = float((datetime.strptime(two_hours_time_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                                
                                if  dif_two_M > 135.0 or dif_two_M < 105.0:
                                    error = [subject, visit, '2-hours post dose, Time', two_hours_time_fomr_field_instance,\
                                             'The time selected should be less than 2h15 and greater than 1h45 after the study treatment administration', \
                                                f'2-hours post dose,Time: {two_hours_time_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'PE0130']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision PE0130 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # Revision PE0140
                        if str(time_dosing_miltefosine_administration) != 'nan':

                            try:
                                dif_four_M = float((datetime.strptime(four_hours_time_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                     
                                if  dif_four_M > 255.0 or dif_four_M < 225.0:
                                    
                                    error = [subject, visit, '4-hours post dose, Time', four_hours_time_form_field_isntance,\
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', \
                                                f'4-hours post dose,Time: {four_hours_time_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'PE0140']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision PE0140 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # Revision PE0150
                        if str(time_dosing_miltefosine_administration) != 'nan':
                            try:
                                dif_eight_M = float((datetime.strptime(eight_hours_time_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                                if dif_eight_M > 495.0 or dif_eight_M < 465.0:
                                    error = [subject, visit, '8-hours post dose, Time', eight_hours_time_form_field_instance,\
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', \
                                                f'8-hours post dose,Time: {eight_hours_time_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'PE0150']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision PE0150 --> {e} - Subject: {subject},  Visit: {visit} ')  


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    physical_examination_output = pd.DataFrame(lista_revision, columns=column_names)
    physical_examination_output = physical_examination_output[~physical_examination_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
    
    sheet = excel_writer.create_sheet("Physical Examination")

    for row in dataframe_to_rows(physical_examination_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return physical_examination_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx")
    physical_examination(df_root, path_excel) 