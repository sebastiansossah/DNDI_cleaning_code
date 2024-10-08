import numpy as np
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def prior_concomitant_procedures(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Prior And Concomitant Procedures
    '''

    df= df_root[df_root['name']== 'Prior And Concomitant Procedures']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Inform_consent_date'})

    df_adverse = df_root[df_root['name']=='Adverse Events']
    df_adverse = df_adverse[['Visit','Participante', 'Campo', 'Valor']]
    df_adverse['to_join'] = df_adverse['Valor']
    df_adverse = df_adverse[df_adverse['Campo']== 'Adverse Event ID']
    df_adverse = df_adverse[['Participante', 'to_join' ,'Valor']]
    df_adverse = df_adverse.rename(columns={'Participante':'Subject', 'Valor':'adverse_event_id'})

    df_adverse_start = df_root[df_root['name']=='Adverse Events']
    df_adverse_start = df_adverse_start[['Visit','Participante', 'Campo', 'Valor']]
    df_adverse_start['to_join_start'] = df_adverse_start['Valor']
    df_adverse_start = df_adverse_start[df_adverse_start['Campo']== 'Start Date']
    df_adverse_start = df_adverse_start[['Participante', 'to_join_start' ,'Valor']]
    df_adverse_start = df_adverse_start.rename(columns={'Participante':'Subject', 'Valor':'start_date_adverse'})

    df_adverse_end = df_root[df_root['name']=='Adverse Events']
    df_adverse_end = df_adverse_end[['Visit','Participante', 'Campo', 'Valor']]
    df_adverse_end['to_join_end'] = df_adverse_end['Valor']
    df_adverse_end = df_adverse_end[df_adverse_end['Campo']== 'End Date']
    df_adverse_end = df_adverse_end[['Participante', 'to_join_end' ,'Valor']]
    df_adverse_end = df_adverse_end.rename(columns={'Participante':'Subject', 'Valor':'end_date_adverse'})

    df_adverse_aditional = df_root[df_root['name']=='Adverse Events']
    df_adverse_aditional = df_adverse_aditional[['Visit','Participante', 'Campo', 'Valor']]
    df_adverse_aditional['to_join_adverse_aditional'] = df_adverse_aditional['Valor']
    df_adverse_aditional = df_adverse_aditional[df_adverse_aditional['Campo']== 'Adverse Event ID']
    df_adverse_aditional = df_adverse_aditional[['Participante', 'to_join_adverse_aditional' ,'Valor']]
    df_adverse_aditional = df_adverse_aditional.rename(columns={'Participante':'Subject', 'Valor':'adverse_event_id_aditional'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    lista_logs = ['Prior And Concomitant Procedures']
    lista_revision = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]
        sujeto_principal = sujeto_principal.sort_values(by=['FormFieldInstance Id'], ascending=True)
        sujeto_principal = sujeto_principal.reset_index(drop=True)

        list_procesure_id_review = []
        list_of_tuples_name_procedures_dates = []

        # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, por lo que usamos el siguiente codigo para realizar la particion
        date_indices = sujeto_principal.index[sujeto_principal['Campo'] == 'Procedure ID'].tolist()
        subdatasets = [sujeto_principal.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]


        for subdataset in subdatasets:

            pru = subdataset
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = 'unscheduled'
            pru['status'] = subdataset['activityState'].unique()
            
            try:
                pru['to_join'] = pru['Adverse Event ID'].str.split('|').str[0]
            except:
                pru['to_join'] = ''
            try:
                pru['to_join_start'] = pru['Start date'].str.split('|').str[0]
            except:
                pru['to_join_start'] = ''

            try:
                pru['to_join_end'] = pru['End date'].str.split('|').str[0]
            except:
                pru['to_join_end'] = ''
            
            try:
                pru['to_join_adverse_aditional'] = pru['Aditional Adverse Event ID'].str.split('|').str[0]
            except:
                pru['to_join_adverse_aditional'] = ''
            

            pru = pru.merge(df_adverse, on=['Subject', 'to_join'], how='left')
            pru = pru.merge(df_adverse_start, on=['Subject', 'to_join_start'], how='left')
            pru = pru.merge(df_adverse_end, on=['Subject', 'to_join_end'], how='left')
            pru = pru.merge(df_adverse_aditional, on=['Subject', 'to_join_adverse_aditional'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')


            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                adverse_event_id_from_table = row['adverse_event_id']
                adverse_event_start_date = row['start_date_adverse']
                adverse_event_end_date = row['end_date_adverse']

                adverse_event_id_from_table_aditional = row['to_join_adverse_aditional']
                inform_consent_date = row['Inform_consent_date']
                end_study_date = row['end_study_date']

                if status != '':

                    try: 
                        procedure_id = row['Procedure ID']
                        procedure_id_pure = procedure_id.split('|')[0]
                        procedure_id_form_field_instance = procedure_id.split('|')[1]
                        procedure_id_disname = procedure_id.split('|')[0]
                    except:
                        procedure_id_pure = ''
                        procedure_id_form_field_instance = 'This field does not have any data'
                        procedure_id_disname = 'Empty'
                    
                    try:
                        procedure_name = row['Procedure Name']
                        procedure_name_pure = procedure_name.split('|')[0]
                        procedure_name_form_field_instance = procedure_name.split('|')[1]
                        procedure_name_disname = procedure_name.split('|')[0]
                    except:
                        procedure_name_pure = ''
                        procedure_name_form_field_instance = 'This field does not have any data'
                        procedure_name_disname = 'Empty'
                    
                    try:
                        indication = row['Indication']
                        indication_pure = indication.split('|')[0]
                        indication_form_field_instance = indication.split('|')[1]
                        indication_disname = indication.split('|')[2]
                    except:
                        indication_pure = ''
                        indication_form_field_instance = 'This field does not have any data'
                        indication_disname = 'Empty'
                    
                    try:
                        indication_category = row['Indication Category']
                        indication_category_pure = indication_category.split('|')[0]
                        indication_category_form_field_instance = indication_category.split('|')[1]
                        indication_category_disname = indication_category.split('|')[2]
                    except:
                        indication_category_pure = ''
                        indication_category_form_field_instance = 'This field does not have any data'
                        indication_category_disname = 'Empty'
                    
                    try:
                        adverse_event_id = row['Adverse Event ID']
                        adverse_event_id_pure = adverse_event_id.split('|')[0]
                        adverse_event_id_form_field_instance = adverse_event_id.split('|')[1]
                        adverse_event_id_disname = adverse_event_id.split('|')[0]
                    except:
                        adverse_event_id_pure = ''
                        adverse_event_id_form_field_instance = 'This field does not have any data'
                        adverse_event_id_disname = 'Empty'
                    
                    try:
                        start_date = row['Start date']
                        start_date_pure = start_date.split('|')[0]
                        start_date_form_field_instance = start_date.split('|')[1]
                        start_date_disname = start_date.split('|')[0]
                    except:
                        start_date_pure = ''
                        start_date_form_field_instance = 'This field does not have any data'
                        start_date_disname = 'Empty'
                    
                    try:
                        ongoing = row['Ongoing']
                        ongoing_pure = ongoing.split('|')[0]
                        ongoing_form_field_instance = ongoing.split('|')[1]
                        ongoing_disname = ongoing.split('|')[2]
                    except:
                        ongoing_pure = ''
                        ongoing_form_field_instance = 'This field does not have any data'
                        ongoing_disname = 'Empty'
                    
                    try:
                        end_date = row['End date']
                        end_date_pure = end_date.split('|')[0]
                        end_date_form_field_instance = end_date.split('|')[1]
                        end_date_disname = end_date.split('|')[0]
                    except:
                        end_date_pure = ''
                        end_date_form_field_instance = 'This field does not have any data'
                        end_date_disname = 'Empty'

                    try:
                        aditional_adverse_event = row['Aditional Adverse Event ID']
                        aditional_adverse_event_pure = aditional_adverse_event.split('|')[0]
                        aditional_adverse_event_form_field_instance = aditional_adverse_event.split('|')[1]
                        aditional_adverse_event_disname = aditional_adverse_event.split('|')[0]
                    except:
                        aditional_adverse_event_pure = ''
                        aditional_adverse_event_form_field_instance = 'This field does not have any data'
                        aditional_adverse_event_disname = 'Empty'
                    

                    # ------------------------------------------------------------------------------------------------------------------
                    if start_date_pure == '':
                        pass
                    else:   
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(start_date_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Start date', start_date_form_field_instance,\
                                        f , start_date_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    if end_date_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(end_date_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'End date', end_date_form_field_instance,\
                                        f , end_date_disname, 'GE0020']
                                lista_revision.append(error)     
                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision PR0010
                    try:
                        if procedure_id_pure in list_procesure_id_review:
                            error =  [subject, visit, 'Procedure ID', procedure_id_form_field_instance, \
                                        'This value should be unique, it can not be repeated', \
                                            procedure_id_disname, 'PR0010']
                            lista_revision.append(error)
                        else:
                            list_procesure_id_review.append(procedure_id_pure)
                    except:
                        lista_logs.append(f'Revision CM0010 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision PR0020
                    try:
                        procedure_name_compare = procedure_name_pure.lower().strip()
                        tuple_to_compare = (procedure_name_compare, start_date_pure ,start_date_pure)

                        if tuple_to_compare in list_of_tuples_name_procedures_dates:
                            error =  [subject, visit, 'Procedure Name', procedure_name_form_field_instance, \
                                        'There are two procedures that have the same name, and the dates overlap', \
                                            procedure_name_disname, 'PR0020']
                            lista_revision.append(error)
                        else:
                            list_of_tuples_name_procedures_dates.append(tuple_to_compare)
                    except Exception as e:
                        lista_logs.append(f'Revision CM0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    
                    # Revision PR0040
                    if adverse_event_id_pure != '':
                        try:
                            if adverse_event_id_from_table != '-' or adverse_event_id_from_table != np.nan or  \
                                str(adverse_event_id_from_table) != 'nan' or float(adverse_event_id_from_table) !=0.0 or str(adverse_event_id_from_table) != '':
                                pass
                            else:
                                error = [subject, visit, 'Adverse Event ID', adverse_event_id_form_field_instance, \
                                            'If Adverse Event is selected, the value should be an existing AE ID', \
                                                adverse_event_id_disname, 'PR0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PR0040 --> {e} - Subject: {subject},  Visit: {visit} ')
                    else:
                        pass

                    # Este contador sirve para identificar cuando el dataset original tiene daverse event id, start date y end date creado
                    # Si en el dataset original no existe estos datos, el codigo no ejecuta la limpieza 
                    contador_para_validar = 0
                    if adverse_event_id_pure != '':
                        if start_date_pure != '':    
                            contador_para_validar +=1
                        if end_date_pure != '':
                            contador_para_validar +=1

                    # Revision PR0050
                    try:
                        if contador_para_validar ==0:
                            pass
                        elif contador_para_validar == 1:

                            if adverse_event_start_date != '-' or adverse_event_start_date != np.nan \
                                or str(adverse_event_start_date) != 'nan' or float(adverse_event_start_date) !=0.0 or str(adverse_event_start_date) != '':
                                pass
                            else:
                                error = [subject, visit, 'Adverse Event ID', adverse_event_id_form_field_instance, \
                                         'The start and end dates of procedure do not correspond to the Indication AE ID start and end date.', adverse_event_id_disname, 'PR0050']
                                lista_revision.append(error)

                        elif contador_para_validar == 2:
                            if adverse_event_start_date != '-' or adverse_event_start_date != np.nan \
                                or str(adverse_event_start_date) != 'nan' or float(adverse_event_start_date) !=0.0 or str(adverse_event_start_date) != '':

                                if adverse_event_end_date != '-' or adverse_event_end_date != np.nan \
                                    or str(adverse_event_end_date) != 'nan' or float(adverse_event_end_date) !=0.0 or str(adverse_event_end_date) != '':
                                    pass
                                else:
                                    error = [subject, visit, 'Adverse Event ID', adverse_event_id_form_field_instance, \
                                             'The start and end dates of medication do not correspond to the Indication AE ID start and end date.', \
                                               adverse_event_id_disname, 'PR0050']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision PR0050 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision PR0080
                    if aditional_adverse_event_pure != '':
                        try:
                            if adverse_event_id_from_table_aditional != '-' or adverse_event_id_from_table_aditional != np.nan or  \
                                str(adverse_event_id_from_table_aditional) != 'nan' or float(adverse_event_id_from_table_aditional) !=0.0 or str(adverse_event_id_from_table_aditional) != '':
                                pass
                            else:
                                error = [subject, visit, 'Aditional Adverse Event ID', aditional_adverse_event_form_field_instance, \
                                            'The value should be an existing AE ID', \
                                                aditional_adverse_event_disname, 'PR0080']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PR0080 --> {e} - Subject: {subject},  Visit: {visit} ')
                    else:
                        pass

                    # Revision -> PR0140
                    if str(end_study_date) != 'nan':
                        try:
                            if datetime.strptime(str(start_date_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Start Date', start_date_form_field_instance,\
                                        'Start Date must be before the End of study/Early withdrawal date. ', start_date_disname, 'PR0140']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PR0140 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision PR0150
                    try:
                        if datetime.strptime(str(end_date_pure), '%d-%b-%Y') >= datetime.strptime(str(start_date_pure), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'End date', end_date_form_field_instance, \
                                        'The date should be equal or greater than the start date', end_date_disname, 'PR0150']
                            lista_revision.append(error)
                    except Exception as e:
                            lista_logs.append(f'Revision PR0150 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> PR0160
                    if str(end_study_date) != 'nan':
                        try:
                            if datetime.strptime(str(end_date_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'End Date', end_date_form_field_instance,\
                                        'End Date must be before the End of study/Early withdrawal date. ', end_date_disname, 'PR0160']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision PR0160 --> {e} - Subject: {subject},  Visit: {visit}  ')

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    prior_concomitant_procedures_output = pd.DataFrame(lista_revision, columns=column_names)
    prior_concomitant_procedures_output = prior_concomitant_procedures_output[~prior_concomitant_procedures_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]

    sheet = excel_writer.create_sheet('Prior Concomitant Procedures')

    for row in dataframe_to_rows(prior_concomitant_procedures_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return prior_concomitant_procedures_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)


if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx")
    prior_concomitant_procedures(df_root, path_excel ) 



