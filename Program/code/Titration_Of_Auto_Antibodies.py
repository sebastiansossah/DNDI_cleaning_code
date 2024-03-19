import numpy as np
import math
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def titration_of_auto_antibodies(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Titration Of Auto-Antibodies
    '''

    df= df_root[df_root['name']== 'Titration Of Auto-Antibodies']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_logs = ['Titration Of Auto-Antibodies']
    lista_revision = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                if status != '':
                    try:
                        Anti_ssDNA_Comments_if_any = row["Anti-ssDNA, Comments if any"]
                        Anti_ssDNA_Comments_if_any_pure = Anti_ssDNA_Comments_if_any.split('|')[0]
                        Anti_ssDNA_Comments_if_any_form_field_instance = Anti_ssDNA_Comments_if_any.split('|')[1]
                    except Exception as e:
                        Anti_ssDNA_Comments_if_any_pure = math.nan
                        Anti_ssDNA_Comments_if_any_form_field_instance = 'This field does not have any data'

                    try:
                        date_collected = row["Date Sample Collected"]
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1]
                    except Exception as e:
                        date_collected_pure = '' 
                        date_collected_form_field_instance = 'This field does not have any data'


                    # --------------------------------------------------------------------   
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance,\
                                        f , date_collected_pure, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision TI0020
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance ,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form' , f'{date_collected_pure} - {date_of_visit}', 'TI0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision TI0020--> {e} - Subject: {subject},  Visit: {visit} ')
                    

                    # Revision TI0030
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance ,\
                                        'The date/time of test performed cant be before the informed consent date/time', f'{date_collected_pure} - {date_inform_consent}', 'TI0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision TI0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> TI0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance,\
                                        'Date Sample Collected must be before the End of study/Early withdrawal date. ', date_collected_pure, 'TI0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision TI0040 --> {e} - Subject: {subject},  Visit: {visit}  ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    titration_of_auto_antibodies_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("Titration Of AutoAntibodies")

    for row in dataframe_to_rows(titration_of_auto_antibodies_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)
    return titration_of_auto_antibodies_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\db3bea42-add2-4185-9a08-d8a8a2622277.xlsx")
    titration_of_auto_antibodies(df_root, path_excel ) 
                    