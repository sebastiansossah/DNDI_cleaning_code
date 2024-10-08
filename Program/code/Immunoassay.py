import pandas as pd
import os
import math
from datetime import datetime
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def immunoassay(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Immunoassay
    '''
    # Normals ranges file

    script_directory = os.path.dirname(os.path.abspath(__file__)) if '__file__' in locals() else os.getcwd()
    relative_folder_path = r"data\rangos_normales"
    folder_path = os.path.join(script_directory.replace('\code', ''), relative_folder_path)
    file = os.listdir(folder_path)
    path = f"{folder_path}\{[x for x in file if 'Immunoassay' in x][0]}" 
    df_normal_ranges = pd.read_csv(path, sep=';')

    df= df_root[df_root['name']== 'Immunoassay (Thyroid Stimulating Hormone)']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Immunoassay (Thyroid Stimulating Hormone)']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")


    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]
        sujeto_principal = sujeto_principal.sort_values(by=['FormFieldInstance Id'], ascending=True)
        sujeto_principal = sujeto_principal.reset_index(drop=True)

        adverse_events_id_review = []
        list_of_tuples_adverse_id = []

        # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, 
        # por lo que usamos el siguiente codigo para realizar la particion

        date_indices = sujeto_principal.index[sujeto_principal['Campo'] == 'Blood Sample Collected'].tolist()
        subdatasets = [sujeto_principal.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]

        for subdataset in subdatasets:

            pru_1 = subdataset
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = 'unscheduled'
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            # print(pru)
            # print('----------------')
            
            for index, row in pru.iterrows():

                if index != 0:

                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']


                if visit != 'unscheduled':
                    was_DV_performed = row['was_DV_performed']
                    was_DV_performed_pure = was_DV_performed.split('|')[0]
                    was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                else:
                    was_DV_performed_pure='unsch'
                    was_DV_performed_form_field_instance = 'Empty'
                
   
                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                if status != '':
                    try:
                        blood_sample_collected = row['Blood Sample Collected']
                        blood_sample_collected_pure = blood_sample_collected.split('|')[0]
                        blood_sample_collected_form_field_instance = blood_sample_collected.split('|')[1]
                        blood_sample_collected_disname = blood_sample_collected.split('|')[2]
                    except Exception as e:
                        blood_sample_collected_pure = math.nan  
                        blood_sample_collected_form_field_instance  = 'This field does not have any data'
                        blood_sample_collected_disname = 'Empty'

                    try:
                        date_collected = row['Date Sample Collected']
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1]
                        date_collected_disname = date_collected.split('|')[0]
                    except Exception as e:
                        date_collected_pure = ''
                        date_collected_form_field_instance = 'This field does not have any data'
                        date_collected_disname = 'Empty'

                    try:
                        provide_reason = row['Provide the reason']
                        provide_reason_pure = provide_reason.split('|')[0]
                        provide_reason_form_field_instance = provide_reason.split('|')[1]
                        provide_reason_disname = provide_reason.split('|')[0]
                    except Exception as e:
                        provide_reason_pure = math.nan  
                        provide_reason_form_field_instance = 'This field does not have any data'
                        provide_reason_disname = 'Empty'

                    try:
                        TSH = row['TSH']
                        TSH_pure = TSH.split('|')[0]
                        TSH_form_field_instance = TSH.split('|')[1]
                        TSH_disname = TSH.split('|')[0] 
                    except Exception as e:
                        TSH_pure = ''
                        TSH_form_field_instance = 'This field does not have any data'
                        TSH_disname = 'Empty'
                    

                    try:
                        TSH_specify = row['TSH, If abnormal, Specify']
                        TSH_specify_pure = TSH_specify.split('|')[0]
                        TSH_specify_form_field_instnace = TSH_specify.split('|')[1]
                        TSH_specify_disname = TSH_specify.split('|')[0]
                    except Exception as e:
                        TSH_specify_pure = math.nan  
                        TSH_specify_form_field_instnace = 'This field does not have any data'
                        TSH_specify_disname = 'Empty'

                    try:
                        TSH_out_normal = row['TSH, Out of normal range?']
                        TSH_out_normal_pure = TSH_out_normal.split('|')[0]
                        TSH_out_normal_form_field_instance = TSH_out_normal.split('|')[1]
                        TSH_out_normal_disname = TSH_out_normal.split('|')[2]
                    except Exception as e:
                        TSH_out_normal_pure = math.nan  
                        TSH_out_normal_form_field_instance = 'This field does not have any data'
                        TSH_out_normal_disname = 'Empty'

                    try:
                        TSH_result = row['TSH, Result (uIU/mL)']
                        TSH_result_pure = TSH_result.split('|')[0]
                        TSH_result_form_field_instance = TSH_result.split('|')[1]
                        TSH_result_disname = TSH_result.split('|')[0]
                    except Exception as e:
                        TSH_result_pure = math.nan  
                        TSH_result_form_field_instance = 'This field does not have any data'
                        TSH_result_disname = 'Empty'

                    # -----------------------------------------------------------------------------------------------------------------

                    # Revision GE0070
                    if str(was_DV_performed_pure) !=  'unsch':
                        if float(was_DV_performed_pure) !=  1.0:
                            error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                            lista_revision.append(error)

                    if date_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance ,f , date_collected_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IM0010
                    if date_collected_pure != '' and str(date_of_visit) != 'nan':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_collected_f != date_of_visit_f:
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form',\
                                            f'{date_collected_disname} - {date_of_visit}', 'IM0010']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision IM0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IM0020
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_collected_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance, \
                                        'The date/time of test performed can not be before the informed consent date/time',\
                                            f'{date_collected_disname} - {date_inform_consent}', 'IM0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision IM0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> IM0030
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date Sample Collected', date_collected_form_field_instance,\
                                         'Date Sample Collected must be before the End of study/Early withdrawal date. ', date_collected_disname, 'IM0030']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IM0030 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision IM0050
                    try:
                        if float(blood_sample_collected_pure) == 1.0:
                            if  float(TSH_result_pure) == 0.0 or str(TSH_result_pure) =='' or str(TSH_result_pure) == float('nan'):
                                error = [subject, visit, 'TSH', TSH_form_field_instance, \
                                        'It does not seem right that the TSH was not done but the sample was collected, please review', \
                                            f'Blood sample collected: {blood_sample_collected_disname} - TSH: {TSH_disname}', 'IM0050']
                                lista_revision.append(error)
                            else:
                                pass
                    except Exception as e:
                        lista_logs.append(f'Revision IM0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision IM0060
                        if float(TSH_out_normal_pure) == 1.0:
                            if float(TSH_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="TSH, Result (uIU/mL)"]['min'].iloc[0]) \
                                and float(TSH_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="TSH, Result (uIU/mL)"]['max'].iloc[0]) :
                                error = [subject, visit, 'TSH, Out of normal range?', TSH_out_normal_form_field_instance,\
                                         'According to the result, the value is not out of range, please review.', \
                                            f"TSH Result: {TSH_result_disname} - TSH Out of Normal: {TSH_out_normal_pure}", 'IM0060']
                                lista_revision.append(error)

                        # Revision IM0070
                        elif float(TSH_out_normal_pure) == 0.0:
                            #if float(TSH_result_pure) <  0.35  or float(TSH_result_pure) > 4.94 :
                            if float(TSH_result_pure) < float(df_normal_ranges[df_normal_ranges['field']=="TSH, Result (uIU/mL)"]['min'].iloc[0]) \
                                or float(TSH_result_pure) > float(df_normal_ranges[df_normal_ranges['field']=="TSH, Result (uIU/mL)"]['max'].iloc[0]) :
                                error = [subject, visit, 'TSH, Out of normal range?', TSH_out_normal_disname,\
                                         'According to the result, the value is out of range, please review.', \
                                            f"TSH Result: {TSH_result_disname} - TSH Out of Normal: {TSH_out_normal_pure}", 'IM0070']
                                lista_revision.append(error)   
                    except Exception as e:
                        lista_logs.append(f'Revision IM0060--> {e} - Subject: {subject},  Visit: {visit} ')

    excel_writer = load_workbook(path_excel_writer)
    column_names =  ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    immunoassay_output = pd.DataFrame(lista_revision, columns=column_names)
    immunoassay_output = immunoassay_output[~immunoassay_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
    sheet = excel_writer.create_sheet("Immunoassay")

    for row in dataframe_to_rows(immunoassay_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return immunoassay_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx')
    immunoassay(df_root, path_excel) 