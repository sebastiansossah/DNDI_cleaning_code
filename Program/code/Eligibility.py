from datetime import datetime
import numpy as np
import math
import pandas as pd
from revision_fechas import revision_fecha
from log_writer import log_writer
import warnings
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')


def eligibility(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Eligibility
    '''
    script_directory = os.path.dirname(os.path.abspath(__file__)) if '__file__' in locals() else os.getcwd()
    relative_folder_path = r"data\rangos_normales"
    folder_path = os.path.join(script_directory.replace('\code', ''), relative_folder_path)
    file = os.listdir(folder_path)
    path = f"{folder_path}\{[x for x in file if 'Eligibility' in x][0]}" 
    df_normal_ranges = pd.read_csv(path, sep=';')


    df= df_root[df_root['name']=='Eligibility']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)


    df_demographic_age = df_root[df_root['name']=='Demographics']
    df_demographic_age = df_demographic_age[['Participante', 'Campo', 'Valor']]
    df_demographic_age = df_demographic_age[df_demographic_age['Campo']=='Age at consent']
    df_demographic_age = df_demographic_age[['Participante','Valor']]
    df_demographic_age = df_demographic_age.rename(columns={'Participante':'Subject', 'Valor':'age_participant'})

    df_covid = df_root[df_root['name']=='Covid 19 testing']
    df_covid = df_covid[['Visit','Participante', 'Campo', 'Valor']]
    df_covid = df_covid[df_covid['Campo']=='Result']
    df_covid = df_covid[['Visit','Participante','Valor']]
    df_covid = df_covid.rename(columns={'Participante':'Subject', 'Valor':'covid_result'})


    df_vein = df_root[df_root['name']=='Vein assessment']
    df_vein = df_vein[['Visit','Participante', 'Campo', 'Valor']]
    df_vein = df_vein[df_vein['Campo']=='Suitable veins for multiple venepunctures/cannulations found?']
    df_vein = df_vein[['Visit','Participante','Valor']]
    df_vein = df_vein.rename(columns={'Participante':'Subject', 'Valor':'vein_assesment'})

    df_urinary = df_root[df_root['name']=='Urinary Drug Screen']
    df_urinary = df_urinary[['Visit','Participante', 'Campo', 'Valor']]
    df_urinary = df_urinary[df_urinary['Campo']=='Check below trace/positive results']
    df_urinary = df_urinary[['Visit','Participante','Valor']]
    df_urinary = df_urinary.rename(columns={'Participante':'Subject', 'Valor':'urinary_test'})

    df_virology_HIV = df_root[df_root['name']=='Virology']
    df_virology_HIV = df_virology_HIV[['Visit','Participante', 'Campo', 'Valor']]
    df_virology_HIV = df_virology_HIV[df_virology_HIV['Campo']=='HIV-1, Result']
    df_virology_HIV = df_virology_HIV[['Visit','Participante','Valor']]
    df_virology_HIV = df_virology_HIV.rename(columns={'Participante':'Subject', 'Valor':'HIV1_result'})

    df_virology_HIV2 = df_root[df_root['name']=='Virology']
    df_virology_HIV2 = df_virology_HIV2[['Visit','Participante', 'Campo', 'Valor']]
    df_virology_HIV2 = df_virology_HIV2[df_virology_HIV2['Campo']=='HIV-2, Result']
    df_virology_HIV2 = df_virology_HIV2[['Visit','Participante','Valor']]
    df_virology_HIV2 = df_virology_HIV2.rename(columns={'Participante':'Subject', 'Valor':'HIV2_result'})

    df_virology_HbsAg  = df_root[df_root['name']=='Virology']
    df_virology_HbsAg = df_virology_HbsAg[['Visit','Participante', 'Campo', 'Valor']]
    df_virology_HbsAg = df_virology_HbsAg[df_virology_HbsAg['Campo']=='HbsAg (Hepatitis B surface antigen), Result']
    df_virology_HbsAg = df_virology_HbsAg[['Visit','Participante','Valor']]
    df_virology_HbsAg = df_virology_HbsAg.rename(columns={'Participante':'Subject', 'Valor':'hbsag_result'})

    df_virology_Hcv  = df_root[df_root['name']=='Virology']
    df_virology_Hcv = df_virology_Hcv[['Visit','Participante', 'Campo', 'Valor']]
    df_virology_Hcv = df_virology_Hcv[df_virology_Hcv['Campo']=='HCV Ab (Hepatitis C virus antibody), Result']
    df_virology_Hcv = df_virology_Hcv[['Visit','Participante','Valor']]
    df_virology_Hcv = df_virology_Hcv.rename(columns={'Participante':'Subject', 'Valor':'hcv_result'})

    df__lession_count  = df_root[df_root['name']=='Lesion Measurement']
    df__lession_count = df__lession_count[['Visit','Participante', 'Campo', 'Valor']]
    df__lession_count = df__lession_count[df__lession_count['Visit']=='Screening Visit']
    df__lession_count = df__lession_count[df__lession_count['Campo']=='Anatomical Location']
    df__lession_count['Count'] = df__lession_count.groupby(by='Participante')['Valor'].transform('count')
    df__lession_count = df__lession_count[['Participante','Count']].drop_duplicates()
    df__lession_count = df__lession_count.rename(columns={'Participante':'Subject'})

    df__lession_dia  = df_root[df_root['name']=='Lesion Measurement']
    df__lession_dia = df__lession_dia[['Visit','Participante', 'Campo', 'Valor']]
    df__lession_dia = df__lession_dia[df__lession_dia['Campo']=='Longest diameter of lesion in mm']
    df__lession_dia = df__lession_dia[['Visit','Participante','Valor']]
    df__lession_dia = df__lession_dia.groupby('Participante', as_index=False).max()
    df__lession_dia = df__lession_dia.rename(columns={'Participante':'Subject', 'Valor':'Max_diameter'})
    df__lession_dia = df__lession_dia[['Visit','Subject','Max_diameter']]

    df_vital = df_root[df_root['name']=='Vital Signs']
    df_vital = df_vital[['Visit','Participante', 'Campo', 'Valor']]
    df_vital = df_vital[df_vital['Campo']=='Undefined, Diastolic Blood Pressure (mmHg)']
    df_vital = df_vital[['Visit','Participante','Valor']]
    df_vital = df_vital.rename(columns={'Participante':'Subject', 'Valor':'diastolic_preasure'})

    df_vital_sis = df_root[df_root['name']=='Vital Signs']
    df_vital_sis = df_vital_sis[['Visit','Participante', 'Campo', 'Valor']]
    df_vital_sis = df_vital_sis[df_vital_sis['Campo']=='Undefined, Systolic Blood Pressure  (mmHg)']
    df_vital_sis = df_vital_sis[['Visit','Participante','Valor']]
    df_vital_sis = df_vital_sis.rename(columns={'Participante':'Subject', 'Valor':'systolic_preasure'})

    df_lead_egc = df_root[df_root['name']=='12-Lead ECG']
    df_lead_egc = df_lead_egc[['Visit','Participante', 'Campo', 'Valor']]
    df_lead_egc = df_lead_egc[df_lead_egc['Campo']=='Undefined, HR (bpm)']
    df_lead_egc = df_lead_egc[['Visit','Participante','Valor']]
    df_lead_egc = df_lead_egc.rename(columns={'Participante':'Subject', 'Valor':'HR'})

    df_lead_egc_if_abnormal = df_root[df_root['name']=='12-Lead ECG']
    df_lead_egc_if_abnormal = df_lead_egc_if_abnormal[['Visit','Participante', 'Campo', 'Valor']]
    df_lead_egc_if_abnormal = df_lead_egc_if_abnormal[df_lead_egc_if_abnormal['Campo']=='Undefined, If abnormal, specify']
    df_lead_egc_if_abnormal = df_lead_egc_if_abnormal[['Visit','Participante','Valor']]
    df_lead_egc_if_abnormal = df_lead_egc_if_abnormal.rename(columns={'Participante':'Subject', 'Valor':'abnormal_specify'})

    # df_informed = df_root[df_root['name']=='Informed Consent']
    # df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    # df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    # df_informed = df_informed.rename(columns={'Participante':'Subject'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})


    df_lead_egc_undefined = df_root[df_root['name']=='12-Lead ECG']
    df_lead_egc_undefined = df_lead_egc_undefined[['Visit','Participante', 'Campo', 'Valor']]
    df_lead_egc_undefined = df_lead_egc_undefined[df_lead_egc_undefined['Campo']=='Undefined, QTcF (msec)']
    df_lead_egc_undefined = df_lead_egc_undefined[['Visit','Participante','Valor']]
    df_lead_egc_undefined = df_lead_egc_undefined.rename(columns={'Participante':'Subject', 'Valor':'QTCF_undefined'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []

    lista_logs = ['Eligibility']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()

            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_demographic_age, on=['Subject'], how='left')
            pru = pru.merge(df_covid, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_vein, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_urinary, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_virology_HbsAg, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_virology_Hcv, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_virology_HIV, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_virology_HIV2, on=['Subject', 'Visit'], how='left') #
            pru = pru.merge(df__lession_dia, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df__lession_count, on=['Subject'], how='left')
            pru = pru.merge(df_vital, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_vital_sis, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_lead_egc, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_lead_egc_if_abnormal, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_lead_egc_undefined, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            # if sujeto == '011001':
            #     print(pru)
            #     print('-------------------')



            lista_revision_I_E = []

      
            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']
                age_participant = row['age_participant']
                covid_result = row['covid_result']
                vein_assesment = row['vein_assesment']
                urinary_test = row['urinary_test']
                cuenta_lesiones = row['Count']
                diametro_lesiones = row['Max_diameter']
                diastolic_preasure = row['diastolic_preasure']
                systolic_preasure = row['systolic_preasure']
                HR_EGC = row['HR']
                abnormal_specify = row['abnormal_specify']
                QTCF = row['QTCF_undefined']
                informed_consent_date = row['Informed_consent_date']
                hiv1_result = row['HIV1_result']
                hiv2_result = row['HIV2_result']
                hbsag_result = row['hbsag_result']
                hcv_result = row['hcv_result']

                try:
                    was_DV_performed = row['was_DV_performed']
                    was_DV_performed_pure = was_DV_performed.split('|')[0]
                    was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                except:
                    was_DV_performed_pure = math.nan
            
                if status != '':

                    try:
                        subject_eligible_for_study = row['Is the subject eligible for the study?']
                        subject_eligible_for_study_pure = subject_eligible_for_study.split('|')[0]
                        subject_eligible_for_study_form_field_instance = subject_eligible_for_study.split('|')[1]
                        subject_eligible_for_study_disname = subject_eligible_for_study.split('|')[2]
                    except Exception as e:
                        subject_eligible_for_study_pure = math.nan
                        subject_eligible_for_study_form_field_instance = 'This field doesnt have any data'
                        subject_eligible_for_study_disname = 'Empty'
                    

                    try:
                        participant_randomization = row['Is the participant eligible to randomization?']
                        participant_randomization_pure = participant_randomization.split('|')[0]
                        participant_randomization_form_field_instance = participant_randomization.split('|')[1]
                        participant_randomization_disname = participant_randomization.split('|')[2]
                    except Exception as e:
                        participant_randomization_pure = math.nan
                        participant_randomization_form_field_instance = 'This field doesnt have any data'
                        participant_randomization_disname = 'Empty'

                    try:
                        will_randomized = row['Will the participant be randomized?']
                        will_randomized_pure = will_randomized.split('|')[0]
                        will_randomized_form_field_instance = will_randomized.split('|')[1]
                        will_randomized_disname = will_randomized.split('|')[2]
                    except Exception as e:
                        will_randomized_pure = math.nan
                        will_randomized_form_field_instance = 'This field doesnt have any data'
                        will_randomized_disname = 'Empty'

                    try:
                        subject_enrolled_study = row['Is the subject enrolled in the study?']
                        subject_enrolled_study_pure = subject_enrolled_study.split('|')[0]
                        subject_enrolled_study_form_field_instance = subject_enrolled_study.split('|')[1]
                        subject_enrolled_study_disname = subject_enrolled_study.split('|')[2]
                    except Exception as e:
                        subject_enrolled_study_pure = math.nan
                        subject_enrolled_study_form_field_instance = 'This field doesnt have any data'
                        subject_enrolled_study_disname = 'Empty'

                    
                    try:
                        date_of_decision = row['Date of decision to not go beyond screening']
                        date_of_decision_pure = date_of_decision.split('|')[0]
                        date_of_decision_form_field_instance = date_of_decision.split('|')[1]
                        date_of_decision_disname = date_of_decision.split('|')[0]
                    except Exception as e:
                        date_of_decision_pure = ''
                        date_of_decision_form_field_instance = 'This field doesnt have any data'
                        date_of_decision_disname = 'Empty'

                    try:
                        date_decision_not_randomize = row['Date of decision to not randomize the participant']
                        date_decision_not_randomize_pure = date_decision_not_randomize.split('|')[0]
                        date_decision_not_randomize_form_field_instance = date_decision_not_randomize.split('|')[1]
                        date_decision_not_randomize_disname = date_decision_not_randomize.split('|')[0]
                    except Exception as e:
                        date_decision_not_randomize_pure = ''
                        date_decision_not_randomize_form_field_instance = 'This field doesnt have any data'
                        date_decision_not_randomize_disname = 'Empty'

                    try:
                        randomization_number = row['Randomization number allocated to the replacement subject']
                        randomization_number_pure = randomization_number.split('|')[0]
                        randomization_number_form_field_isntance = randomization_number.split('|')[1]
                        randomization_number_disname = randomization_number.split('|')[0]
                    except Exception as e:
                        randomization_number_pure = ''
                        randomization_number_form_field_isntance = 'This field doesnt have any data'
                        randomization_number_disname = 'Empty'

                    try:
                        subject_replacing = row['Is the subject replacing a subject previously been randomized in the current study?']
                        subject_replacing_pure = subject_replacing.split('|')[0]
                        subject_replacing_form_field_instance = subject_replacing.split('|')[1]
                        subject_replacing_disname = subject_replacing.split('|')[2]
                    except Exception as e:
                        subject_replacing_pure = ''
                        subject_replacing_form_field_instance = 'This field doesnt have any data'
                        subject_replacing_disname = 'Empty'

                    try:
                        eligibility_criteria_number = row['Eligibility criteria number']
                        eligibility_criteria_number_pure = eligibility_criteria_number.split('|')[0]
                        eligibility_criteria_number_form_field_instance = eligibility_criteria_number.split('|')[1]
                        eligibility_criteria_number_disname = eligibility_criteria_number.split('|')[0]
                    except:
                        eligibility_criteria_number_pure = math.nan
                        eligibility_criteria_number_form_field_instance = 'This field doesnt have any data'
                        eligibility_criteria_number_disname = 'Empty'

                    try:
                        eligibility_criteria_type = row['Eligibility criteria type']
                        eligibility_criteria_type_pure = eligibility_criteria_type.split('|')[0]
                        eligibility_criteria_type_form_field_instance = eligibility_criteria_type.split('|')[1]
                        eligibility_criteria_type_disname = eligibility_criteria_type.split('|')[2]
                    except:
                        eligibility_criteria_type_pure = math.nan
                        eligibility_criteria_type_form_field_instance = 'This field doesnt have any data'
                        eligibility_criteria_type_disname = 'Empty'

                    try:
                        eligibility_specify = row['Specify']
                        eligibility_specify_pure = eligibility_specify.split('|')[0]
                        eligibility_specify_form_field_instance = eligibility_specify.split('|')[1]
                        eligibility_specify_disname = eligibility_specify.split('|')[2]
                    except:
                        eligibility_specify_pure = math.nan
                        eligibility_specify_form_field_instance = 'This field doesnt have any data'
                        eligibility_specify_disname = 'Empty'
                    


                    #---------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)
                
                    if float(participant_randomization_pure) == 0.0 or float(subject_eligible_for_study_pure) == 0.0:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_of_decision_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of decision to not go beyond screening', date_of_decision_form_field_instance ,f , date_of_decision_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    #Revision para los que son solo de screening visit -----------------------------------------------------------------------------------------------
                    if visit == 'Screening Visit':

                        #Revision  IE0060
                        if date_of_decision_pure == '':
                            pass
                        else:
                            try:
                                if datetime.strptime(str(date_of_decision_pure), '%d-%b-%Y') >= datetime.strptime(str(informed_consent_date), '%d-%b-%Y'):
                                    pass
                                else: 
                                    error = [subject, visit, 'Date of decision to not go beyond screening', date_of_decision_form_field_instance, \
                                            'The date must not be before the informed consent date', date_of_decision_disname, 'IE0060']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision IE0100 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                        # Revision para IE0130
                        try:
                            if float(subject_eligible_for_study_pure) == 0.0:
                                if float(subject_enrolled_study_pure) == 1.0:
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'How come the participant is not eligible for the study, but the question "Is the "Is the participant enrolled" is "Yes", Please check', \
                                                f"Subject Elegible: {subject_eligible_for_study_disname} - Subject Enrolled: {subject_enrolled_study_disname}", 'IE0130']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0130 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                        # Revision para IE0180 
                        try:
                            if float(will_randomized_pure) == 0.0:
                                if math.isnan(float(eligibility_criteria_number_pure)) == False or math.isnan(float(eligibility_criteria_type_pure)) == False or math.isnan(float(eligibility_specify_pure)) == False :
                                    pass
                                else:
                                    error = [subject, visit, 'Will the participant be randomized?', will_randomized_form_field_instance, \
                                            'If "Is the participant eligible to randomization?" ="No" at least one section of "Provide unfulfilled eligibility criteria" should be added', \
                                                will_randomized_disname, 'IE0180']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0180 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0190
                        try:
                            if float(will_randomized_pure) == 1.0:
                                    
                                if math.isnan(float(eligibility_criteria_number_pure)) == False or math.isnan(float(eligibility_criteria_type_pure)) == False or math.isnan(float(eligibility_specify_pure)) == False :
                                    error = [subject, visit, 'Will the participant be randomized?', will_randomized_form_field_instance, \
                                            'If "Is the participant eligible to randomization?" ="Yes", no sections of "Provide unfulfilled eligibility criteria" should be added',\
                                                will_randomized_disname, 'IE0190']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0190 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0220
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                if float(age_participant) < float(df_normal_ranges[df_normal_ranges['field']=="age_eligibility"]['min'].iloc[0]) or \
                                    float(age_participant) > float(df_normal_ranges[df_normal_ranges['field']=="age_eligibility"]['max'].iloc[0]) :
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The Subject can not be eligible because he/she is not between 18 and 50 years old, please review', 
                                            f"Subject Elegible: {subject_eligible_for_study_disname} - Age Participant: {age_participant}", 'IE0220']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0220 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0270
                        if visit == 'D-1':
                            try:
                                if float(subject_eligible_for_study_pure) == 1.0:
                                    if float(covid_result) == 1.0:
                                        error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                                'The Subject can not be eligible because he/she had a positive COVID-19 test result, please review', 
                                                f"Subject Elegible: {subject_eligible_for_study_disname} - Covid Result: {covid_result}", 'IE0270']
                                        lista_revision.append(error)
                                    else:
                                        pass
                            except Exception as e:
                                lista_logs.append(f'Revision IE0270 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0350
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                if float(urinary_test) == 1.0:
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The participant can not be eligible because he/she has trace/positive results in the urinary drug screen',
                                              f"Subject Elegible: {subject_eligible_for_study_disname} - Urinary test result: {urinary_test}", 'IE0350']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0350 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0441
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                if float(hiv1_result) == 1.0 or float(hiv2_result) == 1.0 or float(hbsag_result) == 1.0  or float(hcv_result) == 1.0:
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The participant can not be eligible because he/she has positive virology results', subject_eligible_for_study_disname, 'IE0441']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0441 --> {e} - Subject: {subject},  Visit: {visit} ')


                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                if float(cuenta_lesiones) > float(df_normal_ranges[df_normal_ranges['field']=="cuenta_lesiones"]['max'].iloc[0]) or \
                                    float(diametro_lesiones) > float(df_normal_ranges[df_normal_ranges['field']=="diametro_lesiones"]['max'].iloc[0]):

                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The participant has more than 4 lesions, or lesions over 4cm long in diameter or lesions with mucosal involvement in the Lesion Measurement form, he/she should not be eligible for randomization', \
                                                f"Subject Elegible for study: {subject_eligible_for_study_disname} - Number of lesions: {cuenta_lesiones} - Diameter of lesion: {diametro_lesiones}", 'IE0447']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE447 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0451
                        
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                #if float(diastolic_preasure) > 90.0:
                                if float(diastolic_preasure) > float(df_normal_ranges[df_normal_ranges['field']=="diastolic_preasure"]['max'].iloc[0]):
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance,\
                                            'The participant has a Diastolic Blood Pressure over 90 mmHg ,he/she should not be eligible for the study', \
                                                f"Subject Elegible for study: {subject_eligible_for_study_disname} - Distolic Preasure: {diastolic_preasure}", 'IE0451']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0451 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0455
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                # if float(systolic_preasure) < 100.0 or float(systolic_preasure) > 140.0:
                                if float(systolic_preasure) < float(df_normal_ranges[df_normal_ranges['field']=="systolic_preasure"]['min'].iloc[0]) or\
                                      float(systolic_preasure) > float(df_normal_ranges[df_normal_ranges['field']=="systolic_preasure"]['max'].iloc[0]):
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The participant has a Systolic Blood Pressure that is not between 100 and 140 mmHg,he/she should not be eligible for the study',\
                                                  f"Subject Elegible for study: {subject_eligible_for_study_disname} - Systolic Preasure: {systolic_preasure}", 'IE0455']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0455 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0459
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                #if float(HR_EGC) < 45.0 or float(HR_EGC) > 90.0:
                                if float(HR_EGC) < float(df_normal_ranges[df_normal_ranges['field']=="HR_EGC"]['min'].iloc[0]) or\
                                      float(HR_EGC) > float(df_normal_ranges[df_normal_ranges['field']=="HR_EGC"]['max'].iloc[0]) :

                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The participant has a HR that is not between 45 and 90 bpm,he/she should not be eligible for the study',\
                                                  f"Subject Elegible for study: {subject_eligible_for_study_disname} - HR Result: {HR_EGC}", 'IE0459']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0459 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0463
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                if float(abnormal_specify) == 2.0:
                                    error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                            'The participant has a Clinically significan not ECG,he/she should not be eligible for the study',\
                                                f"Subject Elegible for study: {subject_eligible_for_study_disname} - Abnormal specify: {abnormal_specify}", 'IE0463']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0463 --> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision para los que son solo D-1 ---------------------------------------------------------------------------------------------------------
                    if visit == 'D-1':
                        
                        # Revision IE0100
                        if date_decision_not_randomize_pure == '':
                            pass
                        else:
                            try:
                                if datetime.strptime(str(date_decision_not_randomize_pure), '%d-%b-%Y') >= datetime.strptime(str(informed_consent_date), '%d-%b-%Y'):
                                    pass
                                else: 
                                    error = [subject, visit, 'Date of decision to not randomize the participant' , date_decision_not_randomize_form_field_instance, \
                                            'The date must not be before the informed consent date', date_decision_not_randomize_disname, 'IE0100']
                                    lista_revision.append(error)
                            except Exception as e:
                                lista_logs.append(f'Revision IE0100 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0120
                        try:
                            if float(participant_randomization_pure) == 0.0:
                                if float(will_randomized_pure) == 1.0:
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                            'How come the participant is not eligible for randomization, but the question "Is the participant randomized" is "Yes", Please check',\
                                                f"Participant Randomization: {participant_randomization_disname} - Will Participate randomize: {will_randomized_disname}", 'IE0120']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0120 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0140
                        try:
                            if float(participant_randomization_pure) == 0.0:
                                if math.isnan(float(eligibility_criteria_number_pure)) == False or math.isnan(float(eligibility_criteria_type_pure)) == False or math.isnan(float(eligibility_specify_pure)) == False :
                                    pass
                                else:
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                             'If "Is the participant eligible to randomization?" ="No" at least one section of "Provide unfulfilled eligibility criteria" should be added',\
                                                  subject_eligible_for_study, 'IE0140']
                                    lista_revision.append(error)
                        except Exception as e :
                            lista_logs.append(f'Revision IE0140 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0150
                        try:
                            if float(subject_eligible_for_study_pure) == 1.0:
                                if math.isnan(float(eligibility_criteria_number_pure)) == False or math.isnan(float(eligibility_criteria_type_pure)) == False or math.isnan(float(eligibility_specify_pure)) == False :
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                             'If "Is the participant eligible to randomization?" ="Yes", no sections of "Provide unfulfilled eligibility criteria" should be added', \
                                                f"{subject_eligible_for_study_pure}", 'IE0150']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0150 --> {e} - Subject: {subject},  Visit: {visit}  ')

                        # Revision para IE210
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                # if float(age_participant) < 18.0 or float(age_participant) > 50:
                                if float(age_participant) < float(df_normal_ranges[df_normal_ranges['field']=="age_eligibility"]['min'].iloc[0]) or \
                                    float(age_participant) > float(df_normal_ranges[df_normal_ranges['field']=="age_eligibility"]['max'].iloc[0]):
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance,
                                            'The Subject can not be eligible because he/she is not between 18 and 50 years old, please review',
                                            f"Participant Randomization: {participant_randomization_disname} - Age Participant: {age_participant}", 'IE0210']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE210 --> {e} - Subject: {subject},  Visit: {visit} ')
               
                        # Revision para IE0260
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                if float(covid_result) == 1.0:
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                            'The Subject can not be eligible because he/she had a positive COVID-19 test result, please review', 
                                            f"Participant Randomization: {participant_randomization_disname} - Covid Result: {covid_result}", 'IE0260']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0260 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0440
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                if float(hiv1_result) == 1.0 or float(hiv2_result) == 1.0 or float(hbsag_result) == 1.0  or float(hcv_result) == 1.0:
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                            'The participant can not be eligible because he/she has positive virology results', participant_randomization_disname, 'IE0440']
                                    lista_revision.append(error)
                                else:
                                    pass
                        except Exception as e:
                            lista_logs.append(f'Revision IE0440 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                        # Revision para IE0446
                        
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                #if float(cuenta_lesiones) > 4.0 or float(diametro_lesiones) > 400.0:
                                if float(cuenta_lesiones) > float(df_normal_ranges[df_normal_ranges['field']=="cuenta_lesiones"]['max'].iloc[0]) or \
                                    float(diametro_lesiones) > float(df_normal_ranges[df_normal_ranges['field']=="diametro_lesiones"]['max'].iloc[0]):

                                    error = [subject, visit, 'Is the participant eligible to randomization?', \
                                            participant_randomization_form_field_instance,\
                                                'The participant has more than 4 lesions, or lesions over 4cm long in diameter or lesions with mucosal involvement in the Lesion Measurement form, he/she should not be eligible for randomization', \
                                                    f"Participant Randomization: {participant_randomization_disname} - Number of lesions: {cuenta_lesiones} - Diameter of lesion: {diametro_lesiones}", 'IE0446']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0446 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0450
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                if float(diastolic_preasure) > float(df_normal_ranges[df_normal_ranges['field']=="diastolic_preasure"]['max'].iloc[0]):
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_disname, \
                                            'The participant has a Diastolic Blood Pressure over 90 mmHg ,he/she should not be eligible for randomization',\
                                                  f"Participant Randomization: {participant_randomization_disname} - Diastolic Preasure: {diastolic_preasure}", 'IE0450']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0450 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0454
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                #if float(systolic_preasure) < 100.0 or float(systolic_preasure) > 140.0:
                                if float(systolic_preasure) < float(df_normal_ranges[df_normal_ranges['field']=="systolic_preasure"]['min'].iloc[0]) or\
                                      float(systolic_preasure) > float(df_normal_ranges[df_normal_ranges['field']=="systolic_preasure"]['max'].iloc[0]):
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                            'The participant has a Systolic Blood Pressure that is not between 100 and 140 mmHg, he/she should not be eligible for randomization',\
                                                  f"Participant Randomization: {participant_randomization_disname} - Systolic Preasure Result: {systolic_preasure}", 'IE0454']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0454 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0458
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                #if float(HR_EGC) < 45.0 or float(HR_EGC) > 90.0:
                                if float(HR_EGC) < float(df_normal_ranges[df_normal_ranges['field']=="HR_EGC"]['min'].iloc[0]) or\
                                      float(HR_EGC) > float(df_normal_ranges[df_normal_ranges['field']=="HR_EGC"]['max'].iloc[0]) :
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                            'The participant has a HR that is not between 45 and 90 bpm,he/she should not be eligible for randomization', \
                                            f"Participant Randomization: {participant_randomization_disname} - HR EGC Result: {HR_EGC}", 'IE0458']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0458 --> {e} - Subject: {subject},  Visit: {visit} ')

                        # Revision para IE0462
                        try:
                            if float(participant_randomization_pure) == 1.0:
                                if float(abnormal_specify) == 2.0:
                                    error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                            'The participant has a Clinically significan not ECG,he/she should not be eligible for randomization', \
                                            f"Participant Randomization: {participant_randomization_pure} - abnormal_specify: {abnormal_specify}", 'IE0462']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IE0462 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision para los que estan en los dos formularios-------------------------------------------------------------------------------------------
                    # Revision IE0020
                    try:
                        if float(eligibility_criteria_type_pure) == 1.0:
                            if float(eligibility_criteria_number_pure) >= 1.0 and float(eligibility_criteria_number_pure) <= 14.0:
                                pass
                            else:
                                error = [subject, visit, 'Eligibility criteria number', eligibility_criteria_number_form_field_instance, \
                                         'if Eligibility criteria type is Inclusion, the number has to be from 1 to 14', eligibility_criteria_number_disname, 'IE0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision IE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision IE0030
                    try:
                        if float(eligibility_criteria_type_pure) == 2.0:
                            if float(eligibility_criteria_number_pure) >= 1.0 and float(eligibility_criteria_number_pure) <= 25.0:
                                pass
                            else:
                                error = [subject, visit, 'Eligibility criteria number', eligibility_criteria_number_form_field_instance, \
                                         'if Eligibility criteria type is Exclusion, the number has to be from 1 to 25', eligibility_criteria_number_disname, 'IE0030']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision IE0030 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IE0040 
                    try:
                        tuple_review_inclusion_exclusion = (eligibility_criteria_type_pure, eligibility_criteria_number_pure)

                        if math.isnan(float(tuple_review_inclusion_exclusion[0])) or math.isnan(float(tuple_review_inclusion_exclusion[1])):
                            pass
                        else:
                            if tuple_review_inclusion_exclusion in lista_revision_I_E:
                                error = [subject, visit, 'Eligibility criteria number', eligibility_criteria_number_form_field_instance, \
                                        'The same criteria (Inclusion, Exclusion) and number, must not be duplicated', eligibility_criteria_number_disname, 'IE0040']
                                lista_revision.append(error)
                            else:
                                if '' in tuple_review_inclusion_exclusion:
                                    pass
                                else:
                                    lista_revision_I_E.append(tuple_review_inclusion_exclusion)
                    except Exception as e:
                        lista_logs.append(f'Revision IE0040 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision para IE0310
                    try:
                        if float(subject_eligible_for_study_pure) == 1.0:
                            if float(vein_assesment) == 0.0:
                                error = [subject, visit, 'Is the subject eligible for the study?', subject_eligible_for_study_form_field_instance, \
                                         'The participant can not be eligible because he/she doesnt have suitable veins for multiple venepunctures/cannulations found (Vein assessment)',\
                                              f"Subject Elefgible study: {subject_eligible_for_study_disname} - Vein Assesment: {vein_assesment}", 'IE0310']
                                lista_revision.append(error)
                            else:
                                pass
                    except Exception as e:
                        lista_logs.append(f'Revision IE0310 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision para IE0468
                    try:
                        if float(participant_randomization_pure) == 1.0:
                            if float(QTCF) > float(df_normal_ranges[df_normal_ranges['field']=="QTCF"]['max'].iloc[0]):
                                
                                error = [subject, visit, 'Is the participant eligible to randomization?', participant_randomization_form_field_instance, \
                                         'The participant has a QTcF interval (>450msec), he/she should not be eligible for randomization',\
                                             f"Participant Randomization: {participant_randomization_disname} - QTCF Result: {QTCF}" ,'IE0468']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision IE0468 --> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    eligibility_output = pd.DataFrame(lista_revision, columns=column_names).drop_duplicates()
    eligibility_output = eligibility_output[~eligibility_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]

    sheet = excel_writer.create_sheet("Eligibility")

    for row in dataframe_to_rows(eligibility_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    
    log_writer(lista_logs)

    return eligibility_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(R'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\data\f90c554a-361a-4054-a2bb-a64a72c9b621.xlsx')
    df_root.rename(columns = {'Instancia':'FormFieldInstance Id'}, inplace = True)
    df_root = df_root[(df_root['activityState']== 'DATA_VERIFIED') | (df_root['activityState']== 'DATA_ENTRY_COMPLETE')]
    eligibility(df_root, path_excel) 