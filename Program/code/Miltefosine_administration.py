import numpy as np
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')
pd.set_option('display.max_columns', None)

def miltefosine_administration(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Miltefosine Administration
    '''

    df = df_root[df_root['name']== 'Miltefosine Administration'] 
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName', 'Variable']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    
    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Inform_consent_date'})

    df_date_visit = df_root[df_root['name']== 'Date of visit']
    df_date_visit = df_date_visit[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit = df_date_visit[df_date_visit['Campo']== 'Visit Date']
    df_date_visit = df_date_visit[df_date_visit['Visit']== 'D-1']
    df_date_visit = df_date_visit[['Participante','Valor']]
    df_date_visit = df_date_visit.rename(columns={'Participante':'Subject', 'Valor':'Visita_randomization'})

    # df_adverse = df_root[df_root['name']=='Adverse Events']
    # df_adverse = df_adverse[['Visit','Participante', 'Campo', 'Valor']]
    # df_adverse = df_adverse[df_adverse['Campo']== 'Action taken with study treatment (Miltefosine)']
    # df_adverse = df_adverse[df_adverse['Valor'].isin([1,2,3,4,5,6,7,8,9,99,'1','2','3','4','5','6','7','8','9','99',1.0,2.0,3.0,4.0,5.0,6.0,7.0,8.0,9.0,99.0 ])]
    # df_adverse = df_adverse[['Participante','Valor']]
    # df_adverse = df_adverse.rename(columns={'Participante':'Subject', 'Valor':'Action_taken_miltefosine'})

    dataframe_join_adverse = pd.DataFrame()
    for sujeto_adverse in set(df_root[df_root['name']=='Adverse Events']['Participante'].tolist()):

        df_adverse = df_root[df_root['Participante']==sujeto_adverse]
        df_adverse = df_adverse[df_adverse['name']=='Adverse Events']

        df_adverse = df_adverse.sort_values(by=['FormFieldInstance Id'], ascending=True)
        df_adverse = df_adverse.reset_index(drop=True)

        date_indices = df_adverse.index[df_adverse['Campo'] == 'Adverse Event Reported Term'].tolist()
        subdatasets = [df_adverse.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]

        for subdataset in subdatasets:

            pru_adverse = subdataset[['Campo', 'Valor']].T
            new_columns_adverse = pru_adverse.iloc[0]
            pru_adverse = pru_adverse[1:].set_axis(new_columns_adverse, axis=1)
            pru_adverse['Subject'] = sujeto_adverse
            pru_adverse = pru_adverse[pru_adverse['Action taken with study treatment (Miltefosine)'].isin([1,2,3,4,5,6,7,8,9,99,'1','2','3','4','5','6','7','8','9','99',1.0,2.0,3.0,4.0,5.0,6.0,7.0,8.0,9.0,99.0 ])]
            pru_adverse =pru_adverse[['Subject','Start Date', 'Action taken with study treatment (Miltefosine)']]
            pru_adverse = pru_adverse.rename(columns={'Start Date':'to_join', 'Action taken with study treatment (Miltefosine)': 'Action_taken_miltefosine' })
    
            dataframe_join_adverse = pd.concat([dataframe_join_adverse, pru_adverse], ignore_index=True)
    
    dataframe_join_adverse = dataframe_join_adverse.drop_duplicates()

    df_date_visit_adverse = df_root[df_root['name']== 'Adverse Events']
    df_date_visit_adverse = df_date_visit_adverse[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit_adverse = df_date_visit_adverse[df_date_visit_adverse['Campo']== 'Start Date']
    df_date_visit_adverse['to_join'] = df_date_visit_adverse['Valor']
    df_date_visit_adverse = df_date_visit_adverse[['Participante', 'to_join','Valor']].drop_duplicates()
    df_date_visit_adverse = df_date_visit_adverse.rename(columns={'Participante':'Subject' ,'Valor':'Fecha_adverse'})

    df_date_visit_d1 = df_root[df_root['name']== 'Date of visit']
    df_date_visit_d1 = df_date_visit_d1[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit_d1 = df_date_visit_d1[df_date_visit_d1['Campo']== 'Visit Date']
    df_date_visit_d1 = df_date_visit_d1[df_date_visit_d1['Visit']== 'D1']
    df_date_visit_d1 = df_date_visit_d1[['Participante','Valor']]
    df_date_visit_d1 = df_date_visit_d1.rename(columns={'Participante':'Subject', 'Valor':'fecha_visita_d1'})

    df_date_visit_d28 = df_root[df_root['name']== 'Date of visit']
    df_date_visit_d28 = df_date_visit_d28[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit_d28 = df_date_visit_d28[df_date_visit_d28['Campo']== 'Visit Date']
    df_date_visit_d28 = df_date_visit_d28[df_date_visit_d28['Visit']== 'D28']
    df_date_visit_d28 = df_date_visit_d28[['Participante','Valor']]
    df_date_visit_d28 = df_date_visit_d28.rename(columns={'Participante':'Subject', 'Valor':'fecha_visita_d28'})

    df_date_visit_maxima = df_root[df_root['name']== 'Date of visit']
    df_date_visit_maxima = df_date_visit_maxima[['Visit','Participante', 'Campo', 'Valor']]
    df_date_visit_maxima = df_date_visit_maxima[df_date_visit_maxima['Campo']== 'Visit Date']
    df_date_visit_maxima['Valor'] = pd.to_datetime(df_date_visit_maxima['Valor'], format='%d-%b-%Y')
    max_per_participante = df_date_visit_maxima.groupby('Participante')['Valor'].transform('max')
    df_date_visit_maxima = df_date_visit_maxima[df_date_visit_maxima['Valor'] == max_per_participante]
    df_date_visit_maxima['Valor'] = pd.to_datetime(df_date_visit_maxima['Valor']).dt.strftime('%d-%b-%Y').str.upper()
    df_date_visit_maxima = df_date_visit_maxima[['Participante', 'Valor' ]]
    df_date_visit_maxima = df_date_visit_maxima.rename(columns={'Participante':'Subject', 'Valor':'fecha_visita_maxima'})

    warnings.filterwarnings('ignore')

    lista_revision = []
    lista_logs = ['Miltefosine Administration']
    

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]
        sujeto_principal = sujeto_principal.sort_values(by=['FormFieldInstance Id'], ascending=True)
        sujeto_principal = sujeto_principal.reset_index(drop=True)


        # miltefosine_administration_id_pure = sujeto_principal[sujeto_principal['Campo']== 'Miltefosine Administration ID']['Valor'][0]
        # miltefosine_administration_id_form_field_instance = sujeto_principal[sujeto_principal['Campo']== 'Miltefosine Administration ID']['FormFieldInstance Id'][0]


        date_dosing_historico_list = []

        visit_dictionary = {'Screening Visit': '', 'D-1':'', 'D1':'', 'D2':'', 
        'D3':'', 'D4':'', 'D7':'', 'D14':'','D15':'', 'D16':'', 'D17':'', 
        'D18':'', 'D21':'', 'D28':'', 'D29':'', 'D30':'', 'D31':'', 'D32':'', 
        'D35':'', 'D42':'', 'D63':'', 'D90':'', 'D105':''}

    # Los formularios que estan clasificados como unscheduled, no se pueden iterar con la visita, por lo que usamos el siguiente codigo para realizar la particion

        date_indices = sujeto_principal.index[sujeto_principal['Variable'] == 'ECMLTDAT'].tolist()
        subdatasets = [sujeto_principal.iloc[start:end] for start, end in zip(date_indices, date_indices[1:] + [None])]

        for subdataset in subdatasets:
            pru_1 = subdataset
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = 'unscheduled'
            pru['status'] = pru_1['activityState'].unique()


            pru['to_join'] = pru['Date of dosing'].str.split('|',expand= True)[0]

            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(dataframe_join_adverse, on=['Subject', 'to_join'], how='left')

            pru = pru.merge(df_date_visit, on=['Subject'], how='left')
            pru = pru.merge(df_date_visit_adverse, on=['Subject', 'to_join'], how='left')
            pru = pru.merge(df_date_visit_d1, on=['Subject'], how='left')
            pru = pru.merge(df_date_visit_d28 ,on=['Subject'], how='left')
            pru = pru.merge(df_date_visit_maxima ,on=['Subject'], how='left')
            # print(pru)
            # print('----------------------')

            
            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']
                inform_consent_date = row['Inform_consent_date']
                action_taken_miltefosine = row['Action_taken_miltefosine']
                visita_randomization = row['Visita_randomization']
                fecha_adverse = row['Fecha_adverse']
                fecha_visita_d1 = row['fecha_visita_d1']
                fecha_visita_d28 = row['fecha_visita_d28']
                fecha_visita_maxima = row['fecha_visita_maxima']


                if status != '':
                    
                    try:
                        date_dosing = row['Date of dosing']
                        date_dosing_pure = date_dosing.split('|')[0]
                        date_dosing_form_field_instance = date_dosing.split('|')[1]
                        date_dosing_disname = date_dosing.split('|')[0]
                    except:
                        date_dosing_pure = ''
                        date_dosing_form_field_instance = 'This field does not have any data'
                        date_dosing_disname = 'Empty'

                    try:                    
                        time_dosing = row['Time of Dosing']
                        time_dosing_pure = time_dosing.split('|')[0]
                        time_dosing_form_field_instance = time_dosing.split('|')[1]
                        time_dosing_disname = time_dosing.split('|')[0]
                    except:
                        time_dosing_pure = ''
                        time_dosing_form_field_instance = 'This field does not have any data'
                        time_dosing_disname = 'Empty'
                    
                    try:
                        dose_mg = row['Dose (mg)']
                        dose_mg_pure = dose_mg.split('|')[0]
                        dose_mg_form_field_instance = dose_mg.split('|')[1]
                        dose_mg_disname = dose_mg.split('|')[0]
                    except:
                        dose_mg_pure = ''
                        dose_mg_form_field_instance = 'This field does not have any data'
                        dose_mg_disname = 'Empty'
                        
                    try:
                        fasting_status = row['Fasting status']
                        fasting_status_pure = fasting_status.split('|')[0]
                        fasting_status_form_field_instance = fasting_status.split('|')[1]
                        fasting_status_disname = fasting_status.split('|')[2]
                    except:
                        fasting_status_pure = ''
                        fasting_status_form_field_instance = 'This field does not have any data'
                        fasting_status_disname = 'Empty'

                    # try:
                    #     miltefosine_administration_id = row['Miltefosine Administration ID']
                    #     miltefosine_administration_id_pure = miltefosine_administration_id.split('|')[0]
                    #     miltefosine_administration_id_form_field_instance = miltefosine_administration_id.split('|')[1]
                    #     miltefosine_administration_id_disname = miltefosine_administration_id.split('|')[0]
                    # except:
                    #     miltefosine_administration_id_pure = ''
                    #     miltefosine_administration_id_form_field_instance = 'This field does not have any data'
                    #     miltefosine_administration_id_disname = 'Empty'
                    
                    try:
                        dosing_event = row['Dosing Event']
                        dosing_event_pure = dosing_event.split('|')[0]
                        dosing_event_form_field_instance = dosing_event.split('|')[1]
                        dosing_event_disname = dosing_event.split('|')[2]
                    except:
                        dosing_event_pure = ''
                        dosing_event_form_field_instance = 'This field does not have any data'
                        dosing_event_disname = 'Empty'

                    try:
                        reason_adjustment = row['Reason for dose adjustment']
                        reason_adjustment_pure = reason_adjustment.split('|')[0]
                        reason_adjustment_form_field_instance = reason_adjustment.split('|')[1]
                        reason_adjustment_disname = reason_adjustment.split('|')[0]
                    except:
                        reason_adjustment_pure = ''
                        reason_adjustment_form_field_instance = 'This field does not have any data'
                        reason_adjustment_disname = 'Empty'
                    
                    try:
                        subject_vomited = row['Subject vomited']
                        subject_vomited_pure = subject_vomited.split('|')[0]
                        subject_vomited_form_field_instance = subject_vomited.split('|')[1]
                        subject_vomited_disname = subject_vomited.split('|')[2]
                    except:
                        subject_vomited_pure = ''
                        subject_vomited_form_field_instance = 'This field does not have any data'
                        subject_vomited_disname = 'Empty'
                 #--------------------------------------------------------------------------------------------------------
                    if date_dosing_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_dosing_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of dosing', date_dosing_form_field_instance,\
                                        f , date_dosing_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision ECML0020
                    try:
                        if str(fecha_visita_d1)!= 'nan' and str(fecha_visita_d28) != 'nan':
                            if datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') < datetime.strptime(str(fecha_visita_d1), '%d-%b-%Y')  or \
                            datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') > datetime.strptime(str(fecha_visita_d28), '%d-%b-%Y'):
                                error =  [subject, visit, 'Date of dosing', date_dosing_form_field_instance, \
                                            'The date must be between the D1 and the D28 date', date_dosing_disname, 'ECML0020']
                                lista_revision.append(error)

                                

                        elif str(fecha_visita_d1)!= 'nan' and str(fecha_visita_maxima) != 'nan':

                            if datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') < datetime.strptime(str(fecha_visita_d1), '%d-%b-%Y') : 
                                
                                error =  [subject, visit, 'Date of dosing', date_dosing_form_field_instance, \
                                            'The date must not be between the D1 and the D28 date', date_dosing_disname, 'ECML0020']
                                lista_revision.append(error)
                            
                        
                    except Exception as e:
                            lista_logs.append(f'Revision ECML0020 --> {e} - Subject: {subject},  Visit: {visit} ')




                    # Revision ECML0030
                    date_hour_tuple  =(date_dosing_pure, time_dosing_pure)
                    #print(date_hour_tuple)
                    try:
                        if date_hour_tuple in date_dosing_historico_list:
                            error = [subject, visit, 'Date of dosing', date_dosing_form_field_instance, \
                                            'The dosing date can not be repeated', date_dosing_disname, 'ECML0030']
                            lista_revision.append(error)
                        else:
                            date_dosing_historico_list.append(date_hour_tuple)
                    except Exception as e:
                        lista_logs.append(f'Revision ECML0030 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision ECML0040
                    try:
                        if datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') >= datetime.strptime(str(inform_consent_date), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'Date of decision to not go beyond screening', date_dosing_form_field_instance, \
                                        'The date must not be before the informed consent date', date_dosing_disname, 'ECML0040']
                            lista_revision.append(error)
                    except Exception as e:
                            lista_logs.append(f'Revision ECML0040 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision ECML0050
                    try:
                        if datetime.strptime(str(date_dosing_pure), '%d-%b-%Y') >= datetime.strptime(str(visita_randomization), '%d-%b-%Y'):
                            pass
                        else: 
                            error = [subject, visit, 'The date/time of dosing can not be before the randomization date/time', \
                                     date_dosing_form_field_instance, \
                                        'The date must not be before the informed consent date', \
                                            f'{date_dosing_disname} - {visita_randomization}', 'ECML0050']
                            lista_revision.append(error)
                    except Exception as e:
                            lista_logs.append(f'Revision ECML0050 --> {e} - Subject: {subject},  Visit: {visit} ')
                              
                    # Revision ECML0080
                    if str(action_taken_miltefosine) == 'nan' and str(action_taken_miltefosine) == '':
                        try:
                            if float(dosing_event_pure) == 2.0:
                                if float(reason_adjustment_pure) == 1.0:
                                    if float(action_taken_miltefosine) == 3.0:
                                        pass
                                    else:
                                        error = [subject, visit, 'Dosing Event', dosing_event_form_field_instance,\
                                                'If dosing event is Temporarily discontinued and the reason for adjustment is "Adverse event" there should be an adverse event created where the action taken (Miltefosine) should be CT  drug stopped (temporarily)', \
                                                    dosing_event_disname, 'ECML0080']
                                        lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision ECML0080 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision ECML0090
                    if str(action_taken_miltefosine) == 'nan' and str(action_taken_miltefosine) == '':
                        try:
                            if float(dosing_event_pure) == 3.0:
                                if float(reason_adjustment_pure) == 1.0:
                                    if float(action_taken_miltefosine) == 4.0:
                                        pass
                                    else:
                                        error = [subject, visit, 'Dosing Event', dosing_event_form_field_instance,\
                                                'If dosing event is Permanently discontinued and the reason for adjustment is "Adverse event" there should be an adverse event created where the action taken (Miltefosine) should be CT  drug stopped (permanently)', \
                                                    dosing_event_disname, 'ECML0090']
                                        lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision ECML0090 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision ECML0100
                    try:
                        if float(dose_mg_pure) == 0.0:
                            if float(reason_adjustment_pure) == 1.0:
                                if float(action_taken_miltefosine) == 2.0:
                                    pass
                                else:
                                    error = [subject, visit, 'Dosing Event', dosing_event_form_field_instance,\
                                             'If dosing is 0 and the reason for adjustment is "Adverse event" there should be an adverse event created where the action taken (Miltefosine) should be CT  dose reduced', \
                                                 dosing_event_disname, 'ECML0100']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision ECML0100 --> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision ECML0110
                    if str(fecha_adverse) == 'nan' and str(fecha_adverse) == '':
                        try:
                            if float(subject_vomited_pure) == 1.0:
                                if str(date_dosing_pure) == str(fecha_adverse):
                                    pass
                                else: 
                                    error = [subject, visit, 'The date/time of dosing can not be before the randomization date/time', \
                                            date_dosing_form_field_instance, \
                                                'The date must not be before the informed consent date', date_dosing_disname, 'ECML0050']
                                    lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision ECML0110 --> {e} - Subject: {subject},  Visit: {visit} ')

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    miltefosine_administration_output = pd.DataFrame(lista_revision, columns=column_names).drop_duplicates()

 
    sheet = excel_writer.create_sheet('Miltefosine Administration')

    for row in dataframe_to_rows(miltefosine_administration_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)

    log_writer(lista_logs)

    return miltefosine_administration_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx")
    miltefosine_administration(df_root, path_excel ) 






