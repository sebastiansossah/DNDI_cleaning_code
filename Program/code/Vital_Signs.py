import numpy as np
import math
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
warnings.filterwarnings('ignore')


def vital_signs(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Vital Signs
    '''

    # Normals ranges file
    script_directory = os.path.dirname(os.path.abspath(__file__)) if '__file__' in locals() else os.getcwd()
    relative_folder_path = r"data\rangos_normales"
    folder_path = os.path.join(script_directory.replace('\code', ''), relative_folder_path)
    file = os.listdir(folder_path)
    path = f"{folder_path}\{[x for x in file if 'Vital_' in x][0]}" 
    df_normal_ranges = pd.read_csv(path, sep=';')

    df= df_root[df_root['name']== 'Vital Signs']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    df_time_dosing1 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1 = df_time_dosing1[(df_time_dosing1['Campo']=='Date of dosing') | (df_time_dosing1['Campo']=='Time of Dosing')]
    df_time_dosing = df_time_dosing1[df_time_dosing1['Campo']=='Date of dosing']
    df_time_dosing['time_dosing_cpg_administration'] =  df_time_dosing1[df_time_dosing1['FormFieldInstance Id'].isin(df_time_dosing['FormFieldInstance Id'] + 1) & (df_time_dosing1['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing =df_time_dosing[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing = df_time_dosing.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})

    df_time_milteosine1 = df_root[df_root['name']== 'Miltefosine Administration'].sort_values(by='FormFieldInstance Id')
    df_time_milteosine1 = df_time_milteosine1[(df_time_milteosine1['Campo']=='Date of dosing') | (df_time_milteosine1['Campo']=='Time of Dosing') | (df_time_milteosine1['Campo']=='Dose (mg)')]
    df_time_milteosine = df_time_milteosine1[df_time_milteosine1['Campo']=='Time of Dosing']
    df_time_milteosine['date_ex_to_join'] =  df_time_milteosine1[df_time_milteosine1['FormFieldInstance Id'].isin(df_time_milteosine['FormFieldInstance Id'] - 1) & (df_time_milteosine1['Campo'] == 'Date of dosing')]['Valor'].values
    df_time_milteosine = df_time_milteosine[df_time_milteosine['Valor'].str.split(':').str[0] != '00']
    df_time_milteosine['time_dosing_miltefosine_administration'] = df_time_milteosine.groupby(['Participante', 'date_ex_to_join'])['Valor'].transform(lambda x: x.min())
    df_time_milteosine =df_time_milteosine[['Participante','date_ex_to_join', 'time_dosing_miltefosine_administration']].drop_duplicates()
    df_time_milteosine = df_time_milteosine.rename(columns={'Participante':'Subject'})



    lista_revision = []
    lista_logs = ['Vital Signs']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()

            try:
                pru['date_ex_to_join'] = pru['Date of assessment performed'].str.split('|',expand= True)[0]
            except:
                pru['date_ex_to_join'] = 'Nothing to join'

            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')

            if visita != 'Unscheduled Visits':
                pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
                pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            pru = pru.merge(df_time_dosing, on=['Subject', 'date_ex_to_join'], how='left')
            pru = pru.merge(df_time_milteosine, on=['Subject', 'date_ex_to_join'], how='left')
            # if sujeto =='011002':
            # print(pru)
            # print('---------------------------------------')

            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    print(pru)
                    print('---------------------------------------')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']
                time_dosing_cpg_administration = row['time_dosing_cpg_administration']
                time_dosing_miltefosine_administration = row['time_dosing_miltefosine_administration']

                
                if visita == 'Unscheduled Visits':
                    was_DV_performed_pure = 1.0
                    date_of_visit = ''
                else:
                    was_DV_performed = row['was_DV_performed']
                    was_DV_performed_pure = was_DV_performed.split('|')[0]
                    was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
                    date_of_visit = row['Date_of_visit']

                
                if status != '':
                    try: 
                        was_vital_signs_performed = row['Was the vital signs assessment performed?']
                        was_vital_signs_performed_pure = was_vital_signs_performed.split('|')[0]
                        was_vital_signs_performed_form_field_instance = was_vital_signs_performed.split('|')[1]
                        was_vital_signs_performed_disname = was_vital_signs_performed.split('|')[2]
                    except Exception as e:
                        was_vital_signs_performed_pure = math.nan
                        was_vital_signs_performed_form_field_instance = 'This field does not have any data'
                        was_vital_signs_performed_disname = 'Empty'
                    
                    try:
                        date_assesment = row['Date of assessment performed']
                        date_assesment_pure = date_assesment.split('|')[0]
                        date_assesment_form_field_instance = date_assesment.split('|')[1]
                    except Exception as e:
                        date_assesment_pure = ''
                        date_assesment_form_field_instance = 'This field does not have any data'
                    
                    try:
                        BMI = row['Undefined, BMI']
                        BMI_pure = BMI.split('|')[0]
                        BMI_form_field_instance = BMI.split('|')[1]
                        BMI_form_disname = BMI.split('|')[2]
                    except Exception as e:
                        BMI_pure = math.nan
                        BMI_form_field_instance = 'This field does not have any data'
                        BMI_form_disname = 'Empty'
                    
                    try:
                        height = row['Undefined, Height (cm)']
                        height_pure = height.split('|')[0]
                        height_form_field_instance = height.split('|')[1]
                        height_disname = height.split('|')[0]
                    except Exception as e:
                        height_pure = math.nan
                        height_form_field_instance = 'This field does not have any data'
                        height_disname = 'Empty'
                    
                    try: 
                        weight = row['Undefined, Weight (kg)']
                        weight_pure = weight.split('|')[0]
                        weight_form_field_instance = weight.split('|')[1]
                        weight_disname = weight.split('|')[0]                    
                    except Exception as e: 
                        weight_pure = math.nan
                        weight_form_field_instance = 'This field does not have any data'
                        weight_disname = 'Empty'
                    
                    try:
                        Pre_dose_Diastolic_Blood_Pressure = row["Pre dose, Diastolic Blood Pressure"]
                        Pre_dose_Diastolic_Blood_Pressure_pure = Pre_dose_Diastolic_Blood_Pressure.split('|')[0]
                        Pre_dose_Diastolic_Blood_Pressure_form_field_insntance = Pre_dose_Diastolic_Blood_Pressure.split('|')[1]
                        Pre_dose_Diastolic_Blood_Pressure_disname = Pre_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        Pre_dose_Diastolic_Blood_Pressure_pure = math.nan
                        Pre_dose_Diastolic_Blood_Pressure_form_field_insntance = 'This field does not have any data'
                        Pre_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        Pre_dose_Diastolic_Blood_Pressure_value = row["Pre dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        Pre_dose_Diastolic_Blood_Pressure_value_pure = Pre_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        Pre_dose_Diastolic_Blood_Pressure_value_form_field_instance = Pre_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        Pre_dose_Diastolic_Blood_Pressure_value_disname = Pre_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        Pre_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        Pre_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        Pre_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Respiratory_rate = row["60-mins post dose, Respiratory rate"]
                        mins_60_post_dose_Respiratory_rate_pure = mins_60_post_dose_Respiratory_rate.split('|')[0]
                        mins_60_post_dose_Respiratory_rate_form_field_instance = mins_60_post_dose_Respiratory_rate.split('|')[1]
                        mins_60_post_dose_Respiratory_rate_disname = mins_60_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        mins_60_post_dose_Respiratory_rate_pure = math.nan
                        mins_60_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Respiratory_rate_value = row["60-mins post dose, Respiratory rate (breaths/min)"]
                        mins_60_post_dose_Respiratory_rate_value_pure = mins_60_post_dose_Respiratory_rate_value.split('|')[0]
                        mins_60_post_dose_Respiratory_rate_value_form_field_instance = mins_60_post_dose_Respiratory_rate_value.split('|')[1]
                        mins_60_post_dose_Respiratory_rate_value_disname = mins_60_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        mins_60_post_dose_Respiratory_rate_value_pure = math.nan
                        mins_60_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        Undefined_Respiratory_rate = row["Undefined, Respiratory rate"]
                        Undefined_Respiratory_rate_pure = Undefined_Respiratory_rate.split('|')[0]
                        Undefined_Respiratory_rate_form_field_isntance = Undefined_Respiratory_rate.split('|')[1]
                        Undefined_Respiratory_rate_disname = Undefined_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        Undefined_Respiratory_rate_pure = math.nan
                        Undefined_Respiratory_rate_form_field_isntance = 'This field does not have any data'
                        Undefined_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        Undefined_Respiratory_rate_value = row["Undefined, Respiratory rate (breaths/min)"]
                        Undefined_Respiratory_rate_value_pure = Undefined_Respiratory_rate_value.split('|')[0]
                        Undefined_Respiratory_rate_value_form_field_instance = Undefined_Respiratory_rate_value.split('|')[1]
                        Undefined_Respiratory_rate_value_disname = Undefined_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        Undefined_Respiratory_rate_value_pure = math.nan
                        Undefined_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        Undefined_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        Pre_dose_Oral_Temperature = row["Pre dose, Body Temperature"]
                        Pre_dose_Oral_Temperature_pure = Pre_dose_Oral_Temperature.split('|')[0]
                        Pre_dose_Oral_Temperature_form_field_instance = Pre_dose_Oral_Temperature.split('|')[1]
                        Pre_dose_Oral_Temperature_disname = Pre_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        Pre_dose_Oral_Temperature_pure = math.nan
                        Pre_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        Pre_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        Pre_dose_Oral_Temperature_value = row["Pre dose, Oral Temperature (°C)"]
                        Pre_dose_Oral_Temperature_value_pure = Pre_dose_Oral_Temperature_value.split('|')[0]
                        Pre_dose_Oral_Temperature_value_form_field_instance = Pre_dose_Oral_Temperature_value.split('|')[1]
                        Pre_dose_Oral_Temperature_value_disname = Pre_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        Pre_dose_Oral_Temperature_value_pure = math.nan
                        Pre_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        Pre_dose_Oral_Temperature_value_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Respiratory_rate = row["2-hours post dose, Respiratory rate"]
                        hours_2_post_dose_Respiratory_rate_pure = hours_2_post_dose_Respiratory_rate.split('|')[0]
                        hours_2_post_dose_Respiratory_rate_form_field_instance = hours_2_post_dose_Respiratory_rate.split('|')[1]
                        hours_2_post_dose_Respiratory_rate_disname = hours_2_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        hours_2_post_dose_Respiratory_rate_pure = math.nan
                        hours_2_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Respiratory_rate_value = row["2-hours post dose, Respiratory rate (breaths/min)"]
                        hours_2_post_dose_Respiratory_rate_value_pure = hours_2_post_dose_Respiratory_rate_value.split('|')[0]
                        hours_2_post_dose_Respiratory_rate_value_form_field_instance = hours_2_post_dose_Respiratory_rate_value.split('|')[1]
                        hours_2_post_dose_Respiratory_rate_value_disname = hours_2_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        hours_2_post_dose_Respiratory_rate_value_pure = math.nan
                        hours_2_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Diastolic_Blood_Pressure = row["8-hours post dose, Diastolic Blood Pressure"]
                        hours_8_post_dose_Diastolic_Blood_Pressure_pure = hours_8_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        hours_8_post_dose_Diastolic_Blood_Pressure_form_field_instance = hours_8_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        hours_8_post_dose_Diastolic_Blood_Pressure_disname = hours_8_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_8_post_dose_Diastolic_Blood_Pressure_pure = math.nan 
                        hours_8_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Diastolic_Blood_Pressure_value = row["8-hours post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        hours_8_post_dose_Diastolic_Blood_Pressure_value_pure = hours_8_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        hours_8_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = hours_8_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        hours_8_post_dose_Diastolic_Blood_Pressure_value_disname = hours_8_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_8_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        hours_8_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        mins_15_post_dose_Diastolic_Blood_Pressure = row["15-mins post dose, Diastolic Blood Pressure"]
                        mins_15_post_dose_Diastolic_Blood_Pressure_pure = mins_15_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        mins_15_post_dose_Diastolic_Blood_Pressure_form_field_instance = mins_15_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        mins_15_post_dose_Diastolic_Blood_Pressure_disname = mins_15_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        mins_15_post_dose_Diastolic_Blood_Pressure_pure = math.nan
                        mins_15_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        mins_15_post_dose_Diastolic_Blood_Pressure_value = row["15-mins post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        mins_15_post_dose_Diastolic_Blood_Pressure_value_pure = mins_15_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        mins_15_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = mins_15_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        mins_15_post_dose_Diastolic_Blood_Pressure_value_disname = mins_15_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        mins_15_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        mins_15_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        Undefined_Diastolic_Blood_Pressure = row["Undefined, Diastolic Blood Pressure"]
                        Undefined_Diastolic_Blood_Pressure_pure = Undefined_Diastolic_Blood_Pressure.split('|')[0]
                        Undefined_Diastolic_Blood_Pressure_form_field_instance = Undefined_Diastolic_Blood_Pressure.split('|')[1]
                        Undefined_Diastolic_Blood_Pressure_disname = Undefined_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        Undefined_Diastolic_Blood_Pressure_pure = math.nan
                        Undefined_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        Undefined_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        Undefined_Diastolic_Blood_Pressure_value = row['Undefined, Diastolic Blood Pressure (mmHg)']
                        Undefined_Diastolic_Blood_Pressure_value_pure = Undefined_Diastolic_Blood_Pressure_value.split('|')[0]
                        Undefined_Diastolic_Blood_Pressure_value_form_field_instance = Undefined_Diastolic_Blood_Pressure_value.split('|')[1]
                        Undefined_Diastolic_Blood_Pressure_value_disname = Undefined_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        Undefined_Diastolic_Blood_Pressure_value_pure = math.nan
                        Undefined_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        Undefined_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Diastolic_Blood_Pressure = row["60-mins post dose, Diastolic Blood Pressure"]
                        mins_60_post_dose_Diastolic_Blood_Pressure_pure = mins_60_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        mins_60_post_dose_Diastolic_Blood_Pressure_form_field_instance = mins_60_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        mins_60_post_dose_Diastolic_Blood_Pressure_disname = mins_60_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        mins_60_post_dose_Diastolic_Blood_Pressure_pure = math.nan 
                        mins_60_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Diastolic_Blood_Pressure_value = row["60-mins post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        mins_60_post_dose_Diastolic_Blood_Pressure_value_pure = mins_60_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        mins_60_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = mins_60_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        mins_60_post_dose_Diastolic_Blood_Pressure_value_disname = mins_60_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        mins_60_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        mins_60_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Pulse_rate = row["8-hours post dose, Pulse rate"]
                        hours_8_post_dose_Pulse_rate_pure = hours_8_post_dose_Pulse_rate.split('|')[0]
                        hours_8_post_dose_Pulse_rate_form_field_instance = hours_8_post_dose_Pulse_rate.split('|')[1]
                        hours_8_post_dose_Pulse_rate_disname = hours_8_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        hours_8_post_dose_Pulse_rate_pure = math.nan
                        hours_8_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Pulse_rate_value = row["8-hours post dose, Pulse rate (beats/min)"]
                        hours_8_post_dose_Pulse_rate_value_pure = hours_8_post_dose_Pulse_rate_value.split('|')[0]
                        hours_8_post_dose_Pulse_rate_value_form_field_instance = hours_8_post_dose_Pulse_rate_value.split('|')[1]
                        hours_8_post_dose_Pulse_rate_value_disname = hours_8_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        hours_8_post_dose_Pulse_rate_value_pure = math.nan
                        hours_8_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        Undefined_Systolic_Blood_Pressure = row["Undefined, Systolic Blood Pressure"]
                        Undefined_Systolic_Blood_Pressure_pure = Undefined_Systolic_Blood_Pressure.split('|')[0]
                        Undefined_Systolic_Blood_Pressure_form_field_instance = Undefined_Systolic_Blood_Pressure.split('|')[1]
                        Undefined_Systolic_Blood_Pressure_disname = Undefined_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        Undefined_Systolic_Blood_Pressure_pure = math.nan
                        Undefined_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        Undefined_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        Undefined_Systolic_Blood_Pressure_value = row["Undefined, Systolic Blood Pressure  (mmHg)"]
                        Undefined_Systolic_Blood_Pressure_value_pure = Undefined_Systolic_Blood_Pressure_value.split('|')[0]
                        Undefined_Systolic_Blood_Pressure_value_form_field_instance = Undefined_Systolic_Blood_Pressure_value.split('|')[1]
                        Undefined_Systolic_Blood_Pressure_value_disname = Undefined_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        Undefined_Systolic_Blood_Pressure_value_pure = math.nan
                        Undefined_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        Undefined_Systolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Diastolic_Blood_Pressure = row["2-hours post dose, Diastolic Blood Pressure"]
                        hours_2_post_dose_Diastolic_Blood_Pressure_pure = hours_2_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        hours_2_post_dose_Diastolic_Blood_Pressure_form_field_instance = hours_2_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        hours_2_post_dose_Diastolic_Blood_Pressure_disname = hours_2_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_2_post_dose_Diastolic_Blood_Pressure_pure = math.nan
                        hours_2_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Diastolic_Blood_Pressure_value = row["2-hours post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        hours_2_post_dose_Diastolic_Blood_Pressure_value_pure = hours_2_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        hours_2_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = hours_2_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        hours_2_post_dose_Diastolic_Blood_Pressure_value_disname = hours_2_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_2_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        hours_2_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Pulse_rate = row["2-hours post dose, Pulse rate"]
                        hours_2_post_dose_Pulse_rate_pure = hours_2_post_dose_Pulse_rate.split('|')[0]
                        hours_2_post_dose_Pulse_rate_form_field_instance = hours_2_post_dose_Pulse_rate.split('|')[1]
                        hours_2_post_dose_Pulse_rate_disname = hours_2_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        hours_2_post_dose_Pulse_rate_pure = math.nan
                        hours_2_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Pulse_rate_value = row["2-hours post dose, Pulse rate (beats/min)"]
                        hours_2_post_dose_Pulse_rate_value_pure = hours_2_post_dose_Pulse_rate_value.split('|')[0]
                        hours_2_post_dose_Pulse_rate_value_form_field_instance = hours_2_post_dose_Pulse_rate_value.split('|')[1]
                        hours_2_post_dose_Pulse_rate_value_disname = hours_2_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        hours_2_post_dose_Pulse_rate_value_pure = math.nan
                        hours_2_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Pulse_rate = row["60-mins post dose, Pulse rate"]
                        mins_60_post_dose_Pulse_rate_pure = mins_60_post_dose_Pulse_rate.split('|')[0]
                        mins_60_post_dose_Pulse_rate_form_field_instance = mins_60_post_dose_Pulse_rate.split('|')[1]
                        mins_60_post_dose_Pulse_rate_disname = mins_60_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        mins_60_post_dose_Pulse_rate_pure = math.nan
                        mins_60_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Pulse_rate_value = row["60-mins post dose, Pulse rate (beats/min)"]
                        mins_60_post_dose_Pulse_rate_value_pure = mins_60_post_dose_Pulse_rate_value.split('|')[0]
                        mins_60_post_dose_Pulse_rate_value_form_field_instance = mins_60_post_dose_Pulse_rate_value.split('|')[1]
                        mins_60_post_dose_Pulse_rate_value_disname = mins_60_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        mins_60_post_dose_Pulse_rate_value_pure = math.nan
                        mins_60_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        Undefined_Oral_Temperature = row["Undefined, Body Temperature"]
                        Undefined_Oral_Temperature_pure = Undefined_Oral_Temperature.split('|')[0]
                        Undefined_Oral_Temperature_form_field_instance = Undefined_Oral_Temperature.split('|')[1]
                        Undefined_Oral_Temperature_disname = Undefined_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        Undefined_Oral_Temperature_pure = math.nan
                        Undefined_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        Undefined_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        Undefined_Oral_Temperature_value = row["Undefined, Oral Temperature (°C)"]
                        Undefined_Oral_Temperature_value_pure = Undefined_Oral_Temperature_value.split('|')[0]
                        Undefined_Oral_Temperature_value_form_field_instance = Undefined_Oral_Temperature_value.split('|')[1]
                        Undefined_Oral_Temperature_value_disname = Undefined_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        Undefined_Oral_Temperature_value_pure = math.nan
                        Undefined_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        Undefined_Oral_Temperature_value_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Oral_Temperature = row["2-hours post dose, Body Temperature"]
                        hours_2_post_dose_Oral_Temperature_pure = hours_2_post_dose_Oral_Temperature.split('|')[0]
                        hours_2_post_dose_Oral_Temperature_form_field_instance = hours_2_post_dose_Oral_Temperature.split('|')[1]
                        hours_2_post_dose_Oral_Temperature_disname = hours_2_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        hours_2_post_dose_Oral_Temperature_pure = math.nan
                        hours_2_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Oral_Temperature_value = row["2-hours post dose, Oral Temperature (°C)"]
                        hours_2_post_dose_Oral_Temperature_value_pure = hours_2_post_dose_Oral_Temperature_value.split('|')[0]
                        hours_2_post_dose_Oral_Temperature_value_form_field_instance = hours_2_post_dose_Oral_Temperature_value.split('|')[1]
                        hours_2_post_dose_Oral_Temperature_value_disname = hours_2_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        hours_2_post_dose_Oral_Temperature_value_pure = math.nan
                        hours_2_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Oral_Temperature_value_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Oral_Temperature = row["60-mins post dose, Body Temperature"]
                        mins_60_post_dose_Oral_Temperature_pure = mins_60_post_dose_Oral_Temperature.split('|')[0]
                        mins_60_post_dose_Oral_Temperature_form_field_instance = mins_60_post_dose_Oral_Temperature.split('|')[1]
                        mins_60_post_dose_Oral_Temperature_disname = mins_60_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        mins_60_post_dose_Oral_Temperature_pure = math.nan
                        mins_60_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Oral_Temperature_value = row["60-mins post dose, Oral Temperature (°C)"]
                        mins_60_post_dose_Oral_Temperature_value_pure = mins_60_post_dose_Oral_Temperature_value.split('|')[0]
                        mins_60_post_dose_Oral_Temperature_value_form_field_instance = mins_60_post_dose_Oral_Temperature_value.split('|')[1]
                        mins_60_post_dose_Oral_Temperature_value_disname = mins_60_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        mins_60_post_dose_Oral_Temperature_value_pure = math.nan
                        mins_60_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Oral_Temperature_value_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Respiratory_rate = row["8-hours post dose, Respiratory rate"]
                        hours_8_post_dose_Respiratory_rate_pure = hours_8_post_dose_Respiratory_rate.split('|')[0]
                        hours_8_post_dose_Respiratory_rate_form_field_instance = hours_8_post_dose_Respiratory_rate.split('|')[1]
                        hours_8_post_dose_Respiratory_rate_disname = hours_8_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        hours_8_post_dose_Respiratory_rate_pure = math.nan
                        hours_8_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Respiratory_rate_value = row["8-hours post dose, Respiratory rate (breaths/min)"]
                        hours_8_post_dose_Respiratory_rate_value_pure = hours_8_post_dose_Respiratory_rate_value.split('|')[0]
                        hours_8_post_dose_Respiratory_rate_value_form_field_instance = hours_8_post_dose_Respiratory_rate_value.split('|')[1]
                        hours_8_post_dose_Respiratory_rate_value_disname = hours_8_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        hours_8_post_dose_Respiratory_rate_value_pure = math.nan
                        hours_8_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Oral_Temperature = row["12-hours post dose, Body Temperature"]
                        hours_12_post_dose_Oral_Temperature_pure = hours_12_post_dose_Oral_Temperature.split('|')[0]
                        hours_12_post_dose_Oral_Temperature_form_field_instance = hours_12_post_dose_Oral_Temperature.split('|')[1]
                        hours_12_post_dose_Oral_Temperature_disname = hours_12_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        hours_12_post_dose_Oral_Temperature_pure = math.nan
                        hours_12_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Oral_Temperature_value = row["12-hours post dose, Oral Temperature (°C)"]
                        hours_12_post_dose_Oral_Temperature_value_pure = hours_12_post_dose_Oral_Temperature_value.split('|')[0]
                        hours_12_post_dose_Oral_Temperature_value_form_field_instance = hours_12_post_dose_Oral_Temperature_value.split('|')[1]
                        hours_12_post_dose_Oral_Temperature_value_disname = hours_12_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        hours_12_post_dose_Oral_Temperature_value_pure = math.nan
                        hours_12_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Oral_Temperature_value_disname = 'Empty' 
                        
                    
                    try:
                        Pre_dose_Pulse_rate = row["Pre dose, Pulse rate"]
                        Pre_dose_Pulse_rate_pure = Pre_dose_Pulse_rate.split('|')[0]
                        Pre_dose_Pulse_rate_form_field_instance = Pre_dose_Pulse_rate.split('|')[1]
                        Pre_dose_Pulse_rate_disname = Pre_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        Pre_dose_Pulse_rate_pure = math.nan
                        Pre_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        Pre_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        Pre_dose_Pulse_rate_value = row["Pre dose, Pulse rate (beats/min)"]
                        Pre_dose_Pulse_rate_value_pure = Pre_dose_Pulse_rate_value.split('|')[0]
                        Pre_dose_Pulse_rate_value_form_field_instance = Pre_dose_Pulse_rate_value.split('|')[1]
                        Pre_dose_Pulse_rate_value_disname = Pre_dose_Pulse_rate_value.split('|')[2]
                    except Exception as e:
                        Pre_dose_Pulse_rate_value_pure = math.nan
                        Pre_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        Pre_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Systolic_Blood_Pressure = row["8-hours post dose, Systolic Blood Pressure"]
                        hours_8_post_dose_Systolic_Blood_Pressure_pure = hours_8_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        hours_8_post_dose_Systolic_Blood_Pressure_form_field_instance = hours_8_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        hours_8_post_dose_Systolic_Blood_Pressure_disname = hours_8_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_8_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        hours_8_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Systolic_Blood_Pressure_value = row["8-hours post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        hours_8_post_dose_Systolic_Blood_Pressure_value_pure = hours_8_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        hours_8_post_dose_Systolic_Blood_Pressure_value_form_field_instance = hours_8_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        hours_8_post_dose_Systolic_Blood_Pressure_value_disname = hours_8_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_8_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        hours_8_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Systolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Systolic_Blood_Pressure = row["30-mins post dose, Systolic Blood Pressure"]
                        mins_30_post_dose_Systolic_Blood_Pressure_pure = mins_30_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        mins_30_post_dose_Systolic_Blood_Pressure_form_field_instance = mins_30_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        mins_30_post_dose_Systolic_Blood_Pressure_disname = mins_30_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        mins_30_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        mins_30_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Systolic_Blood_Pressure_value = row["30-mins post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        mins_30_post_dose_Systolic_Blood_Pressure_value_pure = mins_30_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        mins_30_post_dose_Systolic_Blood_Pressure_value_form_field_instance = mins_30_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        mins_30_post_dose_Systolic_Blood_Pressure_value_disname = mins_30_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        mins_30_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        mins_30_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Systolic_Blood_Pressure_value_disname = 'Empty'
 
                    try:
                        hours_4_post_dose_Respiratory_rate = row["4-hours post dose, Respiratory rate"]
                        hours_4_post_dose_Respiratory_rate_pure = hours_4_post_dose_Respiratory_rate.split('|')[0]
                        hours_4_post_dose_Respiratory_rate_form_field_instance = hours_4_post_dose_Respiratory_rate.split('|')[1]
                        hours_4_post_dose_Respiratory_rate_disname = hours_4_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        hours_4_post_dose_Respiratory_rate_pure = math.nan
                        hours_4_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        hours_4_post_dose_Respiratory_rate_value = row["4-hours post dose, Respiratory rate (breaths/min)"]
                        hours_4_post_dose_Respiratory_rate_value_pure = hours_4_post_dose_Respiratory_rate_value.split('|')[0]
                        hours_4_post_dose_Respiratory_rate_value_form_field_instance = hours_4_post_dose_Respiratory_rate_value.split('|')[1]
                        hours_4_post_dose_Respiratory_rate_value_disname = hours_4_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        hours_4_post_dose_Respiratory_rate_value_pure = math.nan
                        hours_4_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Diastolic_Blood_Pressure = row["30-mins post dose, Diastolic Blood Pressure"]
                        mins_30_post_dose_Diastolic_Blood_Pressure_pure = mins_30_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        mins_30_post_dose_Diastolic_Blood_Pressure_form_field_instance = mins_30_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        mins_30_post_dose_Diastolic_Blood_Pressure_disname = mins_30_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        mins_30_post_dose_Diastolic_Blood_Pressure_pure = math.nan
                        mins_30_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Diastolic_Blood_Pressure_value = row["30-mins post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        mins_30_post_dose_Diastolic_Blood_Pressure_value_pure = mins_30_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        mins_30_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = mins_30_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        mins_30_post_dose_Diastolic_Blood_Pressure_value_disname = mins_30_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        mins_30_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        mins_30_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Pulse_rate = row["30-mins post dose, Pulse rate"]
                        mins_30_post_dose_Pulse_rate_pure = mins_30_post_dose_Pulse_rate.split('|')[0]
                        mins_30_post_dose_Pulse_rate_form_field_instance = mins_30_post_dose_Pulse_rate.split('|')[1]
                        mins_30_post_dose_Pulse_rate_disname = mins_30_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        mins_30_post_dose_Pulse_rate_pure = math.nan
                        mins_30_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Pulse_rate_value = row["30-mins post dose, Pulse rate (beats/min)"]
                        mins_30_post_dose_Pulse_rate_value_pure = mins_30_post_dose_Pulse_rate_value.split('|')[0]
                        mins_30_post_dose_Pulse_rate_value_form_field_instance = mins_30_post_dose_Pulse_rate_value.split('|')[1]
                        mins_30_post_dose_Pulse_rate_value_disname = mins_30_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        mins_30_post_dose_Pulse_rate_value_pure = math.nan
                        mins_30_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        Pre_dose_Respiratory_rate = row["Pre dose, Respiratory rate"]
                        Pre_dose_Respiratory_rate_pure = Pre_dose_Respiratory_rate.split('|')[0]
                        Pre_dose_Respiratory_rate_form_field_instance = Pre_dose_Respiratory_rate.split('|')[1]
                        Pre_dose_Respiratory_rate_disname = Pre_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        Pre_dose_Respiratory_rate_pure = math.nan
                        Pre_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        Pre_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        Pre_dose_Respiratory_rate_value = row["Pre dose, Respiratory rate (breaths/min)"]
                        Pre_dose_Respiratory_rate_value_pure = Pre_dose_Respiratory_rate_value.split('|')[0]
                        Pre_dose_Respiratory_rate_value_form_field_instance = Pre_dose_Respiratory_rate_value.split('|')[1]
                        Pre_dose_Respiratory_rate_value_disname = Pre_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        Pre_dose_Respiratory_rate_value_pure = math.nan
                        Pre_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        Pre_dose_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Respiratory_rate = row["12-hours post dose, Respiratory rate"]
                        hours_12_post_dose_Respiratory_rate_pure = hours_12_post_dose_Respiratory_rate.split('|')[0]
                        hours_12_post_dose_Respiratory_rate_form_field_instance = hours_12_post_dose_Respiratory_rate.split('|')[1]
                        hours_12_post_dose_Respiratory_rate_disname = hours_12_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        hours_12_post_dose_Respiratory_rate_pure = math.nan
                        hours_12_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Respiratory_rate_value = row["12-hours post dose, Respiratory rate (breaths/min)"]
                        hours_12_post_dose_Respiratory_rate_value_pure = hours_12_post_dose_Respiratory_rate_value.split('|')[0]
                        hours_12_post_dose_Respiratory_rate_value_form_field_instance = hours_12_post_dose_Respiratory_rate_value.split('|')[1]
                        hours_12_post_dose_Respiratory_rate_value_disname = hours_12_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        hours_12_post_dose_Respiratory_rate_value_pure = math.nan
                        hours_12_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Respiratory_rate_value_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Pulse_rate = row["12-hours post dose, Pulse rate"]
                        hours_12_post_dose_Pulse_rate_pure = hours_12_post_dose_Pulse_rate.split('|')[0]
                        hours_12_post_dose_Pulse_rate_form_field_instance = hours_12_post_dose_Pulse_rate.split('|')[1]
                        hours_12_post_dose_Pulse_rate_disname = hours_12_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        hours_12_post_dose_Pulse_rate_pure = math.nan
                        hours_12_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Pulse_rate_value = row["12-hours post dose, Pulse rate (beats/min)"]
                        hours_12_post_dose_Pulse_rate_value_pure = hours_12_post_dose_Pulse_rate_value.split('|')[0]
                        hours_12_post_dose_Pulse_rate_value_form_field_instance = hours_12_post_dose_Pulse_rate_value.split('|')[1]
                        hours_12_post_dose_Pulse_rate_value_disname = hours_12_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        hours_12_post_dose_Pulse_rate_value_pure = math.nan
                        hours_12_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Systolic_Blood_Pressure = row["2-hours post dose, Systolic Blood Pressure"]
                        hours_2_post_dose_Systolic_Blood_Pressure_pure = hours_2_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        hours_2_post_dose_Systolic_Blood_Pressure_form_field_instance = hours_2_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        hours_2_post_dose_Systolic_Blood_Pressure_disname = hours_2_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_2_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        hours_2_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        hours_2_post_dose_Systolic_Blood_Pressure_value = row["2-hours post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        hours_2_post_dose_Systolic_Blood_Pressure_value_pure = hours_2_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        hours_2_post_dose_Systolic_Blood_Pressure_value_form_field_instance = hours_2_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        hours_2_post_dose_Systolic_Blood_Pressure_value_empty = hours_2_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_2_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        hours_2_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_2_post_dose_Systolic_Blood_Pressure_value_empty = 'Empty'
                    
                    try:
                        hours_8_post_dose_Oral_Temperature = row["8-hours post dose, Body Temperature"]
                        hours_8_post_dose_Oral_Temperature_pure = hours_8_post_dose_Oral_Temperature.split('|')[0]
                        hours_8_post_dose_Oral_Temperature_form_field_instance = hours_8_post_dose_Oral_Temperature.split('|')[1]
                        hours_8_post_dose_Oral_Temperature_disname = hours_8_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        hours_8_post_dose_Oral_Temperature_pure = math.nan 
                        hours_8_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        hours_8_post_dose_Oral_Temperature_value = row["8-hours post dose, Oral Temperature (°C)"]
                        hours_8_post_dose_Oral_Temperature_value_pure = hours_8_post_dose_Oral_Temperature_value.split('|')[0]
                        hours_8_post_dose_Oral_Temperature_value_form_field_instance = hours_8_post_dose_Oral_Temperature_value.split('|')[1]
                        hours_8_post_dose_Oral_Temperature_value_disname = hours_8_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        hours_8_post_dose_Oral_Temperature_value_pure = math.nan
                        hours_8_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        hours_8_post_dose_Oral_Temperature_value_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Systolic_Blood_Pressure = row["60-mins post dose, Systolic Blood Pressure"]
                        mins_60_post_dose_Systolic_Blood_Pressure_pure = mins_60_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        mins_60_post_dose_Systolic_Blood_Pressure_form_field_instance = mins_60_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        mins_60_post_dose_Systolic_Blood_Pressure_disname = mins_60_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        mins_60_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        mins_60_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        mins_60_post_dose_Systolic_Blood_Pressure_value = row["60-mins post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        mins_60_post_dose_Systolic_Blood_Pressure_value_pure = mins_60_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        mins_60_post_dose_Systolic_Blood_Pressure_value_form_field_instance = mins_60_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        mins_60_post_dose_Systolic_Blood_Pressure_value_disname = mins_60_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        mins_60_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        mins_60_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        mins_60_post_dose_Systolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        mins_15_post_dose_Oral_Temperature = row["15-mins post dose, Body Temperature"]
                        mins_15_post_dose_Oral_Temperature_pure = mins_15_post_dose_Oral_Temperature.split('|')[0]
                        mins_15_post_dose_Oral_Temperature_form_field_instance = mins_15_post_dose_Oral_Temperature.split('|')[1]
                        mins_15_post_dose_Oral_Temperature_disname = mins_15_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        mins_15_post_dose_Oral_Temperature_pure = math.nan
                        mins_15_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        mins_15_post_dose_Oral_Temperature_value = row["15-mins post dose, Oral Temperature (°C)"]
                        mins_15_post_dose_Oral_Temperature_value_pure = mins_15_post_dose_Oral_Temperature_value.split('|')[0]
                        mins_15_post_dose_Oral_Temperature_value_form_field_instance = mins_15_post_dose_Oral_Temperature_value.split('|')[1]
                        mins_15_post_dose_Oral_Temperature_value_disname = mins_15_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        mins_15_post_dose_Oral_Temperature_value_pure = math.nan
                        mins_15_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Oral_Temperature_value_disname = 'Empty'
                    
                    try:
                        hours_4_post_dose_Pulse_rate = row["4-hours post dose, Pulse rate"]
                        hours_4_post_dose_Pulse_rate_pure = hours_4_post_dose_Pulse_rate.split('|')[0]
                        hours_4_post_dose_Pulse_rate_form_field_instance = hours_4_post_dose_Pulse_rate.split('|')[1]
                        hours_4_post_dose_Pulse_rate_disname = hours_4_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        hours_4_post_dose_Pulse_rate_pure = math.nan
                        hours_4_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        hours_4_post_dose_Pulse_rate_value = row["4-hours post dose, Pulse rate (beats/min)"]
                        hours_4_post_dose_Pulse_rate_value_pure = hours_4_post_dose_Pulse_rate_value.split('|')[0]
                        hours_4_post_dose_Pulse_rate_value_form_field_instance = hours_4_post_dose_Pulse_rate_value.split('|')[1]
                        hours_4_post_dose_Pulse_rate_value_disname = hours_4_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        hours_4_post_dose_Pulse_rate_value_pure = math.nan
                        hours_4_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        Pre_dose_Systolic_Blood_Pressure = row["Pre dose, Systolic Blood Pressure"]
                        Pre_dose_Systolic_Blood_Pressure_pure = Pre_dose_Systolic_Blood_Pressure.split('|')[0]
                        Pre_dose_Systolic_Blood_Pressure_form_field_instance = Pre_dose_Systolic_Blood_Pressure.split('|')[1]
                        Pre_dose_Systolic_Blood_Pressure_disname = Pre_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        Pre_dose_Systolic_Blood_Pressure_pure = math.nan
                        Pre_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        Pre_dose_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        Pre_dose_Systolic_Blood_Pressure_value = row["Pre dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        Pre_dose_Systolic_Blood_Pressure_value_pure = Pre_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        Pre_dose_Systolic_Blood_Pressure_value_form_field_instance = Pre_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        Pre_dose_Systolic_Blood_Pressure_value_disname = Pre_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        Pre_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        Pre_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        Pre_dose_Systolic_Blood_Pressure_value_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Systolic_Blood_Pressure = row["12-hours post dose, Systolic Blood Pressure"]
                        hours_12_post_dose_Systolic_Blood_Pressure_pure = hours_12_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        hours_12_post_dose_Systolic_Blood_Pressure_form_field_instance = hours_12_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        hours_12_post_dose_Systolic_Blood_Pressure_disname = hours_12_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_12_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        hours_12_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        hours_12_post_dose_Systolic_Blood_Pressure_value = row["12-hours post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        hours_12_post_dose_Systolic_Blood_Pressure_value_pure = hours_12_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        hours_12_post_dose_Systolic_Blood_Pressure_value_form_field_instance = hours_12_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        hours_12_post_dose_Systolic_Blood_Pressure_value_disname = hours_12_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_12_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        hours_12_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Systolic_Blood_Pressure_value_disname = 'Empty'

                    try:
                        mins_30_post_dose_Oral_Temperature = row["30-mins post dose, Body Temperature"]
                        mins_30_post_dose_Oral_Temperature_pure = mins_30_post_dose_Oral_Temperature.split('|')[0]
                        mins_30_post_dose_Oral_Temperature_form_field_instance = mins_30_post_dose_Oral_Temperature.split('|')[1]
                        mins_30_post_dose_Oral_Temperature_disname = mins_30_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        mins_30_post_dose_Oral_Temperature_pure = math.nan 
                        mins_30_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Oral_Temperature_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Oral_Temperature_value = row["30-mins post dose, Oral Temperature (°C)"]
                        mins_30_post_dose_Oral_Temperature_value_pure = mins_30_post_dose_Oral_Temperature_value.split('|')[0]
                        mins_30_post_dose_Oral_Temperature_value_form_field_instance = mins_30_post_dose_Oral_Temperature_value.split('|')[1]
                        mins_30_post_dose_Oral_Temperature_value_disname = mins_30_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        mins_30_post_dose_Oral_Temperature_value_pure = math.nan
                        mins_30_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Oral_Temperature_value_disname = 'Empty'
                   
                    try:
                        hours_4_post_dose_Oral_Temperature = row["4-hours post dose, Body Temperature"]
                        hours_4_post_dose_Oral_Temperature_pure = hours_4_post_dose_Oral_Temperature.split('|')[0]
                        hours_4_post_dose_Oral_Temperature_form_field_instance = hours_4_post_dose_Oral_Temperature.split('|')[1]
                        hours_4_post_dose_Oral_Temperature_disname = hours_4_post_dose_Oral_Temperature.split('|')[2]
                    except Exception as e:
                        hours_4_post_dose_Oral_Temperature_pure = math.nan
                        hours_4_post_dose_Oral_Temperature_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Oral_Temperature_disname = 'Empty'
                   
                    try:
                        hours_4_post_dose_Oral_Temperature_value = row["4-hours post dose, Oral Temperature (°C)"]
                        hours_4_post_dose_Oral_Temperature_value_pure = hours_4_post_dose_Oral_Temperature_value.split('|')[0]
                        hours_4_post_dose_Oral_Temperature_value_form_field_instance = hours_4_post_dose_Oral_Temperature_value.split('|')[1]
                        hours_4_post_dose_Oral_Temperature_value_disname = hours_4_post_dose_Oral_Temperature_value.split('|')[0]
                    except Exception as e:
                        hours_4_post_dose_Oral_Temperature_value_pure = math.nan
                        hours_4_post_dose_Oral_Temperature_value_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Oral_Temperature_value_disname = 'Empty'
                   
                    try:
                        hours_12_post_dose_Diastolic_Blood_Pressure = row["12-hours post dose, Diastolic Blood Pressure"]
                        hours_12_post_dose_Diastolic_Blood_Pressure_pure = hours_12_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        hours_12_post_dose_Diastolic_Blood_Pressure_form_field_instance = hours_12_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        hours_12_post_dose_Diastolic_Blood_Pressure_disname = hours_12_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_12_post_dose_Diastolic_Blood_Pressure_pure = math.nan 
                        hours_12_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                   
                    try:
                        hours_12_post_dose_Diastolic_Blood_Pressure_value = row["12-hours post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        hours_12_post_dose_Diastolic_Blood_Pressure_value_pure = hours_12_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        hours_12_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = hours_12_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        hours_12_post_dose_Diastolic_Blood_Pressure_value_disname = hours_12_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_12_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan
                        hours_12_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_12_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                   
                    try:
                        mins_15_post_dose_Systolic_Blood_Pressure = row["15-mins post dose, Systolic Blood Pressure"]
                        mins_15_post_dose_Systolic_Blood_Pressure_pure = mins_15_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        mins_15_post_dose_Systolic_Blood_Pressure_form_field_instance = mins_15_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        mins_15_post_dose_Systolic_Blood_Pressure_disname = mins_15_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        mins_15_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        mins_15_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                   
                    try:
                        mins_15_post_dose_Systolic_Blood_Pressure_value = row["15-mins post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        mins_15_post_dose_Systolic_Blood_Pressure_value_pure = mins_15_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        mins_15_post_dose_Systolic_Blood_Pressure_value_form_field_instance = mins_15_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        mins_15_post_dose_Systolic_Blood_Pressure_value_disname = mins_15_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        mins_15_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        mins_15_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Systolic_Blood_Pressure_value_disname = 'Empty'
                   
                    try:
                        hours_4_post_dose_Diastolic_Blood_Pressure = row["4-hours post dose, Diastolic Blood Pressure"]
                        hours_4_post_dose_Diastolic_Blood_Pressure_pure = hours_4_post_dose_Diastolic_Blood_Pressure.split('|')[0]
                        hours_4_post_dose_Diastolic_Blood_Pressure_form_field_instance = hours_4_post_dose_Diastolic_Blood_Pressure.split('|')[1]
                        hours_4_post_dose_Diastolic_Blood_Pressure_disname = hours_4_post_dose_Diastolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_4_post_dose_Diastolic_Blood_Pressure_pure =math.nan
                        hours_4_post_dose_Diastolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Diastolic_Blood_Pressure_disname = 'Empty'
                    
                    try:
                        hours_4_post_dose_Diastolic_Blood_Pressure_value = row["4-hours post dose, Diastolic Blood Pressure (Sitting) (mmHg)"]
                        hours_4_post_dose_Diastolic_Blood_Pressure_value_pure = hours_4_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                        hours_4_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = hours_4_post_dose_Diastolic_Blood_Pressure_value.split('|')[1]
                        hours_4_post_dose_Diastolic_Blood_Pressure_value_disname = hours_4_post_dose_Diastolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_4_post_dose_Diastolic_Blood_Pressure_value_pure = math.nan 
                        hours_4_post_dose_Diastolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Diastolic_Blood_Pressure_value_disname = 'Empty'
                  
                    try:
                        Undefined_Pulse_rate = row["Undefined, Pulse rate"]
                        Undefined_Pulse_rate_pure = Undefined_Pulse_rate.split('|')[0]
                        Undefined_Pulse_rate_form_field_instance = Undefined_Pulse_rate.split('|')[1]
                        Undefined_Pulse_rate_disname = Undefined_Pulse_rate.split('|')[2]
                    except Exception as e:
                        Undefined_Pulse_rate_pure = math.nan
                        Undefined_Pulse_rate_form_field_instance = 'This field does not have any data'
                        Undefined_Pulse_rate_disname = 'Empty'
                   
                    try:
                        Undefined_Pulse_rate_value = row["Undefined, Pulse rate (beats/min)"]
                        Undefined_Pulse_rate_value_pure = Undefined_Pulse_rate_value.split('|')[0]
                        Undefined_Pulse_rate_value_form_field_instance = Undefined_Pulse_rate_value.split('|')[1]
                        Undefined_Pulse_rate_value_disname = Undefined_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        Undefined_Pulse_rate_value_pure = math.nan
                        Undefined_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        Undefined_Pulse_rate_value_disname = 'Empty'
                    
                    try:
                        hours_4_post_dose_Systolic_Blood_Pressure = row["4-hours post dose, Systolic Blood Pressure"]
                        hours_4_post_dose_Systolic_Blood_Pressure_pure = hours_4_post_dose_Systolic_Blood_Pressure.split('|')[0]
                        hours_4_post_dose_Systolic_Blood_Pressure_form_field_instance = hours_4_post_dose_Systolic_Blood_Pressure.split('|')[1]
                        hours_4_post_dose_Systolic_Blood_Pressure_disname = hours_4_post_dose_Systolic_Blood_Pressure.split('|')[2]
                    except Exception as e:
                        hours_4_post_dose_Systolic_Blood_Pressure_pure = math.nan
                        hours_4_post_dose_Systolic_Blood_Pressure_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Systolic_Blood_Pressure_disname = 'Empty'
                   
                    try:
                        hours_4_post_dose_Systolic_Blood_Pressure_value = row["4-hours post dose, Systolic Blood Pressure  (Sitting) (mmHg)"]
                        hours_4_post_dose_Systolic_Blood_Pressure_value_pure = hours_4_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                        hours_4_post_dose_Systolic_Blood_Pressure_value_form_field_instance = hours_4_post_dose_Systolic_Blood_Pressure_value.split('|')[1]
                        hours_4_post_dose_Systolic_Blood_Pressure_value_disname = hours_4_post_dose_Systolic_Blood_Pressure_value.split('|')[0]
                    except Exception as e:
                        hours_4_post_dose_Systolic_Blood_Pressure_value_pure = math.nan
                        hours_4_post_dose_Systolic_Blood_Pressure_value_form_field_instance = 'This field does not have any data'
                        hours_4_post_dose_Systolic_Blood_Pressure_value_disname = 'Empty'
                  
                    try:
                        mins_15_post_dose_Respiratory_rate = row["15-mins post dose, Respiratory rate"]
                        mins_15_post_dose_Respiratory_rate_pure = mins_15_post_dose_Respiratory_rate.split('|')[0]
                        mins_15_post_dose_Respiratory_rate_form_field_instance = mins_15_post_dose_Respiratory_rate.split('|')[1]
                        mins_15_post_dose_Respiratory_rate_disname = mins_15_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        mins_15_post_dose_Respiratory_rate_pure = math.nan
                        mins_15_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Respiratory_rate_disname = 'Empty'
                  
                    try:
                        mins_15_post_dose_Respiratory_rate_value = row["15-mins post dose, Respiratory rate (breaths/min)"]
                        mins_15_post_dose_Respiratory_rate_value_pure = mins_15_post_dose_Respiratory_rate_value.split('|')[0]
                        mins_15_post_dose_Respiratory_rate_value_form_field_instance = mins_15_post_dose_Respiratory_rate_value.split('|')[1]
                        mins_15_post_dose_Respiratory_rate_value_disname = mins_15_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        mins_15_post_dose_Respiratory_rate_value_pure = math.nan
                        mins_15_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Respiratory_rate_value_disname = 'Empty'
                   
                    try:
                        mins_30_post_dose_Respiratory_rate = row["30-mins post dose, Respiratory rate"]
                        mins_30_post_dose_Respiratory_rate_pure = mins_30_post_dose_Respiratory_rate.split('|')[0]
                        mins_30_post_dose_Respiratory_rate_form_field_instance = mins_30_post_dose_Respiratory_rate.split('|')[1]
                        mins_30_post_dose_Respiratory_rate_disname = mins_30_post_dose_Respiratory_rate.split('|')[2]
                    except Exception as e:
                        mins_30_post_dose_Respiratory_rate_pure = math.nan
                        mins_30_post_dose_Respiratory_rate_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Respiratory_rate_disname = 'Empty'
                    
                    try:
                        mins_30_post_dose_Respiratory_rate_value = row["30-mins post dose, Respiratory rate (breaths/min)"]
                        mins_30_post_dose_Respiratory_rate_value_pure = mins_30_post_dose_Respiratory_rate_value.split('|')[0]
                        mins_30_post_dose_Respiratory_rate_value_form_field_instance = mins_30_post_dose_Respiratory_rate_value.split('|')[1]
                        mins_30_post_dose_Respiratory_rate_value_disname = mins_30_post_dose_Respiratory_rate_value.split('|')[0]
                    except Exception as e:
                        mins_30_post_dose_Respiratory_rate_value_pure = math.nan
                        mins_30_post_dose_Respiratory_rate_value_form_field_instance = 'This field does not have any data'
                        mins_30_post_dose_Respiratory_rate_value_disname = 'Empty'
                        
                    try:
                        mins_15_post_dose_Pulse_rate = row["15-mins post dose, Pulse rate"]
                        mins_15_post_dose_Pulse_rate_pure = mins_15_post_dose_Pulse_rate.split('|')[0]
                        mins_15_post_dose_Pulse_rate_form_field_instance = mins_15_post_dose_Pulse_rate.split('|')[1]
                        mins_15_post_dose_Pulse_rate_disname = mins_15_post_dose_Pulse_rate.split('|')[2]
                    except Exception as e:
                        mins_15_post_dose_Pulse_rate_pure = math.nan 
                        mins_15_post_dose_Pulse_rate_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Pulse_rate_disname = 'Empty'
                    
                    try:
                        mins_15_post_dose_Pulse_rate_value = row["15-mins post dose, Pulse rate (beats/min)"]
                        mins_15_post_dose_Pulse_rate_value_pure = mins_15_post_dose_Pulse_rate_value.split('|')[0]
                        mins_15_post_dose_Pulse_rate_value_form_field_instance = mins_15_post_dose_Pulse_rate_value.split('|')[1]
                        mins_15_post_dose_Pulse_rate_value_disname = mins_15_post_dose_Pulse_rate_value.split('|')[0]
                    except Exception as e:
                        mins_15_post_dose_Pulse_rate_value_pure = math.nan
                        mins_15_post_dose_Pulse_rate_value_form_field_instance = 'This field does not have any data'
                        mins_15_post_dose_Pulse_rate_value_disname = 'Empty'
                    
                    #----------- TIMES
                    try:
                        predose_time = row['Pre dose, Time']
                        predose_time_pure = predose_time.split('|')[0]
                        predose_time_form_field_definition = predose_time.split('|')[1]
                    except Exception as e:
                        predose_time_pure = math.nan
                        predose_time_form_field_definition = 'This field does not have any data'
                    
                    try:
                        post_dose_15 = row['15-mins post dose, Time']
                        post_dose_15_pure = post_dose_15.split('|')[0]
                        post_dose_15_form_field_instance = post_dose_15.split('|')[1]
                    except:
                        post_dose_15_pure = math.nan
                        post_dose_15_form_field_instance = 'This field does not have any data'
                    
                    try:
                        post_dose_30 = row['30-mins post dose, Time']
                        post_dose_30_pure = post_dose_30.split('|')[0]
                        post_dose_30_form_field_instance = post_dose_30.split('|')[1]
                    except:
                        post_dose_30_pure = math.nan
                        post_dose_30_form_field_instance = 'This field does not have any data'
                    
                    try:
                        post_dose_60 = row['60-mins post dose, Time']
                        post_dose_60_pure = post_dose_60.split('|')[0]
                        post_dose_60_form_field_instance =  post_dose_60.split('|')[1]
                    except:
                        post_dose_60_pure = math.nan
                        post_dose_60_form_field_instance = 'This field does not have any data'

                    try:
                        post_dose_2H = row['2-hours post dose, Time']
                        post_dose_2H_pure = post_dose_2H.split('|')[0]
                        post_dose_2H_form_field_instance = post_dose_2H.split('|')[1]
                    except:
                        post_dose_2H_pure = math.nan 
                        post_dose_2H_form_field_instance  = 'This field does not have any data'

                    try:
                        post_dose_4H = row['4-hours post dose, Time']
                        post_dose_4H_pure = post_dose_4H.split('|')[0]
                        post_dose_4H_form_field_instance = post_dose_4H.split('|')[1]
                    except:
                        post_dose_4H_pure = math.nan 
                        post_dose_4H_form_field_instance = 'This field does not have any data'

                    try:
                        post_dose_8H = row['8-hours post dose, Time']
                        post_dose_8H_pure = post_dose_8H.split('|')[0]
                        post_dose_8H_form_field_instance = post_dose_8H.split('|')[1]
                    except:
                        post_dose_8H_pure = math.nan 
                        post_dose_8H_form_field_instance = 'This field does not have any data'  

                    try:
                        post_dose_12H = row['12-hours post dose, Time']
                        post_dose_12H_pure = post_dose_12H.split('|')[0]
                        post_dose_12H_form_field_instance = post_dose_12H.split('|')[1]
                    except:
                        post_dose_12H_pure = math.nan 
                        post_dose_12H_form_field_instance = 'This field does not have any data'  

                    # --------------------------------------------------------------------------------------------------------   
                    if date_assesment_pure == '':
                        pass
                    else:         
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_assesment_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of assessment performed', date_assesment_form_field_instance ,f , date_assesment_pure, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision VS0020
                    if date_assesment_pure != '' and date_of_visit != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_assesment_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of assessment performed', date_assesment_form_field_instance,
                                        'The date should be the same as the visit date in the "Date of Visit" Form', f'{date_assesment_pure} - {date_of_visit}', 'VS0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision VS0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision VS0030
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_assesment_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_assesment_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else:
                                error = [subject, visit, 'Date of assessment performed', date_assesment_form_field_instance,
                                        'The date/time of test performed can not be before the informed consent date/time', f'{date_assesment_pure} - {date_inform_consent}', 'VS0030']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision VS0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> VS0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_assesment_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_assesment_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of assessment performed', date_assesment_form_field_instance ,'Visit Date must be before the End of study/Early withdrawal date. ', date_assesment_pure, 'VS0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision VS0040 --> {e}  - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision VS0050
                    try:
                        if  math.isnan(float(BMI_pure)) == False:
                            if visita == 'Screening Visit':
                                pass
                            else:
                                error = [subject, visit, 'Undefined, BMI', BMI_form_field_instance, 'Only at Screening', BMI_form_disname, 'VS0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision VS0050--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision VS0060
                    try:
                        if  math.isnan(float(height_pure)) == False:
                            if visita == 'Screening Visit':
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Height (cm)', height_form_field_instance,'Only at Screening', height_disname, 'VS0060']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision VS0060--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision VS0070
                    try:
                        if  math.isnan(float(weight_pure)) == False:
                            if visita == 'Screening Visit' or visita == 'D-1' or visita == 'D42' or visita == 'Unscheduled':
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Weight (kg)', weight_form_field_instance, 'Only at Screening', weight_disname, 'VS0070']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision VS0070--> {e} - Subject: {subject},  Visit: {visit} ')


                    try:
                        # Revision VS0080
                        if float(Undefined_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(Undefined_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(Undefined_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Diastolic Blood Pressure', Undefined_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"Undefined, Diastolic Blood Pressure: {Undefined_Diastolic_Blood_Pressure_disname} - Undefined, Diastolic Blood Pressure Result: {Undefined_Diastolic_Blood_Pressure_value_disname}", 'VS0080']
                                lista_revision.append(error)

                        # Revision VS0170
                        elif float(Undefined_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(Undefined_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(Undefined_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, 'Undefined, Diastolic Blood Pressure', Undefined_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"Undefined, Diastolic Blood Pressure interpretation: {Undefined_Diastolic_Blood_Pressure_disname} - Undefined, Diastolic Blood Pressure Result: {Undefined_Diastolic_Blood_Pressure_value_disname}", 'VS0170']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0080--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0090
                        if float(Pre_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(Pre_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(Pre_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose, Diastolic Blood Pressure', Pre_dose_Diastolic_Blood_Pressure_form_field_insntance ,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Diastolic Blood Pressure interpretation: {Pre_dose_Diastolic_Blood_Pressure_disname} - Pre dose, Diastolic Blood Pressure Result: {Pre_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0090']
                                lista_revision.append(error)

                        # Revision VS0180
                        elif float(Pre_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(Pre_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and\
                                  float(Pre_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, 'Pre dose, Diastolic Blood Pressure', Pre_dose_Diastolic_Blood_Pressure_form_field_insntance ,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                         f"Pre dose, Diastolic Blood Pressure interpretation: {Pre_dose_Diastolic_Blood_Pressure_disname} - Pre dose, Diastolic Blood Pressure Result: {Pre_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0180']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0090--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0100
                        if float(mins_15_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(mins_15_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and\
                                  float(mins_15_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '15-mins post dose, Diastolic Blood Pressure', mins_15_post_dose_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"15-mins post dose, Diastolic Blood Pressure Interpretation: {mins_15_post_dose_Diastolic_Blood_Pressure_disname} - 15-mins post dose, Diastolic Blood Pressure Result: {mins_15_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0100']
                                lista_revision.append(error)

                        # Revision VS0190
                        elif float(mins_15_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(mins_15_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '15-mins post dose, Diastolic Blood Pressure', mins_15_post_dose_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"15-mins post dose, Diastolic Blood Pressure Interpretation: {mins_15_post_dose_Diastolic_Blood_Pressure_disname} - 15-mins post dose, Diastolic Blood Pressure Result: {mins_15_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0190']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0100--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0110
                        if float(mins_30_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(mins_30_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '30-mins post dose, Diastolic Blood Pressure', mins_30_post_dose_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"30-mins post dose, Diastolic Blood Pressure interpretation: {mins_30_post_dose_Diastolic_Blood_Pressure_disname} - 30-mins post dose, Diastolic Blood Pressure Result: {mins_30_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0110']
                                lista_revision.append(error)

                        # Revision VS0200
                        elif float(mins_30_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(mins_30_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '30-mins post dose, Diastolic Blood Pressure', mins_30_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"30-mins post dose, Diastolic Blood Pressure interpretation: {mins_30_post_dose_Diastolic_Blood_Pressure_disname} - 30-mins post dose, Diastolic Blood Pressure Result: {mins_30_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0200']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0110--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0120
                        if float(mins_60_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(mins_60_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '60-mins post dose, Diastolic Blood Pressure', mins_60_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"60-mins post dose, Diastolic Blood Pressure Interpretation: {mins_60_post_dose_Diastolic_Blood_Pressure_disname} - 60-mins post dose, Diastolic Blood Pressure Result: {mins_60_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0120']
                                lista_revision.append(error)

                        # Revision VS0210
                        elif float(mins_60_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(mins_60_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '60-mins post dose, Diastolic Blood Pressure', mins_60_post_dose_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"60-mins post dose, Diastolic Blood Pressure Interpretation: {mins_60_post_dose_Diastolic_Blood_Pressure_disname} - 60-mins post dose, Diastolic Blood Pressure Result: {mins_60_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0210']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0120--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0130
                        if float(hours_2_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_2_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Diastolic_Blood_Pressure_value_pure) <=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '2-hours post dose, Diastolic Blood Pressure', hours_2_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"2-hours post dose, Diastolic Blood Pressure Interpretation: {hours_2_post_dose_Diastolic_Blood_Pressure_disname} - 2-hours post dose, Diastolic Blood Pressure Result: {hours_2_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0130']
                                lista_revision.append(error)

                        # Revision VS0220
                        elif float(hours_2_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_2_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '2-hours post dose, Diastolic Blood Pressure', hours_2_post_dose_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"2-hours post dose, Diastolic Blood Pressure Interpretation: {hours_2_post_dose_Diastolic_Blood_Pressure_disname} - 2-hours post dose, Diastolic Blood Pressure Result: {hours_2_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0220']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0130--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0140
                        if float(hours_4_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_4_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '4-hours post dose, Diastolic Blood Pressure', hours_4_post_dose_Diastolic_Blood_Pressure_form_field_instance,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"4-hours post dose, Diastolic Blood Pressure Interpretation: {hours_4_post_dose_Diastolic_Blood_Pressure_disname} - 4-hours post dose, Diastolic Blood Pressure Result: {hours_4_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0140']
                                lista_revision.append(error)

                        # Revision VS0230
                        elif float(hours_4_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_4_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '4-hours post dose, Diastolic Blood Pressure', hours_4_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"4-hours post dose, Diastolic Blood Pressure Interpretation: {hours_4_post_dose_Diastolic_Blood_Pressure_disname} - 4-hours post dose, Diastolic Blood Pressure Result: {hours_4_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0230']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0140--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0150
                        if float(hours_8_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_8_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '8-hours post dose, Diastolic Blood Pressure', hours_8_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"8-hours post dose, Diastolic Blood Pressure Interpretation: {hours_8_post_dose_Diastolic_Blood_Pressure_disname} - 8-hours post dose, Diastolic Blood Pressure Result: {hours_8_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0150']
                                lista_revision.append(error)

                        # Revision VS0240
                        elif float(hours_8_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_8_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '8-hours post dose, Diastolic Blood Pressure', hours_8_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"8-hours post dose, Diastolic Blood Pressure Interpretation: {hours_8_post_dose_Diastolic_Blood_Pressure_disname} - 8-hours post dose, Diastolic Blood Pressure Result: {hours_8_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0240']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0150--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0160
                        if float(hours_12_post_dose_Diastolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_12_post_dose_Diastolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '12-hours post dose, Diastolic Blood Pressure', hours_12_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is not within expected range (50 to 90), therefore the Interpretation can not be Normal.', 
                                            f"12-hours post dose, Diastolic Blood Pressure Interpretation: {hours_12_post_dose_Diastolic_Blood_Pressure_disname} - 12-hours post dose, Diastolic Blood Pressure Result: {hours_12_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0160']
                                lista_revision.append(error)

                        # Revision VS0250
                        elif float(hours_12_post_dose_Diastolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_12_post_dose_Diastolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Diastolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Diastolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '12-hours post dose, Diastolic Blood Pressure', hours_12_post_dose_Diastolic_Blood_Pressure_form_field_instance ,
                                         'The Diastolic Blood Pressure is within expected range (50 to 90), the Interpretation should not be Abnormal.', 
                                            f"12-hours post dose, Diastolic Blood Pressure Interpretation: {hours_12_post_dose_Diastolic_Blood_Pressure_disname} - 12-hours post dose, Diastolic Blood Pressure Result: {hours_12_post_dose_Diastolic_Blood_Pressure_value_disname}", 'VS0250']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0160--> {e} - Subject: {subject},  Visit: {visit} ')
# ---------------------------------------------------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    try:
                        # Revision VS0260
                        if float(Undefined_Oral_Temperature_pure) == 1.0:
                            if float(Undefined_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(Undefined_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Oral Temperature', Undefined_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"Undefined, Oral Temperature Interpretation: {Undefined_Oral_Temperature_disname} - Undefined, Oral Temperature Result: {Undefined_Oral_Temperature_value_disname}", 'VS0260']
                                lista_revision.append(error)

                        # Revision VS0350
                        elif float(Undefined_Oral_Temperature_pure) == 2.0:
                            if float(Undefined_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(Undefined_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                error = [subject, visit, 'Undefined, Oral Temperature', Undefined_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"Undefined, Oral Temperature Interpretation: {Undefined_Oral_Temperature_disname} - Undefined, Oral Temperature Result: {Undefined_Oral_Temperature_value_disname}", 'VS0350']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0260--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0280
                        if float(mins_15_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(mins_15_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '15-mins post dose, Body Temperature', mins_15_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Body Temperature Interpretation: {mins_15_post_dose_Oral_Temperature_disname} - 15-mins post dose, Body Temperature Result: {mins_15_post_dose_Oral_Temperature_value_disname}", 'VS0280']
                                lista_revision.append(error)

                        # Revision VS0370
                        elif float(mins_15_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(mins_15_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '15-mins post dose, Body Temperature', mins_15_post_dose_Oral_Temperature_form_field_instance,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Body Temperature Interpretation: {mins_15_post_dose_Oral_Temperature_disname} - 15-mins post dose, Body Temperature Result: {mins_15_post_dose_Oral_Temperature_value_disname}", 'VS0370']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0280--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0270
                        if float(Pre_dose_Oral_Temperature_pure) == 1.0:
                            if float(Pre_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(Pre_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose, Body Temperature', Pre_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Body Temperature Interpretation: {Pre_dose_Oral_Temperature_disname} - Pre dose, Body Temperature Result: {Pre_dose_Oral_Temperature_value_disname}", 'VS0270']
                                lista_revision.append(error)

                        # Revision VS0360
                        elif float(Pre_dose_Oral_Temperature_pure) == 2.0:
                            if float(Pre_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(Pre_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                error = [subject, visit, 'Pre dose, Body Temperature', Pre_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Body Temperature Interpretation: {Pre_dose_Oral_Temperature_disname} - Pre dose, Body Temperature Result: {Pre_dose_Oral_Temperature_value_disname}", 'VS0360']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0270--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0290
                        if float(mins_30_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(mins_30_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '30-mins post dose, Body Temperature', mins_30_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"30-mins post dose, Body Temperature Interpretation: {mins_30_post_dose_Oral_Temperature_disname} - 30-mins post dose, Body Temperature Result: {mins_30_post_dose_Oral_Temperature_value_disname}", 'VS0290']
                                lista_revision.append(error)

                        # Revision VS0380
                        elif float(mins_30_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(mins_30_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '30-mins post dose, Body Temperature', mins_30_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"30-mins post dose, Body Temperature Interpretation: {mins_30_post_dose_Oral_Temperature_disname} - 30-mins post dose, Body Temperature Result: {mins_30_post_dose_Oral_Temperature_value_disname}", 'VS0380']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0290--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0300
                        if float(mins_60_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(mins_60_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '60-mins post dose, Oral Temperature', mins_60_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"60-mins post dose, Oral Temperature Interpretation: {mins_60_post_dose_Oral_Temperature_disname} - 60-mins post dose, Oral Temperature Result: {mins_60_post_dose_Oral_Temperature_value_disname}", 'VS0300']
                                lista_revision.append(error)

                        # Revision VS0390
                        elif float(mins_60_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(mins_60_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '60-mins post dose, Oral Temperature', mins_60_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"60-mins post dose, Oral Temperature Interpretation: {mins_60_post_dose_Oral_Temperature_disname} - 60-mins post dose, Oral Temperature Result: {mins_60_post_dose_Oral_Temperature_value_disname}", 'VS0390']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0300--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0310
                        if float(hours_2_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(hours_2_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '2-hours post dose, Oral Temperature', hours_2_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Oral Temperature Interpretation: {hours_2_post_dose_Oral_Temperature_disname} - 2-hours post dose, Oral Temperature Result: {hours_2_post_dose_Oral_Temperature_value_disname}", 'VS0310']
                                lista_revision.append(error)

                        # Revision VS0400
                        elif float(hours_2_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(hours_2_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '2-hours post dose, Oral Temperature', hours_2_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Oral Temperature Interpretation: {hours_2_post_dose_Oral_Temperature_disname} - 2-hours post dose, Oral Temperature Result: {hours_2_post_dose_Oral_Temperature_value_disname}", 'VS0400']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0310--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0330
                        if float(hours_8_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(hours_8_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '8-hours post dose, Oral Temperature', hours_8_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"8-hours post dose, Oral Temperature Interpretation: {hours_8_post_dose_Oral_Temperature_disname} - 8-hours post dose, Oral Temperature Result: {hours_8_post_dose_Oral_Temperature_value_disname}", 'VS0330']
                                lista_revision.append(error)

                        # Revision VS0420
                        elif float(hours_8_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(hours_8_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '8-hours post dose, Oral Temperature', hours_8_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"8-hours post dose, Oral Temperature Interpretation: {hours_8_post_dose_Oral_Temperature_disname} - 8-hours post dose, Oral Temperature Result: {hours_8_post_dose_Oral_Temperature_value_disname}", 'VS0420']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0330--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0320
                        if float(hours_4_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(hours_4_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '4-hours post dose, Oral Temperature', hours_4_post_dose_Oral_Temperature_form_field_instance,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"4-hours post dose, Oral Temperature interpretation: {hours_4_post_dose_Oral_Temperature_disname} - 4-hours post dose, Oral Temperature Result: {hours_4_post_dose_Oral_Temperature_value_disname}", 'VS0320']
                                lista_revision.append(error)

                        # Revision VS0410
                        elif float(hours_4_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(hours_4_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '4-hours post dose, Oral Temperature', hours_4_post_dose_Oral_Temperature_form_field_instance,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"4-hours post dose, Oral Temperature interpretation: {hours_4_post_dose_Oral_Temperature_disname} - 4-hours post dose, Oral Temperature Result: {hours_4_post_dose_Oral_Temperature_value_disname}", 'VS0410']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0320--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0340
                        if float(hours_12_post_dose_Oral_Temperature_pure) == 1.0:
                            if float(hours_12_post_dose_Oral_Temperature_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '12-hours post dose, Oral Temperature', hours_12_post_dose_Oral_Temperature_form_field_instance,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"12-hours post dose, Oral Temperature Interpretation: {hours_12_post_dose_Oral_Temperature_disname} - 12-hours post dose, Oral Temperature Result: {hours_12_post_dose_Oral_Temperature_value_disname}", 'VS0340']
                                lista_revision.append(error)

                        # Revision VS0430
                        elif float(hours_12_post_dose_Oral_Temperature_pure) == 2.0:
                            if float(hours_12_post_dose_Oral_Temperature_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Oral_Temperature_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Oral Temperature")]['max'].iloc[0])  :
                                error = [subject, visit, '12-hours post dose, Oral Temperature', hours_12_post_dose_Oral_Temperature_form_field_instance ,
                                         'The Oral Temperature is not within expected range (35 to 37.5) , therefore the Interpretation can not be Normal.', 
                                         f"12-hours post dose, Oral Temperature Interpretation: {hours_12_post_dose_Oral_Temperature_disname} - 12-hours post dose, Oral Temperature Result: {hours_12_post_dose_Oral_Temperature_value_disname}", 'VS0430']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0340--> {e} - Subject: {subject},  Visit: {visit} ')

# -----------------------------------
                    try:
                        # Revision VS0440
                        if float(Undefined_Pulse_rate_pure) == 1.0:
                            if float(Undefined_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(Undefined_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Pulse rate', Undefined_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected0 range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"Undefined, Pulse rate Interpretation: {Undefined_Pulse_rate_disname} - Undefined, Pulse rate Result: {Undefined_Pulse_rate_value_disname}", 'VS0440']
                                lista_revision.append(error)

                        # Revision VS0530
                        elif float(Undefined_Pulse_rate_pure) == 2.0:
                            if float(Undefined_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(Undefined_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, 'Undefined, Pulse rate', Undefined_Pulse_rate_form_field_instance,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"Undefined, Pulse rate Interpretation: {Undefined_Pulse_rate_disname} - Undefined, Pulse rate Result: {Undefined_Pulse_rate_value_disname}", 'VS0530']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0440--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0450
                        if float(Pre_dose_Pulse_rate_pure) == 1.0:
                            if float(Pre_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(Pre_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose, Pulse rate', Pre_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Pulse rate interpretation: {Pre_dose_Pulse_rate_disname} - Pre dose, Pulse rate Result: {Pre_dose_Pulse_rate_value_disname}", 'VS0450']
                                lista_revision.append(error)

                        # Revision VS0540
                        elif float(Pre_dose_Pulse_rate_pure) == 2.0:
                            if float(Pre_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(Pre_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, 'Pre dose, Pulse rate', Pre_dose_Pulse_rate_form_field_instance,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                          f"Pre dose, Pulse rate interpretation: {Pre_dose_Pulse_rate_disname} - Pre dose, Pulse rate Result: {Pre_dose_Pulse_rate_value_disname}", 'VS0540']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0450--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0460
                        if float(mins_15_post_dose_Pulse_rate_pure) == 1.0:
                            if float(mins_15_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '15-mins post dose, Pulse rate', mins_15_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Pulse rate Interpretation: {mins_15_post_dose_Pulse_rate_disname} - 15-mins post dose, Pulse rate Result: {mins_15_post_dose_Pulse_rate_value_disname}", 'VS0460']
                                lista_revision.append(error)

                        # Revision VS0550
                        elif float(mins_15_post_dose_Pulse_rate_pure) == 2.0:
                            if float(mins_15_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '15-mins post dose, Pulse rate', mins_15_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Pulse rate Interpretation: {mins_15_post_dose_Pulse_rate_disname} - 15-mins post dose, Pulse rate Result: {mins_15_post_dose_Pulse_rate_value_disname}", 'VS0550']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0460--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0470
                        if float(mins_30_post_dose_Pulse_rate_pure) == 1.0:
                            if float(mins_30_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '15-mins post dose, Pulse rate', mins_30_post_dose_Pulse_rate_form_field_instance,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Pulse rate Interpretation: {mins_30_post_dose_Pulse_rate_disname} - 15-mins post dose, Pulse rate Result: {mins_30_post_dose_Pulse_rate_value_disname}", 'VS0470']
                                lista_revision.append(error)

                        # Revision VS0560
                        elif float(mins_30_post_dose_Pulse_rate_pure) == 2.0:
                            if float(mins_30_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '15-mins post dose, Pulse rate', mins_30_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Pulse rate Interpretation: {mins_30_post_dose_Pulse_rate_disname} - 15-mins post dose, Pulse rate Result: {mins_30_post_dose_Pulse_rate_value_disname}", 'VS0560']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0470--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0480
                        if float(mins_60_post_dose_Pulse_rate_pure) == 1.0:
                            if float(mins_60_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '60-mins post dose, Pulse rate', mins_60_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"60-mins post dose, Pulse rate Interpretation: {mins_60_post_dose_Pulse_rate_disname} - 60-mins post dose, Pulse rate Result: {mins_60_post_dose_Pulse_rate_value_disname}", 'VS0480']
                                lista_revision.append(error)

                        # Revision VS0570
                        elif float(mins_60_post_dose_Pulse_rate_pure) == 2.0:
                            if float(mins_60_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '60-mins post dose, Pulse rate', mins_60_post_dose_Pulse_rate_form_field_instance,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"60-mins post dose, Pulse rate Interpretation: {mins_60_post_dose_Pulse_rate_disname} - 60-mins post dose, Pulse rate Result: {mins_60_post_dose_Pulse_rate_value_disname}", 'VS0570']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0480--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0490
                        if float(hours_2_post_dose_Pulse_rate_pure) == 1.0:
                            if float(hours_2_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '2-hours post dose, Pulse rate', hours_2_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Pulse rate Interpretation: {hours_2_post_dose_Pulse_rate_disname} - 2-hours post dose, Pulse rate Result: {hours_2_post_dose_Pulse_rate_value_disname}", 'VS0490']
                                lista_revision.append(error)

                        # Revision VS0580
                        elif float(hours_2_post_dose_Pulse_rate_pure) == 2.0:
                            if float(hours_2_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '2-hours post dose, Pulse rate', hours_2_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Pulse rate Interpretation: {hours_2_post_dose_Pulse_rate_disname} - 2-hours post dose, Pulse rate Result: {hours_2_post_dose_Pulse_rate_value_disname}", 'VS0580']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0490--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0490
                        if float(hours_2_post_dose_Pulse_rate_pure) == 1.0:
                            if float(hours_2_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '2-hours post dose, Pulse rate', hours_2_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Pulse rate Interpretation: {hours_2_post_dose_Pulse_rate_disname} - 2-hours post dose, Pulse rate Result: {hours_2_post_dose_Pulse_rate_value_disname}", 'VS0490']
                                lista_revision.append(error)

                        # Revision VS0580
                        elif float(hours_2_post_dose_Pulse_rate_pure) == 2.0:
                            if float(hours_2_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '2-hours post dose, Pulse rate', hours_2_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Pulse rate Interpretation: {hours_2_post_dose_Pulse_rate_disname} - 2-hours post dose, Pulse rate Result: {hours_2_post_dose_Pulse_rate_value_disname}", 'VS0580']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0490--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0500
                        if float(hours_4_post_dose_Pulse_rate_pure) == 1.0:
                            if float(hours_4_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '4-hours post dose, Pulse rate', hours_4_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"4-hours post dose, Pulse rate Interpretation: {hours_4_post_dose_Pulse_rate_disname} - 4-hours post dose, Pulse rate Result: {hours_4_post_dose_Pulse_rate_value_disname}", 'VS0500']
                                lista_revision.append(error)

                        # Revision VS0590
                        elif float(hours_4_post_dose_Pulse_rate_pure) == 2.0:
                            if float(hours_4_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '4-hours post dose, Pulse rate', hours_4_post_dose_Pulse_rate_form_field_instance ,
                                         '	The Pulse rate is within expected range (45 to 90), the Interpretation should not be Abnormal.', 
                                         f"4-hours post dose, Pulse rate Interpretation: {hours_4_post_dose_Pulse_rate_disname} - 4-hours post dose, Pulse rate Result: {hours_4_post_dose_Pulse_rate_value_disname}", 'VS0590']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0500--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0510
                        if float(hours_8_post_dose_Pulse_rate_pure) == 1.0:
                            if float(hours_8_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '8-hours post dose, Pulse rate', hours_8_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"8-hours post dose, Pulse rate Interpretation: {hours_8_post_dose_Pulse_rate_disname} - 8-hours post dose, Pulse rate Result: {hours_8_post_dose_Pulse_rate_value_disname}", 'VS0510']
                                lista_revision.append(error)

                        # Revision VS0600
                        elif float(hours_8_post_dose_Pulse_rate_pure) == 2.0:
                            if float(hours_8_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '8-hours post dose, Pulse rate', hours_8_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"8-hours post dose, Pulse rate Interpretation: {hours_8_post_dose_Pulse_rate_disname} - 8-hours post dose, Pulse rate Result: {hours_8_post_dose_Pulse_rate_value_disname}", 'VS0600']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0510--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0520
                        if float(hours_12_post_dose_Pulse_rate_pure) == 1.0:
                            if float(hours_12_post_dose_Pulse_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '12-hours post dose, Pulse rate', hours_12_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"12-hours post dose, Pulse rate Interpretation: {hours_12_post_dose_Pulse_rate_disname} - 12-hours post dose, Pulse rate Result: {hours_12_post_dose_Pulse_rate_value_disname}", 'VS0520']
                                lista_revision.append(error)

                        # Revision VS0610
                        elif float(hours_12_post_dose_Pulse_rate_pure) == 2.0:
                            if float(hours_12_post_dose_Pulse_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Pulse_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Pulse rate")]['max'].iloc[0])  :
                                error = [subject, visit, '12-hours post dose, Pulse rate', hours_12_post_dose_Pulse_rate_form_field_instance ,
                                         'The Pulse rate is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', 
                                         f"12-hours post dose, Pulse rate Interpretation: {hours_12_post_dose_Pulse_rate_disname} - 12-hours post dose, Pulse rate Result: {hours_12_post_dose_Pulse_rate_value_disname}", 'VS0610']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0520--> {e} - Subject: {subject},  Visit: {visit} ')
# -----------------------------------

                    try:
                        # Revision VS0620
                        if float(Undefined_Respiratory_rate_pure) == 1.0:
                            if float(Undefined_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(Undefined_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Respiratory rate', Undefined_Respiratory_rate_form_field_isntance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                            f"Undefined, Respiratory rate Interpretation: {Undefined_Respiratory_rate_disname} - Undefined, Respiratory rate Result: {Undefined_Respiratory_rate_value_disname}", 'VS0620']
                                lista_revision.append(error)

                        # Revision VS0710
                        elif float(Undefined_Respiratory_rate_pure) == 2.0:
                            if float(Undefined_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(Undefined_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, 'Undefined, Respiratory rate', Undefined_Respiratory_rate_form_field_isntance ,
                                         'The Respiratory rate is within expected range (12 to 18), the Interpretation should not be Abnormal.', 
                                         f"Undefined, Respiratory rate Interpretation: {Undefined_Respiratory_rate_disname} - Undefined, Respiratory rate Result: {Undefined_Respiratory_rate_value_disname}", 'VS0710']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0620--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0630
                        if float(Pre_dose_Respiratory_rate_pure) == 1.0:
                            if float(Pre_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(Pre_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose, Respiratory rate', Pre_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Respiratory rate Interpretation: {Pre_dose_Respiratory_rate_pure} - Pre dose, Respiratory rate Result: {Pre_dose_Respiratory_rate_value_disname}", 'VS0630']
                                lista_revision.append(error)

                        # Revision VS0720
                        elif float(Pre_dose_Respiratory_rate_pure) == 2.0:
                            if float(Pre_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(Pre_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, 'Pre dose, Respiratory rate', Pre_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Respiratory rate Interpretation: {Pre_dose_Respiratory_rate_pure} - Pre dose, Respiratory rate Result: {Pre_dose_Respiratory_rate_value_disname}", 'VS0720']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0630--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0640
                        if float(mins_15_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(mins_15_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '15-mins post dose, Respiratory rate', mins_15_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Respiratory rate Interpretation: {mins_15_post_dose_Respiratory_rate_disname} - 15-mins post dose, Respiratory rate Result: {mins_15_post_dose_Respiratory_rate_value_disname}", 'VS0640']
                                lista_revision.append(error)

                        # Revision VS0730
                        elif float(mins_15_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(mins_15_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '15-mins post dose, Respiratory rate', mins_15_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"15-mins post dose, Respiratory rate Interpretation: {mins_15_post_dose_Respiratory_rate_disname} - 15-mins post dose, Respiratory rate Result: {mins_15_post_dose_Respiratory_rate_value_disname}", 'VS0730']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0640--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0650
                        if float(mins_30_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(mins_30_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '30-mins post dose, Respiratory rate', mins_30_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"30-mins post dose, Respiratory rate Interpretation: {mins_30_post_dose_Respiratory_rate_disname} - 30-mins post dose, Respiratory rate Result: {mins_30_post_dose_Respiratory_rate_value_disname}", 'VS0650']
                                lista_revision.append(error)

                        # Revision VS0740
                        elif float(mins_30_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(mins_30_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '30-mins post dose, Respiratory rate', mins_30_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"30-mins post dose, Respiratory rate Interpretation: {mins_30_post_dose_Respiratory_rate_disname} - 30-mins post dose, Respiratory rate Result: {mins_30_post_dose_Respiratory_rate_value_disname}", 'VS0740']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0650--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0660
                        if float(mins_60_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(mins_60_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '60-mins post dose, Respiratory rate', mins_60_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"60-mins post dose, Respiratory rate Interpretation: {mins_60_post_dose_Respiratory_rate_disname} - 60-mins post dose, Respiratory rate Result: {mins_60_post_dose_Respiratory_rate_value_disname}", 'VS0660']
                                lista_revision.append(error)

                        # Revision VS0750
                        elif float(mins_60_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(mins_60_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '60-mins post dose, Respiratory rate', mins_60_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"60-mins post dose, Respiratory rate Interpretation: {mins_60_post_dose_Respiratory_rate_disname} - 60-mins post dose, Respiratory rate Result: {mins_60_post_dose_Respiratory_rate_value_disname}", 'VS0750']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0660--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0670
                        if float(hours_2_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(hours_2_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '2-hours post dose, Respiratory rate', hours_2_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Respiratory rate Interpretation: {hours_2_post_dose_Respiratory_rate_disname} - 2-hours post dose, Respiratory rate Result: {hours_2_post_dose_Respiratory_rate_value_disname}", 'VS0670']
                                lista_revision.append(error)

                        # Revision VS0760
                        elif float(hours_2_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(hours_2_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '2-hours post dose, Respiratory rate', hours_2_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"2-hours post dose, Respiratory rate Interpretation: {hours_2_post_dose_Respiratory_rate_disname} - 2-hours post dose, Respiratory rate Result: {hours_2_post_dose_Respiratory_rate_value_disname}", 'VS0760']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0670--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0680
                        if float(hours_4_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(hours_4_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '4-hours post dose, Respiratory rate', hours_4_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"4-hours post dose, Respiratory rate Interpretation: {hours_4_post_dose_Respiratory_rate_disname} - 4-hours post dose, Respiratory rate Result: {hours_4_post_dose_Respiratory_rate_value_disname}", 'VS0680']
                                lista_revision.append(error)

                        # Revision VS0770
                        elif float(hours_4_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(hours_4_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '4-hours post dose, Respiratory rate', hours_4_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"4-hours post dose, Respiratory rate Interpretation: {hours_4_post_dose_Respiratory_rate_disname} - 4-hours post dose, Respiratory rate Result: {hours_4_post_dose_Respiratory_rate_value_disname}", 'VS0770']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0680--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0690
                        if float(hours_8_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(hours_8_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '8-hours post dose, Respiratory rate', hours_8_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"8-hours post dose, Respiratory rate Interpretation: {hours_8_post_dose_Respiratory_rate_disname} - 8-hours post dose, Respiratory rate Result: {hours_8_post_dose_Respiratory_rate_value_disname}", 'VS0690']
                                lista_revision.append(error)

                        # Revision VS0780
                        elif float(hours_8_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(hours_8_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '8-hours post dose, Respiratory rate', hours_8_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"8-hours post dose, Respiratory rate Interpretation: {hours_8_post_dose_Respiratory_rate_disname} - 8-hours post dose, Respiratory rate Result: {hours_8_post_dose_Respiratory_rate_value_disname}", 'VS0780']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0690--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0700
                        if float(hours_12_post_dose_Respiratory_rate_pure) == 1.0:
                            if float(hours_12_post_dose_Respiratory_rate_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '12-hours post dose, Respiratory rate', hours_12_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"12-hours post dose, Respiratory rate Interpretation: {hours_12_post_dose_Respiratory_rate_disname} - 12-hours post dose, Respiratory rate Result: {hours_12_post_dose_Respiratory_rate_value_disname}", 'VS0700']
                                lista_revision.append(error)

                        # Revision VS0790
                        elif float(hours_12_post_dose_Respiratory_rate_pure) == 2.0:
                            if float(hours_12_post_dose_Respiratory_rate_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Respiratory_rate_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Respiratory rate")]['max'].iloc[0])  :
                                error = [subject, visit, '12-hours post dose, Respiratory rate', hours_12_post_dose_Respiratory_rate_form_field_instance ,
                                         'The Respiratory rate is not within expected range (12 to 18), therefore the Interpretation can not be Normal.', 
                                         f"12-hours post dose, Respiratory rate Interpretation: {hours_12_post_dose_Respiratory_rate_disname} - 12-hours post dose, Respiratory rate Result: {hours_12_post_dose_Respiratory_rate_value_disname}", 'VS0790']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0700--> {e} - Subject: {subject},  Visit: {visit} ')
#---------------------------------------------------------------------------

                    try:
                        # Revision VS0800
                        if float(Undefined_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(Undefined_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(Undefined_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, 'Undefined, Systolic Blood Pressure', Undefined_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.' , 
                                         f"Undefined, Systolic Blood Pressure Interpretation: {Undefined_Systolic_Blood_Pressure_disname} - Undefined, Systolic Blood Pressure Result: {Undefined_Systolic_Blood_Pressure_value_disname}", 'VS0800']
                                lista_revision.append(error)

                        # Revision VS0890
                        elif float(Undefined_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(Undefined_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(Undefined_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, 'Undefined, Systolic Blood Pressure', Undefined_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                         f"Undefined, Systolic Blood Pressure Interpretation: {Undefined_Systolic_Blood_Pressure_disname} - Undefined, Systolic Blood Pressure Result: {Undefined_Systolic_Blood_Pressure_value_disname}", 'VS0890']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0800--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # # Revision VS0800
                        # if float(Pre_dose_Systolic_Blood_Pressure) == 1.0:
                        #     if float(Pre_dose_Systolic_Blood_Pressure_value) >= 12.0 and float(Pre_dose_Systolic_Blood_Pressure_value) <= 18.0 :
                        #         error = [subject, visit, 'Pre dose, Systolic Blood Pressure' ,'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.' , Pre_dose_Systolic_Blood_Pressure_value, 'VS0800']
                        #         lista_revision.append(error)

                        # Revision VS0890
                        if float(Pre_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(Pre_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(Pre_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, 'Pre dose, Systolic Blood Pressure', Pre_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                         f"Pre dose, Systolic Blood Pressure Interpretation: {Pre_dose_Systolic_Blood_Pressure_disname} - Pre dose, Systolic Blood Pressure Result: {Pre_dose_Systolic_Blood_Pressure_value_disname}", 'VS0890']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0890--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0820
                        if float(mins_15_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(mins_15_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '15-mins post dose, Systolic Blood Pressure', mins_15_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.',
                                              f"15-mins post dose, Systolic Blood Pressure Interpretation: {mins_15_post_dose_Systolic_Blood_Pressure_disname} - 15-mins post dose, Systolic Blood Pressure Result: {mins_15_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0820']
                                lista_revision.append(error)

                        # Revision VS0910
                        elif float(mins_15_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(mins_15_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_15_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '15-mins post dose, Systolic Blood Pressure', mins_15_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                             f"15-mins post dose, Systolic Blood Pressure Interpretation: {mins_15_post_dose_Systolic_Blood_Pressure_disname} - 15-mins post dose, Systolic Blood Pressure Result: {mins_15_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0910']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0820--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0830
                        if float(mins_30_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(mins_30_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '30-mins post dose, Systolic Blood Pressure', mins_30_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"30-mins post dose, Systolic Blood Pressure Interpretation: {mins_30_post_dose_Systolic_Blood_Pressure_disname} - 30-mins post dose, Systolic Blood Pressure Result: {mins_30_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0830']
                                lista_revision.append(error)

                        # Revision VS0920
                        elif float(mins_30_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(mins_30_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_30_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '30-mins post dose, Systolic Blood Pressure', mins_30_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"30-mins post dose, Systolic Blood Pressure Interpretation: {mins_30_post_dose_Systolic_Blood_Pressure_disname} - 30-mins post dose, Systolic Blood Pressure Result: {mins_30_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0920']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0830--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0840
                        if float(mins_60_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(mins_60_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '60-mins post dose, Systolic Blood Pressure', mins_60_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"60-mins post dose, Systolic Blood Pressure Interpretation: {mins_60_post_dose_Systolic_Blood_Pressure_disname} - 60-mins post dose, Systolic Blood Pressure Result: {mins_60_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0840']
                                lista_revision.append(error)

                        # Revision VS0930
                        elif float(mins_60_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(mins_60_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(mins_60_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '60-mins post dose, Systolic Blood Pressure', mins_60_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is within expected range (100 to 140), the Interpretation should not be Abnormal.', 
                                            f"60-mins post dose, Systolic Blood Pressure Interpretation: {mins_60_post_dose_Systolic_Blood_Pressure_disname} - 60-mins post dose, Systolic Blood Pressure Result: {mins_60_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0930']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0840--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0850
                        if float(hours_2_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_2_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '2-hours post dose, Systolic Blood Pressure', hours_2_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"2-hours post dose, Systolic Blood Pressure Interpretation: {hours_2_post_dose_Systolic_Blood_Pressure_disname} - 2-hours post dose, Systolic Blood Pressure Result: {hours_2_post_dose_Systolic_Blood_Pressure_value_empty}", 'VS0850']
                                lista_revision.append(error)

                        # Revision VS0940
                        elif float(hours_2_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_2_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_2_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '2-hours post dose, Systolic Blood Pressure', hours_2_post_dose_Systolic_Blood_Pressure_form_field_instance,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                             f"2-hours post dose, Systolic Blood Pressure Interpretation: {hours_2_post_dose_Systolic_Blood_Pressure_disname} - 2-hours post dose, Systolic Blood Pressure Result: {hours_2_post_dose_Systolic_Blood_Pressure_value_empty}", 'VS0940']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0850--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0860
                        if float(hours_4_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_4_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '4-hours post dose, Systolic Blood Pressure', hours_4_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"4-hours post dose, Systolic Blood Pressure Interpretation: {hours_4_post_dose_Systolic_Blood_Pressure_disname} - 4-hours post dose, Systolic Blood Pressure Result: {hours_4_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0860']
                                lista_revision.append(error)

                        # Revision VS0950
                        elif float(hours_4_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_4_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_4_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '4-hours post dose, Systolic Blood Pressure', hours_4_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"4-hours post dose, Systolic Blood Pressure Interpretation: {hours_4_post_dose_Systolic_Blood_Pressure_disname} - 4-hours post dose, Systolic Blood Pressure Result: {hours_4_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0950']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0860--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0870
                        if float(hours_8_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_8_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '8-hours post dose, Systolic Blood Pressure', hours_8_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"8-hours post dose, Systolic Blood Pressure Interpretation: {hours_8_post_dose_Systolic_Blood_Pressure_disname} - 8-hours post dose, Systolic Blood Pressure Result: {hours_8_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0870']
                                lista_revision.append(error)

                        # Revision VS0960
                        elif float(hours_8_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_8_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_8_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '8-hours post dose, Systolic Blood Pressure', hours_8_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"8-hours post dose, Systolic Blood Pressure Interpretation: {hours_8_post_dose_Systolic_Blood_Pressure_disname} - 8-hours post dose, Systolic Blood Pressure Result: {hours_8_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0960']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0870--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision VS0880
                        if float(hours_12_post_dose_Systolic_Blood_Pressure_pure) == 1.0:
                            if float(hours_12_post_dose_Systolic_Blood_Pressure_value_pure) >= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0]) :
                                pass
                            else:
                                error = [subject, visit, '12-hours post dose, Systolic Blood Pressure', hours_12_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"12-hours post dose, Systolic Blood Pressure Interpretation: {hours_12_post_dose_Systolic_Blood_Pressure_disname} 12-hours post dose, Systolic Blood Pressure Result: {hours_12_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0880']
                                lista_revision.append(error)

                        # Revision VS0970
                        elif float(hours_12_post_dose_Systolic_Blood_Pressure_pure) == 2.0:
                            if float(hours_12_post_dose_Systolic_Blood_Pressure_value_pure) >=  float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['min'].iloc[0]) and \
                                float(hours_12_post_dose_Systolic_Blood_Pressure_value_pure) <= float(df_normal_ranges[(df_normal_ranges['field']== "Systolic Blood Pressure")]['max'].iloc[0])  :
                                error = [subject, visit, '12-hours post dose, Systolic Blood Pressure', hours_12_post_dose_Systolic_Blood_Pressure_form_field_instance ,
                                         'The Systolic Blood Pressure is not within expected range (100 to 140), therefore the Interpretation can not be Normal.', 
                                            f"12-hours post dose, Systolic Blood Pressure Interpretation: {hours_12_post_dose_Systolic_Blood_Pressure_disname} 12-hours post dose, Systolic Blood Pressure Result: {hours_12_post_dose_Systolic_Blood_Pressure_value_disname}", 'VS0970']
                                lista_revision.append(error)                     
                    except Exception as e:
                        lista_logs.append(f'Revision VS0880--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision VS0980
                    try:
                        if  float(was_vital_signs_performed_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the vital signs assessment performed?', was_vital_signs_performed_form_field_instance ,
                                         'The "Not Required" option can only be selected if visit is D-1 and D-1 date=Screening visit date', was_vital_signs_performed_disname, 'VS0980']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision VS0980--> {e} - Subject: {subject},  Visit: {visit} ')

                    # -------------------------------------------------------------------------------- Revision para CPG ----------------------------------------------------
                    # Revision VS0990 
                    if str(time_dosing_cpg_administration) != 'nan' and str(predose_time_pure) != 'nan':
                            
                        try:
                            dif = float((datetime.strptime(time_dosing_cpg_administration, '%H:%M') - datetime.strptime(predose_time_pure, '%H:%M')).total_seconds() / 60)
                            if dif < 0.0 or dif > 90.0:
                                    
                                error = [subject, visit, 'Pre dose, Time', predose_time_form_field_definition,
                                             'The time selected should be less than 60 min before the study treatment administration', 
                                                f'Pre dose, Time: {predose_time_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS0990']
                                lista_revision.append(error)

                        except Exception as e:
                            lista_logs.append(f'Revision VS0990 --> {e} - Subject: {subject},  Visit: {visit} ')  
                    
    
                        # Revision VS1000
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_15_pure) != 'nan':
                            
                            try:
                                dif_15 = float((datetime.strptime(post_dose_15_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_15 > 23.0 or dif_15 < 7.0:
                                    
                                    error = [subject, visit, '15-mins post dose, Time', post_dose_15_form_field_instance,
                                             'The time selected should be less than 23min and greater than 7 min after the study treatment administration', 
                                                f'15-mins post dose, Time: {post_dose_15_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1000']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1000 --> {e} - Subject: {subject},  Visit: {visit} ')  
                            
    
                        # Revision VS1010
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_30_pure) != 'nan':
                            
                            try:
                                dif_30 = float((datetime.strptime(post_dose_30_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_30 > 38.0 or dif_30 < 22.0:
                                    
                                    error = [subject, visit, '30-mins post dose, Time', post_dose_30_form_field_instance,
                                             'The time selected should be less than 38 min and greater than 22 min after the study treatment administration', 
                                                f'30-mins post dose, Time: {post_dose_30_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1010']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1010 --> {e} - Subject: {subject},  Visit: {visit} ')  
    

                        # Revision VS1020
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_60_pure) != 'nan':
                            
                            try:
                                dif_60 = float((datetime.strptime(post_dose_60_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_60 > 68.0 or dif_60 < 52.0:
                                    
                                    error = [subject, visit, '60-mins post dose, Time', post_dose_60_form_field_instance,
                                             'The time selected should be less than 68 min and greater than 52 min after the study treatment administration', 
                                                f'60-mins post dose, Time: {post_dose_60_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1020']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1020 --> {e} - Subject: {subject},  Visit: {visit} ')  
    

                        # Revision VS1030
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_2H_pure) != 'nan':
                            
                            try:
                                dif_2H = float((datetime.strptime(post_dose_2H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_2H > 135.0 or dif_2H < 105.0:
                                    
                                    error = [subject, visit, '2-hours post dose, Time', post_dose_2H_form_field_instance,
                                             'The time selected should be less than 2h15 and greater than 1h45 after the study treatment administration', 
                                                f'2-hours post dose, Time: {post_dose_2H_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1030']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1030 --> {e} - Subject: {subject},  Visit: {visit} ')  
    

                        # Revision VS1040
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_4H_pure) != 'nan':
                            
                            try:
                                dif_4H = float((datetime.strptime(post_dose_4H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_4H > 255.0 or dif_4H < 225.0:
                                    
                                    error = [subject, visit, '4-hours post dose, Time', post_dose_4H_form_field_instance,
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', 
                                                f'4-hours post dose, Time: {post_dose_4H_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1040']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1040 --> {e} - Subject: {subject},  Visit: {visit} ')
    

                        # Revision VS1050
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_8H_pure) != 'nan':
                            
                            try:
                                dif_8H = float((datetime.strptime(post_dose_8H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_8H > 495.0 or dif_8H < 465.0:
                                    
                                    error = [subject, visit, '8-hours post dose, Time', post_dose_8H_form_field_instance,
                                             'The time selected should be less than 8h15 and greater than 7h45 after the study treatment administration', 
                                                f'8-hours post dose, Time: {post_dose_8H_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1050']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1050 --> {e} - Subject: {subject},  Visit: {visit} ')
    

                        # Revision VS1060
                        if str(time_dosing_cpg_administration) != 'nan'  and str(post_dose_12H_pure) != 'nan':
                            
                            try:
                                dif_12H = float((datetime.strptime(post_dose_12H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_12H > 735.0 or dif_12H < 705.0:
                                    
                                    error = [subject, visit, '12-hours post dose, Time', post_dose_12H_form_field_instance,
                                             'The time selected should be less than 12h15 and greater than 11h45 after the study treatment administration', 
                                                f'12-hours post dose, Time: {post_dose_8H_pure} - dose time administration: {time_dosing_cpg_administration}', 'VS1060']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1060 --> {e} - Subject: {subject},  Visit: {visit} ')
                        
                        # -------------------------------------------------------------------------------- Revision para Miltefosine ----------------------------------------------------
                    # Revision VS0990
                    if str(time_dosing_miltefosine_administration) != 'nan'  and str(predose_time_pure) != 'nan':
                            
                        try:
                            dif_M = float((datetime.strptime(time_dosing_miltefosine_administration, '%H:%M') - datetime.strptime(predose_time_pure, '%H:%M')).total_seconds() / 60)
                          
                            if dif_M < 0.0 or dif_M > 90.0:
                                
                                error = [subject, visit, 'Pre dose, Time', predose_time_form_field_definition,
                                             'The time selected should be less than 60 min before the study treatment administration', 
                                                f'Pre dose, Time: {predose_time_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS0990']
                                lista_revision.append(error)

                        except Exception as e:
                            lista_logs.append(f'Revision VS0990 --> {e} - Subject: {subject},  Visit: {visit} ')  
                    
    
                        # Revision VS1000
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_15_pure) != 'nan':
                            
                            try:
                                dif_15_M = float((datetime.strptime(post_dose_15_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                  
                                if dif_15_M > 23.0 or dif_15_M < 7.0:
                                    
                                    error = [subject, visit, '15-mins post dose, Time', post_dose_15_form_field_instance,
                                             'The time selected should be less than 23min and greater than 7 min after the study treatment administration', 
                                                f'15-mins post dose, Time: {post_dose_15_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1000']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1000 --> {e} - Subject: {subject},  Visit: {visit} ')  
                            
    
                        # Revision VS1010
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_30_pure) != 'nan':
                            
                            try:
                                dif_30_M = float((datetime.strptime(post_dose_30_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                                if dif_30_M > 38.0 or dif_30_M < 22.0:
                          
                                    error = [subject, visit, '30-mins post dose, Time', post_dose_30_form_field_instance,
                                             'The time selected should be less than 38 min and greater than 22 min after the study treatment administration', 
                                                f'30-mins post dose, Time: {post_dose_30_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1010']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1010 --> {e} - Subject: {subject},  Visit: {visit} ')  
    

                        # Revision VS1020
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_60_pure) != 'nan':
                            
                            try:
                                dif_60_M = float((datetime.strptime(post_dose_60_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                 
                                if dif_60_M > 68.0 or dif_60_M < 52.0:
                                    
                                    error = [subject, visit, '60-mins post dose, Time', post_dose_60_form_field_instance,
                                             'The time selected should be less than 68 min and greater than 52 min after the study treatment administration', 
                                                f'60-mins post dose, Time: {post_dose_60_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1020']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1020 --> {e} - Subject: {subject},  Visit: {visit} ')  
    

                        # Revision VS1030
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_2H_pure) != 'nan':
                            
                            try:
                                dif_2H_M = float((datetime.strptime(post_dose_2H_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                    
                                if dif_2H_M > 135.0 or dif_2H_M < 105.0:
                                    
                                    error = [subject, visit, '2-hours post dose, Time', post_dose_2H_form_field_instance,
                                             'The time selected should be less than 2h15 and greater than 1h45 after the study treatment administration', 
                                                f'2-hours post dose, Time: {post_dose_2H_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1030']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1030 --> {e} - Subject: {subject},  Visit: {visit} ')  
    

                        # Revision VS1040
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_4H_pure) != 'nan':
                            
                            try:
                                dif_4H_M = float((datetime.strptime(post_dose_4H_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                         
                                if dif_4H_M > 255.0 or dif_4H_M < 225.0:
                                    
                                    error = [subject, visit, '4-hours post dose, Time', post_dose_4H_form_field_instance,
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', 
                                                f'4-hours post dose, Time: {post_dose_4H_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1040']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1040 --> {e} - Subject: {subject},  Visit: {visit} ')
    

                        # Revision VS1050
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_8H_pure) != 'nan':
                            
                            try:
                                dif_8H_M = float((datetime.strptime(post_dose_8H_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                        
                                if dif_8H_M > 495.0 or dif_8H_M < 465.0:
                                    
                                    error = [subject, visit, '8-hours post dose, Time', post_dose_8H_form_field_instance,
                                             'The time selected should be less than 8h15 and greater than 7h45 after the study treatment administration', 
                                                f'8-hours post dose, Time: {post_dose_8H_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1050']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1050 --> {e} - Subject: {subject},  Visit: {visit} ')
    

                        # Revision VS1060
                        if str(time_dosing_miltefosine_administration) != 'nan' and str(post_dose_12H_pure):
                            
                            try:
                                dif_12H_M = float((datetime.strptime(post_dose_12H_pure, '%H:%M') - datetime.strptime(time_dosing_miltefosine_administration, '%H:%M')).total_seconds() / 60)
                  
                                if dif_12H_M > 735.0 or dif_12H_M < 705.0:
                                    
                                    error = [subject, visit, '12-hours post dose, Time', post_dose_12H_form_field_instance,
                                             'The time selected should be less than 12h15 and greater than 11h45 after the study treatment administration', 
                                                f'12-hours post dose, Time: {post_dose_8H_pure} - dose time administration: {time_dosing_miltefosine_administration}', 'VS1060']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision VS1060 --> {e} - Subject: {subject},  Visit: {visit} ')


    excel_writer = load_workbook(path_excel_writer)
    column_names =  ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    vital_signs_output = pd.DataFrame(lista_revision, columns=column_names)
    vital_signs_output = vital_signs_output[~vital_signs_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
    
    sheet = excel_writer.create_sheet("Vital Signs")

    for row in dataframe_to_rows(vital_signs_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return vital_signs_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    vital_signs(df_root, path_excel)