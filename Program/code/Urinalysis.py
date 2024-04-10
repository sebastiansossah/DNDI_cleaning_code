from datetime import datetime
from log_writer import log_writer
import math
import numpy as np
from revision_fechas import revision_fecha
import warnings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')


def urinalysis(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Urinalysis
    '''

    df= df_root[df_root['name']== 'Urinalysis']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Urinalysis']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')

            for index, row in pru.iterrows():

                if index != 0:
                    lista_logs.append('Duplicados en la data, revisar subdataset')
                    
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                if status != '':
                    try:
                        date_collected = row['Date sample collected']
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1]
                    except Exception as e:
                        date_collected_pure = ''   
                        date_collected_form_field_instance = 'This field does not have any data'

                    try:
                        urine_sample_collected = row['Urine Sample collected']
                        urine_sample_collected_pure = urine_sample_collected.split('|')[0]
                        urine_sample_collected_form_field_instance = urine_sample_collected.split('|')[1]
                        urine_sample_collected_disname = urine_sample_collected.split('|')[2]
                    except Exception as e:
                        urine_sample_collected_pure = math.nan
                        urine_sample_collected_form_field_instance = 'This field does not have any data'
                        urine_sample_collected_disname = 'Empty' 

                    # --------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,f , date_collected_pure, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UR0010
                    try:
                        if  float(urine_sample_collected_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Urine Sample collected', urine_sample_collected_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and the D-1 visit date =Screening visit date or normal and done in the previous 10 days', \
                                            urine_sample_collected_disname, 'UR0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision UR0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UR0020
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_collected_f != date_of_visit_f:
                                error = [subject, visit, 'Date sample collected', date_collected_form_field_instance,\
                                        'The date should be the same as the visit date in the "Date of Visit" form', \
                                            f'{date_collected_pure} - {date_of_visit}', 'UR0020']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision UR0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision UR0030
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_collected_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance,\
                                        'The date/time of test performed can not be before the informed consent date/time',\
                                        f'{date_collected_pure} - {date_inform_consent}', 'UR0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision UR0030--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision -> UR0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,'Date Collected must be before the End of study/Early withdrawal date. ', date_collected_pure, 'UR0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision UR0040 --> {e}  - Subject: {subject},  Visit: {visit} ')

                    lista_validacion = [
                        'Bilirubin',
                        'Blood',
                        'Glucose',
                        'Ketones',
                        'Leukocytes',
                        'Nitrite',
                        'pH',
                        'Protein',
                        'Specific gravity',
                        'Urobilinogen',
                    ]
                    mi_cuenta= 0

                    for validador_raw in lista_validacion:
                        try: 
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
                        
                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta += 1
                        else:
                            pass
                        
                    # Revision UR0060
                    try:
                        if float(urine_sample_collected_pure) == 1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Urine Sample collected', urine_sample_collected_form_field_instance,\
                                         'If Urine Sample Collected is checked as "Yes", validate that at least one of the Laboratory Tests has been completed.', \
                                            urine_sample_collected_disname, 'UR0060']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision UR0060--> {e} - Subject: {subject},  Visit: {visit} ')

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    urinalysis_output = pd.DataFrame(lista_revision, columns=column_names)
    urinalysis_output = urinalysis_output[~urinalysis_output['Form Field Instance ID'].isin(lista_instancias_abiertas)]
    
    sheet = excel_writer.create_sheet("Urinalysis")

    for row in dataframe_to_rows(urinalysis_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return urinalysis_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI.xlsx')
    urinalysis(df_root, path_excel) 