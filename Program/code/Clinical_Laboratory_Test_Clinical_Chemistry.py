from datetime import datetime
import math
from log_writer import log_writer
import numpy as np
from revision_fechas import revision_fecha
import warnings
import pandas as pd
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def clinical_laboratory_test_clinical_chemistry(df_root, path_excel_writer, lista_instancias_abiertas):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Clinical Laboratory Test - Clinical Chemistry
    '''

    df= df_root[df_root['name']== 'Clinical Laboratory Test - Clinical Chemistry']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]   
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_demographic = df_root[df_root['name']=='Demographics']
    df_demographic = df_demographic[['Participante', 'Campo', 'Valor']]
    df_demographic = df_demographic[df_demographic['Campo']=='Gender']
    df_demographic = df_demographic[['Participante','Valor']]
    df_demographic = df_demographic.rename(columns={'Participante':'Subject', 'Valor':'Genero'})

    df_demographic_age = df_root[df_root['name']=='Demographics']
    df_demographic_age = df_demographic_age[['Participante', 'Campo', 'Valor']]
    df_demographic_age = df_demographic_age[df_demographic_age['Campo']=='Age at consent']
    df_demographic_age = df_demographic_age[['Participante','Valor']]
    df_demographic_age = df_demographic_age.rename(columns={'Participante':'Subject', 'Valor':'age_participant'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['Clinical Laboratory Test - Clinical Chemistry']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_demographic, on=['Subject'], how='left')
            pru = pru.merge(df_demographic_age, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            # print(pru)
            # print('-----------------------')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                genero = row['Genero']
                subject_age = row['age_participant']
                end_study_date = row['end_study_date']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                if status != '':
                    try:
                        Alanine_out_normal = row['Alanine Aminotransferase (ALT), Out of normal range?']
                        Alanine_out_normal_pure = Alanine_out_normal.split('|')[0]
                        Alanine_out_normal_form_field_instance = Alanine_out_normal.split('|')[1]
                        Alanine_out_normal_disname = Alanine_out_normal.split('|')[2]
                    except Exception as e:
                        Alanine_out_normal_pure = math.nan
                        Alanine_out_normal_form_field_instance = 'This field does not have any data'
                        Alanine_out_normal_disname = 'Empty'

                    try:
                        Alanine_result = row['Alanine Aminotransferase (ALT), Result (U/L)']
                        Alanine_result_pure = Alanine_result.split('|')[0]
                        Alanine_result_form_field_isntance = Alanine_result.split('|')[1]
                        Alanine_result_disname = Alanine_result.split('|')[0]
                    except Exception as e:
                        Alanine_result_pure = math.nan
                        Alanine_result_form_field_isntance = 'This field does not have any data'
                        Alanine_result_disname = 'Empty'

                    try:
                        Albumin_out_normal = row['Albumin, Out of normal range?']
                        Albumin_out_normal_pure = Albumin_out_normal.split('|')[0]
                        Albumin_out_normal_form_field_isntance = Albumin_out_normal.split('|')[1]
                        Albumin_out_normal_disname = Albumin_out_normal.split('|')[2]
                    except Exception as e:
                        Albumin_out_normal_pure = math.nan
                        Albumin_out_normal_form_field_isntance = 'This field does not have any data'
                        Albumin_out_normal_disname = 'Empty'

                    try:
                        Albumin_result = row['Albumin, Result (g/dL)']
                        Albumin_result_pure = Albumin_result.split('|')[0]
                        Albumin_result_form_field_instance = Albumin_result.split('|')[1]
                        Albumin_result_disname = Albumin_result.split('|')[0]
                    except Exception as e:
                        Albumin_result_pure = math.nan
                        Albumin_result_form_field_instance = 'This field does not have any data'
                        Albumin_result_disname = 'Empty'

                    try:
                        Alkaline_out_normal = row['Alkaline Phosphatase, Out of normal range?']
                        Alkaline_out_normal_pure = Alkaline_out_normal.split('|')[0]
                        Alkaline_out_normal_form_field_instance = Alkaline_out_normal.split('|')[1]
                        Alkaline_out_normal_disname = Alkaline_out_normal.split('|')[2] 
                    except Exception as e:
                        Alkaline_out_normal_pure = math.nan
                        Alkaline_out_normal_form_field_instance = 'This field does not have any data'
                        Alkaline_out_normal_disname = 'Empty'

                    try:
                        Alkaline_result = row['Alkaline Phosphatase, Result (U/L)']
                        Alkaline_result_pure = Alkaline_result.split('|')[0]
                        Alkaline_result_form_field_instance = Alkaline_result.split('|')[1]
                        Alkaline_result_disname = Alkaline_result.split('|')[0]
                    except Exception as e:
                        Alkaline_result_pure = math.nan
                        Alkaline_result_form_field_instance = 'This field does not have any data'
                        Alkaline_result_disname = 'Empty'

                    try:
                        Aspartate_out_normal =  row['Aspartate Aminotransferase (AST), Out of normal range?']
                        Aspartate_out_normal_pure = Aspartate_out_normal.split('|')[0]
                        Aspartate_out_normal_form_field_instance = Aspartate_out_normal.split('|')[1]
                        Aspartate_out_normal_disname = Aspartate_out_normal.split('|')[2]
                    except Exception as e:
                        Aspartate_out_normal_pure = math.nan
                        Aspartate_out_normal_form_field_instance = 'This field does not have any data'
                        Aspartate_out_normal_disname = 'Empty'

                    try:
                        Aspartate_result =  row['Aspartate Aminotransferase (AST), Result (U/L)']
                        Aspartate_result_pure = Aspartate_result.split('|')[0]
                        Aspartate_result_form_field_instance = Aspartate_result.split('|')[1]
                        Aspartate_result_disname = Aspartate_result.split('|')[0]
                    except Exception as e:
                        Aspartate_result_pure = math.nan
                        Aspartate_result_form_field_instance = 'This field does not have any data'
                        Aspartate_result_disname = 'Empty'

                    try:
                        Bicarbonate_out_normal =  row['Bicarbonate, Out of normal range?']
                        Bicarbonate_out_normal_pure = Bicarbonate_out_normal.split('|')[0]
                        Bicarbonate_out_normal_form_field_instance = Bicarbonate_out_normal.split('|')[1]
                        Bicarbonate_out_normal_disname = Bicarbonate_out_normal.split('|')[2]
                    except Exception as e:
                        Bicarbonate_out_normal_pure = math.nan
                        Bicarbonate_out_normal_form_field_instance = 'This field does not have any data'
                        Bicarbonate_out_normal_disname = 'Empty'

                    try:
                        Bicarbonate_result =  row['Bicarbonate, Result (mEq/L)']
                        Bicarbonate_result_pure = Bicarbonate_result.split('|')[0]
                        Bicarbonate_result_form_field_instance = Bicarbonate_result.split('|')[1]
                        Bicarbonate_result_disname = Bicarbonate_result.split('|')[0]
                    except Exception as e:
                        Bicarbonate_result_pure = math.nan
                        Bicarbonate_result_form_field_instance = 'This field does not have any data'
                        Bicarbonate_result_disname = 'Empty'

                    try:
                        Bilirubin_direct_out_normal = row['Bilirubin (Direct) (only if Total is elevated), Out of normal range?']
                        Bilirubin_direct_out_normal_pure = Bilirubin_direct_out_normal.split('|')[0]
                        Bilirubin_direct_out_normal_form_field_instance = Bilirubin_direct_out_normal.split('|')[1]
                        Bilirubin_direct_out_normal_disname = Bilirubin_direct_out_normal.split('|')[2]
                    except Exception as e:
                        Bilirubin_direct_out_normal_pure = math.nan
                        Bilirubin_direct_out_normal_form_field_instance = 'This field does not have any data'
                        Bilirubin_direct_out_normal_disname = 'Empty'

                    try:
                        Bilirubin_direct_result = row['Bilirubin (Direct) (only if Total is elevated), Result (mg/dL)']
                        Bilirubin_direct_result_pure = Bilirubin_direct_result.split('|')[0]
                        Bilirubin_direct_result_form_field_instance = Bilirubin_direct_result.split('|')[1]
                        Bilirubin_direct_result_disname = Bilirubin_direct_result.split('|')[0]                   
                    except Exception as e:
                        Bilirubin_direct_result_pure = math.nan
                        Bilirubin_direct_result_form_field_instance = 'This field does not have any data'
                        Bilirubin_direct_result_disname = 'Empty'

                    try:
                        Bilirubin_out_normal = row['Bilirubin (Total), Out of normal range?']
                        Bilirubin_out_normal_pure = Bilirubin_out_normal.split('|')[0]
                        Bilirubin_out_normal_form_field_instance = Bilirubin_out_normal.split('|')[1]
                        Bilirubin_out_normal_disname = Bilirubin_out_normal.split('|')[2]
                    except Exception as e:
                        Bilirubin_out_normal_pure = math.nan
                        Bilirubin_out_normal_form_field_instance = 'This field does not have any data'
                        Bilirubin_out_normal_disname = 'Empty'

                    try:
                        Bilirubin_result = row['Bilirubin (Total), Result (mg/dL)']
                        Bilirubin_result_pure = Bilirubin_result.split('|')[0]
                        Bilirubin_result_form_field_instance = Bilirubin_result.split('|')[1]
                        Bilirubin_result_disname = Bilirubin_result.split('|')[0]   
                    except Exception as e:
                        Bilirubin_result_pure = math.nan
                        Bilirubin_result_form_field_instance = 'This field does not have any data'
                        Bilirubin_result_disname = 'Empty'

                    try:
                        blood_sample_collected = row['Blood Sample Collected']
                        blood_sample_collected_pure = blood_sample_collected.split('|')[0]
                        blood_sample_collected_form_field_instance = blood_sample_collected.split('|')[1]
                        blood_sample_collected_disname = blood_sample_collected.split('|')[2]
                    except Exception as e:
                        blood_sample_collected_pure = math.nan
                        blood_sample_collected_form_field_instance = 'This field does not have any data'
                        blood_sample_collected_disname = 'Empty'

                    try:
                        C_Reactive_out_normal = row['C-Reactive Protein, Out of normal range?']
                        C_Reactive_out_normal_pure = C_Reactive_out_normal.split('|')[0]
                        C_Reactive_out_normal_form_field_instance = C_Reactive_out_normal.split('|')[1]
                        C_Reactive_out_normal_disname = C_Reactive_out_normal.split('|')[2]
                    except Exception as e:
                        C_Reactive_out_normal_pure = math.nan
                        C_Reactive_out_normal_form_field_instance = 'This field does not have any data'
                        C_Reactive_out_normal_disname = 'Empty'

                    try:
                        C_Reactive_result = row['C-Reactive Protein, Result (mg/L)']
                        C_Reactive_result_pure = C_Reactive_result.split('|')[0]
                        C_Reactive_result_form_field_instance = C_Reactive_result.split('|')[1]
                        C_Reactive_result_disname = C_Reactive_result.split('|')[0]
                    except Exception as e:
                        C_Reactive_result_pure = math.nan
                        C_Reactive_result_form_field_instance = 'This field does not have any data'
                        C_Reactive_result_disname = 'Empty'

                    try:
                        Calcium_out_normal = row['Calcium, Out of normal range?']
                        Calcium_out_normal_pure = Calcium_out_normal.split('|')[0]
                        Calcium_out_normal_form_field_instance = Calcium_out_normal.split('|')[1]
                        Calcium_out_normal_disname = Calcium_out_normal.split('|')[2]
                    except Exception as e:
                        Calcium_out_normal_pure = math.nan
                        Calcium_out_normal_form_field_instance = 'This field does not have any data'
                        Calcium_out_normal_disname = 'Empty'

                    try:
                        Calcium_result = row['Calcium, Result (mEq/L)']
                        Calcium_result_pure = Calcium_result.split('|')[0]
                        Calcium_result_form_field_instance = Calcium_result.split('|')[1]
                        Calcium_result_disname = Calcium_result.split('|')[0]
                    except Exception as e:
                        Calcium_result_pure = math.nan
                        Calcium_result_form_field_instance = 'This field does not have any data'
                        Calcium_result_disname = 'Empty'

                    try:
                        Chloride_out_normal = row['Chloride, Out of normal range?']
                        Chloride_out_normal_pure = Chloride_out_normal.split('|')[0]
                        Chloride_out_normal_form_field_instance = Chloride_out_normal.split('|')[1]
                        Chloride_out_normal_disname = Chloride_out_normal.split('|')[0]
                    except Exception as e:
                        Chloride_out_normal_pure = math.nan
                        Chloride_out_normal_form_field_instance = 'This field does not have any data'
                        Chloride_out_normal_disname = 'Empty'

                    try:
                        Chloride_result = row['Chloride, Result (mmol/L)']
                        Chloride_result_pure = Chloride_result.split('|')[0]
                        Chloride_result_form_field_instacnce = Chloride_result.split('|')[1]
                        Chloride_result_disname = Chloride_result.split('|')[0]
                    except Exception as e:
                        Chloride_result_pure = math.nan
                        Chloride_result_form_field_instacnce = 'This field does not have any data'
                        Chloride_result_disname = 'Empty'

                    try:
                        Cholesterol_out_normal = row['Cholesterol (Total), Out of normal range?']
                        Cholesterol_out_normal_pure = Cholesterol_out_normal.split('|')[0]
                        Cholesterol_out_normal_form_field_instance = Cholesterol_out_normal.split('|')[1]
                        Cholesterol_out_normal_disname = Cholesterol_out_normal.split('|')[2]
                    except Exception as e:
                        Cholesterol_out_normal_pure = math.nan
                        Cholesterol_out_normal_form_field_instance = 'This field does not have any data'
                        Cholesterol_out_normal_disname = 'Empty'

                    try:
                        Cholesterol_result = row['Cholesterol (Total), Result (mg/dL)']
                        Cholesterol_result_pure = Cholesterol_result.split('|')[0]
                        Cholesterol_result_form_field_instance = Cholesterol_result.split('|')[1]
                        Cholesterol_result_disname = Cholesterol_result.split('|')[0]
                    except Exception as e:
                        Cholesterol_result_pure = math.nan
                        Cholesterol_result_form_field_instance = 'This field does not have any data'
                        Cholesterol_result_disname = 'Empty'

                    try:
                        Creatine_out_normal = row['Creatine Kinase (CK), Out of normal range?']
                        Creatine_out_normal_pure = Creatine_out_normal.split('|')[0]
                        Creatine_out_normal_form_field_instance = Creatine_out_normal.split('|')[1]
                        Creatine_out_normal_disname = Creatine_out_normal.split('|')[2]
                    except Exception as e:
                        Creatine_out_normal_pure = math.nan
                        Creatine_out_normal_form_field_instance = 'This field does not have any data'
                        Creatine_out_normal_disname = 'Empty'

                    try:
                        Creatine_result = row['Creatine Kinase (CK), Result (U/L)']
                        Creatine_result_pure = Creatine_result.split('|')[0]
                        Creatine_result_form_field_instance = Creatine_result.split('|')[1]
                        Creatine_result_disname = Creatine_result.split('|')[0]
                    except Exception as e:
                        Creatine_result_pure = math.nan
                        Creatine_result_form_field_instance = 'This field does not have any data'
                        Creatine_result_disname = 'Empty'

                    try:
                        date_collected = row['Date Collected']
                        date_collected_pure = date_collected.split('|')[0]
                        date_collected_form_field_instance = date_collected.split('|')[1]
                        date_collected_disname = date_collected.split('|')[0]
                    except Exception as e:
                        date_collected_pure = ''
                        date_collected_form_field_instance = 'This field does not have any data'
                        date_collected_disname = 'Empty'

                    try:
                        GGT_out_normal = row['Gamma Glutamyl Transferase (GGT), Out of normal range?']
                        GGT_out_normal_pure = GGT_out_normal.split('|')[0]
                        GGT_out_normal_form_field_instance = GGT_out_normal.split('|')[1]
                        GGT_out_normal_disname = GGT_out_normal.split('|')[2]
                    except Exception as e:
                        GGT_out_normal_pure = math.nan
                        GGT_out_normal_form_field_instance = 'This field does not have any data'
                        GGT_out_normal_disname = 'Empty'

                    try:
                        GGT_result = row['Gamma Glutamyl Transferase (GGT), Result (U/L)']
                        GGT_result_pure = GGT_result.split('|')[0]
                        GGT_result_form_instance = GGT_result.split('|')[1]
                        GGT_result_disname = GGT_result.split('|')[0]
                    except Exception as e:
                        GGT_result_pure = math.nan
                        GGT_result_form_instance = 'This field does not have any data'
                        GGT_result_disname = 'Empty'

                    try:
                        Potassium_out_normal = row['Potassium, Out of normal range?']
                        Potassium_out_normal_pure = Potassium_out_normal.split('|')[0]
                        Potassium_out_normal_form_field_instance = Potassium_out_normal.split('|')[1]
                        Potassium_out_normal_disname = Potassium_out_normal.split('|')[2]
                    except Exception as e:
                        Potassium_out_normal_pure = math.nan
                        Potassium_out_normal_form_field_instance = 'This field does not have any data'
                        Potassium_out_normal_disname = 'Empty'

                    try:
                        Potassium_result = row['Potassium, Result (mmol/L)']
                        Potassium_result_pure = Potassium_result.split('|')[0]
                        Potassium_result_form_field_instance = Potassium_result.split('|')[1]
                        Potassium_result_disname = Potassium_result.split('|')[0]
                    except Exception as e:
                        Potassium_result_pure = math.nan
                        Potassium_result_form_field_instance = 'This field does not have any data'
                        Potassium_result_disname = 'Empty'

                    try:
                        Protein_out_normal = row['Protein (Total), Out of normal range?']
                        Protein_out_normal_pure = Protein_out_normal.split('|')[0]
                        Protein_out_normal_form_field_instance = Protein_out_normal.split('|')[1]
                        Protein_out_normal_disname = Protein_out_normal.split('|')[2]
                    except Exception as e:
                        Protein_out_normal_pure = math.nan
                        Protein_out_normal_form_field_instance = 'This field does not have any data'
                        Protein_out_normal_disname = 'Empty'

                    try:
                        Protein_total_result = row['Protein (Total), Result (mg/dL)']
                        Protein_total_result_pure = Protein_total_result.split('|')[0]
                        Protein_total_result_form_field_instance = Protein_total_result.split('|')[1]
                        Protein_total_result_disname = Protein_total_result.split('|')[0]
                    except Exception as e:
                        Protein_total_result_pure = math.nan
                        Protein_total_result_form_field_instance = 'This field does not have any data'
                        Protein_total_result_disname = 'Empty'

                    try:
                        Serum_out_normal = row['Serum Creatinine, Out of normal range?']
                        Serum_out_normal_pure = Serum_out_normal.split('|')[0]
                        Serum_out_normal_form_field_instance = Serum_out_normal.split('|')[1]
                        Serum_out_normal_disname = Serum_out_normal.split('|')[2]
                    except Exception as e:
                        Serum_out_normal_pure = math.nan
                        Serum_out_normal_form_field_instance = 'This field does not have any data'
                        Serum_out_normal_disname = 'Empty'

                    try:
                        Serum_result = row['Serum Creatinine, Result  (mg/dL)']
                        Serum_result_pure = Serum_result.split('|')[0]
                        Serum_result_form_field_instance = Serum_result.split('|')[1]
                        Serum_result_disname = Serum_result.split('|')[0]
                    except Exception as e:
                        Serum_result_pure = math.nan
                        Serum_result_form_field_instance = 'This field does not have any data'
                        Serum_result_disname = 'Empty'

                    try:
                        Sodium_out_normal = row['Sodium, Out of normal range?']
                        Sodium_out_normal_pure = Sodium_out_normal.split('|')[0]
                        Sodium_out_normal_form_field_instance = Sodium_out_normal.split('|')[1]
                        Sodium_out_normal_disname = Sodium_out_normal.split('|')[2]
                    except Exception as e:
                        Sodium_out_normal_pure = math.nan
                        Sodium_out_normal_form_field_instance = 'This field does not have any data'
                        Sodium_out_normal_disname = 'Empty'

                    try:
                        Sodium_result = row['Sodium, Result (mmol/L)']
                        Sodium_result_pure = Sodium_result.split('|')[0]
                        Sodium_result_form_field_instance = Sodium_result.split('|')[1]
                        Sodium_result_disname = Sodium_result.split('|')[0]
                    except Exception as e:
                        Sodium_result_pure = math.nan
                        Sodium_result_form_field_instance = 'This field does not have any data'
                        Sodium_result_disname = 'Empty'

                    try:
                        Triglycerides_out_normal = row['Triglycerides, Out of normal range?']
                        Triglycerides_out_normal_pure = Triglycerides_out_normal.split('|')[0]
                        Triglycerides_out_normal_form_field_instance = Triglycerides_out_normal.split('|')[1]
                        Triglycerides_out_normal_disname = Triglycerides_out_normal.split('|')[2]
                    except Exception as e:
                        Triglycerides_out_normal_pure = math.nan
                        Triglycerides_out_normal_form_field_instance = 'This field does not have any data'
                        Triglycerides_out_normal_disname = 'Empty'

                    try:
                        Triglycerides_result = row['Triglycerides, Result (mg/dL)']
                        Triglycerides_result_pure = Triglycerides_result.split('|')[0]
                        Triglycerides_result_form_field_instance = Triglycerides_result.split('|')[1]
                        Triglycerides_result_disname = Triglycerides_result.split('|')[0]
                    except Exception as e:
                        Triglycerides_result_pure = math.nan
                        Triglycerides_result_form_field_instance = 'This field does not have any data'
                        Triglycerides_result_disname = 'Empty'

                    try:
                        Urea_BUN_out_normal = row['Urea (BUN), Out of normal range?']
                        Urea_BUN_out_normal_pure = Urea_BUN_out_normal.split('|')[0]
                        Urea_BUN_out_normal_form_field_instance = Urea_BUN_out_normal.split('|')[1]
                        Urea_BUN_out_normal_disname = Urea_BUN_out_normal.split('|')[2]
                    except Exception as e:
                        Urea_BUN_out_normal_pure = math.nan
                        Urea_BUN_out_normal_form_field_instance = 'This field does not have any data'
                        Urea_BUN_out_normal_disname = 'Empty'

                    try:
                        Urea_BUN_result = row['Urea (BUN), Result (mg/dL)']
                        Urea_BUN_result_pure = Urea_BUN_result.split('|')[0]
                        Urea_BUN_result_form_field_instance = Urea_BUN_result.split('|')[1]
                        Urea_BUN_result_disname = Urea_BUN_result.split('|')[0]
                    except Exception as e:
                        Urea_BUN_result_pure = math.nan
                        Urea_BUN_result_form_field_instance = 'This field does not have any data'
                        Urea_BUN_result_disname = 'Empty'

                    try:
                        Uric_Acid_out_normal = row['Uric Acid, Out of normal range?']
                        Uric_Acid_out_normal_pure = Uric_Acid_out_normal.split('|')[0]
                        Uric_Acid_out_normal_form_field_instance = Uric_Acid_out_normal.split('|')[1]
                        Uric_Acid_out_normal_disname = Uric_Acid_out_normal.split('|')[2]
                    except Exception as e:
                        Uric_Acid_out_normal_pure = math.nan
                        Uric_Acid_out_normal_form_field_instance = 'This field does not have any data'
                        Uric_Acid_out_normal_disname = 'Empty'

                    try:
                        Uric_Acid_result = row['Uric Acid, Result (mmol/L)']
                        Uric_Acid_result_pure = Uric_Acid_result.split('|')[0]
                        Uric_Acid_result_form_field_instance = Uric_Acid_result.split('|')[1]
                        Uric_Acid_result_disname = Uric_Acid_result.split('|')[0]
                    except Exception as e:
                        Uric_Acid_result_pure = math.nan
                        Uric_Acid_result_form_field_instance = 'This field does not have any data'
                        Uric_Acid_result_disname = 'Empty'

                    try:
                        ferretin_result = row['Ferritin, Result (mg/L)']
                        ferretin_result_pure = ferretin_result.split('|')[0]
                        ferretin_result_form_field_instance = ferretin_result.split('|')[1]
                    except:
                        ferretin_result_pure = math.nan
                        ferretin_result_form_field_instance = 'Empty'
                    
                    try:
                        ferretin_out_normal = row['Ferritin, Out of normal range?']
                        ferretin_out_normal_pure = ferretin_out_normal.split('|')[0]
                        ferretin_out_normal_form_field_instance = ferretin_out_normal.split('|')[1]
                    except:
                        ferretin_out_normal_pure = math.nan
                        ferretin_out_normal_form_field_instance = 'Empty'

                    #-------------------------------------------------------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)
                    
                    if date_collected_pure == '':
                        pass
                    else:
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_collected_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of examination performed', date_collected_form_field_instance ,f , date_collected_pure, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBC0010
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_collected_f != date_of_visit_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance,\
                                        'The date should be the same as the visit date in the "Date of Visit" Form' , f'{date_collected_pure} - {date_of_visit}', 'LBT0010']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBC0010--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBC0030
                    if date_collected_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_collected_f = datetime.strptime(date_collected_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_collected_f < date_inform_consent_f:
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance,\
                                        'The date/time of test performed can not be before the informed consent date/time' ,f'{date_collected_pure} - {date_inform_consent}', 'LBC0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LBC0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LBC0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_collected_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_collected_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date Collected', date_collected_form_field_instance ,'Date Collected must be before the End of study/Early withdrawal date. ', date_collected_pure, 'LBC0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LBC0040 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # Revision LBC0050
                    try:
                        if  float(blood_sample_collected_pure) == 9.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance,\
                                         'The "Not Required" option can only be selected if visit is D-1 and the D-1 visit date =Screening visit date or normal and done in the previous 10 days', \
                                            blood_sample_collected_pure, 'LBC0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision  LBC0260
                        if float(Alkaline_out_normal_pure) == 1.0:
                            if float(Alkaline_result_pure) > 40.0 and float(Alkaline_result_pure) < 150.0:
                                error = [subject, visit, 'Alkaline Phosphatase, Out of normal range?', Alkaline_result_form_field_instance,\
                                         'According to the result, the value is not out of range, please review. (40.0 - 150.0)', Alkaline_result_pure, 'LBC0260']
                                lista_revision.append(error)

                        # Revision LBC0060
                        elif float(Alkaline_out_normal_pure) == 0.0:
                            if float(Alkaline_result_pure) < 40.0 or float(Alkaline_result_pure) > 150.0:
                                error = [subject, visit, 'Alkaline Phosphatase, Out of normal range?', Alkaline_result_form_field_instance,\
                                         'According to the result, the value is out of range, please review. (40.0 - 150.0)', Alkaline_result_pure, 'LBC0060']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0060--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0270
                        if float(GGT_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(GGT_result_pure) > 12.0 and float(GGT_result_pure) < 64.0 :
                                    error = [subject, visit, 'Gamma Glutamyl Transferase (GGT), Out of normal range? ', GGT_result_form_instance  ,\
                                             'According to the result, the value is not out of range, please review. (12.0 - 64.0)' , f'{GGT_out_normal_disname} - {GGT_result_pure}', 'LBC0270']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(GGT_result_pure) > 9.0 and float(GGT_result_pure) < 36.0:
                                    error = [subject, visit, 'Gamma Glutamyl Transferase (GGT), Out of normal range? ',GGT_result_form_instance,\
                                             'According to the result, the value is not out of range, please review. (9.0 - 36.0)' ,  f'{GGT_out_normal_disname} - {GGT_result_pure}', 'LBC0270']
                                    lista_revision.append(error)

                        # Revision LBC0070
                        elif float(GGT_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(GGT_result_pure) < 12.0 or float(GGT_result_pure) > 64.0 :
                                    error = [subject, visit, 'Gamma Glutamyl Transferase (GGT), Out of normal range? ', GGT_result_form_instance ,\
                                             'According to the result, the value is out of range, please review. (12.0 - 64.0)' , f'{GGT_out_normal_disname} - {GGT_result_pure}', 'LBC0070']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(GGT_result_pure) < 9.0 or float(GGT_result_pure) > 36.0:
                                    error = [subject, visit, 'Gamma Glutamyl Transferase (GGT), Out of normal range? ', GGT_result_form_instance ,\
                                             'According to the result, the value is out of range, please review. (9.0 - 36.0)' ,  f'{GGT_out_normal_disname} - {GGT_result_pure}', 'LBC0070']
                                    lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0070--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0280
                        if float(Bilirubin_out_normal_pure) == 1.0:
                            if float(Bilirubin_result_pure) > 0.3 and float(Bilirubin_result_pure) < 1.2:
                                error = [subject, visit, 'Bilirubin (Total), Out of normal range?', Bilirubin_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (0.3 - 1.2)',  f' {Bilirubin_out_normal_disname} - {Bilirubin_result_pure}', 'LBC0280']
                                lista_revision.append(error)

                        # Revision LBC0080
                        elif float(Bilirubin_out_normal_pure) == 0.0:
                            if float(Bilirubin_result_pure) < 0.3 or float(Bilirubin_result_pure) > 1.2:
                                error = [subject, visit, 'Bilirubin (Total), Out of normal range?', Bilirubin_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (0.3 - 1.2)' ,  f' {Bilirubin_out_normal_disname} - {Bilirubin_result_pure}', 'LBC0080']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0280--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0290
                        if float(Bilirubin_direct_out_normal_pure) == 1.0:
                            if float(Bilirubin_direct_result_pure) > 0.0 and float(Bilirubin_direct_result_pure) < 0.5:
                                error = [subject, visit, \
                                         'Bilirubin (Direct) (only if Total is elevated), Out of normal range?'\
                                            , Bilirubin_direct_result_form_field_instance,\
                                                'According to the result, the value is not out of range, please review. (0.0 - 0.5)', \
                                                    f'{Bilirubin_direct_out_normal_disname} - {Bilirubin_direct_result_pure}', 'LBC0290']
                                lista_revision.append(error)

                        # Revision LBC0090
                        elif float(Bilirubin_direct_out_normal_pure) == 0.0:
                            if float(Bilirubin_direct_result_pure) < 0.0 or float(Bilirubin_direct_result_pure) > 0.5:
                                error = [subject, visit, 'Bilirubin (Direct) (only if Total is elevated), Out of normal range?e?',\
                                           Bilirubin_direct_result_form_field_instance,\
                                            'According to the result, the value is out of range, please review.(0.0 - 0.5)', \
                                                    f'{Bilirubin_direct_out_normal_disname} - {Bilirubin_direct_result_pure}', 'LBC0090']
                                lista_revision.append(error)
                                          
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0290--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0300
                        if float(Albumin_out_normal_pure) == 1.0:
                            if float(Albumin_result_pure) > 3.5 and float(Albumin_result_pure) < 5.0:
                                error = [subject, visit, 'Albumin, Out of normal range?', Albumin_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (3.5 - 5.0)', \
                                            f'{Albumin_out_normal_disname} - {Albumin_result_pure}', 'LBC0300']
                                lista_revision.append(error)

                        # Revision LBC0100
                        elif float(Albumin_out_normal_pure) == 0.0:
                            if float(Albumin_result_pure) < 3.5 or float(Albumin_result_pure) > 56.0:
                                error = [subject, visit, 'Albumin, Out of normal range?', Albumin_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (3.5 - 5.0)', \
                                            f'{Albumin_out_normal_disname} - {Albumin_result_pure}', 'LBC0100']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0300--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0310
                        if float(Aspartate_out_normal_pure) == 1.0:
                            if float(Aspartate_result_pure) > 11.0 and float(Aspartate_result_pure) < 34.0:
                                error = [subject, visit, 'Aspartate Aminotransferase (AST), Out of normal range?', Aspartate_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (11.0 - 34.0)' , f'{Aspartate_out_normal_disname} - {Aspartate_result_pure}', 'LBC0310']
                                lista_revision.append(error)

                        # Revision LBC0110
                        elif float(Aspartate_out_normal_pure) == 0.0:
                            if float(Aspartate_result_pure) < 11.0 or float(Aspartate_result_pure) > 34.0:
                                error = [subject, visit, 'Aspartate Aminotransferase (AST), Out of normal range?', Aspartate_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (11.0 - 34.0)' , f'{Aspartate_out_normal_disname} - {Aspartate_result_pure}', 'LBC0110']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0310--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LBC0320
                    try:
                        if float(Alanine_out_normal_pure) == 1.0:

                            if genero == 1.0:
                                if float(Alanine_result_pure) > 0.0 and \
                                float(Alanine_result_pure) < 45.0:
                                    
                                    error = [subject, visit, 'Alanine Aminotransferase (ALT), Out of normal range?', Alanine_result_form_field_isntance ,\
                                            'If Out of range="No"and Alanine Aminotransferase (ALT), Result (U/L) is not between 0 and 34 (Female) or is not between 0 and 45 (Male)' , \
                                                Alanine_result_pure, 'LBC0320']
                                    lista_revision.append(error)

                            elif genero == 2.0:
                                if float(Alanine_result_pure) > 0.0 and \
                                float(Alanine_result_pure) < 34.0:
                                    
                                    error = [subject, visit, 'Alanine Aminotransferase (ALT), Out of normal range?', Alanine_result_form_field_isntance ,\
                                            'If Out of range="No"and Alanine Aminotransferase (ALT), Result (U/L) is not between 0 and 34 (Female) or is not between 0 and 45 (Male)' , \
                                                Alanine_result_pure, 'LBC0320']
                                    lista_revision.append(error)
                        
                        # Revision LBC0120
                        elif float(Alanine_out_normal_pure) == 0.0:

                            if genero == 1.0:
                                if float(Alanine_result_pure) < 0.0 or \
                                float(Alanine_result_pure) > 45.0:
                                    error = [subject, visit, 'Alanine Aminotransferase (ALT), Out of normal range?', Alanine_result_form_field_isntance ,\
                                            'If Out of range="No"and Alanine Aminotransferase (ALT), Result (U/L) is not between 0 and 34 (Female) or is not between 0 and 45 (Male)' , \
                                                Alanine_result_pure, 'LBC0120']
                                    lista_revision.append(error)       

                            elif genero == 1.0:
                                if float(Alanine_result_pure) < 0.0 or \
                                float(Alanine_result_pure) > 34.0:
                                    error = [subject, visit, 'Alanine Aminotransferase (ALT), Out of normal range?', Alanine_result_form_field_isntance ,\
                                            'If Out of range="No"and Alanine Aminotransferase (ALT), Result (U/L) is not between 0 and 34 (Female) or is not between 0 and 45 (Male)' , \
                                                Alanine_result_pure, 'LBC0120']
                                    lista_revision.append(error)              

                    except Exception as e:
                        lista_logs.append(f'Revision LBC0120--> {e} - Subject: {subject},  Visit: {visit} ')


                    try:
                        # Revision LBC0330
                        if float(Protein_out_normal_pure) == 1.0:
                            if float(Protein_total_result_pure) > 6.4 and float(Protein_total_result_pure) < 8.3:
                                error = [subject, visit, 'Protein (Total), Out of normal range?', Protein_total_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (6.4 - 8.3)', \
                                            f'{Protein_out_normal_disname} - {Protein_total_result_pure}', 'LBC0330']
                                lista_revision.append(error)

                        # Revision LBC0130
                        elif float(Protein_out_normal_pure) == 0.0:
                            if float(Protein_total_result_pure) < 6.4 or float(Protein_total_result_pure) > 8.3:
                                error = [subject, visit, 'Protein (Total), Out of normal range?', Protein_total_result_form_field_instance,\
                                         'According to the result, the value is out of range, please review. (6.4 - 8.3)', \
                                            f'{Protein_out_normal_disname} - {Protein_total_result_pure}', 'LBC0130']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0330--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0340
                        if float(Creatine_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Creatine_result_pure) > 30.0 and float(Creatine_result_pure) < 200.0 :
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', Creatine_result_form_field_instance,\
                                             'According to the result, the value is not out of range, please review. (30.0 - 200.0)', \
                                                f'{Creatine_out_normal_disname} - {Creatine_result_pure}', 'LBC0340']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Creatine_result_pure) > 29.0 and float(Creatine_result_pure) < 168.0:
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', Creatine_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (29.0 - 168.0)', \
                                                f'{Creatine_out_normal_disname} - {Creatine_result_pure}', 'LBC0340']
                                    lista_revision.append(error)

                        # Revision LBC0140
                        elif float(Creatine_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Creatine_result_pure) < 30.0 or float(Creatine_result_pure) > 200.0 :
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', Creatine_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review. (30.0 - 200.0)', \
                                                f'{Creatine_out_normal_disname} - {Creatine_result_pure}', 'LBC0140']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Creatine_result_pure) < 29.0 or float(Creatine_result_pure) > 168.0:
                                    error = [subject, visit, 'Creatine Kinase (CK), Out of normal range?', Creatine_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review. (29.0 - 168.0)', \
                                                f'{Creatine_out_normal_disname} - {Creatine_result_pure}', 'LBC0140']
                                    lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBC0340 --> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0350
                        if float(Sodium_out_normal_pure) == 1.0:
                            if float(Sodium_result_pure) > 136.0 and float(Sodium_result_pure) < 145.0:
                                error = [subject, visit, 'Sodium, Out of normal range?', Sodium_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (136.0 - 145.0)', f'{Sodium_out_normal_disname} - {Sodium_result_pure}', 'LBC0350']
                                lista_revision.append(error)

                        # Revision LBC0150
                        elif float(Sodium_out_normal_pure) == 0.0:
                            if float(Sodium_result_pure) < 136.0 or float(Sodium_result_pure) > 145.0:
                                error = [subject, visit, 'Sodium, Out of normal range?', Sodium_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (136.0 - 145.0)',  f'{Sodium_out_normal_disname} - {Sodium_result_pure}', 'LBC0150']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0350--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0360
                        if float(Potassium_out_normal_pure) == 1.0:
                            if float(Potassium_result_pure) > 3.5 and float(Potassium_result_pure) < 5.1 :
                                error = [subject, visit, 'Potassium, Out of normal range?', Potassium_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (3.5 - 5.1)', f'{Potassium_out_normal_disname} - {Potassium_result_pure}', 'LBC0360']
                                lista_revision.append(error)

                        # Revision LBC0160
                        elif float(Potassium_out_normal_pure) == 0.0:
                            if float(Potassium_result_pure) <  3.5  or float(Potassium_result_pure) > 5.1 :
                                error = [subject, visit, 'Potassium, Out of normal range?', Potassium_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (3.5 - 5.1)', f'{Potassium_out_normal_disname} - {Potassium_result_pure}', 'LBC0160']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0360--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0370
                        if float(Calcium_out_normal_pure) == 1.0:

                            if float(Calcium_result_pure) > 8.4 and float(Calcium_result_pure) < 10.2:
                                error = [subject, visit, 'Calcium, Out of normal range?', Calcium_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (8.4 - 10.2)', \
                                                f'{Calcium_out_normal_disname} - {Calcium_result_pure}', 'LBC0370']
                                lista_revision.append(error)

                        # Revision LBC0170
                        elif float(Calcium_out_normal_pure) == 0.0:

                            if float(Calcium_result_pure) < 8.4 or float(Calcium_result_pure) > 10.2 :
                                error = [subject, visit, 'Calcium, Out of normal range?', Calcium_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (8.4 - 10.2)', f'{Calcium_out_normal_disname} - {Calcium_result_pure}', 'LBC0170']
                                lista_revision.append(error)
                                

                    except Exception as e:
                        lista_logs.append(f'Revision LBC0370 --> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0380
                        if float(Bicarbonate_out_normal_pure) == 1.0:
                            if float(Bicarbonate_result_pure) > 22.0 and float(Bicarbonate_result_pure) < 31.0 :
                                error = [subject, visit, 'Bicarbonate, Out of normal range?', Bicarbonate_result_form_field_instance,\
                                         'According to the result, the value is not out of range, please review. (22.0 - 31.0)', \
                                            f'{Bicarbonate_out_normal_disname} - {Bicarbonate_result_pure}', 'LBC0380']
                                lista_revision.append(error)

                        # Revision LBC0180
                        elif float(Bicarbonate_out_normal_pure) == 0.0:
                            if float(Bicarbonate_result_pure) <  22.0  or float(Bicarbonate_result_pure) > 31.0 :
                                error = [subject, visit, 'Bicarbonate, Out of normal range?', Bicarbonate_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (22.0 - 31.0)', \
                                            f'{Bicarbonate_out_normal_disname} - {Bicarbonate_result_pure}', 'LBC0180']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0380--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0390
                        if float(Chloride_out_normal_pure) == 1.0:
                            if float(Chloride_result_pure) > 98.0 and float(Chloride_result_pure) < 107.0 :
                                error = [subject, visit, 'Chloride, Out of normal range?', Chloride_result_form_field_instacnce ,\
                                         'According to the result, the value is not out of range, please review. (98.0 - 107.0)', f'{Chloride_out_normal_disname} - {Chloride_result_pure}', 'LBC0390']
                                lista_revision.append(error)

                        # Revision LBC0190
                        elif float(Chloride_out_normal_pure) == 0.0:
                            if float(Chloride_result_pure) <  98.0  or float(Chloride_result_pure) > 107.0 :
                                error = [subject, visit, 'Chloride, Out of normal range?', Chloride_result_form_field_instacnce ,\
                                         'According to the result, the value is out of range, please review. (98.0 - 107.0)', f'{Chloride_out_normal_disname} - {Chloride_result_pure}', 'LBC0190']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0390--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0400
                        if float(Serum_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Serum_result_pure) >  0.73 and float(Serum_result_pure) < 1.18 :
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', Serum_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (0.73 - 1.18)', f'{Serum_out_normal_disname} - {Serum_result_pure}', 'LBC0400']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Serum_result_pure) > 0.55 and float(Serum_result_pure) <  1.02:
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', Serum_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (0.55 - 1.02)', f'{Serum_out_normal_disname} - {Serum_result_pure}', 'LBC0400']
                                    lista_revision.append(error)

                        # Revision LBC0200
                        elif float(Serum_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Serum_result_pure) < 0.73 or float(Serum_result_pure) > 1.18 :
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', Serum_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (0.73 - 1.18)', f'{Serum_out_normal_disname} - {Serum_result_pure}', 'LBC0200']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Serum_result_pure) < 0.55 or float(Serum_result_pure) > 1.02:
                                    error = [subject, visit, 'Serum Creatinine, Out of normal range?', Serum_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review. (0.55 - 1.02)', f'{Serum_out_normal_disname} - {Serum_result_pure}', 'LBC0200']
                                    lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBC0400--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    try:
                        # Revision LBC0410
                        if float(Urea_BUN_out_normal_pure) == 1.0:
                            if float(Urea_BUN_result_pure) >  18.84 and float(Urea_BUN_result_pure) < 42.80:
                                error = [subject, visit, 'Urea (BUN), Out of normal range?', Urea_BUN_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (18.84 - 42.80)', f'{Urea_BUN_out_normal_disname} - {Urea_BUN_result_pure}', 'LBC0410']
                                lista_revision.append(error)

                        # Revision LBC0210
                        elif float(Urea_BUN_out_normal_pure) == 0.0:
     
                            if float(Urea_BUN_result_pure) < 18.84 or float(Urea_BUN_result_pure) > 42.80:
                                error = [subject, visit, 'Urea (BUN), Out of normal range?', Urea_BUN_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review. (18.84 - 42.80)', f'{Urea_BUN_out_normal_disname} - {Urea_BUN_result_pure}', 'LBC0210']
                                lista_revision.append(error)


                    except Exception as e:
                        lista_logs.append(f'Revision LBC0410--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0420
                        if float(Uric_Acid_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(Uric_Acid_result_pure) >  3.7 and float(Uric_Acid_result_pure) < 7.7 :
                                    error = [subject, visit, 'Uric Acid, Out of normal range?', Uric_Acid_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (3.7 - 7.7)', f'{Uric_Acid_out_normal_disname} - {Uric_Acid_result_pure}', 'LBC0420']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Uric_Acid_result_pure) > 2.5 and float(Uric_Acid_result_pure) <  6.2:
                                    error = [subject, visit, 'Uric Acid, Out of normal range?', Uric_Acid_result_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (2.5 - 6.2)', f'{Uric_Acid_out_normal_disname} - {Uric_Acid_result_pure}', 'LBC0420']
                                    lista_revision.append(error)

                        # Revision LBC0220
                        elif float(Uric_Acid_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(Uric_Acid_result_pure) < 3.7 or float(Uric_Acid_result_pure) > 7.7 :
                                    error = [subject, visit, 'Uric Acid, Out of normal range?', Uric_Acid_result_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (3.7 - 7.7)', f'{Uric_Acid_out_normal_disname} - {Uric_Acid_result_pure}', 'LBC0220']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(Uric_Acid_result_pure) < 2.5 or float(Uric_Acid_result_pure) > 6.5:
                                    error = [subject, visit, 'Uric Acid, Out of normal range?', Uric_Acid_result_form_field_instance ,\
                                             'According to the result, the value is out of range, please review. (2.5 - 6.2)', f'{Uric_Acid_out_normal_disname} - {Uric_Acid_result_pure}', 'LBC0220']
                                    lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBC0420--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0430
                        if float(Cholesterol_out_normal_pure) == 1.0:
                            if float(Cholesterol_result_pure) <=  200.0 :
                                error = [subject, visit, 'Cholesterol (Total), Out of normal range?', Cholesterol_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (< 200.0)', f'{Cholesterol_out_normal_disname} - {Cholesterol_result_pure}', 'LBC0430']
                                lista_revision.append(error)

                        # Revision LBC0230
                        elif float(Cholesterol_out_normal_pure) == 0.0:
                            if float(Cholesterol_result_pure) >=  200.0  :
                                error = [subject, visit, 'Cholesterol (Total), Out of normal range?', Cholesterol_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (< 200.0)', f'{Cholesterol_out_normal_disname} - {Cholesterol_result_pure}', 'LBC0230']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0430--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0440
                        if float(Triglycerides_out_normal_pure) == 1.0:
                            if float(Triglycerides_result_pure) <= 150.0 :
                                error = [subject, visit, 'Triglycerides, Out of normal range?', Triglycerides_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (< 150.0)', f'{Triglycerides_out_normal_disname} - {Triglycerides_result_pure}', 'LBC0440']
                                lista_revision.append(error)

                        # Revision LBC0240
                        elif float(Triglycerides_out_normal_pure) == 0.0:
                            if float(Triglycerides_result_pure) >= 150.0 :
                                error = [subject, visit, 'Triglycerides, Out of normal range?', Triglycerides_result_form_field_instance,\
                                         'According to the result, the value is out of range, please review. (50.0 - 150.0)', f'{Triglycerides_out_normal_disname} - {Triglycerides_result_pure}', 'LBC0240']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0440--> {e} - Subject: {subject},  Visit: {visit} ')

                    try:
                        # Revision LBC0450
                        if float(C_Reactive_out_normal_pure) == 1.0:
                            if float(C_Reactive_result_pure) <= 0.5 :
                                error = [subject, visit, 'C-Reactive Protein, Out of normal range?', C_Reactive_result_form_field_instance ,\
                                         'According to the result, the value is not out of range, please review. (< 0.5)', f'{C_Reactive_out_normal_disname} - {C_Reactive_result_pure}', 'LBC0450']
                                lista_revision.append(error)

                        # Revision LBC0250
                        elif float(C_Reactive_out_normal_pure) == 0.0:
                            if  float(C_Reactive_result_pure) >= 0.5 :
                                error = [subject, visit, 'C-Reactive Protein, Out of normal range?', C_Reactive_result_form_field_instance ,\
                                         'According to the result, the value is out of range, please review. (< 0.5)', f'{C_Reactive_out_normal_disname} - {C_Reactive_result_pure}', 'LBC0250']
                                lista_revision.append(error)
                                            
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0450--> {e} - Subject: {subject},  Visit: {visit}')
                    

                    try:
                        # Revision LBC0470
                        if float(ferretin_out_normal_pure) == 1.0:
                            if float(genero) == 1.0:
                                if float(ferretin_result_pure) >  21.81 and float(ferretin_result_pure) < 274.66 :
                                    error = [subject, visit, 'Ferritin, Out of normal range?', ferretin_out_normal_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (21.81 - 274.66)', ferretin_result_pure, 'LBC0470']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(ferretin_result_pure) > 4.63 and float(ferretin_result_pure) < 204.0 :
                                    error = [subject, visit, 'Ferritin, Out of normal range?', ferretin_out_normal_form_field_instance ,\
                                             'According to the result, the value is not out of range, please review. (4.63 - 204.0)', ferretin_result_pure, 'LBC0470']
                                    lista_revision.append(error)

                        # Revision LBC0480
                        elif float(ferretin_out_normal_pure) == 0.0:
                            if float(genero) == 1.0:
                                if float(ferretin_result_pure) < 21.81 or float(ferretin_result_pure) > 274.66 :
                                    error = [subject, visit, 'Ferritin, Out of normal range?', ferretin_out_normal_form_field_instance,\
                                             'According to the result, the value is out of range, please review. (21.81 - 274.66)', ferretin_result_pure, 'LBC0480']
                                    lista_revision.append(error)
                                
                            elif float(genero) == 2.0:
                                if float(ferretin_result_pure) < 4.63 or float(ferretin_result_pure) > 204.0:
                                    error = [subject, visit, 'Ferritin, Out of normal range?', ferretin_out_normal_form_field_instance ,\
                                             'According to the result, the value is out of range, please review. (4.63 - 204.0)', ferretin_result_pure, 'LBC0480']
                                    lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LBC0470--> {e} - Subject: {subject},  Visit: {visit} ')


                
                    lista_validacion = ['Alkaline Phosphatase',
                                        'Alkaline Phosphatase, Result (U/L)',
                                        'Alkaline Phosphatase, Out of normal range?',
                                        'Alkaline Phosphatase, If abnormal, Specify',
                                        'Gamma Glutamyl Transferase (GGT)',
                                        'Gamma Glutamyl Transferase (GGT), Result (U/L)',
                                        'Gamma Glutamyl Transferase (GGT), Out of normal range?',
                                        'Gamma Glutamyl Transferase (GGT), If abnormal, Specify',
                                        'Bilirubin (Total)',
                                        'Bilirubin (Total), Result (mg/dL)',
                                        'Bilirubin (Total), Out of normal range?',
                                        'Bilirubin (Total), If abnormal, Specify',
                                        'Bilirubin (Direct) (only if Total is elevated)',
                                        'Bilirubin (Direct) (only if Total is elevated), Result (mg/dL)',
                                        'Bilirubin (Direct) (only if Total is elevated), Out of normal range?',
                                        'Bilirubin (Direct) (only if Total is elevated) , If abnormal, Specify',
                                        'Albumin',
                                        'Albumin, Result (g/dL)',
                                        'Albumin, Out of normal range?',
                                        'Albumin, If abnormal, Specify',
                                        'Aspartate Aminotransferase (AST)',
                                        'Aspartate Aminotransferase (AST), Result (U/L)',
                                        'Aspartate Aminotransferase (AST), Out of normal range?',
                                        'Aspartate Aminotransferase (AST), If abnormal, Specify',
                                        'Alanine Aminotransferase (ALT)',
                                        'Alanine Aminotransferase (ALT), Result (U/L)',
                                        'Alanine Aminotransferase (ALT), Out of normal range?',
                                        'Alanine Aminotransferase (ALT), If abnormal, Specify',
                                        'Protein (Total)',
                                        'Protein (Total), Result (mg/dL)',
                                        'Protein (Total), Out of normal range?',
                                        'Protein (Total), If abnormal, Specify',
                                        'Creatine Kinase (CK)',
                                        'Creatine Kinase (CK), Result (U/L)',
                                        'Creatine Kinase (CK), Out of normal range?',
                                        'Creatine Kinase (CK), If abnormal, Specify',
                                        'Sodium',
                                        'Sodium, Result (mmol/L)',
                                        'Sodium, Out of normal range?',
                                        'Sodium, If abnormal, Specify',
                                        'Potassium',
                                        'Potassium, Result (mmol/L)',
                                        'Potassium, Out of normal range?',
                                        'Potassium, If abnormal, Specify',
                                        'Calcium',
                                        'Calcium, Result (mEq/L)',
                                        'Calcium, Out of normal range?',
                                        'Calcium, If abnormal, Specify',
                                        'Bicarbonate',
                                        'Bicarbonate, Result (mEq/L)',
                                        'Bicarbonate, Out of normal range?',
                                        'Bicarbonate, If abnormal, Specify',
                                        'Chloride',
                                        'Chloride, Result (mmol/L)',
                                        'Chloride, Out of normal range?',
                                        'Chloride, If abnormal, Specify',
                                        'Serum Creatinine',
                                        'Serum Creatinine, Result  (mg/dL)',
                                        'Serum Creatinine, Out of normal range?',
                                        'Serum Creatinine, If abnormal, Specify',
                                        'Urea (BUN)',
                                        'Urea (BUN), Result (mg/dL)',
                                        'Urea (BUN), Out of normal range?',
                                        'Urea, If abnormal, Specify',
                                        'Uric Acid',
                                        'Uric Acid, Result (mmol/L)',
                                        'Uric Acid, Out of normal range?',
                                        'Uric Acid, If abnormal, Specify',
                                        'Cholesterol (Total)',
                                        'Cholesterol (Total), Result (mg/dL)',
                                        'Cholesterol (Total), Out of normal range?',
                                        'Cholesterol (Total), If abnormal, Specify',
                                        'Glucose',
                                        'Glucose, Result  (mg/dL)',
                                        'Glucose, Out of normal range?',
                                        'Glucose, If abnormal, Specify',
                                        'Triglycerides',
                                        'Triglycerides, Result (mg/dL)',
                                        'Triglycerides, Out of normal range?',
                                        'Triglycerides, If abnormal, Specify',
                                        'C-Reactive Protein',
                                        'C-Reactive Protein, Result (mg/L)',
                                        'C-Reactive Protein, Out of normal range?',
                                        'C-Reactive Protein, If abnormal, Specify',
                                        'Ferritin, Result (mg/L)',
                                        'Ferritin, Out of normal range?',
                                        'Ferritin, If abnormal, Specify']
                    
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try:
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan

                        if  math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass
 
                    # Revision LBC0460
                    try:
                        if float(blood_sample_collected_pure) ==1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Blood Sample Collected', blood_sample_collected_form_field_instance ,\
                                         'If Blood Sample Collected is checked as "Yes", not all laboratory tests can be "not done"', \
                                            blood_sample_collected_pure, 'LBC0460']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LBC0460--> {e} - Subject: {subject},  Visit: {visit} ')
                    
    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    clinical_laboratory_test_clinical_chemistry_output = pd.DataFrame(lista_revision, columns=column_names)

    
    sheet = excel_writer.create_sheet("CL - Clinical Chemistry")
    
    for row in dataframe_to_rows(clinical_laboratory_test_clinical_chemistry_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)
    return clinical_laboratory_test_clinical_chemistry_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    clinical_laboratory_test_clinical_chemistry(df_root, path_excel) 