import numpy as np
import math
import pandas as pd
from datetime import datetime
from log_writer import log_writer
from revision_fechas import revision_fecha
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

def lead_ECG(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de 12-Lead ECG
    '''

    df= df_root[df_root['name']== '12-Lead ECG']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)


    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Visit','Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})

    lista_revision = []
    lista_logs = ['12-Lead ECG']

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita
            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')


            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]
   
                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']
                time_format = '%H:%M'

                
                if status != '':

                    try:
                        were_ECG_performed = row['Were the ECG performed?']
                        were_ECG_performed_pure  = were_ECG_performed.split('|')[0]
                        were_ECG_performed_form_field_instance = were_ECG_performed.split('|')[1]
                        were_ECG_performed_disname = were_ECG_performed.split('|')[2]
                    except Exception as e:
                        were_ECG_performed_pure = math.nan
                        were_ECG_performed_form_field_instance = 'This field does not have any data'
                        were_ECG_performed_disname = 'Empty'
                    
                    try: 
                        date_of_egc = row['Date of ECG Performed']
                        date_of_egc_pure = date_of_egc.split('|')[0]
                        date_of_egc_form_field_instance = date_of_egc.split('|')[1]
                        date_of_egc_disname = date_of_egc.split('|')[0]
                    except Exception as e:
                        date_of_egc_pure = ''
                        date_of_egc_form_field_instance = 'This field does not have any data'
                        date_of_egc_disname = 'Empty'
                    
                    try:                    
                        Undefined_HR_bpm = row["Undefined, HR (bpm)"] 
                        Undefined_HR_bpm_pure = Undefined_HR_bpm.split('|')[0]
                        Undefined_HR_bpm_form_field_instance = Undefined_HR_bpm.split('|')[1]
                        Undefined_HR_bpm_disname = Undefined_HR_bpm.split('|')[0]
                    except Exception as e:
                        Undefined_HR_bpm_pure = math.nan
                        Undefined_HR_bpm_form_field_instance = 'This field does not have any data'
                        Undefined_HR_bpm_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1_Interpretation = row['Pre dose triplicate 1, Interpretation']
                        Pre_dose_triplicate_1_Interpretation_pure = Pre_dose_triplicate_1_Interpretation.split('|')[0]
                        Pre_dose_triplicate_1_Interpretation_form_field_instance = Pre_dose_triplicate_1_Interpretation.split('|')[1]
                        Pre_dose_triplicate_1_Interpretation_disname = Pre_dose_triplicate_1_Interpretation.split('|')[2]
                    except Exception as e:
                        Pre_dose_triplicate_1_Interpretation_pure = math.nan 
                        Pre_dose_triplicate_1_Interpretation_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1_Interpretation_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1_HR_bpm = row["Pre dose triplicate 1, HR (bpm)"]
                        Pre_dose_triplicate_1_HR_bpm_pure = Pre_dose_triplicate_1_HR_bpm.split('|')[0]
                        Pre_dose_triplicate_1_HR_bpm_form_field_instance = Pre_dose_triplicate_1_HR_bpm.split('|')[1]
                        Pre_dose_triplicate_1_HR_bpm_disname = Pre_dose_triplicate_1_HR_bpm.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_1_HR_bpm_pure = math.nan
                        Pre_dose_triplicate_1_HR_bpm_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1_HR_bpm_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_HR_bpm = row["Pre dose triplicate 2, HR (bpm)"]
                        Pre_dose_triplicate_2_HR_bpm_pure = Pre_dose_triplicate_2_HR_bpm.split('|')[0]
                        Pre_dose_triplicate_2_HR_bpm_form_field_instance = Pre_dose_triplicate_2_HR_bpm.split('|')[1]
                        Pre_dose_triplicate_2_HR_bpm_disname = Pre_dose_triplicate_2_HR_bpm.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_HR_bpm_pure = math.nan
                        Pre_dose_triplicate_2_HR_bpm_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_HR_bpm_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_3_HR_bpm = row["Pre dose triplicate 3, HR (bpm)"]
                        Pre_dose_triplicate_3_HR_bpm_pure = Pre_dose_triplicate_3_HR_bpm.split('|')[0]
                        Pre_dose_triplicate_3_HR_bpm_form_field_instance = Pre_dose_triplicate_3_HR_bpm.split('|')[1]
                        Pre_dose_triplicate_3_HR_bpm_disname = Pre_dose_triplicate_3_HR_bpm.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_HR_bpm_pure = math.nan
                        Pre_dose_triplicate_3_HR_bpm_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_HR_bpm_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1__RR_msec = row['Pre dose triplicate 1, RR (msec)']
                        Pre_dose_triplicate_1__RR_msec_pure = Pre_dose_triplicate_1__RR_msec.split('|')[0]
                        Pre_dose_triplicate_1__RR_msec_form_field_instance = Pre_dose_triplicate_1__RR_msec.split('|')[1]
                        Pre_dose_triplicate_1__RR_msec_disname = Pre_dose_triplicate_1__RR_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_1__RR_msec_pure = math.nan
                        Pre_dose_triplicate_1__RR_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1__RR_msec_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_HR_bpm = row["15-min post dose, HR (bpm)"]
                        min_15_post_dose_HR_bpm_pure = min_15_post_dose_HR_bpm.split('|')[0]
                        min_15_post_dose_HR_bpm_form_field_instance = min_15_post_dose_HR_bpm.split('|')[1]
                        min_15_post_dose_HR_bpm_disname = min_15_post_dose_HR_bpm.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_HR_bpm_pure = math.nan
                        min_15_post_dose_HR_bpm_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_HR_bpm_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_HR_bpm = row["30-min post dose, HR (bpm)"]
                        min_30_post_dose_HR_bpm_pure = min_30_post_dose_HR_bpm.split('|')[0]
                        min_30_post_dose_HR_bpm_form_field_instance = min_30_post_dose_HR_bpm.split('|')[1]
                        min_30_post_dose_HR_bpm_disname = min_30_post_dose_HR_bpm.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_HR_bpm_pure = math.nan
                        min_30_post_dose_HR_bpm_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_HR_bpm_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_HR_bpm = row["60-min post dose, HR (bpm)"]
                        min_60_post_dose_HR_bpm_pure = min_60_post_dose_HR_bpm.split('|')[0]
                        min_60_post_dose_HR_bpm_form_field_instance = min_60_post_dose_HR_bpm.split('|')[1]
                        min_60_post_dose_HR_bpm_disname = min_60_post_dose_HR_bpm.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_HR_bpm_pure = math.nan 
                        min_60_post_dose_HR_bpm_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_HR_bpm_disname = 'Empty'
                    
                    try:
                        Undefined_PR_msec = row["Undefined, PR (msec)"]
                        Undefined_PR_msec_pure = Undefined_PR_msec.split('|')[0]
                        Undefined_PR_msec_form_field_instance = Undefined_PR_msec.split('|')[1]
                        Undefined_PR_msec_disname = Undefined_PR_msec.split('|')[0]
                    except Exception as e:
                        Undefined_PR_msec_pure = math.nan
                        Undefined_PR_msec_form_field_instance = 'This field does not have any data'
                        Undefined_PR_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_PR_msec = row["Pre dose triplicate 2, PR (msec)"]
                        Pre_dose_triplicate_2_PR_msec_pure = Pre_dose_triplicate_2_PR_msec.split('|')[0]
                        Pre_dose_triplicate_2_PR_msec_form_field_instance = Pre_dose_triplicate_2_PR_msec.split('|')[1]
                        Pre_dose_triplicate_2_PR_msec_disname = Pre_dose_triplicate_2_PR_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_PR_msec_pure = math.nan
                        Pre_dose_triplicate_2_PR_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_PR_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_3_PR_msec = row["Pre dose triplicate 3, PR (msec)"]
                        Pre_dose_triplicate_3_PR_msec_pure = Pre_dose_triplicate_3_PR_msec.split('|')[0]
                        Pre_dose_triplicate_3_PR_msec_form_field_instance = Pre_dose_triplicate_3_PR_msec.split('|')[1]
                        Pre_dose_triplicate_3_PR_msec_disname = Pre_dose_triplicate_3_PR_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_PR_msec_pure = math.nan
                        Pre_dose_triplicate_3_PR_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_PR_msec_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_PR_msec = row["15-min post dose, PR (msec)"]
                        min_15_post_dose_PR_msec_pure = min_15_post_dose_PR_msec.split('|')[0]
                        min_15_post_dose_PR_msec_form_field_instance = min_15_post_dose_PR_msec.split('|')[1]
                        min_15_post_dose_PR_msec_disname = min_15_post_dose_PR_msec.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_PR_msec_pure = math.nan
                        min_15_post_dose_PR_msec_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_PR_msec_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_PR_msec = row["30-min post dose, PR (msec)"]
                        min_30_post_dose_PR_msec_pure = min_30_post_dose_PR_msec.split('|')[0]
                        min_30_post_dose_PR_msec_form_field_instance = min_30_post_dose_PR_msec.split('|')[1]
                        min_30_post_dose_PR_msec_disname = min_30_post_dose_PR_msec.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_PR_msec_pure = math.nan
                        min_30_post_dose_PR_msec_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_PR_msec_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_PR_msec = row["60-min post dose, PR (msec)"]
                        min_60_post_dose_PR_msec_pure = min_60_post_dose_PR_msec.split('|')[0]
                        min_60_post_dose_PR_msec_form_field_instance = min_60_post_dose_PR_msec.split('|')[1]
                        min_60_post_dose_PR_msec_disname = min_60_post_dose_PR_msec.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_PR_msec_pure = math.nan
                        min_60_post_dose_PR_msec_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_PR_msec_disname = 'Empty'
                    
                    try:
                        Undefined_QRS_msec = row["Undefined, QRS (msec)"]
                        Undefined_QRS_msec_pure = Undefined_QRS_msec.split('|')[0]
                        Undefined_QRS_msec_form_field_instance = Undefined_QRS_msec.split('|')[1]
                        Undefined_QRS_msec_disname = Undefined_QRS_msec.split('|')[0]
                    except Exception as e:
                        Undefined_QRS_msec_pure = math.nan
                        Undefined_QRS_msec_form_field_instance = 'This field does not have any data'
                        Undefined_QRS_msec_disname = 'Empty'
                    
                    try:
                        Undefined_RR_msec = row["Undefined, RR (msec)"]
                        Undefined_RR_msec_pure = Undefined_RR_msec.split('|')[0]
                        Undefined_RR_msec_form_field_instance = Undefined_RR_msec.split('|')[1]
                        Undefined_RR_msec_disname = Undefined_RR_msec.split('|')[0]
                    except Exception as e:
                        Undefined_RR_msec_pure = math.nan
                        Undefined_RR_msec_form_field_instance = 'This field does not have any data'
                        Undefined_RR_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1_QRS_msec = row["Pre dose triplicate 1, QRS (msec)"]
                        Pre_dose_triplicate_1_QRS_msec_pure = Pre_dose_triplicate_1_QRS_msec.split('|')[0]
                        Pre_dose_triplicate_1_QRS_msec_form_field_instance = Pre_dose_triplicate_1_QRS_msec.split('|')[1]
                        Pre_dose_triplicate_1_QRS_msec_disname = Pre_dose_triplicate_1_QRS_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_1_QRS_msec_pure = math.nan
                        Pre_dose_triplicate_1_QRS_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1_QRS_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_QRS_msec = row["Pre dose triplicate 2, QRS (msec)"]
                        Pre_dose_triplicate_2_QRS_msec_pure = Pre_dose_triplicate_2_QRS_msec.split('|')[0]
                        Pre_dose_triplicate_2_QRS_msec_form_field_instance = Pre_dose_triplicate_2_QRS_msec.split('|')[1]
                        Pre_dose_triplicate_2_QRS_msec_disname = Pre_dose_triplicate_2_QRS_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_QRS_msec_pure = math.nan
                        Pre_dose_triplicate_2_QRS_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_QRS_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_RR_msec = row['Pre dose triplicate 2, RR (msec)']
                        Pre_dose_triplicate_2_RR_msec_pure = Pre_dose_triplicate_2_RR_msec.split('|')[0]
                        Pre_dose_triplicate_2_RR_msec_form_field_instance = Pre_dose_triplicate_2_RR_msec.split('|')[1]
                        Pre_dose_triplicate_2_RR_msec_disname = Pre_dose_triplicate_2_RR_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_RR_msec_pure = math.nan
                        Pre_dose_triplicate_2_RR_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_RR_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_3_QRS_msec = row["Pre dose triplicate 3, QRS (msec)"]
                        Pre_dose_triplicate_3_QRS_msec_pure = Pre_dose_triplicate_3_QRS_msec.split('|')[0]
                        Pre_dose_triplicate_3_QRS_msec_form_field_instance = Pre_dose_triplicate_3_QRS_msec.split('|')[1]
                        Pre_dose_triplicate_3_QRS_msec_disname = Pre_dose_triplicate_3_QRS_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_QRS_msec_pure = math.nan
                        Pre_dose_triplicate_3_QRS_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_QRS_msec_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_QRS_msec = row["15-min post dose, QRS (msec)"]
                        min_15_post_dose_QRS_msec_pure = min_15_post_dose_QRS_msec.split('|')[0]
                        min_15_post_dose_QRS_msec_form_field_instance = min_15_post_dose_QRS_msec.split('|')[1]
                        min_15_post_dose_QRS_msec_disname = min_15_post_dose_QRS_msec.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_QRS_msec_pure = math.nan
                        min_15_post_dose_QRS_msec_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_QRS_msec_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_QRS_msec = row["30-min post dose, QRS (msec)"]
                        min_30_post_dose_QRS_msec_pure = min_30_post_dose_QRS_msec.split('|')[0]
                        min_30_post_dose_QRS_msec_form_field_instance = min_30_post_dose_QRS_msec.split('|')[1]
                        min_30_post_dose_QRS_msec_disname = min_30_post_dose_QRS_msec.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_QRS_msec_pure = math.nan
                        min_30_post_dose_QRS_msec_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_QRS_msec_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_QRS_msec = row["60-min post dose, QRS (msec)"]
                        min_60_post_dose_QRS_msec_pure = min_60_post_dose_QRS_msec.split('|')[0]
                        min_60_post_dose_QRS_msec_form_field_instance = min_60_post_dose_QRS_msec.split('|')[1]
                        min_60_post_dose_QRS_msec_disname = min_60_post_dose_QRS_msec.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_QRS_msec_pure = math.nan
                        min_60_post_dose_QRS_msec_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_QRS_msec_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_RR_msec = row['15-min post dose, RR (msec)']
                        min_15_post_dose_RR_msec_pure = min_15_post_dose_RR_msec.split('|')[0]
                        min_15_post_dose_RR_msec_form_field_instance = min_15_post_dose_RR_msec.split('|')[1]
                        min_15_post_dose_RR_msec_disname = min_15_post_dose_RR_msec.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_RR_msec_pure = math.nan
                        min_15_post_dose_RR_msec_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_RR_msec_disname = 'Empty'
                    
                    try:
                        Undefined_QT_msec = row["Undefined, QT (msec)"]
                        Undefined_QT_msec_pure = Undefined_QT_msec.split('|')[0]
                        Undefined_QT_msec_form_field_instance = Undefined_QT_msec.split('|')[1]
                        Undefined_QT_msec_disname = Undefined_QT_msec.split('|')[0]
                    except Exception as e:
                        Undefined_QT_msec_pure = math.nan
                        Undefined_QT_msec_form_field_instance = 'This field does not have any data'
                        Undefined_QT_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1_QT_msec = row["Pre dose triplicate 1, QT (msec)"]
                        Pre_dose_triplicate_1_QT_msec_pure = Pre_dose_triplicate_1_QT_msec.split('|')[0]
                        Pre_dose_triplicate_1_QT_msec_form_field_instance = Pre_dose_triplicate_1_QT_msec.split('|')[1]
                        Pre_dose_triplicate_1_QT_msec_disname = Pre_dose_triplicate_1_QT_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_1_QT_msec_pure = math.nan
                        Pre_dose_triplicate_1_QT_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1_QT_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1_PR_msec = row['Pre dose triplicate 1, PR (msec)']
                        Pre_dose_triplicate_1_PR_msec_pure = Pre_dose_triplicate_1_PR_msec.split('|')[0]
                        Pre_dose_triplicate_1_PR_msec_form_field_instance = Pre_dose_triplicate_1_PR_msec.split('|')[1]
                        Pre_dose_triplicate_1_PR_msec_disname = Pre_dose_triplicate_1_PR_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_1_PR_msec_pure = math.nan
                        Pre_dose_triplicate_1_PR_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1_PR_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_QT_msec = row["Pre dose triplicate 2, QT (msec)"]
                        Pre_dose_triplicate_2_QT_msec_pure = Pre_dose_triplicate_2_QT_msec.split('|')[0]
                        Pre_dose_triplicate_2_QT_msec_form_field_instance = Pre_dose_triplicate_2_QT_msec.split('|')[1]
                        Pre_dose_triplicate_2_QT_msec_disname = Pre_dose_triplicate_2_QT_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_QT_msec_pure = math.nan
                        Pre_dose_triplicate_2_QT_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_QT_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_3_QT_msec = row["Pre dose triplicate 3, QT (msec)"]
                        Pre_dose_triplicate_3_QT_msec_pure = Pre_dose_triplicate_3_QT_msec.split('|')[0]
                        Pre_dose_triplicate_3_QT_msec_form_field_instance = Pre_dose_triplicate_3_QT_msec.split('|')[1]
                        Pre_dose_triplicate_3_QT_msec_disname = Pre_dose_triplicate_3_QT_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_QT_msec_pure = math.nan
                        Pre_dose_triplicate_3_QT_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_QT_msec_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_RR_msec = row['30-min post dose, RR (msec)']
                        min_30_post_dose_RR_msec_pure = min_30_post_dose_RR_msec.split('|')[0]
                        min_30_post_dose_RR_msec_form_field_instance = min_30_post_dose_RR_msec.split('|')[1]
                        min_30_post_dose_RR_msec_disname = min_30_post_dose_RR_msec.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_RR_msec_pure = math.nan
                        min_30_post_dose_RR_msec_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_RR_msec_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_QT_msec = row["15-min post dose, QT (msec)"]
                        min_15_post_dose_QT_msec_pure = min_15_post_dose_QT_msec.split('|')[0]
                        min_15_post_dose_QT_msec_form_field_instance = min_15_post_dose_QT_msec.split('|')[1]
                        min_15_post_dose_QT_msec_disname = min_15_post_dose_QT_msec.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_QT_msec_pure = math.nan
                        min_15_post_dose_QT_msec_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_QT_msec_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_RR_msec = row['60-min post dose, RR (msec)']
                        min_60_post_dose_RR_msec_pure = min_60_post_dose_RR_msec.split('|')[0]
                        min_60_post_dose_RR_msec_form_field_instance = min_60_post_dose_RR_msec.split('|')[1]
                        min_60_post_dose_RR_msec_disname = min_60_post_dose_RR_msec.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_RR_msec_pure = math.nan
                        min_60_post_dose_RR_msec_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_RR_msec_disname = 'Empty'
                    
                    try: 
                        Pre_dose_triplicate_3_RR_msec = row['Pre dose triplicate 3, RR (msec)']
                        Pre_dose_triplicate_3_RR_msec_pure = Pre_dose_triplicate_3_RR_msec.split('|')[0]
                        Pre_dose_triplicate_3_RR_msec_form_field_instance = Pre_dose_triplicate_3_RR_msec.split('|')[1]
                        Pre_dose_triplicate_3_RR_msec_disname = Pre_dose_triplicate_3_RR_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_RR_msec_pure = math.nan
                        Pre_dose_triplicate_3_RR_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_RR_msec_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_QT_msec = row["30-min post dose, QT (msec)"]
                        min_30_post_dose_QT_msec_pure = min_30_post_dose_QT_msec.split('|')[0]
                        min_30_post_dose_QT_msec_form_field_instance = min_30_post_dose_QT_msec.split('|')[1]
                        min_30_post_dose_QT_msec_disname = min_30_post_dose_QT_msec.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_QT_msec_pure = math.nan
                        min_30_post_dose_QT_msec_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_QT_msec_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_QT_msec = row["60-min post dose, QT (msec)"]
                        min_60_post_dose_QT_msec_pure = min_60_post_dose_QT_msec.split('|')[0]
                        min_60_post_dose_QT_msec_form_field_instance = min_60_post_dose_QT_msec.split('|')[1]
                        min_60_post_dose_QT_msec_disname = min_60_post_dose_QT_msec.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_QT_msec_pure = math.nan
                        min_60_post_dose_QT_msec_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_QT_msec_disname = 'Empty'
                    
                    try:
                        Undefined_QTcF_msec = row["Undefined, QTcF (msec)"]
                        Undefined_QTcF_msec_pure = Undefined_QTcF_msec.split('|')[0]
                        Undefined_QTcF_msec_form_field_instance = Undefined_QTcF_msec.split('|')[1]
                        Undefined_QTcF_msec_disname = Undefined_QTcF_msec.split('|')[0]
                    except Exception as e:
                        Undefined_QTcF_msec_pure = math.nan
                        Undefined_QTcF_msec_form_field_instance = 'This field does not have any data'
                        Undefined_QTcF_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_1_QTcF_msec = row["Pre dose triplicate 1, QTcF (msec)"]
                        Pre_dose_triplicate_1_QTcF_msec_pure = Pre_dose_triplicate_1_QTcF_msec.split('|')[0]
                        Pre_dose_triplicate_1_QTcF_msec_form_field_instance = Pre_dose_triplicate_1_QTcF_msec.split('|')[1]
                        Pre_dose_triplicate_1_QTcF_msec_disname = Pre_dose_triplicate_1_QTcF_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_1_QTcF_msec_pure = math.nan
                        Pre_dose_triplicate_1_QTcF_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_1_QTcF_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_QTcF_msec = row["Pre dose triplicate 2, QTcF (msec)"]
                        Pre_dose_triplicate_2_QTcF_msec_pure = Pre_dose_triplicate_2_QTcF_msec.split('|')[0]
                        Pre_dose_triplicate_2_QTcF_msec_form_field_instance = Pre_dose_triplicate_2_QTcF_msec.split('|')[1]
                        Pre_dose_triplicate_2_QTcF_msec_disname = Pre_dose_triplicate_2_QTcF_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_QTcF_msec_pure = math.nan
                        Pre_dose_triplicate_2_QTcF_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_QTcF_msec_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_3_QTcF_msec = row["Pre dose triplicate 3, QTcF (msec)"]
                        Pre_dose_triplicate_3_QTcF_msec_pure = Pre_dose_triplicate_3_QTcF_msec.split('|')[0]
                        Pre_dose_triplicate_3_QTcF_msec_form_field_instance = Pre_dose_triplicate_3_QTcF_msec.split('|')[1]
                        Pre_dose_triplicate_3_QTcF_msec_disname = Pre_dose_triplicate_3_QTcF_msec.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_QTcF_msec_pure = math.nan
                        Pre_dose_triplicate_3_QTcF_msec_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_QTcF_msec_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_QTcF_msec = row["15-min post dose, QTcF (msec)"]
                        min_15_post_dose_QTcF_msec_pure = min_15_post_dose_QTcF_msec.split('|')[0]
                        min_15_post_dose_QTcF_msec_form_field_instance = min_15_post_dose_QTcF_msec.split('|')[1]
                        min_15_post_dose_QTcF_msec_disname = min_15_post_dose_QTcF_msec.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_QTcF_msec_pure = math.nan
                        min_15_post_dose_QTcF_msec_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_QTcF_msec_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_QTcF_msec = row["30-min post dose, QTcF (msec)"]
                        min_30_post_dose_QTcF_msec_pure = min_30_post_dose_QTcF_msec.split('|')[0]
                        min_30_post_dose_QTcF_msec_form_field_instance = min_30_post_dose_QTcF_msec.split('|')[1]
                        min_30_post_dose_QTcF_msec_disname = min_30_post_dose_QTcF_msec.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_QTcF_msec_pure = math.nan
                        min_30_post_dose_QTcF_msec_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_QTcF_msec_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_QTcF_msec = row["60-min post dose, QTcF (msec)"]
                        min_60_post_dose_QTcF_msec_pure = min_60_post_dose_QTcF_msec.split('|')[0]
                        min_60_post_dose_QTcF_msec_form_field_instance = min_60_post_dose_QTcF_msec.split('|')[1]
                        min_60_post_dose_QTcF_msec_disname = min_60_post_dose_QTcF_msec.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_QTcF_msec_pure = math.nan
                        min_60_post_dose_QTcF_msec_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_QTcF_msec_disname = 'Empty'
                    
                    try:
                        Undefined_Interpretation = row["Undefined, Interpretation"]
                        Undefined_Interpretation_pure = Undefined_Interpretation.split('|')[0]
                        Undefined_Interpretation_form_field_instance = Undefined_Interpretation.split('|')[1]
                        Undefined_Interpretation_disname = Undefined_Interpretation.split('|')[0]
                    except Exception as e:
                        Undefined_Interpretation_pure = math.nan
                        Undefined_Interpretation_form_field_instance = 'This field does not have any data'
                        Undefined_Interpretation_disname = 'Empty'
                    
                    # try:
                    #     Pre_dose_triplicate_1_Interpretation = row["Pre dose triplicate 1, Interpretation"]
                    #     Pre_dose_triplicate_1_Interpretation_pure = Pre_dose_triplicate_1_Interpretation.split('|')[0]
                    #     Pre_dose_triplicate_1_Interpretation_form_field_instance = Pre_dose_triplicate_1_Interpretation.split('|')[1]
                    #     Pre_dose_triplicate_1_Interpretation_disname = Pre_dose_triplicate_1_Interpretation.split('|')[0]
                    # except Exception as e:
                    #     Pre_dose_triplicate_1_Interpretation_pure = math.nan
                    #     Pre_dose_triplicate_1_Interpretation_form_field_instance = 'This field does not have any data'
                    #     Pre_dose_triplicate_1_Interpretation_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_2_Interpretation = row["Pre dose triplicate 2, Interpretation"]
                        Pre_dose_triplicate_2_Interpretation_pure = Pre_dose_triplicate_2_Interpretation.split('|')[0]
                        Pre_dose_triplicate_2_Interpretation_form_field_instance = Pre_dose_triplicate_2_Interpretation.split('|')[1]
                        Pre_dose_triplicate_2_Interpretation_disname = Pre_dose_triplicate_2_Interpretation.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_2_Interpretation_pure = math.nan
                        Pre_dose_triplicate_2_Interpretation_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_2_Interpretation_disname = 'Empty'
                    
                    try:
                        Pre_dose_triplicate_3_Interpretation = row["Pre dose triplicate 3, Interpretation"]
                        Pre_dose_triplicate_3_Interpretation_pure = Pre_dose_triplicate_3_Interpretation.split('|')[0]
                        Pre_dose_triplicate_3_Interpretation_form_field_instance = Pre_dose_triplicate_3_Interpretation.split('|')[1]
                        Pre_dose_triplicate_3_Interpretation_disname = Pre_dose_triplicate_3_Interpretation.split('|')[0]
                    except Exception as e:
                        Pre_dose_triplicate_3_Interpretation_pure = math.nan
                        Pre_dose_triplicate_3_Interpretation_form_field_instance = 'This field does not have any data'
                        Pre_dose_triplicate_3_Interpretation_disname = 'Empty'
                    
                    try:
                        min_15_post_dose_Interpretation = row["15-min post dose, Interpretation"]
                        min_15_post_dose_Interpretation_pure = min_15_post_dose_Interpretation.split('|')[0]
                        min_15_post_dose_Interpretation_form_field_instance = min_15_post_dose_Interpretation.split('|')[1]
                        min_15_post_dose_Interpretation_disname = min_15_post_dose_Interpretation.split('|')[0]
                    except Exception as e:
                        min_15_post_dose_Interpretation_pure = math.nan
                        min_15_post_dose_Interpretation_form_field_instance = 'This field does not have any data'
                        min_15_post_dose_Interpretation_disname = 'Empty'
                    
                    try:
                        min_30_post_dose_Interpretation = row["30-min post dose, Interpretation"]
                        min_30_post_dose_Interpretation_pure = min_30_post_dose_Interpretation.split('|')[0]
                        min_30_post_dose_Interpretation_form_field_instance = min_30_post_dose_Interpretation.split('|')[1]
                        min_30_post_dose_Interpretation_disname = min_30_post_dose_Interpretation.split('|')[0]
                    except Exception as e:
                        min_30_post_dose_Interpretation_pure = math.nan
                        min_30_post_dose_Interpretation_form_field_instance = 'This field does not have any data'
                        min_30_post_dose_Interpretation_disname = 'Empty'
                    
                    try:
                        min_60_post_dose_Interpretation = row["60-min post dose, Interpretation"]
                        min_60_post_dose_Interpretation_pure = min_60_post_dose_Interpretation.split('|')[0]
                        min_60_post_dose_Interpretation_form_field_instance = min_60_post_dose_Interpretation.split('|')[1]
                        min_60_post_dose_Interpretation_disname = min_60_post_dose_Interpretation.split('|')[0]
                    except Exception as e:
                        min_60_post_dose_Interpretation_pure = math.nan
                        min_60_post_dose_Interpretation_form_field_instance = 'This field does not have any data'
                        min_60_post_dose_Interpretation_disname = 'Empty'
                    
                    try:
                        undefined_time = row['Undefined, Interpretation', 'Pre dose triplicate 1, Time 24 hrs']
                        undefined_time_pure = undefined_time.split('|')[0]
                        undefined_time_form_field_instance = undefined_time.split('|')[1]
                        undefined_time_disname = undefined_time.split('|')[0]
                        undefined_time_formated = datetime.strptime(undefined_time_pure, time_format)
                    except:
                        undefined_time_pure = math.nan
                        undefined_time_form_field_instance = 'This field does not have any data'
                        undefined_time_formated = ''  
                        undefined_time_disname = 'Empty'
                    
                    try:
                        predose_triplicate_1_time = row['Pre dose triplicate 1, Time 24 hrs']
                        predose_triplicate_1_time_pure = predose_triplicate_1_time.split('|')[0]
                        predose_triplicate_1_time_form_field_instance =  predose_triplicate_1_time.split('|')[1]
                        predose_triplicate_1_time_disname =  predose_triplicate_1_time.split('|')[0]
                        predose_triplicate_1_time_formated = datetime.strptime(predose_triplicate_1_time_pure, time_format)
                    except:
                        predose_triplicate_1_time_pure = math.nan
                        predose_triplicate_1_time_form_field_instance = 'This field does not have any data'
                        predose_triplicate_1_time_formated = ''
                        predose_triplicate_1_time_disname = 'Empty'

                    try:
                        predose_triplicate_2_time = row['Pre dose triplicate 2, Time 24 hrs']
                        predose_triplicate_2_time_pure = predose_triplicate_2_time.split('|')[0]
                        predose_triplicate_2_time_form_field_instance =  predose_triplicate_2_time.split('|')[1]
                        predose_triplicate_2_time_disname =  predose_triplicate_2_time.split('|')[0]
                        predose_triplicate_2_time_formated = datetime.strptime(predose_triplicate_2_time_pure, time_format)
                    except:
                        predose_triplicate_2_time_pure = math.nan
                        predose_triplicate_2_time_form_field_instance = 'This field does not have any data'
                        predose_triplicate_2_time_formated = ''
                        predose_triplicate_2_time_disname = 'Empty'  
                        

                    try:
                        predose_triplicate_3_time = row['Pre dose triplicate 3, Time 24 hrs']
                        predose_triplicate_3_time_pure = predose_triplicate_3_time.split('|')[0]
                        predose_triplicate_3_time_form_field_instance =  predose_triplicate_3_time.split('|')[1]
                        predose_triplicate_3_time_disname =  predose_triplicate_3_time.split('|')[0]
                        predose_triplicate_3_time_formated = datetime.strptime(predose_triplicate_3_time_pure, time_format)
                    except:
                        predose_triplicate_3_time_pure = math.nan
                        predose_triplicate_3_time_form_field_instance = 'This field does not have any data'
                        predose_triplicate_3_time_formated = ''
                        predose_triplicate_3_time_disname = 'Empty'  
                    
                    try:
                        min_15_time  = row['15-min post dose, Time 24 hrs']
                        min_15_time_pure = min_15_time.split('|')[0]
                        min_15_time_form_field_instance = min_15_time.split('|')[1]
                        min_15_time_disname = min_15_time.split('|')[0]
                        min_15_time_time_formated = datetime.strptime(min_15_time_pure, time_format)
                    except:
                        min_15_time_pure = math.nan
                        min_15_time_form_field_instance = 'This field does not have any data'
                        min_15_time_time_formated = ''
                        min_15_time_disname = 'Empty'  

                    try:
                        min_30_time  = row['30-min post dose, Time 24 hrs']
                        min_30_time_pure = min_30_time.split('|')[0]
                        min_30_time_form_field_instance = min_30_time.split('|')[1]
                        min_30_time_disname = min_30_time.split('|')[0]
                        min_30_time_time_formated = datetime.strptime(min_30_time_pure, time_format)
                    except:
                        min_30_time_pure = math.nan
                        min_30_time_form_field_instance = 'This field does not have any data'
                        min_30_time_time_formated = ''
                        min_30_time_disname = 'Empty'  

                    try:
                        min_60_time  = row['60-min post dose, Time 24 hrs']
                        min_60_time_pure = min_60_time.split('|')[0]
                        min_60_time_form_field_instance = min_60_time.split('|')[1]
                        min_60_time_disname = min_60_time.split('|')[0]
                        min_60_time_time_formated = datetime.strptime(min_60_time_pure, time_format)
                    except:
                        min_60_time_pure = math.nan
                        min_60_time_form_field_instance = 'This field does not have any data'
                        min_60_time_time_formated = ''
                        min_60_time_disname = 'Empty'       

                    #----------------------------------------------------------------------------------------------------------------------------

                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_of_egc_pure == '':
                        pass
                    else:                            
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_of_egc_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of ECG Performed', date_of_egc_form_field_instance ,f , date_of_egc_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LE0010
                    try:
                        if  float(were_ECG_performed_pure) == 3.0: 
                            if visita == 'D-1':
                                pass
                            else:
                                error = [subject, visit, 'Was the vital signs assessment performed?', were_ECG_performed_form_field_instance ,\
                                         'The "Not Required" option can only be selected if visit is D-1 and D-1 Date=Screening visit date or normal and done in the previous 10 days', \
                                            were_ECG_performed_disname, 'LE0010']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0010--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    lista_validacion = [
                        'Undefined',
                        'Pre dose triplicate 1',
                        'Pre dose triplicate 2',
                        'Pre dose triplicate 3',
                        '15-min post dose',
                        '30-min post dose',
                        '60-min post dose',
                            ]
                    
                    mi_cuenta = 0

                    for validador_raw in lista_validacion:
                        try:
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan
         
                        if math.isnan(float(validador)) == False:
                            mi_cuenta+=1
                        else:
                            pass

                    # Revision LE0020
                    try:
                        if float(were_ECG_performed_pure) == 1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Were the ECG performed?', were_ECG_performed_form_field_instance,\
                                         'If ECG was performed, not all sections can be "not done"' , were_ECG_performed_disname, 'LE0020']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0020--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision LE0040
                    if date_of_egc_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_of_egc_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of ECG Performed', date_of_egc_form_field_instance ,'The date of ECG must be the same as the visit date', \
                                        f'{date_of_egc_disname} - {date_of_visit}', 'LE0040']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LE0040--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision LE0050
                    if date_of_egc_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_of_egc_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of ECG Performed', date_of_egc_form_field_instance,\
                                        'The date/time of ECG can not be before the informed consent date/time' ,f'{date_of_egc_disname} - {date_inform_consent}', 'LE0050']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision LE0050--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> LE0060
                    if date_of_egc_pure != '':
                        try:
                            if datetime.strptime(str(date_of_egc_pure), '%d-%b-%Y') >= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of ECG Performed', date_of_egc_form_field_instance ,'Date of ECG Performed must be before the End of study/Early withdrawal date. ', date_of_egc_disname, 'LE0060']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0060 --> {e} - Subject: {subject},  Visit: {visit}  ')

                    # ------------------------------------------ All undifined ---------------------------------------------------------

                    # Revision LE0070
                    try: 
                        if float(Undefined_Interpretation_pure) == 1.0:
                            
                            if float(Undefined_HR_bpm_pure) < 45.0 or float(Undefined_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, 'Undefined, HR (bpm)', Undefined_HR_bpm_form_field_instance ,\
                                         'The HR is not within expected range (45 to 90), therefore the Interpretation can not be Normal.', Undefined_HR_bpm_disname, 'LE0070']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0070--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0140
                    try: 
                        if float(Undefined_Interpretation_pure) == 1.0:
                            if float(Undefined_RR_msec_pure) < 654.6 or float(Undefined_RR_msec_pure) > 1141.4 :
                                error = [subject, visit, 'Undefined, RR (msec)', Undefined_RR_msec_form_field_instance ,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', Undefined_RR_msec_disname, 'LE0140']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0140--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0210
                    try: 
                        if float(Undefined_Interpretation_pure) == 1.0:
                            if float(Undefined_PR_msec_pure) < 120.0 or float(Undefined_PR_msec_pure) > 200.0 :
                                error = [subject, visit, 'Undefined, PR (msec)', Undefined_PR_msec_form_field_instance,\
                                         'The PR is not within expected range (120 to 200), therefore the Interpretation can not be Normal.', Undefined_PR_msec_disname, 'LE0210']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0210--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0280
                    try: 
                        if float(Undefined_Interpretation_pure) == 1.0:
                            if float(Undefined_QRS_msec_pure) < 70.0 or float(Undefined_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, 'Undefined, QRS (msec)', Undefined_RR_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', Undefined_QRS_msec_disname, 'LE0280']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0280--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0350
                    try: 
                        if float(Undefined_Interpretation_pure) == 1.0:
                            if float(Undefined_QT_msec_pure) > 500.0 :
                                error = [subject, visit, 'Undefined, QT (msec)', Undefined_QT_msec_form_field_instance ,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', Undefined_QT_msec_disname, 'LE0350']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0350--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0420
                    if math.isnan(float(Undefined_QTcF_msec_pure)) == False: 
                        try: 
                            if float(Undefined_QTcF_msec_pure) > 350.0 and float(Undefined_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, 'Undefined, QTcF (msec)', Undefined_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', Undefined_QTcF_msec_disname, 'LE0420']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0420--> {e} - Subject: {subject},  Visit: {visit} ')  

                    # Revision LE0430
                    try: 
                        if float(Undefined_Interpretation_pure) == 1.0:
                            if float(Undefined_QTcF_msec_pure) > 500.0:
                                error = [subject, visit, 'Undefined, QTcF (msec)',  Undefined_QTcF_msec_form_field_instance ,\
                                         'The value is above 500 msec, therefore the Interpretation should be abnormal / clinically significan not. Please consider reporting an adverse event',\
                                              Undefined_QTcF_msec_disname, 'LE0430']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0430--> {e} - Subject: {subject},  Visit: {visit} ')   
                    

                        lista_revisar =[
                                'Undefined, HR (bpm)',
                                'Undefined, RR (msec)',
                                'Undefined, PR (msec)',
                                'Undefined, QRS (msec)',
                                'Undefined, QT (msec)',
                                'Undefined, QTcF (msec)',
                        ]

                    cuenta_validacion_abnormal_undefined = 0
                    

                    try:
                        if float(Undefined_HR_bpm_pure) > 45.0 and float(Undefined_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_undefined +=1  

                                              
                    except:
                        pass
                    
                    try:
                        if float(Undefined_RR_msec_pure) > 654.6 and float(Undefined_RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_undefined +=1                 
                    except:
                        pass

                        
                    try:
                        if float(Undefined_PR_msec_pure) > 12.0 and float(Undefined_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_undefined +=1                
                    except:
                        pass
                    
  
                    try:
                        if float(Undefined_QRS_msec_pure) > 70.0 and float(Undefined_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_undefined +=1                  
                    except:
                        pass

 
                    try:
                        if float(Undefined_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_undefined +=1                 
                    except:
                        pass 
                    

                    try:
                        if float(Undefined_QTcF_msec_pure) > 350.0 and float(Undefined_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_undefined +=1                 
                    except:
                        pass


                    # Revision LE0620
                    try:
                        if float(Undefined_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_undefined == 0:
                                error = [subject, visit, 'Undefined, interpretation', Undefined_Interpretation_form_field_instance,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', Undefined_Interpretation_disname, 'LE0620']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0620--> {e} - Subject: {subject},  Visit: {visit} ')  

                    # ------------------------------------------ All pre dose ---------------------------------------------------------

                    # Revision LE0080
                    try: 
                        
                        if float(Pre_dose_triplicate_1_Interpretation_pure) == 1.0:
                            
                            if float(Pre_dose_triplicate_1_HR_bpm_pure) < 45.0 or float(Pre_dose_triplicate_1_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, 'Pre dose triplicate 1, HR (bpm)', Pre_dose_triplicate_1_HR_bpm_form_field_instance,\
                                         'If the Interpretation is Normal, the range must be between 45 and 90', Pre_dose_triplicate_1_HR_bpm_disname, 'LE0080']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0080--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0150
                    try: 
                        if float(Pre_dose_triplicate_1_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_1__RR_msec_pure) < 654.6 or float(Pre_dose_triplicate_1__RR_msec_pure) > 1141.4 :
                                error = [subject, visit, 'Pre dose triplicate 1, RR (msec)', Pre_dose_triplicate_1__RR_msec_form_field_instance,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_1__RR_msec_disname, 'LE0150']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0150--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0290
                    try: 
                        if float(Pre_dose_triplicate_1_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_1_QRS_msec_pure) < 70.0 or float(Pre_dose_triplicate_1_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, 'Pre dose triplicate 1, QRS (msec)', Pre_dose_triplicate_1_QRS_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_1_QRS_msec_disname, 'LE0290']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0290--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0360
                    try: 
                        if float(Pre_dose_triplicate_1_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_1_QT_msec_pure) > 500.0 :
                                error = [subject, visit, 'Pre dose triplicate 1, QT (msec)', Pre_dose_triplicate_1_QRS_msec_form_field_instance,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_1_QT_msec_disname, 'LE0360']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0360--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0440
                    if math.isnan(float(Pre_dose_triplicate_1_QTcF_msec_pure)) == False:
                        try: 
                            if float(Pre_dose_triplicate_1_QTcF_msec_pure) > 350.0 and float(Pre_dose_triplicate_1_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose triplicate 1, QTcF (msec)', Pre_dose_triplicate_1_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_1_QTcF_msec_disname, 'LE0440']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0440--> {e} - Subject: {subject},  Visit: {visit} ')  


                    cuenta_validacion_abnormal_predose_1 = 0

                    try:
                        if float(Pre_dose_triplicate_1_HR_bpm_pure) > 45.0 and float(Pre_dose_triplicate_1_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_predose_1 +=1                
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_1__RR_msec_pure) > 654.6 and float(Pre_dose_triplicate_1__RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_predose_1 +=1                      
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_1_PR_msec_pure) > 12.0 and float(Pre_dose_triplicate_1_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_predose_1 +=1                      
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_1_QRS_msec_pure) > 70.0 and float(Pre_dose_triplicate_1_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_predose_1 +=1                      
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_1_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_predose_1 +=1                    
                    except:
                        pass 

                    try:
                        if float(Pre_dose_triplicate_1_QTcF_msec_pure) > 350.0 and float(Pre_dose_triplicate_1_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_predose_1 +=1                  
                    except:
                        pass


                    # Revision LE0630
                    try:
                        if float(Pre_dose_triplicate_1_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_predose_1 == 0:
                                error = [subject, visit, 'Pre dose triplicate 1, interpretation', Pre_dose_triplicate_1_Interpretation_form_field_instance ,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', Pre_dose_triplicate_1_Interpretation_disname, 'LE0630']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0630--> {e} - Subject: {subject},  Visit: {visit} ')  

                    # ------------------------------------------ All pre dose triplicate 2 ---------------------------------------------------------


                    # Revision LE0090
                    try: 
                        
                        if float(Pre_dose_triplicate_2_Interpretation_pure) == 1.0:
                            
                            if float(Pre_dose_triplicate_2_HR_bpm_pure) < 45.0 or float(Pre_dose_triplicate_2_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, 'Pre dose triplicate 2, HR (bpm)', Pre_dose_triplicate_2_HR_bpm_form_field_instance ,\
                                         'If the Interpretation is Normal, the range must be between 45 and 90', Pre_dose_triplicate_2_HR_bpm_disname, 'LE0090']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0090--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0160
                    try: 
                        if float(Pre_dose_triplicate_2_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_2_RR_msec_pure) < 654.6 or float(Pre_dose_triplicate_2_RR_msec_pure) > 1141.4 :
                                error = [subject, visit, 'Pre dose triplicate 2, RR (msec)', Pre_dose_triplicate_2_RR_msec_form_field_instance ,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_2_RR_msec_disname, 'LE0160']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0160--> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision LE0230
                    try: 
                        if float(Pre_dose_triplicate_2_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_2_PR_msec_pure) < 120.0 or float(Pre_dose_triplicate_2_PR_msec_pure) > 200.0 :
                                error = [subject, visit, 'Pre dose triplicate 2, PR (msec)', Pre_dose_triplicate_2_Interpretation_form_field_instance ,\
                                         'The PR is not within expected range (120 to 200), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_2_Interpretation_disname, 'LE0230']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0230--> {e}') 

                    # Revision LE0300
                    try: 
                        if float(Pre_dose_triplicate_2_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_2_QRS_msec_pure) < 70.0 or float(Pre_dose_triplicate_2_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, 'Pre dose triplicate 2, QRS (msec)', Pre_dose_triplicate_2_QRS_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_2_QRS_msec_disname, 'LE0300']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0300--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0370
                    try: 
                        if float(Pre_dose_triplicate_2_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_2_QT_msec_pure) > 500.0 :
                                error = [subject, visit, 'Pre dose triplicate 2, QT (msec)', Pre_dose_triplicate_2_QT_msec_form_field_instance ,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_2_QT_msec_disname, 'LE0370']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0370--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0460
                    if math.isnan(float(Pre_dose_triplicate_2_QTcF_msec_pure)) == False:
                        try: 
                            if float(Pre_dose_triplicate_2_QTcF_msec_pure) > 350.0 and float(Pre_dose_triplicate_2_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose triplicate 2, QTcF (msec)', Pre_dose_triplicate_2_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_2_QTcF_msec_disname, 'LE0460']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0460--> {e} - Subject: {subject},  Visit: {visit} ')  


                    cuenta_validacion_abnormal_predose_2 = 0

                    try:
                        if float(Pre_dose_triplicate_2_HR_bpm_pure) > 45.0 and float(Pre_dose_triplicate_2_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_predose_2 +=1               
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_2_RR_msec_pure) > 654.6 and float(Pre_dose_triplicate_2_RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_predose_2 +=1                          
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_2_PR_msec_pure) > 12.0 and float(Pre_dose_triplicate_2_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_predose_2 +=1                           
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_2_QRS_msec_pure) > 70.0 and float(Pre_dose_triplicate_2_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_predose_2 +=1                          
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_2_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_predose_2 +=1                         
                    except:
                        pass 

                    try:
                        if float(Pre_dose_triplicate_2_QTcF_msec_pure) > 350.0 and float(Pre_dose_triplicate_2_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_predose_2 +=1                           
                    except:
                        pass


                    # Revision LE0640
                    try:
                        if float(Pre_dose_triplicate_2_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_predose_2 == 0:
                                error = [subject, visit, 'Pre dose triplicate 2, interpretation', Pre_dose_triplicate_2_Interpretation_form_field_instance ,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', Pre_dose_triplicate_2_Interpretation_disname, 'LE0640']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0640--> {e} - Subject: {subject},  Visit: {visit} ')  


                    # ------------------------------------------ All pre dose triplicate 3 ---------------------------------------------------------

                    # Revision LE00100
                    try: 
                        
                        if float(Pre_dose_triplicate_3_Interpretation_pure) == 1.0:
                            
                            if float(Pre_dose_triplicate_3_HR_bpm_pure) < 45.0 or float(Pre_dose_triplicate_3_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, 'Pre dose triplicate 3, HR (bpm)', Pre_dose_triplicate_3_HR_bpm_form_field_instance ,\
                                         'If the Interpretation is Normal, the range must be between 45 and 90', Pre_dose_triplicate_3_HR_bpm_disname, 'LE00100']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE00100--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0170
                    try: 
                        if float(Pre_dose_triplicate_3_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_3_RR_msec_pure) < 654.6 or float(Pre_dose_triplicate_3_RR_msec_pure) > 1141.4 :
                                error = [subject, visit, 'Pre dose triplicate 3, RR (msec)', Pre_dose_triplicate_3_RR_msec_form_field_instance ,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_3_RR_msec_disname, 'LE0170']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0170--> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision LE0240
                    try: 
                        if float(Pre_dose_triplicate_3_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_3_PR_msec_pure) < 120.0 or float(Pre_dose_triplicate_3_PR_msec_pure) > 200.0 :
                                error = [subject, visit, 'Pre dose triplicate 3, PR (msec)', Pre_dose_triplicate_3_PR_msec_form_field_instance,\
                                         'The PR is not within expected range (120 to 200), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_3_PR_msec_disname, 'LE0240']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0240--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0310
                    try: 
                        if float(Pre_dose_triplicate_3_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_3_QRS_msec_pure) < 70.0 or float(Pre_dose_triplicate_3_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, 'Pre dose triplicate 3, QRS (msec)', Pre_dose_triplicate_3_QRS_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_3_QRS_msec_disname, 'LE0310']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0310--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0380
                    try: 
                        if float(Pre_dose_triplicate_3_Interpretation_pure) == 1.0:
                            if float(Pre_dose_triplicate_3_QT_msec_pure) > 500.0 :
                                error = [subject, visit, 'Pre dose triplicate 3, QT (msec)', Pre_dose_triplicate_3_QT_msec_form_field_instance ,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_3_QT_msec_disname, 'LE0380']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0380--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0480
                    if math.isnan(float(Pre_dose_triplicate_3_QTcF_msec_pure)) == False:
                        try: 
                            if float(Pre_dose_triplicate_3_QTcF_msec_pure) > 350.0 and float(Pre_dose_triplicate_3_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, 'Pre dose triplicate 3, QTcF (msec)', Pre_dose_triplicate_3_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', Pre_dose_triplicate_3_QTcF_msec_disname, 'LE0480']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0480--> {e} - Subject: {subject},  Visit: {visit} ')  


                    cuenta_validacion_abnormal_predose_3 = 0

                    try:
                        if float(Pre_dose_triplicate_3_HR_bpm_pure) > 45.0 and float(Pre_dose_triplicate_3_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_predose_3 +=1                 
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_3_RR_msec_pure) > 654.6 and float(Pre_dose_triplicate_3_RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_predose_3 +=1                    
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_3_PR_msec_pure) > 12.0 and float(Pre_dose_triplicate_3_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_predose_3 +=1                     
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_3_QRS_msec_pure) > 70.0 and float(Pre_dose_triplicate_3_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_predose_3 +=1                    
                    except:
                        pass

                    try:
                        if float(Pre_dose_triplicate_3_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_predose_3 +=1                      
                    except:
                        pass 

                    try:
                        if float(Pre_dose_triplicate_3_QTcF_msec_pure) > 350.0 and float(Pre_dose_triplicate_3_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_predose_3 +=1                      
                    except:
                        pass


                    # Revision LE0650
                    try:
                        if float(Pre_dose_triplicate_3_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_predose_3 == 0:
                                error = [subject, visit, 'Pre dose triplicate 3, interpretation', Pre_dose_triplicate_3_Interpretation_form_field_instance ,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', Pre_dose_triplicate_3_Interpretation_disname, 'LE0650']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0650--> {e} - Subject: {subject},  Visit: {visit} ')  


                    # ------------------------------------------ All 15 min post dose ---------------------------------------------------------

                    # Revision LE00110
                    try: 
                        
                        if float(min_15_post_dose_Interpretation_pure) == 1.0:
                            
                            if float(min_15_post_dose_HR_bpm_pure) < 45.0 or float(min_15_post_dose_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, '15-min post dose, HR (bpm)', min_15_post_dose_HR_bpm_form_field_instance ,\
                                         'If the Interpretation is Normal, the range must be between 45 and 90', min_15_post_dose_HR_bpm_disname, 'LE00110']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE00110--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0180
                    try: 
                        if float(min_15_post_dose_Interpretation_pure) == 1.0:
                            if float(min_15_post_dose_RR_msec_pure) < 654.6 or float(min_15_post_dose_RR_msec_pure) > 1141.4 :
                                error = [subject, visit, '15-min post dose, RR (msec)', min_15_post_dose_RR_msec_form_field_instance ,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', min_15_post_dose_RR_msec_disname, 'LE0180']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0180--> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision LE0250
                    try: 
                        if float(min_15_post_dose_Interpretation_pure) == 1.0:
                            if float(min_15_post_dose_PR_msec_pure) < 120.0 or float(min_15_post_dose_PR_msec_pure) > 200.0 :
                                error = [subject, visit, '15-min post dose, PR (msec)', min_15_post_dose_PR_msec_form_field_instance ,\
                                         'The PR is not within expected range (120 to 200), therefore the Interpretation can not be Normal.', min_15_post_dose_PR_msec_disname, 'LE0250']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0250--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0320
                    try: 
                        if float(min_15_post_dose_Interpretation_pure) == 1.0:
                            if float(min_15_post_dose_QRS_msec_pure) < 70.0 or float(min_15_post_dose_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, '15-min post dose, QRS (msec)', min_15_post_dose_QRS_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', min_15_post_dose_QRS_msec_disname, 'LE0320']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0320--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0390
                    try: 
                        if float(min_15_post_dose_Interpretation_pure) == 1.0:
                            if float(min_15_post_dose_QT_msec_pure) > 500.0 :
                                error = [subject, visit, '15-min post dose, QRS (msec)', min_15_post_dose_QT_msec_form_field_instance ,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', min_15_post_dose_QT_msec_disname, 'LE0390']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0390--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0500
                    if math.isnan(float(min_15_post_dose_QTcF_msec_pure)) == False: 
                        try: 
                            if float(min_15_post_dose_QTcF_msec_pure) > 350.0 and float(min_15_post_dose_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, '15-min post dose, QTcF (msec)', min_15_post_dose_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', min_15_post_dose_QTcF_msec_disname, 'LE0500']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0500--> {e} - Subject: {subject},  Visit: {visit} ')  


                    cuenta_validacion_abnormal_15_min = 0

                    try:
                        if float(min_15_post_dose_HR_bpm_pure) > 45.0 and float(min_15_post_dose_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_15_min +=1                   
                    except:
                        pass

                    try:
                        if float(min_15_post_dose_RR_msec_pure) > 654.6 and float(min_15_post_dose_RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_15_min +=1                   
                    except:
                        pass

                    try:
                        if float(min_15_post_dose_PR_msec_pure) > 12.0 and float(min_15_post_dose_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_15_min +=1                    
                    except:
                        pass

                    try:
                        if float(min_15_post_dose_QRS_msec_pure) > 70.0 and float(min_15_post_dose_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_15_min +=1                   
                    except:
                        pass

                    try:
                        if float(min_15_post_dose_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_15_min +=1                   
                    except:
                        pass 

                    try:
                        if float(min_15_post_dose_QTcF_msec_pure) > 350.0 and float(min_15_post_dose_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_15_min +=1                    
                    except:
                        pass


                    # Revision LE0660
                    try:
                        if float(min_15_post_dose_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_15_min == 0:
                                error = [subject, visit, '15-min post dose, interpretation', min_15_post_dose_Interpretation_form_field_instance ,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', min_15_post_dose_Interpretation_disname, 'LE0660']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0660--> {e} - Subject: {subject},  Visit: {visit} ')  


                    # ------------------------------------------ All 30 min post dose ---------------------------------------------------------

                    # Revision LE00120
                    try: 
                        
                        if float(min_30_post_dose_Interpretation_pure) == 1.0:
                            
                            if float(min_30_post_dose_HR_bpm_pure) < 45.0 or float(min_30_post_dose_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, '30-min post dose, HR (bpm)', min_30_post_dose_HR_bpm_form_field_instance,\
                                         'If the Interpretation is Normal, the range must be between 45 and 90', min_30_post_dose_HR_bpm_disname, 'LE00120']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE00120--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0190
                    try: 
                        if float(min_30_post_dose_Interpretation_pure) == 1.0:
                            if float(min_30_post_dose_RR_msec_pure) < 654.6 or float(min_30_post_dose_RR_msec_pure) > 1141.4 :
                                error = [subject, visit, '30-min post dose, RR (msec)', min_30_post_dose_RR_msec_form_field_instance ,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', min_30_post_dose_RR_msec_disname, 'LE0190']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0190--> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision LE0260
                    try: 
                        if float(min_30_post_dose_Interpretation_pure) == 1.0:
                            if float(min_30_post_dose_PR_msec_pure) < 120.0 or float(min_30_post_dose_PR_msec_pure) > 200.0 :
                                error = [subject, visit, '30-min post dose, PR (msec)', min_30_post_dose_PR_msec_form_field_instance,\
                                         'The PR is not within expected range (120 to 200), therefore the Interpretation can not be Normal.', min_30_post_dose_PR_msec_disname, 'LE0260']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0260--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0330
                    try: 
                        if float(min_30_post_dose_Interpretation_pure) == 1.0:
                            if float(min_30_post_dose_QRS_msec_pure) < 70.0 or float(min_30_post_dose_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, '30-min post dose, QRS (msec)', min_30_post_dose_QRS_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', min_30_post_dose_QRS_msec_disname, 'LE0330']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0330--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0400
                    try: 
                        if float(min_30_post_dose_Interpretation_pure) == 1.0:
                            if float(min_30_post_dose_QT_msec_pure) > 500.0 :
                                error = [subject, visit, '30-min post dose, QRS (msec)', min_30_post_dose_QT_msec_form_field_instance ,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', min_30_post_dose_QT_msec_disname, 'LE0400']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision LE0400--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0520
                    if math.isnan(float(min_30_post_dose_QTcF_msec_pure)) == False:
                        try: 
                            if float(min_30_post_dose_QTcF_msec_pure) > 350.0 and float(min_30_post_dose_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, '30-min post dose, QTcF (msec)', min_30_post_dose_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', min_30_post_dose_QTcF_msec_disname, 'LE0520']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0520--> {e} - Subject: {subject},  Visit: {visit} ')  


                    cuenta_validacion_abnormal_30_min = 0

                    try:
                        if float(min_30_post_dose_HR_bpm_pure) > 45.0 and float(min_30_post_dose_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_30_min +=1                   
                    except:
                        pass

                    try:
                        if float(min_30_post_dose_RR_msec_pure) > 654.6 and float(min_30_post_dose_RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_30_min +=1                    
                    except:
                        pass

                    try:
                        if float(min_30_post_dose_PR_msec_pure) > 12.0 and float(min_30_post_dose_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_30_min +=1                    
                    except:
                        pass

                    try:
                        if float(min_30_post_dose_QRS_msec_pure) > 70.0 and float(min_30_post_dose_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_30_min +=1                      
                    except:
                        pass

                    try:
                        if float(min_30_post_dose_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_30_min +=1                 
                    except:
                        pass 

                    try:
                        if float(min_30_post_dose_QTcF_msec_pure) > 350.0 and float(min_30_post_dose_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_30_min +=1                 
                    except:
                        pass


                    # Revision LE0670
                    try:
                        if float(min_30_post_dose_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_30_min == 0:
                                error = [subject, visit, '30-min post dose, interpretation', min_30_post_dose_Interpretation_form_field_instance ,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', min_30_post_dose_Interpretation_disname, 'LE0670']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0670--> {e} - Subject: {subject},  Visit: {visit} ')  



                    # ------------------------------------------ All 60 min post dose ---------------------------------------------------------

                    # Revision LE00130
                    try: 
                        
                        if float(min_60_post_dose_Interpretation_pure) == 1.0:
                            
                            if float(min_60_post_dose_HR_bpm_pure) < 45.0 or float(min_60_post_dose_HR_bpm_pure) > 90.0 :
                                error = [subject, visit, '60-min post dose, HR (bpm)', min_60_post_dose_HR_bpm_form_field_instance,\
                                         'If the Interpretation is Normal, the range must be between 45 and 90', min_60_post_dose_HR_bpm_disname, 'LE00130']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE00130--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0200
                    try: 
                        if float(min_60_post_dose_Interpretation_pure) == 1.0:
                            if float(min_60_post_dose_RR_msec_pure) < 654.6 or float(min_60_post_dose_RR_msec_pure) > 1141.4 :
                                error = [subject, visit, '60-min post dose, RR (msec)', min_60_post_dose_RR_msec_form_field_instance,\
                                         'The RR is not within expected range (654.6 to 1141.4), therefore the Interpretation can not be Normal.', min_60_post_dose_RR_msec_disname, 'LE0200']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0200--> {e} - Subject: {subject},  Visit: {visit} ') 


                    # Revision LE0270
                    try: 
                        if float(min_60_post_dose_Interpretation_pure) == 1.0:
                            if float(min_60_post_dose_PR_msec_pure) < 120.0 or float(min_60_post_dose_PR_msec_pure) > 200.0 :
                                error = [subject, visit, '60-min post dose, PR (msec)', min_60_post_dose_PR_msec_form_field_instance ,\
                                         'The PR is not within expected range (120 to 200), therefore the Interpretation can not be Normal.', min_60_post_dose_PR_msec_disname, 'LE0270']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0270--> {e} - Subject: {subject},  Visit: {visit} ') 

                    # Revision LE0340
                    try: 
                        if float(min_60_post_dose_Interpretation_pure) == 1.0:
                            if float(min_60_post_dose_QRS_msec_pure) < 70.0 or float(min_60_post_dose_QRS_msec_pure) > 120.0 :
                                error = [subject, visit, '60-min post dose, QRS (msec)', min_60_post_dose_QRS_msec_form_field_instance ,\
                                         'The QRS  is not within expected range (70 to 120), therefore the Interpretation can not be Normal.', min_60_post_dose_QRS_msec_disname, 'LE0340']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0340--> {e} - Subject: {subject},  Visit: {visit} ') 
                        
                    # Revision LE0410
                    try: 
                        if float(min_60_post_dose_Interpretation_pure) == 1.0:
                            if float(min_60_post_dose_QT_msec_pure) > 500.0 :
                                error = [subject, visit, '60-min post dose, QRS (msec)', min_60_post_dose_QT_msec_form_field_instance,\
                                         'The QT is not within expected range (below or equal to 500 msec), therefore the Interpretation can not be Normal.', min_60_post_dose_QT_msec_disname, 'LE0410']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0410--> {e} - Subject: {subject},  Visit: {visit} ')     
                    
                    # Revision LE0540
                    if math.isnan(float(min_60_post_dose_QTcF_msec_pure)) == False:
                        try: 
                            if float(min_60_post_dose_QTcF_msec_pure) > 350.0 and float(min_60_post_dose_QTcF_msec_pure) < 450.0 :
                                pass
                            else:
                                error = [subject, visit, '60-min post dose, QTcF (msec)', min_60_post_dose_QTcF_msec_form_field_instance ,\
                                        'The QTcF is not within expected range (350 to 450), therefore the Interpretation can not be Normal.', min_60_post_dose_QTcF_msec_disname, 'LE0540']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0540--> {e} - Subject: {subject},  Visit: {visit} ')  


                    cuenta_validacion_abnormal_60_min = 0

                    try:
                        if float(min_60_post_dose_HR_bpm_pure) > 45.0 and float(min_60_post_dose_HR_bpm_pure) < 90.0 :
                            cuenta_validacion_abnormal_60_min +=1              
                    except:
                        pass

                    try:
                        if float(min_60_post_dose_RR_msec_pure) > 654.6 and float(min_60_post_dose_RR_msec_pure) < 1141.4 :
                            cuenta_validacion_abnormal_60_min +=1                   
                    except:
                        pass

                    try:
                        if float(min_60_post_dose_PR_msec_pure) > 12.0 and float(min_60_post_dose_PR_msec_pure) < 200.0 :
                            cuenta_validacion_abnormal_60_min +=1                    
                    except:
                        pass

                    try:
                        if float(min_60_post_dose_QRS_msec_pure) > 70.0 and float(min_60_post_dose_QRS_msec_pure) < 120.0 :
                            cuenta_validacion_abnormal_60_min +=1                  
                    except:
                        pass

                    try:
                        if float(min_60_post_dose_QT_msec_pure) <= 500.0 :
                            cuenta_validacion_abnormal_60_min +=1                 
                    except:
                        pass 

                    try:
                        if float(min_60_post_dose_QTcF_msec_pure) > 350.0 and float(min_60_post_dose_QTcF_msec_pure) < 450.0 :
                            cuenta_validacion_abnormal_60_min +=1                
                    except:
                        pass


                    # Revision LE0680
                    try:
                        if float(min_60_post_dose_Interpretation_pure) == 2.0:
                            if cuenta_validacion_abnormal_60_min == 0:
                                error = [subject, visit, '60-min post dose, interpretation', min_60_post_dose_Interpretation_form_field_instance,\
                                         'None of the measurements are out of range, the interpretation can not be abnormal', min_60_post_dose_Interpretation_disname, 'LE0680']
                                lista_revision.append(error)

                    except Exception as e:
                        lista_logs.append(f'Revision LE0680--> {e} - Subject: {subject},  Visit: {visit} ')  
                    
                    # -------------------------------------------- Time Revisions ---------------------------------------------------------------------------------------

                    # Revision LE0570
                    if str(predose_triplicate_2_time_formated) != '' and str(predose_triplicate_1_time_formated) != '':
                        try:
                            if float((predose_triplicate_2_time_formated - predose_triplicate_1_time_formated).total_seconds()/60) > 2.0:
                                error = [subject, visit, 'Pre dose triplicate 2, Time 24 hrs', predose_triplicate_2_time_form_field_instance,\
                                            'Pre dose triplicate 2 Time should be within 2 minutes after Pre dose triplicate 1, Time', predose_triplicate_2_time_disname, 'LE0570']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0570--> {e} - Subject: {subject},  Visit: {visit} ')  

                    # Revision LE0580
                    if  str(predose_triplicate_2_time_formated) != '' and str(predose_triplicate_3_time_formated) != '':
                        try:
                            if float((predose_triplicate_3_time_formated - predose_triplicate_2_time_formated).total_seconds()/60) > 2.0:
                                error = [subject, visit, 'Pre dose triplicate 3, Time 24 hrs', predose_triplicate_3_time_form_field_instance,\
                                            'Pre dose triplicate 3 Time should be within 2 minutes after Pre dose triplicate 2, Time', predose_triplicate_3_time_disname, 'LE0580']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision LE0580--> {e} - Subject: {subject},  Visit: {visit} ')  
                    

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    lead_ECG_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("12-Lead ECG")

    for row in dataframe_to_rows(lead_ECG_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return lead_ECG_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    lead_ECG(df_root, path_excel ) 