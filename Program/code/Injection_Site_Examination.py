import numpy as np
import math
import pandas as pd
from datetime import datetime
from revision_fechas import revision_fecha
import warnings
from log_writer import log_writer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

def injection_site_examination(df_root, path_excel_writer):
    '''
    Esta funcion tiene como finalidad la revision de cada uno de los puntos 
    del edit check para el formulario de Injection Site Examination
    '''

    df= df_root[df_root['name']== 'Injection Site Examination']
    lista_sujetos = df['Participante'].unique()
    df = df[['name', 'Visit', 'activityState', 'Participante', 'Estado del Participante', 'Campo', 'Valor', 'FormFieldInstance Id', 'displayName']]
    df['Value_id'] = df['Valor'].astype(str) + '|' + df['FormFieldInstance Id'].astype(str)  + '|' + df['displayName'].astype(str)

    df_visit_date = df_root[df_root['name']=='Date of visit']
    df_visit_date = df_visit_date[['Visit','Participante', 'Campo', 'Valor']]
    df_visit_date = df_visit_date[df_visit_date['Campo']=='Visit Date']
    df_visit_date = df_visit_date[['Visit','Participante','Valor']]
    df_visit_date = df_visit_date.rename(columns={'Participante':'Subject', 'Valor': 'Date_of_visit'})

    df_informed = df_root[df_root['name']=='Informed Consent']
    df_informed = df_informed[['Visit','Participante', 'Campo', 'Valor']]
    df_informed = df_informed[df_informed['Campo']=='Informed consent signature date']
    df_informed = df_informed[['Participante','Valor']]
    df_informed = df_informed.rename(columns={'Participante':'Subject', 'Valor':'Informed_consent_date'})

    df_end_study_general = df_root[df_root['name']== 'End of Study Treatment (Miltefosine)']
    df_end_study_general = df_end_study_general[['Participante', 'Campo', 'Valor', 'Variable' ]]
    df_end_study_general = df_end_study_general[df_end_study_general['Variable'] == 'DSDAT']
    df_end_study_general = df_end_study_general[['Participante', 'Valor']]
    df_end_study_general = df_end_study_general.rename(columns={'Participante':'Subject', 'Valor':'end_study_date'})

    df_visit_done = df_root[df_root['name']=='Date of visit']
    df_visit_done = df_visit_done[['Visit','Participante', 'Campo', 'Valor', 'FormFieldInstance Id']]
    df_visit_done = df_visit_done[df_visit_done['Campo']=='Was the visit performed?']
    df_visit_done['Valor_completo'] = df_visit_done['Valor'].astype(str) + '|' + df_visit_done['FormFieldInstance Id'].astype(str)
    df_visit_done = df_visit_done[['Visit','Participante','Valor_completo']]
    df_visit_done = df_visit_done.rename(columns={'Participante':'Subject', 'Valor_completo':'was_DV_performed'})


    df_time_dosing1 = df_root[df_root['name']=='CpG ODN D35 Administration'].sort_values(by='FormFieldInstance Id')
    df_time_dosing1 = df_time_dosing1[(df_time_dosing1['Campo']=='Date of dosing') | (df_time_dosing1['Campo']=='Time of Dosing')]
    df_time_dosing = df_time_dosing1[df_time_dosing1['Campo']=='Date of dosing']
    df_time_dosing['time_dosing_cpg_administration'] =  df_time_dosing1[df_time_dosing1['FormFieldInstance Id'].isin(df_time_dosing['FormFieldInstance Id'] + 1) & (df_time_dosing1['Campo'] == 'Time of Dosing')]['Valor'].values
    df_time_dosing =df_time_dosing[['Participante','Valor', 'time_dosing_cpg_administration']]
    df_time_dosing = df_time_dosing.rename(columns={'Participante':'Subject', 'Valor':'date_ex_to_join'})


    lista_logs = ['Injection Site Examination']
    lista_revision = []

    # fecha_inicio = datetime.strptime('19-06-2023', "%d-%m-%Y")
    # fecha_fin =  datetime.strptime('31-10-2023', "%d-%m-%Y")

    for sujeto in lista_sujetos:
        sujeto_principal = df[df['Participante']==sujeto]

        lista_validacion_predose = []
        lista_validacion_2_hours = []
        lista_validacion_4_hours = []
        lista_validacion_8_hours = []

        for visita in sujeto_principal.Visit.unique():
            pru_1 = sujeto_principal[sujeto_principal['Visit']==visita]
            pru = pru_1
            pru = pru[['Campo', 'Value_id']].T
            new_columns = pru.iloc[0]
            pru = pru[1:].set_axis(new_columns, axis=1)
            pru['Subject'] = sujeto
            pru['Visit'] = visita

            try:
                pru['date_ex_to_join'] = pru['Date of the Injection site examination'].str.split('|',expand= True)[0]
            except:
                pru['date_ex_to_join'] = 'Nothing to join'

            pru['status'] = pru_1['activityState'].unique()
            pru = pru.merge(df_visit_date, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_informed, on=['Subject'], how='left')
            pru = pru.merge(df_end_study_general, on=['Subject'], how='left')
            pru = pru.merge(df_visit_done, on=['Subject', 'Visit'], how='left')
            pru = pru.merge(df_time_dosing, on=['Subject', 'date_ex_to_join'], how='left')

            for index, row in pru.iterrows():
                status = row['status']
                subject = row['Subject']
                visit = row['Visit']

                date_of_visit = row['Date_of_visit']
                date_inform_consent = row['Informed_consent_date']
                end_study_date = row['end_study_date']
                time_dosing_cpg_administration = row['time_dosing_cpg_administration']

                was_DV_performed = row['was_DV_performed']
                was_DV_performed_pure = was_DV_performed.split('|')[0]
                was_DV_performed_form_field_instance = was_DV_performed.split('|')[1]

   
                if status != '':
                    try:
                        was_injection_performed = row['Was the Injection site examination performed?']
                        was_injection_performed_pure = was_injection_performed.split('|')[0]
                        was_injection_performed_form_field_instance = was_injection_performed.split('|')[1]
                        was_injection_performed_disname = was_injection_performed.split('|')[2]
                    except Exception as e:
                        was_injection_performed_pure = math.nan
                        was_injection_performed_form_field_instance = 'This field does not have any data'
                        was_injection_performed_disname = 'Empty'

                    try:
                        provide_the_reason = row['Provide the reason']
                        provide_the_reason_pure = provide_the_reason.split('|')[0]
                        provide_the_reason_form_field_instance = provide_the_reason.split('|')[1]
                        provide_the_reason_disname = provide_the_reason.split('|')[0]
                    except Exception as e:
                        provide_the_reason_pure = ''
                        provide_the_reason_form_field_instance = 'This field does not have any data'
                        provide_the_reason_disname = 'Empty'

                    try:
                        date_injection = row['Date of the Injection site examination']
                        date_injection_pure = date_injection.split('|')[0]
                        date_injection_form_field_instace = date_injection.split('|')[1]
                        date_injection_disname = date_injection.split('|')[0]
                    except Exception as e:
                        date_injection_pure = ''
                        date_injection_form_field_instace = 'This field does not have any data'
                        date_injection_disname = 'Empty'

                    try:
                        predose_injection_site = row['Predose, Injection site']
                        predose_injection_site_pure = predose_injection_site.split('|')[0]
                        predose_injection_site_form_field_instance = predose_injection_site.split('|')[1]
                        predose_injection_site_disname = predose_injection_site.split('|')[2]
                    except Exception as e:
                        predose_injection_site_pure = math.nan
                        predose_injection_site_form_field_instance = 'This field does not have any data'
                        predose_injection_site_disname = 'Empty'

                    try:
                        post_dose_2_hours = row['2-hours post dose, Injection site']
                        post_dose_2_hours_pure = post_dose_2_hours.split('|')[0]
                        post_dose_2_hours_form_field_instance = post_dose_2_hours.split('|')[1]
                        post_dose_2_hours_disname = post_dose_2_hours.split('|')[0]
                    except Exception as e:
                        post_dose_2_hours_pure = math.nan
                        post_dose_2_hours_form_field_instance = 'This field does not have any data'
                        post_dose_2_hours_disname = 'Empty'
                        
                    try:
                        post_dose_4_hours = row['4-hours post dose, Injection site']
                        post_dose_4_hours_pure = post_dose_4_hours.split('|')[0]
                        post_dose_4_hours_form_field_instance = post_dose_4_hours.split('|')[1]
                        post_dose_4_hours_disname = post_dose_4_hours.split('|')[0]
                    except Exception as e:
                        post_dose_4_hours_pure = math.nan
                        post_dose_4_hours_form_field_instance = 'This field does not have any data'
                        post_dose_4_hours_disname = 'Empty'

                    try:
                        post_dose_8_hours = row['8-hours post dose, Injection site']
                        post_dose_8_hours_pure = post_dose_8_hours.split('|')[0]
                        post_dose_8_hours_form_field_instance = post_dose_8_hours.split('|')[1]
                        post_dose_8_hours_disname = post_dose_8_hours.split('|')[0]
                    except Exception as e:
                        post_dose_8_hours_pure = math.nan
                        post_dose_8_hours_form_field_instance = 'This field does not have any data'
                        post_dose_8_hours_disname = 'Empty'
                    
                    # ----------TIME variables 
                    try:
                        predose_time = row['Predose, Time']
                        predose_time_pure = predose_time.split('|')[0]
                        predose_time_form_field_instance = predose_time.split('|')[1]
                    except:
                        predose_time_pure = math.nan
                        predose_time_form_field_instance = 'This field does not have any data'
                    
                    try:
                        post_dose_2H = row['2-hours post dose, Time']
                        post_dose_2H_pure = post_dose_2H.split('|')[0]
                        post_dose_2H_form_field_instance = post_dose_2H.split('|')[1]
                    except:
                        post_dose_2H_pure = math.nan
                        post_dose_2H_form_field_instance  = 'This field does not have any data'
                    
                    try:
                        post_dose_4H = row['4-hours post dose, Time']
                        post_dose_4H_pure = post_dose_4H.split('|')[0]
                        post_dose_4H_form_field_instance = post_dose_4H.split('|')[1]
                    except:
                        post_dose_4H_pure  = math.nan
                        post_dose_4H_form_field_instance = 'This field does not have any data'
                    
                    try:
                        post_dose_8H = row['8-hours post dose, Time']
                        post_dose_8H_pure = post_dose_8H.split('|')[0]
                        post_dose_8H_form_field_instance = post_dose_8H.split('|')[1]
                    except:
                        post_dose_8H_pure  = math.nan
                        post_dose_8H_form_field_instance = 'This field does not have any data'
 
                    
                    # ---------------------------------------------------------------------------------------------
                    # Revision GE0070
                    if float(was_DV_performed_pure) !=  1.0:
                        error = [subject, visit, 'Visit Pages', was_DV_performed_form_field_instance , 'This Form will be disabled because the visit was not done', was_DV_performed_pure, 'GE0070']
                        lista_revision.append(error)

                    if date_injection_pure == '':
                        pass
                    else:            
                        try:
                            # Primera  revision general de formato de fecha ->GE0020
                            f = revision_fecha(date_injection_pure)
                            if f == None:
                                pass
                            else:
                                error = [subject, visit, 'Date of the Injection site examination', date_injection_form_field_instace,\
                                        f , date_injection_disname, 'GE0020']
                                lista_revision.append(error)     

                        except Exception as e:
                            lista_logs.append(f'Revision GE0020 --> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IS0020
                    if date_injection_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_injection_pure, date_format)
                            date_of_visit_f = datetime.strptime(date_of_visit, date_format)

                            if date_of_test_f != date_of_visit_f:
                                error = [subject, visit, 'Date of the Injection site examination', date_injection_form_field_instace ,\
                                        'the date should be the same as the visit date' , f'{date_injection_disname} - {date_of_visit}', 'IS0020']
                                lista_revision.append(error)
                            else:
                                pass

                        except Exception as e:
                            lista_logs.append(f'Revision IS0020--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IS0030
                    if date_injection_pure != '':
                        try:
                            date_format = '%d-%b-%Y'
                            date_of_test_f = datetime.strptime(date_injection_pure, date_format)
                            date_inform_consent_f = datetime.strptime(date_inform_consent, date_format)

                            if date_of_test_f < date_inform_consent_f:
                                error = [subject, visit, 'Date of the Injection site examination', date_injection_form_field_instace ,\
                                        'The date/time of the Injection site cant be before the informed consent date/time', f'{date_injection_disname} - {date_inform_consent}', 'IS0030']
                                lista_revision.append(error)
                            else:
                                pass
                        except Exception as e:
                            lista_logs.append(f'Revision IS0030--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision -> IS0040
                    if  str(end_study_date) == 'nan' or end_study_date == '' or date_injection_pure == '':
                        pass
                    else:
                        try:
                            if datetime.strptime(str(date_injection_pure), '%d-%b-%Y') <= datetime.strptime(str(end_study_date), '%d-%b-%Y'):
                                pass
                            else: 
                                error = [subject, visit, 'Date of the Injection site examination', date_injection_form_field_instace ,\
                                        'Date of the Injection site examination must be before the End of study/Early withdrawal date. ', date_injection_disname, 'IS0040']
                                lista_revision.append(error)
                        except Exception as e:
                            lista_logs.append(f'Revision IS0040 --> {e} - Subject: {subject},  Visit: {visit}  ')


                    lista_validacion =[ 'Undefined',
                                        'Undefined, Injection site',
                                        'Undefined, Pain grade',
                                        'Undefined, Tenderness grade',
                                        'Undefined, Erythema/Redness grade',
                                        'Undefined, Erythema/Redness diameter (cm)',
                                        'Undefined, Induration/Swelling grade',
                                        'Undefined, Induration/Swelling diameter (cm)',
                                        'Undefined, Warmth',
                                        'Undefined, Lesion photograph taken?',
                                        'Undefined, Provide the reason',
                                        'Predose, Injection site',
                                        'Predose, Pain grade',
                                        'Predose, Tenderness grade',
                                        'Predose, Erythema/Redness grade',
                                        'Predose, Erythema/Redness diameter (cm)',
                                        'Predose, Induration/Swelling grade',
                                        'Predose, Induration/Swelling diameter (cm)',
                                        'Predose, Warmth',
                                        'Predose, Lesion photograph taken?',
                                        'Predose, Provide the reason',
                                        '2-hours post dose, Injection site',
                                        '2-hours post dose, Pain grade',
                                        '2-hours post dose, Tenderness grade',
                                        '2-hours post dose, Erythema/Redness grade',
                                        '2-hours post dose, Erythema/Redness diameter (cm)',
                                        '2-hours post dose, Induration/Swelling grade',
                                        '2-hours post dose, Induration/Swelling diameter (cm)',
                                        '2-hours post dose, Warmth',
                                        '2-hours post dose, Lesion photograph taken?',
                                        '2-hours post dose, Provide the reason',
                                        '4-hours post dose, Injection site',
                                        '4-hours post dose, Pain grade',
                                        '4-hours post dose, Tenderness grade',
                                        '4-hours post dose, Erythema/Redness grade',
                                        '4-hours post dose, Erythema/Redness diameter (cm)',
                                        '4-hours post dose, Induration/Swelling grade',
                                        '4-hours post dose, Induration/Swelling diameter (cm)',
                                        '4-hours post dose, Warmth',
                                        '4-hours post dose, Lesion photograph taken?',
                                        '4-hours post dose, Provide the reason',
                                        '8-hours post dose',
                                        '8-hours post dose, Injection site',
                                        '8-hours post dose, Pain grade',
                                        '8-hours post dose, Tenderness grade',
                                        '8-hours post dose, Erythema/Redness grade',
                                        '8-hours post dose, Erythema/Redness diameter (cm)',
                                        '8-hours post dose, Induration/Swelling grade',
                                        '8-hours post dose, Induration/Swelling diameter (cm)',
                                        '8-hours post dose, Warmth',
                                        '8-hours post dose, Lesion photograph taken?',
                                        '8-hours post dose, Provide the reason']
                    
                    mi_cuenta= 0
                    for validador_raw in lista_validacion:
                        try:
                            validador = row[validador_raw].split('|')[0]
                        except:
                            validador = math.nan

                        
                        if math.isnan(float(validador)) or validador != '-' or validador != np.nan or  str(validador) != 'nan' or float(validador) !=0.0 or str(validador) != '':
                            mi_cuenta+=1
                        else:
                            pass

                    # ----------------------------------------------------------------------
                    # Revision IS0050
                    try:
                        if float(was_injection_performed_pure) ==1.0: 
                            if mi_cuenta != 0:
                                pass
                            else:
                                error = [subject, visit, 'Was the Injection site examination performed?', was_injection_performed_form_field_instance ,\
                                         'If, Was the Injection site examination performed?="Yes" at least one section per time point must be added' , was_injection_performed_disname, 'IS0050']
                                lista_revision.append(error)
                    except Exception as e:
                        lista_logs.append(f'Revision IS0050--> {e} - Subject: {subject},  Visit: {visit} ')
                    
                    # Revision IS0100
                    if math.isnan(float(predose_injection_site_pure)) == False:
                        try:
                            validador_predose = (predose_injection_site_pure, visit)

                            if validador_predose in lista_validacion_predose:
                                    error = [subject, visit, 'Predose, Injection site', predose_injection_site_form_field_instance ,\
                                            'The injection site should no be reported more than once' , predose_injection_site_disname, 'IS0100']
                                    lista_revision.append(error)
                            else:
                                lista_validacion_predose.append(validador_predose)
                        except Exception as e:
                            lista_logs.append(f'Revision IS0100--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IS0110
                    if math.isnan(float(post_dose_2_hours_pure)) != False:
                        try:
                            validador_2_post_dose = (post_dose_2_hours_pure, visit)

                            if validador_2_post_dose in lista_validacion_2_hours:
                                    error = [subject, visit, '2-hours post dose, Injection site', post_dose_2_hours_form_field_instance ,\
                                            'The injection site should no be reported more than once' , post_dose_2_hours_disname, 'IS0110']
                                    lista_revision.append(error)
                            else:
                                lista_validacion_2_hours.append(validador_2_post_dose)
                        except Exception as e:
                            lista_logs.append(f'Revision IS0110--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IS0120
                    if math.isnan(float(post_dose_4_hours_pure)) == False:
                        try:
                            validador_4_post_dose = (post_dose_4_hours_pure, visit)

                            if validador_4_post_dose in lista_validacion_4_hours:
                                    error = [subject, visit, '4-hours post dose, Injection site', post_dose_4_hours_form_field_instance ,\
                                            'The injection site should no be reported more than once' , post_dose_4_hours_disname, 'IS0120']
                                    lista_revision.append(error)
                            else:
                                lista_validacion_4_hours.append(validador_4_post_dose)
                        except Exception as e:
                            lista_logs.append(f'Revision IS0120--> {e} - Subject: {subject},  Visit: {visit} ')

                    # Revision IS0130
                    if math.isnan(float(post_dose_8_hours_pure)) != False:
                        try:
                            validador_8_post_dose = (post_dose_8_hours_pure, visit)

                            if validador_8_post_dose in lista_validacion_8_hours:
                                    error = [subject, visit, '8-hours post dose, Injection site', post_dose_8_hours_form_field_instance ,\
                                            'The injection site should no be reported more than once', post_dose_8_hours_disname, 'IS0130']
                                    lista_revision.append(error)
                            else:
                                lista_validacion_8_hours.append(validador_8_post_dose)
                        except Exception as e:
                            lista_logs.append(f'Revision IS0130--> {e} - Subject: {subject},  Visit: {visit} ')


                    # Revision IS0060
                    if str(time_dosing_cpg_administration) != 'nan':
                            
                        try:
                            dif = float((datetime.strptime(time_dosing_cpg_administration, '%H:%M') - datetime.strptime(predose_time_pure, '%H:%M')).total_seconds() / 60)
                            if dif < 0.0  or  dif > 60.0:
                                    
                                error = [subject, visit, 'Pre dose, Time', predose_time_form_field_instance,\
                                             'The time selected should be less than 60 min before the study treatment administration', \
                                                f'Pre dose, Time: {predose_time_pure} - dose time administration{time_dosing_cpg_administration}', 'IS0060']
                                lista_revision.append(error)

                        except Exception as e:
                            lista_logs.append(f'Revision IS0060 --> {e} - Subject: {subject},  Visit: {visit} ')  


                        # Revision IS0070
                        if str(time_dosing_cpg_administration) != 'nan':
                            
                            try:
                                dif_2H = float((datetime.strptime(post_dose_2H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_2H > 135.0 or dif_2H < 105.0:
                                    
                                    error = [subject, visit, '2-hours post dose, Time', post_dose_2H_form_field_instance,\
                                             'The time selected should be less than 2h15 and greater than 1h45 after the study treatment administration', \
                                                f'2-hours post dose, Time: {post_dose_2H_pure} - dose time administration{time_dosing_cpg_administration}', 'IS0070']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision IS0070 --> {e} - Subject: {subject},  Visit: {visit} ')  


                        # Revision IS0080
                        if str(time_dosing_cpg_administration) != 'nan':
                            
                            try:
                                dif_4H = float((datetime.strptime(post_dose_4H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_4H > 255.0 or dif_4H < 225.0:
                                    
                                    error = [subject, visit, '4-hours post dose, Time', post_dose_4H_form_field_instance,\
                                             'The time selected should be less than 4h15 and greater than 3h45 after the study treatment administration', \
                                                f'4-hours post dose, Time: {post_dose_4H_pure} - dose time administration{time_dosing_cpg_administration}', 'IS0080']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision IS0080 --> {e} - Subject: {subject},  Visit: {visit} ')


                        # Revision IS0090
                        if str(time_dosing_cpg_administration) != 'nan':
                            
                            try:
                                dif_8H = float((datetime.strptime(post_dose_8H_pure, '%H:%M') - datetime.strptime(time_dosing_cpg_administration, '%H:%M')).total_seconds() / 60)
                                if dif_8H > 495.0 or dif_8H < 465.0:
                                    
                                    error = [subject, visit, '8-hours post dose, Time', post_dose_8H_form_field_instance,\
                                             'The time selected should be less than 8h15 and greater than 7h45 after the study treatment administration', \
                                                f'8-hours post dose, Time: {post_dose_8H_pure} - dose time administration{time_dosing_cpg_administration}', 'IS0090']
                                    lista_revision.append(error)

                            except Exception as e:
                                lista_logs.append(f'Revision IS0090 --> {e} - Subject: {subject},  Visit: {visit} ')

    excel_writer = load_workbook(path_excel_writer)
    column_names = ['Subject', 'Visit', 'Field', 'Form Field Instance ID' ,'Standard Error Message', 'Value', 'Check Number']
    injection_site_examination_output = pd.DataFrame(lista_revision, columns=column_names)
    
    sheet = excel_writer.create_sheet("Injection Site Examination")

    for row in dataframe_to_rows(injection_site_examination_output, index=False, header=True):
        sheet.append(row)

    excel_writer.save(path_excel_writer)
    log_writer(lista_logs)

    return injection_site_examination_output[['Form Field Instance ID' ,'Standard Error Message']].replace({',': '', ';': ''}, regex=True)

if __name__ == '__main__':
    path_excel = r"C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\Program\output\prueba.xlsx"
    df_root = pd.read_excel(r'C:\Users\sebastian sossa\Documents\integraIT\projects_integrait\DNDI\data\newDNDI_v2.xlsx')
    injection_site_examination(df_root, path_excel )